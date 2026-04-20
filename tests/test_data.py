import base64
import json

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
import excel_mcp.server as server_module

from excel_mcp.data import (
    append_table_rows,
    describe_dataset as describe_dataset_impl,
    quick_read as quick_read_impl,
    read_as_table,
    read_excel_range,
    read_excel_range_with_metadata,
    search_cells,
    suggest_read_strategy as suggest_read_strategy_impl,
    update_rows_by_key,
    write_data,
)
from excel_mcp.exceptions import DataError, PreconditionFailedError
from excel_mcp.server import (
    append_table_rows as append_table_rows_tool,
    describe_dataset as describe_dataset_tool,
    list_all_sheets,
    quick_read,
    read_data_from_excel,
    read_excel_as_table as read_excel_as_table_tool,
    search_in_sheet,
    suggest_read_strategy as suggest_read_strategy_tool,
    update_rows_by_key as update_rows_by_key_tool,
)


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    assert "operation" in payload
    assert "message" in payload
    return payload


def _create_chartsheet_first_workbook(tmp_path) -> str:
    filepath = str(tmp_path / "chartsheet-first.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Name", "Value"), ("Alice", 30), ("Bob", 25)]:
        ws.append(row)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb._sheets = [chart_sheet, ws]
    wb.active = 0
    wb.save(filepath)
    wb.close()
    return filepath


def test_read_from_explicit_start_cell(tmp_workbook):
    """Bug #1: read_excel_range should respect explicit start_cell when no end_cell given."""
    result = read_excel_range(tmp_workbook, "Sheet1", start_cell="B2")
    flat = [cell for row in result for cell in row]
    assert "Name" not in flat, "start_cell was ignored — returned data from A1"
    assert result[0][0] == 30


def test_read_from_a1_default_uses_full_range(tmp_workbook):
    """Default A1 start should read the full data range."""
    result = read_excel_range(tmp_workbook, "Sheet1")
    assert result[0][0] == "Name"
    assert len(result) == 6


def test_read_with_metadata_respects_start_cell(tmp_workbook):
    """Bug #1 also affects read_excel_range_with_metadata."""
    result = read_excel_range_with_metadata(tmp_workbook, "Sheet1", start_cell="C4")
    cells = result["cells"]
    values = [c["value"] for c in cells]
    assert "Name" not in values, "start_cell was ignored in metadata read"
    assert cells[0]["value"] == "Turku"


def test_read_as_table_returns_headers_and_rows(tmp_workbook):
    result = read_as_table(tmp_workbook, "Sheet1")
    assert result["headers"] == ["Name", "Age", "City"]
    assert len(result["rows"]) == 5
    assert result["rows"][0] == ["Alice", 30, "Helsinki"]
    assert result["rows"][4] == ["Eve", 32, "Espoo"]


def test_read_as_table_with_max_rows(tmp_workbook):
    result = read_as_table(tmp_workbook, "Sheet1", max_rows=2)
    assert len(result["rows"]) == 2
    assert result["total_rows"] == 5
    assert result["truncated"] is True
    assert result["next_start_row"] == 4


def test_read_as_table_supports_start_row_pagination(tmp_workbook):
    result = read_as_table(tmp_workbook, "Sheet1", start_row=4, max_rows=2)

    assert result["headers"] == ["Name", "Age", "City"]
    assert result["rows"] == [
        ["Carol", 35, "Turku"],
        ["Dave", 28, "Oulu"],
    ]
    assert result["total_rows"] == 5
    assert result["truncated"] is True
    assert result["next_start_row"] == 6


def test_read_as_table_can_omit_headers_for_followup_pages(tmp_workbook):
    result = read_as_table(
        tmp_workbook,
        "Sheet1",
        start_row=4,
        max_rows=2,
        include_headers=False,
    )

    assert "headers" not in result
    assert result["rows"] == [
        ["Carol", 35, "Turku"],
        ["Dave", 28, "Oulu"],
    ]
    assert result["truncated"] is True
    assert result["next_start_row"] == 6


def test_read_as_table_omits_next_start_row_on_final_page(tmp_workbook):
    result = read_as_table(tmp_workbook, "Sheet1", start_row=6, max_rows=2)

    assert result["rows"] == [["Eve", 32, "Espoo"]]
    assert result["truncated"] is False
    assert "next_start_row" not in result


def test_read_as_table_rejects_start_row_at_or_above_header(tmp_workbook):
    with pytest.raises(DataError, match="start_row must be greater than header_row"):
        read_as_table(tmp_workbook, "Sheet1", header_row=1, start_row=1)


def test_read_as_table_rejects_non_positive_max_rows(tmp_workbook):
    with pytest.raises(DataError, match="max_rows must be a positive integer"):
        read_as_table(tmp_workbook, "Sheet1", max_rows=0)


def test_read_as_table_ignores_trailing_rows_outside_selected_columns(tmp_path):
    filepath = tmp_path / "selected-columns.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [
        ("Name", "Age", "City"),
        ("Alice", 30, "Helsinki"),
        ("Bob", 25, "Tampere"),
    ]:
        ws.append(row)
    ws["D20"] = "Ignored outside selected columns"
    wb.save(filepath)
    wb.close()

    result = read_as_table(str(filepath), "Sheet1", start_col="A", end_col="C")

    assert result["rows"] == [
        ["Alice", 30, "Helsinki"],
        ["Bob", 25, "Tampere"],
    ]
    assert result["total_rows"] == 2
    assert result["truncated"] is False


def test_read_as_table_ignores_sparse_trailing_rows_after_large_blank_gap(tmp_path):
    filepath = tmp_path / "sparse-footer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [
        ("Name", "Age", "City"),
        ("Alice", 30, "Helsinki"),
        ("Bob", 25, "Tampere"),
    ]:
        ws.append(row)
    ws["C20"] = "Footer note"
    wb.save(filepath)
    wb.close()

    result = read_as_table(str(filepath), "Sheet1")

    assert result["rows"] == [
        ["Alice", 30, "Helsinki"],
        ["Bob", 25, "Tampere"],
    ]
    assert result["total_rows"] == 2
    assert result["truncated"] is False


def test_read_as_table_custom_header_row(tmp_workbook):
    result = read_as_table(tmp_workbook, "Sheet1", header_row=2)
    assert result["headers"] == ["Alice", 30, "Helsinki"]


def test_read_as_table_can_return_records_and_schema(tmp_workbook):
    result = read_as_table(tmp_workbook, "Sheet1", row_mode="objects", infer_schema=True)

    assert result["records"][0] == {
        "name": "Alice",
        "age": 30,
        "city": "Helsinki",
    }
    assert result["schema"] == [
        {"field": "name", "header": "Name", "type": "string", "nullable": False},
        {"field": "age", "header": "Age", "type": "integer", "nullable": False},
        {"field": "city", "header": "City", "type": "string", "nullable": False},
    ]
    assert result["row_mode"] == "objects"
    assert "rows" not in result


def test_read_as_table_marks_formula_columns_as_formula_schema(tmp_path):
    filepath = tmp_path / "formula-schema.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Qty", "Price", "Total"])
    ws.append(["A", 2, 5, "=B2*C2"])
    ws.append(["B", 3, 7, "=B3*C3"])
    wb.save(filepath)
    wb.close()

    result = read_as_table(str(filepath), "Sheet1", row_mode="objects", infer_schema=True)

    assert result["records"][0]["total"] == "=B2*C2"
    assert result["schema"][3] == {
        "field": "total",
        "header": "Total",
        "type": "formula",
        "nullable": False,
    }


def test_read_as_table_schema_normalizes_and_dedupes_headers(tmp_path):
    from openpyxl import Workbook

    filepath = tmp_path / "schema.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "First Name"
    ws["B1"] = None
    ws["C1"] = "First Name"
    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["C2"] = "Admin"
    wb.save(filepath)
    wb.close()

    result = read_as_table(str(filepath), "Sheet1", row_mode="objects", infer_schema=True)

    assert result["schema"] == [
        {"field": "first_name", "header": "First Name", "type": "string", "nullable": False},
        {"field": "column_2", "header": None, "type": "integer", "nullable": False},
        {"field": "first_name_2", "header": "First Name", "type": "string", "nullable": False},
    ]
    assert result["records"][0] == {
        "first_name": "Alice",
        "column_2": 30,
        "first_name_2": "Admin",
    }


def test_read_as_table_schema_transliterates_accented_headers(tmp_path):
    filepath = tmp_path / "schema-fi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Näyttökerrat"
    ws["B1"] = "Lisäklikit"
    ws["C1"] = "CTR %"
    ws["A2"] = 1200
    ws["B2"] = 45
    ws["C2"] = 3.75
    wb.save(filepath)
    wb.close()

    result = read_as_table(str(filepath), "Sheet1", row_mode="objects", infer_schema=True)

    assert result["schema"] == [
        {"field": "nayttokerrat", "header": "Näyttökerrat", "type": "integer", "nullable": False},
        {"field": "lisaklikit", "header": "Lisäklikit", "type": "integer", "nullable": False},
        {"field": "ctr", "header": "CTR %", "type": "number", "nullable": False},
    ]
    assert result["records"][0] == {
        "nayttokerrat": 1200,
        "lisaklikit": 45,
        "ctr": 3.75,
    }


def test_read_as_table_rejects_chart_sheet_with_clear_error(tmp_path):
    filepath = _create_chartsheet_first_workbook(tmp_path)

    with pytest.raises(DataError, match="Sheet 'Charts' is a chartsheet"):
        read_as_table(filepath, "Charts")


def test_read_excel_range_rejects_chart_sheet_with_clear_error(tmp_path):
    filepath = _create_chartsheet_first_workbook(tmp_path)

    with pytest.raises(DataError, match="Sheet 'Charts' is a chartsheet"):
        read_excel_range(filepath, "Charts", start_cell="A1", end_cell="B2")


def test_read_as_table_rejects_boolean_max_rows(tmp_workbook):
    with pytest.raises(DataError, match="max_rows must be a positive integer"):
        read_as_table(tmp_workbook, "Sheet1", max_rows=True)


def test_read_with_metadata_rejects_boolean_max_rows(tmp_workbook):
    with pytest.raises(DataError, match="max_rows must be a positive integer"):
        read_excel_range_with_metadata(tmp_workbook, "Sheet1", start_cell="A1", end_cell="C5", max_rows=True)


def test_read_with_metadata_rejects_cursor_with_boolean_limits(tmp_workbook):
    payload = {
        "v": 1,
        "start_cell": "A2",
        "end_cell": "A5",
        "max_rows": True,
    }
    encoded = base64.urlsafe_b64encode(
        json.dumps(payload, separators=(",", ":"), sort_keys=True).encode("utf-8")
    ).decode("ascii").rstrip("=")

    with pytest.raises(DataError, match="Invalid cursor"):
        read_excel_range_with_metadata(tmp_workbook, "Sheet1", cursor=encoded)

def test_search_cells_finds_exact_match(tmp_workbook):
    results = search_cells(tmp_workbook, "Sheet1", "Alice")
    assert len(results) == 1
    assert results[0]["cell"] == "A2"
    assert results[0]["value"] == "Alice"


def test_search_cells_finds_partial_match(tmp_workbook):
    results = search_cells(tmp_workbook, "Sheet1", "li", exact=False)
    values = [r["value"] for r in results]
    assert "Alice" in values


def test_search_cells_no_match(tmp_workbook):
    results = search_cells(tmp_workbook, "Sheet1", "Nonexistent")
    assert results == []


def test_search_cells_finds_numbers(tmp_workbook):
    results = search_cells(tmp_workbook, "Sheet1", 30)
    assert len(results) == 1
    assert results[0]["cell"] == "B2"


def test_search_cells_matches_numeric_values_from_string_queries(tmp_workbook):
    results = search_cells(tmp_workbook, "Sheet1", "30")
    assert len(results) == 1
    assert results[0]["cell"] == "B2"


def test_search_in_sheet_accepts_numeric_queries(tmp_workbook):
    payload = _load_tool_payload(search_in_sheet(tmp_workbook, "Sheet1", 30))
    assert payload["operation"] == "search_in_sheet"
    assert len(payload["data"]["matches"]) == 1
    assert payload["data"]["matches"][0]["cell"] == "B2"


def test_read_data_from_excel_preview_only_limits_output(tmp_path):
    filepath = tmp_path / "preview.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Value"
    for row in range(2, 17):
        ws[f"A{row}"] = f"Row {row}"
    wb.save(filepath)
    wb.close()

    payload = _load_tool_payload(read_data_from_excel(str(filepath), "Sheet1", preview_only=True))
    preview_rows = {cell["row"] for cell in payload["data"]["cells"]}

    assert len(preview_rows) == 10
    assert payload["data"]["preview_only"] is True
    assert payload["data"]["truncated"] is True


def test_read_data_from_excel_compact_omits_default_validation(tmp_workbook):
    payload = _load_tool_payload(read_data_from_excel(tmp_workbook, "Sheet1", compact=True))
    first_cell = payload["data"]["cells"][0]

    assert first_cell["address"] == "A1"
    assert "validation" not in first_cell


def test_read_data_from_excel_compact_keeps_real_validation(tmp_path):
    from openpyxl.worksheet.datavalidation import DataValidation

    filepath = tmp_path / "validated.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Status"
    ws["A2"] = "Open"
    validation = DataValidation(type="list", formula1='"Open,Closed"')
    validation.add("A2")
    ws.add_data_validation(validation)
    wb.save(filepath)
    wb.close()

    payload = _load_tool_payload(read_data_from_excel(str(filepath), "Sheet1", compact=True))
    cells_by_address = {cell["address"]: cell for cell in payload["data"]["cells"]}

    assert "validation" in cells_by_address["A2"]
    assert cells_by_address["A2"]["validation"]["has_validation"] is True
    assert cells_by_address["A2"]["validation"]["allowed_values"] == ["Open", "Closed"]


def test_read_excel_range_with_metadata_values_only_returns_2d_values(tmp_workbook):
    result = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        start_cell="A1",
        end_cell="B3",
        values_only=True,
    )

    assert result == {
        "range": "A1:B3",
        "sheet_name": "Sheet1",
        "values": [
            ["Name", "Age"],
            ["Alice", 30],
            ["Bob", 25],
        ],
    }


def test_read_excel_range_with_metadata_supports_row_pagination(tmp_workbook):
    result = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        start_cell="A1",
        end_cell="B6",
        max_rows=2,
    )

    assert result["range"] == "A1:B2"
    assert result["total_rows"] == 6
    assert result["truncated"] is True
    assert result["next_start_row"] == 3
    assert result["next_start_cell"] == "A3"
    assert [cell["address"] for cell in result["cells"]] == ["A1", "B1", "A2", "B2"]


def test_read_excel_range_with_metadata_supports_column_pagination(tmp_workbook):
    result = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        start_cell="A1",
        end_cell="C6",
        max_cols=2,
    )

    assert result["range"] == "A1:B6"
    assert result["total_cols"] == 3
    assert result["truncated"] is True
    assert result["next_start_col"] == "C"
    assert result["next_column_start_cell"] == "C1"
    assert [cell["address"] for cell in result["cells"][:4]] == ["A1", "B1", "A2", "B2"]


def test_read_excel_range_with_metadata_values_only_supports_row_pagination(tmp_workbook):
    result = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        start_cell="A1",
        end_cell="B6",
        max_rows=2,
        values_only=True,
    )

    assert result["range"] == "A1:B2"
    assert result["sheet_name"] == "Sheet1"
    assert result["total_rows"] == 6
    assert result["truncated"] is True
    assert result["next_start_row"] == 3
    assert result["next_start_cell"] == "A3"
    assert result["next_cursor"] == result["continuations"]["down"]["cursor"]
    assert result["values"] == [
        ["Name", "Age"],
        ["Alice", 30],
    ]


def test_read_excel_range_with_metadata_values_only_supports_column_pagination(tmp_workbook):
    result = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        start_cell="A1",
        end_cell="C6",
        max_cols=2,
        values_only=True,
    )

    assert result["range"] == "A1:B6"
    assert result["sheet_name"] == "Sheet1"
    assert result["total_cols"] == 3
    assert result["truncated"] is True
    assert result["next_start_col"] == "C"
    assert result["next_column_start_cell"] == "C1"
    assert result["next_cursor"] == result["continuations"]["right"]["cursor"]
    assert result["values"] == [
        ["Name", "Age"],
        ["Alice", 30],
        ["Bob", 25],
        ["Carol", 35],
        ["Dave", 28],
        ["Eve", 32],
    ]


def test_read_data_from_excel_values_only_preview_limits_rows(tmp_path):
    filepath = tmp_path / "preview-values.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Value"
    for row in range(2, 17):
        ws[f"A{row}"] = f"Row {row}"
    wb.save(filepath)
    wb.close()

    payload = _load_tool_payload(
        read_data_from_excel(str(filepath), "Sheet1", preview_only=True, values_only=True)
    )

    assert "cells" not in payload["data"]
    assert len(payload["data"]["values"]) == 10
    assert payload["data"]["values"][0] == ["Value"]
    assert payload["data"]["preview_only"] is True
    assert payload["data"]["truncated"] is True


def test_read_data_from_excel_supports_row_pagination(tmp_workbook):
    payload = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            start_cell="A1",
            end_cell="B6",
            max_rows=2,
        )
    )

    assert payload["data"]["range"] == "A1:B2"
    assert payload["data"]["total_rows"] == 6
    assert payload["data"]["truncated"] is True
    assert payload["data"]["next_start_row"] == 3
    assert payload["data"]["next_start_cell"] == "A3"
    assert payload["data"]["next_cursor"] == payload["data"]["continuations"]["down"]["cursor"]
    assert [cell["address"] for cell in payload["data"]["cells"]] == ["A1", "B1", "A2", "B2"]


def test_read_data_from_excel_supports_column_pagination(tmp_workbook):
    payload = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            start_cell="A1",
            end_cell="C6",
            max_cols=2,
        )
    )

    assert payload["data"]["range"] == "A1:B6"
    assert payload["data"]["total_cols"] == 3
    assert payload["data"]["truncated"] is True
    assert payload["data"]["next_start_col"] == "C"
    assert payload["data"]["next_column_start_cell"] == "C1"
    assert payload["data"]["next_cursor"] == payload["data"]["continuations"]["right"]["cursor"]
    assert [cell["address"] for cell in payload["data"]["cells"][:4]] == ["A1", "B1", "A2", "B2"]


def test_read_data_from_excel_supports_2d_pagination(tmp_workbook):
    payload = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            start_cell="A1",
            end_cell="C6",
            max_rows=2,
            max_cols=2,
            values_only=True,
        )
    )

    assert payload["data"]["range"] == "A1:B2"
    assert payload["data"]["sheet_name"] == "Sheet1"
    assert payload["data"]["total_rows"] == 6
    assert payload["data"]["total_cols"] == 3
    assert payload["data"]["truncated"] is True
    assert payload["data"]["next_start_row"] == 3
    assert payload["data"]["next_start_cell"] == "A3"
    assert payload["data"]["next_start_col"] == "C"
    assert payload["data"]["next_column_start_cell"] == "C1"
    assert payload["data"]["values"] == [
        ["Name", "Age"],
        ["Alice", 30],
    ]
    assert set(payload["data"]["continuations"]) == {"down", "right"}
    assert "next_cursor" not in payload["data"]


def test_read_excel_range_with_metadata_cursor_resumes_downward_page(tmp_workbook):
    first_page = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        start_cell="A1",
        end_cell="B6",
        max_rows=2,
        values_only=True,
    )

    second_page = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        cursor=first_page["continuations"]["down"]["cursor"],
        values_only=True,
    )

    assert second_page["range"] == "A3:B4"
    assert second_page["sheet_name"] == "Sheet1"
    assert second_page["total_rows"] == 4
    assert second_page["truncated"] is True
    assert second_page["next_start_row"] == 5
    assert second_page["next_start_cell"] == "A5"
    assert second_page["values"] == [
        ["Bob", 25],
        ["Carol", 35],
    ]
    assert set(second_page["continuations"]) == {"down"}
    assert second_page["next_cursor"] == second_page["continuations"]["down"]["cursor"]


def test_read_excel_range_with_metadata_cursor_preserves_values_only_without_repeating_flag(tmp_workbook):
    first_page = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        start_cell="A1",
        end_cell="B6",
        max_rows=2,
        values_only=True,
    )

    second_page = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        cursor=first_page["continuations"]["down"]["cursor"],
    )

    assert "values" in second_page
    assert "cells" not in second_page
    assert second_page["values"] == [
        ["Bob", 25],
        ["Carol", 35],
    ]


def test_read_excel_range_with_metadata_cursor_preserves_compact_without_repeating_flag(tmp_workbook):
    first_page = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        start_cell="A1",
        end_cell="B6",
        max_rows=2,
        compact=True,
    )

    second_page = read_excel_range_with_metadata(
        tmp_workbook,
        "Sheet1",
        cursor=first_page["continuations"]["down"]["cursor"],
    )

    assert "cells" in second_page
    assert all("validation" not in cell for cell in second_page["cells"])


def test_read_data_from_excel_cursor_resumes_rightward_window(tmp_workbook):
    first_page = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            start_cell="A1",
            end_cell="C6",
            max_rows=2,
            max_cols=2,
            values_only=True,
        )
    )

    right_page = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            cursor=first_page["data"]["continuations"]["right"]["cursor"],
            values_only=True,
        )
    )

    assert right_page["data"]["range"] == "C1:C2"
    assert right_page["data"]["sheet_name"] == "Sheet1"
    assert right_page["data"]["total_rows"] == 6
    assert right_page["data"]["total_cols"] == 1
    assert right_page["data"]["truncated"] is True
    assert right_page["data"]["next_start_row"] == 3
    assert right_page["data"]["next_start_cell"] == "C3"
    assert right_page["data"]["values"] == [
        ["City"],
        ["Helsinki"],
    ]
    assert set(right_page["data"]["continuations"]) == {"down"}
    assert right_page["data"]["next_cursor"] == right_page["data"]["continuations"]["down"]["cursor"]


def test_read_data_from_excel_cursor_preserves_values_only_without_repeating_flag(tmp_workbook):
    first_page = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            start_cell="A1",
            end_cell="B6",
            max_rows=2,
            values_only=True,
        )
    )

    second_page = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            cursor=first_page["data"]["next_cursor"],
        )
    )

    assert "values" in second_page["data"]
    assert "cells" not in second_page["data"]
    assert second_page["data"]["values"] == [
        ["Bob", 25],
        ["Carol", 35],
    ]


def test_read_data_from_excel_rejects_invalid_cursor(tmp_workbook):
    payload = json.loads(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            cursor="definitely-not-a-valid-cursor",
        )
    )

    assert payload["ok"] is False
    assert payload["error"]["message"] == "Invalid cursor"


def test_read_data_from_excel_values_only_supports_row_pagination(tmp_workbook):
    payload = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            start_cell="A1",
            end_cell="B6",
            max_rows=2,
            values_only=True,
        )
    )

    assert payload["data"]["range"] == "A1:B2"
    assert payload["data"]["total_rows"] == 6
    assert payload["data"]["truncated"] is True
    assert payload["data"]["next_start_row"] == 3
    assert payload["data"]["next_start_cell"] == "A3"
    assert payload["data"]["next_cursor"] == payload["data"]["continuations"]["down"]["cursor"]
    assert payload["data"]["values"] == [
        ["Name", "Age"],
        ["Alice", 30],
    ]


def test_read_data_from_excel_values_only_handles_out_of_bounds_start(tmp_workbook):
    payload = _load_tool_payload(
        read_data_from_excel(
            tmp_workbook,
            "Sheet1",
            start_cell="Z100",
            values_only=True,
        )
    )

    assert payload["data"]["values"] == []
    assert payload["data"]["range"] == "Z100"
    assert "cells" not in payload["data"]


def test_read_data_from_excel_rejects_non_positive_max_rows(tmp_workbook):
    payload = json.loads(read_data_from_excel(tmp_workbook, "Sheet1", max_rows=0))

    assert payload["ok"] is False
    assert payload["error"]["message"] == "max_rows must be a positive integer"


def test_read_data_from_excel_rejects_non_positive_max_cols(tmp_workbook):
    payload = json.loads(read_data_from_excel(tmp_workbook, "Sheet1", max_cols=0))

    assert payload["ok"] is False
    assert payload["error"]["message"] == "max_cols must be a positive integer"


def test_read_excel_as_table_compact_omits_nonessential_metadata(tmp_workbook):
    payload = _load_tool_payload(read_excel_as_table_tool(tmp_workbook, "Sheet1", compact=True))

    assert payload["operation"] == "read_excel_as_table"
    assert payload["data"]["headers"] == ["Name", "Age", "City"]
    assert payload["data"]["rows"] == [
        ["Alice", 30, "Helsinki"],
        ["Bob", 25, "Tampere"],
        ["Carol", 35, "Turku"],
        ["Dave", 28, "Oulu"],
        ["Eve", 32, "Espoo"],
    ]
    assert payload["data"]["structure_token"].startswith("sf_struct_v1_")
    assert payload["data"]["content_token"].startswith("sf_content_v1_")
    assert payload["data"]["snapshot_metadata"]["token_basis"] == "live_workbook_snapshot"


def test_read_excel_as_table_tool_supports_column_windowing(tmp_workbook):
    payload = _load_tool_payload(
        read_excel_as_table_tool(tmp_workbook, "Sheet1", start_col="B", end_col="C", compact=True)
    )

    assert payload["data"]["headers"] == ["Age", "City"]
    assert payload["data"]["rows"][0] == [30, "Helsinki"]
    assert payload["data"]["rows"][-1] == [32, "Espoo"]


def test_read_excel_as_table_tool_rejects_end_col_before_start_col(tmp_workbook):
    payload = json.loads(
        read_excel_as_table_tool(tmp_workbook, "Sheet1", start_col="C", end_col="B")
    )

    assert payload["ok"] is False
    assert payload["error"]["message"] == "end_col must be greater than or equal to start_col"


def test_read_excel_as_table_compact_preserves_truncation_metadata(tmp_workbook):
    payload = _load_tool_payload(read_excel_as_table_tool(tmp_workbook, "Sheet1", max_rows=2, compact=True))

    assert payload["data"]["headers"] == ["Name", "Age", "City"]
    assert payload["data"]["rows"] == [["Alice", 30, "Helsinki"], ["Bob", 25, "Tampere"]]
    assert payload["data"]["total_rows"] == 5
    assert payload["data"]["truncated"] is True


def test_read_excel_as_table_tool_can_return_records_and_schema(tmp_workbook):
    payload = _load_tool_payload(
        read_excel_as_table_tool(
            tmp_workbook,
            "Sheet1",
            compact=True,
            row_mode="objects",
            infer_schema=True,
        )
    )

    assert payload["data"]["headers"] == ["Name", "Age", "City"]
    assert payload["data"]["records"][0] == {
        "name": "Alice",
        "age": 30,
        "city": "Helsinki",
    }
    assert payload["data"]["schema"][1]["type"] == "integer"
    assert payload["data"]["row_mode"] == "objects"
    assert "rows" not in payload["data"]


def test_read_excel_as_table_tool_rejects_chart_sheet_with_clear_error(tmp_path):
    filepath = _create_chartsheet_first_workbook(tmp_path)

    payload = json.loads(read_excel_as_table_tool(filepath, "Charts"))

    assert payload["ok"] is False
    assert payload["error"]["type"] == "DataError"
    assert "Sheet 'Charts' is a chartsheet" in payload["error"]["message"]


def test_read_data_from_excel_tool_rejects_chart_sheet_with_clear_error(tmp_path):
    filepath = _create_chartsheet_first_workbook(tmp_path)

    payload = json.loads(read_data_from_excel(filepath, "Charts", start_cell="A1", end_cell="B2"))

    assert payload["ok"] is False
    assert payload["error"]["type"] == "DataError"
    assert "Sheet 'Charts' is a chartsheet" in payload["error"]["message"]


def test_list_all_sheets_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(list_all_sheets(tmp_workbook))
    assert payload["operation"] == "list_all_sheets"
    assert payload["data"]["sheets"][0]["name"] == "Sheet1"


def test_quick_read_uses_first_sheet_when_sheet_name_omitted(multi_sheet_workbook):
    result = quick_read_impl(multi_sheet_workbook)
    assert result["sheet_name"] == "Sales"
    assert result["auto_selected_sheet"] is True
    assert result["headers"] == ["Product", "Revenue"]


def test_quick_read_respects_explicit_sheet_name(multi_sheet_workbook):
    result = quick_read_impl(multi_sheet_workbook, sheet_name="Inventory")
    assert result["sheet_name"] == "Inventory"
    assert result["auto_selected_sheet"] is False
    assert result["headers"] == ["Item", "Count"]


def test_quick_read_can_return_records_and_schema(multi_sheet_workbook):
    result = quick_read_impl(
        multi_sheet_workbook,
        sheet_name="Inventory",
        row_mode="objects",
        infer_schema=True,
    )

    assert result["records"] == [{"item": "Widget", "count": 42}]
    assert result["schema"] == [
        {"field": "item", "header": "Item", "type": "string", "nullable": False},
        {"field": "count", "header": "Count", "type": "integer", "nullable": False},
    ]
    assert result["row_mode"] == "objects"
    assert "rows" not in result


def test_quick_read_marks_formula_columns_as_formula_schema(tmp_path):
    filepath = tmp_path / "quick-read-formula-schema.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Qty", "Price", "Total"])
    ws.append(["A", 2, 5, "=B2*C2"])
    ws.append(["B", 3, 7, "=B3*C3"])
    wb.save(filepath)
    wb.close()

    result = quick_read_impl(str(filepath), sheet_name="Sheet1", row_mode="objects", infer_schema=True)

    assert result["records"][1]["total"] == "=B3*C3"
    assert result["schema"][3] == {
        "field": "total",
        "header": "Total",
        "type": "formula",
        "nullable": False,
    }


def test_quick_read_tool_returns_json_envelope(multi_sheet_workbook):
    payload = _load_tool_payload(quick_read(multi_sheet_workbook))
    assert payload["operation"] == "quick_read"
    assert payload["data"]["sheet_name"] == "Sales"
    assert payload["data"]["auto_selected_sheet"] is True


def test_quick_read_tool_supports_start_row_pagination(tmp_workbook):
    payload = _load_tool_payload(
        quick_read(tmp_workbook, sheet_name="Sheet1", start_row=4, max_rows=2)
    )

    assert payload["data"]["rows"] == [
        ["Carol", 35, "Turku"],
        ["Dave", 28, "Oulu"],
    ]
    assert payload["data"]["truncated"] is True
    assert payload["data"]["next_start_row"] == 6


def test_quick_read_supports_column_windowing(tmp_workbook):
    result = quick_read_impl(tmp_workbook, sheet_name="Sheet1", start_col="B", end_col="C")

    assert result["headers"] == ["Age", "City"]
    assert result["rows"][0] == [30, "Helsinki"]
    assert result["rows"][-1] == [32, "Espoo"]


def test_quick_read_ignores_sparse_trailing_rows_after_large_blank_gap(tmp_path):
    filepath = tmp_path / "quick-read-sparse-footer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [
        ("Name", "Age", "City"),
        ("Alice", 30, "Helsinki"),
        ("Bob", 25, "Tampere"),
    ]:
        ws.append(row)
    ws["A18"] = "Footer note"
    wb.save(filepath)
    wb.close()

    result = quick_read_impl(str(filepath), sheet_name="Sheet1")

    assert result["rows"] == [
        ["Alice", 30, "Helsinki"],
        ["Bob", 25, "Tampere"],
    ]
    assert result["total_rows"] == 2


def test_quick_read_tool_supports_column_windowing(tmp_workbook):
    payload = _load_tool_payload(
        quick_read(tmp_workbook, sheet_name="Sheet1", start_col="B", end_col="C")
    )

    assert payload["data"]["headers"] == ["Age", "City"]
    assert payload["data"]["rows"][0] == [30, "Helsinki"]


def test_quick_read_rejects_end_col_before_start_col(tmp_workbook):
    payload = json.loads(
        quick_read(tmp_workbook, sheet_name="Sheet1", start_col="C", end_col="B")
    )

    assert payload["ok"] is False
    assert payload["error"]["message"] == "end_col must be greater than or equal to start_col"


def test_quick_read_tool_can_omit_headers_for_followup_pages(tmp_workbook):
    payload = _load_tool_payload(
        quick_read(
            tmp_workbook,
            sheet_name="Sheet1",
            start_row=4,
            max_rows=2,
            include_headers=False,
        )
    )

    assert "headers" not in payload["data"]
    assert payload["data"]["rows"] == [
        ["Carol", 35, "Turku"],
        ["Dave", 28, "Oulu"],
    ]


def test_quick_read_skips_chartsheet_when_first_sheet_is_not_a_worksheet(tmp_path):
    filepath = _create_chartsheet_first_workbook(tmp_path)

    result = quick_read_impl(filepath)

    assert result["sheet_name"] == "Data"
    assert result["auto_selected_sheet"] is True
    assert result["headers"] == ["Name", "Value"]


def test_quick_read_tool_skips_chartsheet_when_first_sheet_is_not_a_worksheet(tmp_path):
    filepath = _create_chartsheet_first_workbook(tmp_path)

    payload = _load_tool_payload(quick_read(filepath))

    assert payload["operation"] == "quick_read"
    assert payload["data"]["sheet_name"] == "Data"
    assert payload["data"]["auto_selected_sheet"] is True


def test_describe_dataset_summarizes_tabular_worksheet(tmp_workbook):
    result = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1", sample_rows=2)

    assert result["target_kind"] == "worksheet"
    assert result["dataset_kind"] == "worksheet_table"
    assert result["sheet_name"] == "Sheet1"
    assert result["used_range"] == "A1:C6"
    assert result["headers"] == ["Name", "Age", "City"]
    assert result["total_rows"] == 5
    assert result["sample_row_count"] == 2
    assert result["schema"][1]["type"] == "integer"
    assert result["recommended_read_tool"] == "quick_read"
    assert any(candidate["field"] == "name" for candidate in result["key_candidates"])


def test_describe_dataset_returns_tokens_and_snapshot_metadata(tmp_workbook):
    result = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")

    assert result["structure_token"].startswith("sf_struct_v1_")
    assert result["content_token"].startswith("sf_content_v1_")
    assert result["snapshot_metadata"]["token_basis"] == "live_workbook_snapshot"
    assert result["snapshot_metadata"]["file_size"] > 0


def test_quick_read_and_describe_dataset_share_structure_token(tmp_workbook):
    describe_result = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")
    quick_result = quick_read_impl(tmp_workbook, sheet_name="Sheet1")

    assert quick_result["structure_token"] == describe_result["structure_token"]


def test_describe_dataset_reports_sparse_trailing_rows_as_separate_block(tmp_path):
    filepath = tmp_path / "describe-sparse-footer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [
        ("Name", "Age", "City"),
        ("Alice", 30, "Helsinki"),
        ("Bob", 25, "Tampere"),
    ]:
        ws.append(row)
    ws["B15"] = "Footer"
    wb.save(filepath)
    wb.close()

    result = describe_dataset_impl(str(filepath), sheet_name="Sheet1", sample_rows=2)

    assert result["used_range"] == "A1:C15"
    assert result["data_end_row"] == 3
    assert result["ignored_trailing_row_count"] == 1
    assert any("separate block" in observation.lower() for observation in result["observations"])


def test_describe_dataset_summarizes_native_excel_table(tmp_workbook):
    from excel_mcp.tables import create_excel_table

    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    result = describe_dataset_impl(tmp_workbook, table_name="Customers", sample_rows=2)

    assert result["target_kind"] == "excel_table"
    assert result["dataset_kind"] == "structured_table"
    assert result["table_name"] == "Customers"
    assert result["headers"] == ["Name", "Age", "City"]
    assert result["total_rows"] == 5
    assert result["sample_row_count"] == 2
    assert result["recommended_read_tool"] == "read_excel_table"
    assert result["recommended_args"]["table_name"] == "Customers"


def test_describe_dataset_identifies_layout_like_sheet(complex_workbook):
    result = describe_dataset_impl(complex_workbook)

    assert result["target_kind"] == "worksheet"
    assert result["sheet_name"] == "Dashboard"
    assert result["auto_selected_sheet"] is True
    assert result["dataset_kind"] == "layout_like_sheet"
    assert result["recommended_read_tool"] == "profile_workbook"
    assert any("merged cells" in observation.lower() for observation in result["observations"])


def test_describe_dataset_handles_explicit_chart_sheet(tmp_path):
    filepath = _create_chartsheet_first_workbook(tmp_path)

    result = describe_dataset_impl(filepath, sheet_name="Charts")

    assert result["target_kind"] == "chartsheet"
    assert result["recommended_read_tool"] == "list_charts"
    assert result["sheet_name"] == "Charts"


def test_suggest_read_strategy_prefers_quick_read_for_simple_sheet(tmp_workbook):
    result = suggest_read_strategy_impl(tmp_workbook, sheet_name="Sheet1")

    assert result["recommended_tool"] == "quick_read"
    assert result["suggested_args"]["sheet_name"] == "Sheet1"
    assert result["suggested_args"]["row_mode"] == "objects"
    assert result["suggested_args"]["infer_schema"] is True


def test_suggest_read_strategy_prefers_native_table_when_available(tmp_workbook):
    from excel_mcp.tables import create_excel_table

    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    result = suggest_read_strategy_impl(tmp_workbook, sheet_name="Sheet1")

    assert result["recommended_tool"] == "read_excel_table"
    assert result["suggested_args"]["table_name"] == "Customers"
    assert result["confidence"] == "high"


def test_suggest_read_strategy_prefers_profile_workbook_for_layout_sheet(complex_workbook):
    result = suggest_read_strategy_impl(complex_workbook)

    assert result["recommended_tool"] == "profile_workbook"
    assert result["suggested_args"] == {"filepath": complex_workbook}
    assert result["confidence"] == "high"


def test_suggest_read_strategy_can_switch_to_range_preview_for_layout_goal(complex_workbook):
    result = suggest_read_strategy_impl(complex_workbook, goal="layout")

    assert result["recommended_tool"] == "read_data_from_excel"
    assert result["suggested_args"]["sheet_name"] == "Dashboard"
    assert result["suggested_args"]["preview_only"] is True
    assert result["suggested_args"]["values_only"] is True
    assert result["suggested_args"]["end_cell"] == "C2"


def test_suggest_read_strategy_handles_explicit_chart_sheet(tmp_path):
    filepath = _create_chartsheet_first_workbook(tmp_path)

    result = suggest_read_strategy_impl(filepath, sheet_name="Charts")

    assert result["recommended_tool"] == "list_charts"
    assert result["suggested_args"]["sheet_name"] == "Charts"
    assert result["target_kind"] == "chartsheet"


def test_describe_dataset_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(describe_dataset_tool(tmp_workbook, sheet_name="Sheet1", sample_rows=2))

    assert payload["operation"] == "describe_dataset"
    assert payload["data"]["dataset_kind"] == "worksheet_table"


def test_suggest_read_strategy_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(suggest_read_strategy_tool(tmp_workbook, sheet_name="Sheet1"))

    assert payload["operation"] == "suggest_read_strategy"
    assert payload["data"]["recommended_tool"] == "quick_read"


def test_quick_read_rejects_invalid_row_mode(multi_sheet_workbook):
    payload = json.loads(quick_read(multi_sheet_workbook, row_mode="dicts"))
    assert payload["ok"] is False
    assert payload["error"]["message"] == "row_mode must be 'arrays' or 'objects'"


def test_tool_responses_are_compact_json(multi_sheet_workbook):
    raw = quick_read(multi_sheet_workbook)

    assert "\n" not in raw
    assert raw.startswith('{"ok":true')


def test_read_data_from_excel_returns_guided_error_before_oversized_payload(
    tmp_workbook,
    monkeypatch,
):
    monkeypatch.setattr(server_module, "MCP_RESPONSE_CHAR_LIMIT", 200)

    payload = json.loads(read_data_from_excel(tmp_workbook, "Sheet1"))

    assert payload["ok"] is False
    assert payload["error"]["type"] == "ResponseTooLargeError"
    assert payload["error"]["estimated_size"] > payload["error"]["limit"]
    assert any("values_only=True" in hint for hint in payload["error"]["hints"])
    assert any("max_rows" in hint for hint in payload["error"]["hints"])
    assert any("max_cols" in hint for hint in payload["error"]["hints"])
    assert any("preview_only=True" in hint for hint in payload["error"]["hints"])
    assert any("start_cell/end_cell" in hint for hint in payload["error"]["hints"])


def test_quick_read_returns_guided_error_before_oversized_payload(
    tmp_workbook,
    monkeypatch,
):
    monkeypatch.setattr(server_module, "MCP_RESPONSE_CHAR_LIMIT", 120)

    payload = json.loads(quick_read(tmp_workbook, sheet_name="Sheet1"))

    assert payload["ok"] is False
    assert payload["error"]["type"] == "ResponseTooLargeError"
    assert any("max_rows" in hint for hint in payload["error"]["hints"])
    assert any("start_row" in hint for hint in payload["error"]["hints"])
    assert any("start_col/end_col" in hint for hint in payload["error"]["hints"])


def test_write_data_dry_run_does_not_persist(tmp_workbook):
    result = write_data(tmp_workbook, "Sheet1", [["Mallory", 44, "Lahti"]], start_cell="A2", dry_run=True)
    assert result["dry_run"] is True
    assert result["changes"][0]["cell"] == "A2"

    table = read_as_table(tmp_workbook, "Sheet1")
    assert table["rows"][0] == ["Alice", 30, "Helsinki"]


def test_write_data_defaults_to_summary_without_changes(tmp_workbook):
    result = write_data(tmp_workbook, "Sheet1", [["Mallory", 44, "Lahti"]], start_cell="A2")

    assert result["dry_run"] is False
    assert result["changed_cells"] == 3
    assert "changes" not in result


def test_write_data_can_include_changes_explicitly(tmp_workbook):
    result = write_data(
        tmp_workbook,
        "Sheet1",
        [["Mallory", 44, "Lahti"]],
        start_cell="A2",
        include_changes=True,
    )

    assert result["changed_cells"] == 3
    assert result["changes"][0]["cell"] == "A2"


def test_append_table_rows_appends_using_headers(tmp_workbook):
    result = append_table_rows(
        tmp_workbook,
        "Sheet1",
        [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
    )

    assert result["rows_appended"] == 1
    assert result["dry_run"] is False

    table = read_as_table(tmp_workbook, "Sheet1")
    assert table["rows"][-1] == ["Mallory", 44, "Lahti"]


def test_append_table_rows_inserts_before_sparse_footer_after_large_blank_gap(tmp_path):
    filepath = str(tmp_path / "append-before-footer.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [
        ("Name", "Age", "City"),
        ("Alice", 30, "Helsinki"),
        ("Bob", 25, "Tampere"),
    ]:
        ws.append(row)
    ws["A20"] = "Footer note"
    wb.save(filepath)
    wb.close()

    result = append_table_rows(
        filepath,
        "Sheet1",
        [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
    )

    assert result["start_row"] == 4
    table = read_as_table(filepath, "Sheet1")
    assert table["rows"][-1] == ["Mallory", 44, "Lahti"]
    wb = load_workbook(filepath)
    try:
        assert wb["Sheet1"]["A20"].value == "Footer note"
    finally:
        wb.close()


def test_append_table_rows_defaults_to_summary_without_changes(tmp_workbook):
    result = append_table_rows(
        tmp_workbook,
        "Sheet1",
        [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
    )

    assert result["changed_cells"] == 3
    assert "changes" not in result


def test_append_table_rows_requires_structure_change_intent_when_token_is_provided(tmp_workbook):
    dataset = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")

    with pytest.raises(PreconditionFailedError, match="allow_structure_change=True"):
        append_table_rows(
            tmp_workbook,
            "Sheet1",
            [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
            expected_structure_token=dataset["structure_token"],
        )


def test_append_table_rows_rejects_stale_structure_token(tmp_workbook):
    dataset = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")
    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    ws["D1"] = "Country"
    wb.save(tmp_workbook)
    wb.close()

    with pytest.raises(PreconditionFailedError, match="Dataset structure changed"):
        append_table_rows(
            tmp_workbook,
            "Sheet1",
            [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
            expected_structure_token=dataset["structure_token"],
            allow_structure_change=True,
        )


def test_append_table_rows_returns_previous_and_new_tokens(tmp_workbook):
    dataset = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")

    result = append_table_rows(
        tmp_workbook,
        "Sheet1",
        [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
        expected_structure_token=dataset["structure_token"],
        allow_structure_change=True,
        dry_run=True,
    )

    assert result["previous_structure_token"] == dataset["structure_token"]
    assert result["new_structure_token"] != dataset["structure_token"]
    assert result["snapshot_metadata"]["token_basis"] == "live_workbook_snapshot"


def test_append_table_rows_dry_run_does_not_persist(tmp_workbook):
    result = append_table_rows(
        tmp_workbook,
        "Sheet1",
        [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
        dry_run=True,
    )

    assert result["dry_run"] is True
    assert result["start_row"] == 7

    table = read_as_table(tmp_workbook, "Sheet1")
    assert table["total_rows"] == 5


def test_append_table_rows_rejects_native_table_boundary(tmp_workbook):
    from excel_mcp.tables import create_excel_table

    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    with pytest.raises(DataError, match="append_excel_table_rows"):
        append_table_rows(
            tmp_workbook,
            "Sheet1",
            [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
        )


def test_update_rows_by_key_updates_matching_rows_and_reports_missing_keys(tmp_workbook):
    result = update_rows_by_key(
        tmp_workbook,
        "Sheet1",
        "Name",
        [
            {"Name": "Alice", "City": "Vantaa", "Age": 31},
            {"Name": "Missing", "City": "Nowhere"},
        ],
    )

    assert result["updated_rows"] == 1
    assert result["missing_keys"] == ["Missing"]

    table = read_as_table(tmp_workbook, "Sheet1")
    assert table["rows"][0] == ["Alice", 31, "Vantaa"]


def test_update_rows_by_key_ignores_sparse_footer_rows_after_large_blank_gap(tmp_path):
    filepath = str(tmp_path / "update-ignore-footer.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [
        ("Name", "Age", "City"),
        ("Alice", 30, "Helsinki"),
        ("Bob", 25, "Tampere"),
    ]:
        ws.append(row)
    ws["A20"] = "Alice"
    ws["B20"] = "Footer marker"
    wb.save(filepath)
    wb.close()

    result = update_rows_by_key(
        filepath,
        "Sheet1",
        "Name",
        [{"Name": "Alice", "City": "Vantaa"}],
    )

    assert result["updated_rows"] == 1
    table = read_as_table(filepath, "Sheet1")
    assert table["rows"][0] == ["Alice", 30, "Vantaa"]
    wb = load_workbook(filepath)
    try:
        assert wb["Sheet1"]["B20"].value == "Footer marker"
    finally:
        wb.close()


def test_update_rows_by_key_defaults_to_summary_without_changes(tmp_workbook):
    result = update_rows_by_key(
        tmp_workbook,
        "Sheet1",
        "Name",
        [{"Name": "Alice", "City": "Vantaa"}],
    )

    assert result["updated_rows"] == 1
    assert result["changed_cells"] == 1
    assert "changes" not in result


def test_update_rows_by_key_rejects_stale_structure_token(tmp_workbook):
    dataset = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")
    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    ws["D1"] = "Country"
    wb.save(tmp_workbook)
    wb.close()

    with pytest.raises(PreconditionFailedError, match="Dataset structure changed"):
        update_rows_by_key(
            tmp_workbook,
            "Sheet1",
            "Name",
            [{"Name": "Alice", "City": "Vantaa"}],
            expected_structure_token=dataset["structure_token"],
        )


def test_update_rows_by_key_returns_tokens_and_snapshot_metadata(tmp_workbook):
    dataset = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")

    result = update_rows_by_key(
        tmp_workbook,
        "Sheet1",
        "Name",
        [{"Name": "Alice", "City": "Vantaa"}],
        expected_structure_token=dataset["structure_token"],
        dry_run=True,
    )

    assert result["previous_structure_token"] == dataset["structure_token"]
    assert result["new_structure_token"] == dataset["structure_token"]
    assert result["previous_content_token"] != result["new_content_token"]
    assert result["snapshot_metadata"]["file_size"] > 0


def test_update_rows_by_key_dry_run_does_not_persist(tmp_workbook):
    result = update_rows_by_key(
        tmp_workbook,
        "Sheet1",
        "Name",
        [{"Name": "Alice", "City": "Vantaa"}],
        dry_run=True,
    )

    assert result["dry_run"] is True
    assert result["changes"][0]["cell"] == "C2"

    table = read_as_table(tmp_workbook, "Sheet1")
    assert table["rows"][0] == ["Alice", 30, "Helsinki"]


def test_append_table_rows_tool_surfaces_precondition_error_details(tmp_workbook):
    dataset = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")
    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    ws["D1"] = "Country"
    wb.save(tmp_workbook)
    wb.close()

    payload = json.loads(
        append_table_rows_tool(
            tmp_workbook,
            "Sheet1",
            [{"Name": "Mallory", "Age": 44, "City": "Lahti"}],
            expected_structure_token=dataset["structure_token"],
            allow_structure_change=True,
        )
    )

    assert payload["ok"] is False
    assert payload["error"]["code"] == "stale_structure_token"
    assert payload["error"]["suggested_next_tool"] == "describe_dataset"


def test_update_rows_by_key_tool_surfaces_precondition_error_details(tmp_workbook):
    dataset = describe_dataset_impl(tmp_workbook, sheet_name="Sheet1")
    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    ws["D1"] = "Country"
    wb.save(tmp_workbook)
    wb.close()

    payload = json.loads(
        update_rows_by_key_tool(
            tmp_workbook,
            "Sheet1",
            "Name",
            [{"Name": "Alice", "City": "Vantaa"}],
            expected_structure_token=dataset["structure_token"],
        )
    )

    assert payload["ok"] is False
    assert payload["error"]["code"] == "stale_structure_token"
    assert payload["error"]["details"]["sheet_name"] == "Sheet1"

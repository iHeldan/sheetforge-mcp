import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


@pytest.fixture
def tmp_workbook(tmp_path):
    """Create a temporary Excel workbook with sample data."""
    filepath = str(tmp_path / "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "City"

    # Data rows
    data = [
        ("Alice", 30, "Helsinki"),
        ("Bob", 25, "Tampere"),
        ("Carol", 35, "Turku"),
        ("Dave", 28, "Oulu"),
        ("Eve", 32, "Espoo"),
    ]
    for i, (name, age, city) in enumerate(data, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = age
        ws[f"C{i}"] = city

    wb.save(filepath)
    wb.close()
    return filepath


@pytest.fixture
def empty_workbook(tmp_path):
    """Create an empty Excel workbook."""
    filepath = str(tmp_path / "empty.xlsx")
    wb = Workbook()
    wb.save(filepath)
    wb.close()
    return filepath


@pytest.fixture
def multi_sheet_workbook(tmp_path):
    """Create a workbook with multiple sheets."""
    filepath = str(tmp_path / "multi.xlsx")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1["A1"] = "Product"
    ws1["B1"] = "Revenue"
    ws1["A2"] = "Widget"
    ws1["B2"] = 1500

    ws2 = wb.create_sheet("Inventory")
    ws2["A1"] = "Item"
    ws2["B1"] = "Count"
    ws2["A2"] = "Widget"
    ws2["B2"] = 42

    wb.save(filepath)
    wb.close()
    return filepath


@pytest.fixture
def complex_workbook(tmp_path):
    """Create a more realistic workbook with charts, tables, rules, and a chartsheet."""
    filepath = str(tmp_path / "complex.xlsx")
    wb = Workbook()

    dashboard = wb.active
    dashboard.title = "Dashboard"
    data = wb.create_sheet("Data")

    dashboard.merge_cells("A1:C1")
    dashboard["A1"] = "Executive Dashboard"
    dashboard["B2"] = "=SUM(Data!B2:B6)"
    dashboard["C2"] = "=SUM(DataWindow)"
    dashboard.freeze_panes = "A2"

    rows = [
        ("Product", "Sales", "Region", "Target"),
        ("Widget", 12, "North", 10),
        ("Gadget", 24, "South", 20),
        ("Thing", 18, "West", 18),
        ("Device", 30, "East", 28),
        ("Tool", 16, "North", 15),
    ]
    for row in rows:
        data.append(row)

    table = Table(displayName="SalesData", ref="A1:D6")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    data.add_table(table)
    data.freeze_panes = "B2"
    data.auto_filter.ref = "A1:D6"
    data.print_area = "A1:H12"
    data.merge_cells("F2:G2")
    data["F2"] = "Merged note"
    data["H2"] = "=SUM(B2:B6)"

    data_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="100")
    data_validation.add("D2:D6")
    data.add_data_validation(data_validation)
    data.conditional_formatting.add(
        "C2:C6",
        FormulaRule(formula=["$B2>15"]),
    )

    dashboard_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="Data!$B$2",
        formula2="Data!$B$6",
    )
    dashboard_validation.add("A4:A8")
    dashboard.add_data_validation(dashboard_validation)
    dashboard.conditional_formatting.add(
        "B4:B8",
        FormulaRule(formula=["COUNTIF(Data!$B$2:$B$6,B4)>0"]),
    )

    wb.defined_names["DataWindow"] = DefinedName(
        "DataWindow",
        attr_text="Data!$B$2:$D$5",
    )
    data.defined_names.add(
        DefinedName(
            "LocalProducts",
            attr_text="Data!$A$2:$A$4",
        )
    )

    dashboard_chart = BarChart()
    dashboard_chart.title = "Sales by Product"
    dashboard_chart_data = Reference(data, min_col=2, min_row=1, max_row=6)
    dashboard_chart_categories = Reference(data, min_col=1, min_row=2, max_row=6)
    dashboard_chart.add_data(dashboard_chart_data, titles_from_data=True)
    dashboard_chart.set_categories(dashboard_chart_categories)
    dashboard.add_chart(dashboard_chart, "E2")

    chart_sheet_chart = BarChart()
    chart_sheet_chart.title = "Targets"
    chart_sheet_data = Reference(data, min_col=4, min_row=1, max_row=6)
    chart_sheet_categories = Reference(data, min_col=1, min_row=2, max_row=6)
    chart_sheet_chart.add_data(chart_sheet_data, titles_from_data=True)
    chart_sheet_chart.set_categories(chart_sheet_categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart_sheet_chart)

    wb._sheets = [chart_sheet, dashboard, data]
    wb.active = 0
    wb.save(filepath)
    wb.close()
    return filepath

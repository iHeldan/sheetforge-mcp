import json

import pytest
from openpyxl import load_workbook

from excel_mcp.formatting import format_range, format_ranges
from excel_mcp.exceptions import ValidationError, FormattingError
from excel_mcp.server import format_ranges as format_ranges_tool


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    assert "operation" in payload
    assert "message" in payload
    return payload


# --- Basic formatting ---

def test_bold_formatting(tmp_workbook):
    result = format_range(tmp_workbook, "Sheet1", "A1", bold=True)
    assert "Applied" in result["message"]

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].font.bold is True
    wb.close()


def test_italic_and_underline(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", italic=True, underline=True)

    wb = load_workbook(tmp_workbook)
    cell = wb["Sheet1"]["A1"]
    assert cell.font.italic is True
    assert cell.font.underline == "single"
    wb.close()


def test_font_size(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", font_size=16)

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].font.size == 16
    wb.close()


def test_font_color(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", font_color="FF0000")

    wb = load_workbook(tmp_workbook)
    # openpyxl may store as FFFF0000 or 00FF0000 depending on version
    assert wb["Sheet1"]["A1"].font.color.rgb in ("FFFF0000", "00FF0000")
    wb.close()


def test_background_color(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", bg_color="00FF00")

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].fill.start_color.rgb == "FF00FF00"
    wb.close()


# --- Range formatting ---

def test_format_range_multiple_cells(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", end_cell="C1", bold=True, font_size=14)

    wb = load_workbook(tmp_workbook)
    for col in ["A", "B", "C"]:
        cell = wb["Sheet1"][f"{col}1"]
        assert cell.font.bold is True
        assert cell.font.size == 14
    wb.close()


# --- Border ---

def test_border_formatting(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", border_style="thin")

    wb = load_workbook(tmp_workbook)
    border = wb["Sheet1"]["A1"].border
    assert border.left.style == "thin"
    assert border.right.style == "thin"
    wb.close()


def test_format_range_accepts_hash_prefixed_colors(tmp_workbook):
    format_range(
        tmp_workbook,
        "Sheet1",
        "A1",
        font_color="#FF0000",
        bg_color="#00FF00",
        border_style="thin",
        border_color="#1F4E78",
    )

    wb = load_workbook(tmp_workbook)
    cell = wb["Sheet1"]["A1"]
    assert cell.font.color.rgb in ("FFFF0000", "00FF0000")
    assert cell.fill.start_color.rgb == "FF00FF00"
    assert cell.border.left.color.rgb == "FF1F4E78"
    wb.close()


def test_format_range_invalid_color_error_has_actionable_hint(tmp_workbook):
    with pytest.raises(FormattingError, match="#1F4E78"):
        format_range(tmp_workbook, "Sheet1", "A1", font_color="blue")


# --- Alignment & wrap ---

def test_alignment_center(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", alignment="center")

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].alignment.horizontal == "center"
    wb.close()


def test_wrap_text(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", wrap_text=True)

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].alignment.wrap_text is True
    wb.close()


# --- Number format ---

def test_number_format(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "B2", number_format="#,##0.00")

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["B2"].number_format == "#,##0.00"
    wb.close()


# --- Merge cells ---

def test_merge_cells(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", end_cell="C1", merge_cells=True)

    wb = load_workbook(tmp_workbook)
    merged = list(wb["Sheet1"].merged_cells.ranges)
    assert any("A1:C1" in str(r) for r in merged)
    wb.close()


# --- Dry run ---

def test_dry_run_does_not_persist(tmp_workbook):
    result = format_range(tmp_workbook, "Sheet1", "A1", bold=True, dry_run=True)
    assert result["changes"][0]["range"] == "A1"

    wb = load_workbook(tmp_workbook)
    # dry_run should NOT save changes
    assert wb["Sheet1"]["A1"].font.bold is not True
    wb.close()


def test_format_range_defaults_to_summary_without_changes(tmp_workbook):
    result = format_range(tmp_workbook, "Sheet1", "A1", bold=True)
    assert result["dry_run"] is False
    assert "changes" not in result


def test_format_range_can_include_changes_explicitly(tmp_workbook):
    result = format_range(tmp_workbook, "Sheet1", "A1", bold=True, include_changes=True)
    assert result["changes"][0]["range"] == "A1"


# --- Conditional formatting ---

def test_conditional_format_cell_is(tmp_workbook):
    cond = {
        "type": "cell_is",
        "params": {
            "operator": "greaterThan",
            "formula": ["30"],
            "fill": {"fgColor": "FFC7CE"},
        },
    }
    result = format_range(
        tmp_workbook, "Sheet1", "B2", end_cell="B6", conditional_format=cond
    )
    assert "Applied" in result["message"]


def test_conditional_format_cell_is_accepts_hash_prefixed_fill_color(tmp_workbook):
    cond = {
        "type": "cell_is",
        "params": {
            "operator": "greaterThan",
            "formula": ["30"],
            "fill": {"fgColor": "#FFC7CE"},
        },
    }
    result = format_range(
        tmp_workbook, "Sheet1", "B2", end_cell="B6", conditional_format=cond
    )
    assert "Applied" in result["message"]


def test_conditional_format_data_bar_accepts_top_level_params(tmp_workbook):
    cond = {
        "type": "data_bar",
        "start_type": "min",
        "end_type": "max",
        "color": "2E86C1",
    }
    result = format_range(
        tmp_workbook, "Sheet1", "B2", end_cell="B6", conditional_format=cond
    )
    assert "Applied" in result["message"]

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert len(ws.conditional_formatting) == 1
    wb.close()


def test_conditional_format_missing_type(tmp_workbook):
    with pytest.raises(FormattingError, match="type not specified"):
        format_range(
            tmp_workbook, "Sheet1", "A1", conditional_format={"params": {}}
        )


def test_format_ranges_applies_multiple_ranges(tmp_workbook):
    result = format_ranges(
        tmp_workbook,
        "Sheet1",
        [
            {"start_cell": "A1", "bold": True},
            {"start_cell": "B2", "end_cell": "C2", "bg_color": "00FF00"},
        ],
    )
    assert result["ranges_formatted"] == 2
    assert result["ranges"] == ["A1", "B2:C2"]
    assert "changes" not in result

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws["A1"].font.bold is True
    assert ws["B2"].fill.start_color.rgb == "FF00FF00"
    assert ws["C2"].fill.start_color.rgb == "FF00FF00"
    wb.close()


def test_format_ranges_dry_run_does_not_persist(tmp_workbook):
    result = format_ranges(
        tmp_workbook,
        "Sheet1",
        [{"start_cell": "A1", "bold": True}, {"start_cell": "B2", "font_size": 18}],
        dry_run=True,
    )
    assert result["dry_run"] is True
    assert len(result["changes"]) == 2

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws["A1"].font.bold is not True
    assert ws["B2"].font.size != 18
    wb.close()


def test_format_ranges_allows_partial_success(tmp_workbook):
    result = format_ranges(
        tmp_workbook,
        "Sheet1",
        [
            {"start_cell": "A1", "bold": True},
            {
                "start_cell": "B2",
                "end_cell": "B6",
                "conditional_format": {
                    "type": "data_bar",
                    "start_type": "bogus",
                    "end_type": "max",
                    "color": "2E86C1",
                },
            },
        ],
    )
    assert result["ranges_formatted"] == 1
    assert result["ranges_failed"] == 1
    assert result["ranges"] == ["A1"]
    assert result["errors"][0]["range"] == "B2:B6"
    assert "changes" not in result

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws["A1"].font.bold is True
    assert len(ws.conditional_formatting) == 0
    wb.close()


def test_format_ranges_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(
        format_ranges_tool(
            tmp_workbook,
            "Sheet1",
            [{"start_cell": "A1", "bold": True}, {"start_cell": "B2", "font_size": 18}],
            dry_run=True,
        )
    )
    assert payload["operation"] == "format_ranges"
    assert payload["dry_run"] is True
    assert payload["data"]["ranges_formatted"] == 2


def test_format_ranges_tool_surfaces_partial_success_metadata(tmp_workbook):
    payload = _load_tool_payload(
        format_ranges_tool(
            tmp_workbook,
            "Sheet1",
            [
                {"start_cell": "A1", "bold": True},
                {
                    "start_cell": "B2",
                    "end_cell": "B6",
                    "conditional_format": {
                        "type": "data_bar",
                        "start_type": "bogus",
                        "end_type": "max",
                        "color": "2E86C1",
                    },
                },
            ],
        )
    )
    assert payload["operation"] == "format_ranges"
    assert payload["data"]["ranges_formatted"] == 1
    assert payload["data"]["ranges_failed"] == 1
    assert payload["warnings"][0] == "1 range(s) failed during batch formatting"


# --- Error cases ---

def test_format_invalid_sheet(tmp_workbook):
    with pytest.raises(ValidationError, match="not found"):
        format_range(tmp_workbook, "NoSheet", "A1", bold=True)


def test_format_invalid_start_cell(tmp_workbook):
    with pytest.raises(ValidationError, match="Invalid start cell"):
        format_range(tmp_workbook, "Sheet1", "123", bold=True)


def test_format_ranges_rejects_non_object_operations(tmp_workbook):
    with pytest.raises(FormattingError, match="must be an object"):
        format_ranges(tmp_workbook, "Sheet1", ["A1"])

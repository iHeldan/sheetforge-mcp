import json

import pytest
from openpyxl import Workbook, load_workbook

from excel_mcp.calculations import apply_formula
from excel_mcp.exceptions import CalculationError, ValidationError
from excel_mcp.server import (
    validate_excel_range as validate_excel_range_tool,
    validate_formula_syntax,
)
from excel_mcp.validation import (
    validate_formula,
    validate_formula_in_cell_operation,
    validate_range_in_sheet_operation,
)


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    return payload


@pytest.mark.parametrize(
    "formula",
    [
        '=webservice("https://example.invalid")',
        '=HyPeRlInK("https://example.invalid","open")',
        '=_xlfn.INDIRECT("A1")',
    ],
)
def test_validate_formula_rejects_unsafe_functions_case_insensitively(formula):
    is_valid, message = validate_formula(formula)

    assert is_valid is False
    assert "Unsafe function" in message


@pytest.mark.parametrize("formula", ["=1+", "=1+*2", "=()", "=SUM(1,2"])
def test_validate_formula_rejects_structurally_incomplete_formulas(formula):
    is_valid, _ = validate_formula(formula)

    assert is_valid is False


@pytest.mark.parametrize(
    "formula",
    [
        "=SUM(1,2)",
        '=IF(A1="(",1,0)',
        "=1+-2",
        "=SUM()",
        "=A1 B1",
        "=A1#",
        "=LAMBDA(x,x+1)(2)",
    ],
)
def test_validate_formula_accepts_supported_excel_structures(formula):
    is_valid, message = validate_formula(formula)

    assert is_valid is True, message


def test_validate_formula_rejects_whitespace_between_numeric_operands():
    is_valid, _ = validate_formula("=1 2")

    assert is_valid is False


def test_apply_formula_rejects_unsafe_formula_without_mutating_workbook(tmp_path):
    filepath = str(tmp_path / "unsafe-formula.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "keep"
    wb.save(filepath)
    wb.close()

    with pytest.raises(CalculationError, match="Unsafe function"):
        apply_formula(filepath, "Sheet1", "A1", '=webservice("https://example.invalid")')

    loaded = load_workbook(filepath)
    assert loaded["Sheet1"]["A1"].value == "keep"
    loaded.close()


def test_validate_formula_reports_exact_match_and_tool_data(tmp_path):
    filepath = str(tmp_path / "formula-match.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=SUM(B1:B2)"
    wb.save(filepath)
    wb.close()

    result = validate_formula_in_cell_operation(
        filepath,
        "Sheet1",
        "A1",
        "=SUM(B1:B2)",
    )
    assert result["matches"] is True

    payload = _load_tool_payload(
        validate_formula_syntax(filepath, "Sheet1", "A1", "=SUM(B1:B2)")
    )
    assert payload["data"]["matches"] is True


def test_validate_range_allows_empty_cells_inside_excel_limits(tmp_workbook):
    result = validate_range_in_sheet_operation(tmp_workbook, "Sheet1", "D10")

    assert result["valid"] is True
    assert result["extends_beyond_data"] is True


def test_validate_range_accepts_excel_bottom_right_cell(tmp_workbook):
    result = validate_range_in_sheet_operation(tmp_workbook, "Sheet1", "XFD1048576")

    assert result["valid"] is True
    assert result["extends_beyond_data"] is True


@pytest.mark.parametrize("start_cell", ["XFE1", "A1048577"])
def test_validate_range_rejects_coordinates_outside_excel_limits(tmp_workbook, start_cell):
    with pytest.raises(ValidationError, match="outside Excel limits"):
        validate_range_in_sheet_operation(tmp_workbook, "Sheet1", start_cell)


def test_validate_range_rejects_reversed_range(tmp_workbook):
    with pytest.raises(ValidationError, match="End row cannot be before start row"):
        validate_range_in_sheet_operation(tmp_workbook, "Sheet1", "A2", "A1")


def test_validate_range_tool_preserves_explicit_end_cell(tmp_workbook):
    payload = _load_tool_payload(
        validate_excel_range_tool(tmp_workbook, "Sheet1", "A1", "D10")
    )

    assert payload["data"]["range"] == "A1:D10"
    assert payload["data"]["extends_beyond_data"] is True

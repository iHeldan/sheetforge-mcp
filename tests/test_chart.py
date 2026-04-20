import json

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.cell import range_boundaries

from excel_mcp.chart import (
    ChartType,
    create_chart_from_series,
    create_chart_in_sheet,
    find_free_canvas_slots,
    list_charts,
)
from excel_mcp.exceptions import ValidationError, ChartError
from excel_mcp.sheet import copy_sheet, rename_sheet
from excel_mcp.server import (
    create_chart as create_chart_tool,
    create_chart_from_series as create_chart_from_series_tool,
    find_free_canvas as find_free_canvas_tool,
    list_charts as list_charts_tool,
)


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    assert "operation" in payload
    assert "message" in payload
    return payload


def _ranges_intersect(first: str, second: str) -> bool:
    first_min_col, first_min_row, first_max_col, first_max_row = range_boundaries(first)
    second_min_col, second_min_row, second_max_col, second_max_row = range_boundaries(second)
    return not (
        first_max_row < second_min_row
        or second_max_row < first_min_row
        or first_max_col < second_min_col
        or second_max_col < first_min_col
    )


@pytest.fixture
def chart_workbook(tmp_path):
    """Workbook with numeric data suitable for charting."""
    from openpyxl import Workbook

    filepath = str(tmp_path / "chart.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"
    ws["C1"] = "Cost"
    for i, (month, rev, cost) in enumerate(
        [("Jan", 100, 60), ("Feb", 150, 80), ("Mar", 200, 90), ("Apr", 180, 85)],
        start=2,
    ):
        ws[f"A{i}"] = month
        ws[f"B{i}"] = rev
        ws[f"C{i}"] = cost
    wb.save(filepath)
    wb.close()
    return filepath


# --- ChartType enum ---

def test_chart_type_enum_has_five_members():
    assert len(ChartType) == 5
    assert set(ChartType) == {
        ChartType.LINE,
        ChartType.BAR,
        ChartType.PIE,
        ChartType.SCATTER,
        ChartType.AREA,
    }


# --- Successful chart creation ---

@pytest.mark.parametrize("chart_type", ["line", "bar", "pie", "area"])
def test_create_chart_supported_types(chart_workbook, chart_type):
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "A1:B5", chart_type, "E1", title=f"Test {chart_type}"
    )
    assert "successfully" in result["message"].lower()
    assert result["details"]["type"] == chart_type


def test_create_scatter_chart(chart_workbook):
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "B1:C5", "scatter", "E1", title="Scatter"
    )
    assert result["details"]["type"] == "scatter"

    charts = list_charts(chart_workbook, sheet_name="Sales")
    created_chart = next(chart for chart in charts if chart["anchor"] == "E1")
    assert created_chart["series"][0]["title"] == "Cost"
    assert created_chart["series"][0]["x_values"].endswith("$B$2:$B$5")
    assert created_chart["series"][0]["y_values"].endswith("$C$2:$C$5")


def test_chart_with_style_options(chart_workbook):
    style = {
        "show_legend": True,
        "legend_position": "b",
        "show_data_labels": True,
        "data_label_options": {"show_val": True, "show_percent": False},
        "grid_lines": True,
    }
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "A1:B5", "bar", "E1", title="Styled", style=style
    )
    assert "successfully" in result["message"].lower()


def test_chart_without_legend(chart_workbook):
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "A1:B5", "line", "E1",
        style={"show_legend": False, "show_data_labels": False},
    )
    assert "successfully" in result["message"].lower()


def test_chart_with_axis_labels(chart_workbook):
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "A1:B5", "bar", "E1",
        title="Revenue", x_axis="Month", y_axis="EUR",
    )
    assert result["details"]["data_range"] == "A1:B5"


def test_chart_omits_empty_axis_titles(chart_workbook):
    create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:B5",
        "bar",
        "E1",
        title="Revenue",
        x_axis="",
        y_axis="",
    )

    charts = list_charts(chart_workbook, sheet_name="Sales")
    created_chart = next(chart for chart in charts if chart["anchor"] == "E1")
    assert "x_axis_title" not in created_chart
    assert "y_axis_title" not in created_chart


def test_create_chart_can_reference_data_from_another_sheet(chart_workbook):
    wb = load_workbook(chart_workbook)
    source = wb.create_sheet("Source")
    source["A1"] = "Month"
    source["B1"] = "Users"
    source["A2"] = "Jan"
    source["B2"] = 10
    source["A3"] = "Feb"
    source["B3"] = 15
    wb.save(chart_workbook)
    wb.close()

    result = create_chart_in_sheet(
        chart_workbook, "Sales", "Source!A1:B3", "bar", "J1", title="Users"
    )

    assert result["details"]["data_range"] == "Source!A1:B3"


def test_chart_dimensions_can_be_set_with_top_level_params(chart_workbook):
    result = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:B5",
        "bar",
        "E1",
        width=12,
        height=8,
    )

    assert result["details"]["width"] == 12
    assert result["details"]["height"] == 8

    charts = list_charts(chart_workbook, sheet_name="Sales")
    created_chart = next(chart for chart in charts if chart["anchor"] == "E1")
    assert created_chart["width"] == 12
    assert created_chart["height"] == 8
    assert created_chart["occupied_range"].startswith("E1:")


def test_chart_dimensions_fallback_to_legacy_style_keys(chart_workbook):
    create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:B5",
        "line",
        "E1",
        style={"width": 11, "height": 6.5},
    )

    charts = list_charts(chart_workbook, sheet_name="Sales")
    created_chart = next(chart for chart in charts if chart["anchor"] == "E1")
    assert created_chart["width"] == 11
    assert created_chart["height"] == 6.5


def test_list_charts_omits_zero_dimensions_for_chartsheets(tmp_path):
    filepath = str(tmp_path / "chartsheet-dimensions.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Month", "Value"), ("Jan", 10), ("Feb", 20)]:
        ws.append(row)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb.save(filepath)
    wb.close()

    charts = list_charts(filepath, sheet_name="Charts")

    assert len(charts) == 1
    assert "width" not in charts[0]
    assert "height" not in charts[0]


def test_find_free_canvas_rejects_boolean_limit(chart_workbook):
    with pytest.raises(ValidationError, match="limit must be a positive integer"):
        find_free_canvas_slots(chart_workbook, "Sales", limit=True)


def test_create_chart_from_series_supports_non_contiguous_ranges(chart_workbook):
    wb = load_workbook(chart_workbook)
    ws = wb["Sales"]
    ws["G1"] = "Clicks"
    ws["G2"] = 12
    ws["G3"] = 18
    ws["G4"] = 20
    ws["G5"] = 16
    wb.save(chart_workbook)
    wb.close()

    result = create_chart_from_series(
        chart_workbook,
        "Sales",
        "bar",
        "J1",
        series=[
            {"title": "Revenue", "values_range": "B2:B5"},
            {"title": "Clicks", "values_range": "G2:G5"},
        ],
        categories_range="A2:A5",
        title="Quick Wins",
    )

    assert result["details"]["series_count"] == 2

    charts = list_charts(chart_workbook, sheet_name="Sales")
    created_chart = next(chart for chart in charts if chart["anchor"] == "J1")
    assert created_chart["title"] == "Quick Wins"
    assert len(created_chart["series"]) == 2
    assert created_chart["series"][0]["values"].endswith("$B$2:$B$5")
    assert created_chart["series"][1]["values"].endswith("$G$2:$G$5")
    assert created_chart["series"][0]["categories"].endswith("$A$2:$A$5")


def test_create_chart_can_use_explicit_series_definitions(chart_workbook):
    wb = load_workbook(chart_workbook)
    ws = wb["Sales"]
    ws["G1"] = "Clicks"
    ws["G2"] = 12
    ws["G3"] = 18
    ws["G4"] = 20
    ws["G5"] = 16
    wb.save(chart_workbook)
    wb.close()

    result = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        None,
        "bar",
        "L1",
        title="Unified",
        series=[
            {"title": "Revenue", "values_range": "B2:B5"},
            {"title": "Clicks", "values_range": "G2:G5"},
        ],
        categories_range="A2:A5",
    )

    assert result["details"]["series_count"] == 2
    assert result["details"]["categories_range"] == "A2:A5"

    charts = list_charts(chart_workbook, sheet_name="Sales")
    created_chart = next(chart for chart in charts if chart["anchor"] == "L1")
    assert created_chart["title"] == "Unified"
    assert len(created_chart["series"]) == 2


def test_create_chart_can_auto_place_to_right_of_data_range(chart_workbook):
    result = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:B5",
        "bar",
        placement={"relative_to": "data_range", "direction": "right", "padding_columns": 2},
    )

    assert result["details"]["location"] == "E1"
    assert result["details"]["placement"]["relative_to"] == "data_range"
    assert result["details"]["occupied_range"].startswith("E1:")


def test_create_chart_from_series_can_auto_place_below_used_range(chart_workbook):
    result = create_chart_from_series(
        chart_workbook,
        "Sales",
        "bar",
        series=[{"title": "Revenue", "values_range": "B2:B5"}],
        categories_range="A2:A5",
        placement={"relative_to": "used_range", "direction": "below", "padding_rows": 1},
    )

    assert result["details"]["location"] == "A7"
    assert result["details"]["placement"]["relative_to"] == "used_range"


def test_create_chart_can_auto_place_below_table(chart_workbook):
    from excel_mcp.tables import create_excel_table

    create_excel_table(chart_workbook, "Sales", "A1:C5", table_name="SalesData")
    result = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:C5",
        "line",
        placement={"relative_to": "table:SalesData", "direction": "below", "padding_rows": 2},
    )

    assert result["details"]["location"] == "A8"
    assert result["details"]["placement"]["relative_to"] == "table:SalesData"


def test_content_placement_accounts_for_existing_chart_footprint(chart_workbook):
    first_chart = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:B5",
        "bar",
        "E1",
        width=12,
        height=8,
    )

    _, _, first_max_col, _ = range_boundaries(first_chart["details"]["occupied_range"])
    second_chart = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:C5",
        "line",
        placement={"relative_to": "content", "direction": "right", "padding_columns": 1},
    )

    second_min_col, second_min_row, _, _ = range_boundaries(second_chart["details"]["occupied_range"])
    assert second_min_row == 1
    assert second_min_col == first_max_col + 2


def test_find_free_canvas_returns_non_overlapping_slots(chart_workbook):
    existing = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:B5",
        "bar",
        "E1",
        width=10,
        height=6,
    )

    result = find_free_canvas_slots(
        chart_workbook,
        "Sales",
        width=10,
        height=6,
        limit=2,
    )

    suggestions = result["suggestions"]
    assert len(suggestions) == 2
    assert result["occupancy"]["charts"] == 1
    assert not _ranges_intersect(suggestions[0]["occupied_range"], "A1:C5")
    assert not _ranges_intersect(
        suggestions[0]["occupied_range"],
        existing["details"]["occupied_range"],
    )
    assert not _ranges_intersect(
        suggestions[0]["occupied_range"],
        suggestions[1]["occupied_range"],
    )


def test_find_free_canvas_defaults_to_chart_dimensions(chart_workbook):
    result = find_free_canvas_slots(chart_workbook, "Sales", limit=1)

    suggestion = result["suggestions"][0]
    assert suggestion["width"] == 15
    assert suggestion["height"] == 7.5


def test_find_free_canvas_can_scan_cell_grid_blocks(chart_workbook):
    result = find_free_canvas_slots(
        chart_workbook,
        "Sales",
        min_rows=2,
        min_cols=2,
        limit=1,
    )

    suggestion = result["suggestions"][0]
    assert suggestion["occupied_range"] == "D1:E2"
    assert suggestion["row_span"] == 2
    assert suggestion["column_span"] == 2


def test_create_chart_can_auto_place_into_free_canvas_gap(chart_workbook):
    existing = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:B5",
        "bar",
        "E1",
        width=10,
        height=6,
    )

    result = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "A1:C5",
        "line",
        width=10,
        height=6,
        placement={"relative_to": "free_canvas", "padding_columns": 1, "padding_rows": 1},
    )

    assert result["details"]["placement"]["relative_to"] == "free_canvas"
    assert "search_window" in result["details"]["placement"]
    assert not _ranges_intersect(result["details"]["occupied_range"], "A1:C5")
    assert not _ranges_intersect(
        result["details"]["occupied_range"],
        existing["details"]["occupied_range"],
    )


def test_default_placement_falls_back_to_target_sheet_content_for_cross_sheet_data(chart_workbook):
    wb = load_workbook(chart_workbook)
    source = wb.create_sheet("Source")
    source["A1"] = "Month"
    source["B1"] = "Users"
    source["A2"] = "Jan"
    source["B2"] = 10
    source["A3"] = "Feb"
    source["B3"] = 15
    wb.save(chart_workbook)
    wb.close()

    result = create_chart_in_sheet(
        chart_workbook,
        "Sales",
        "Source!A1:B3",
        "bar",
        placement={"direction": "below"},
    )

    assert result["details"]["location"] == "A7"
    assert result["details"]["placement"]["relative_to"] == "content"


def test_create_scatter_chart_from_series(chart_workbook):
    result = create_chart_from_series(
        chart_workbook,
        "Sales",
        "scatter",
        "J1",
        series=[
            {"title": "Revenue vs Cost", "x_range": "B2:B5", "y_range": "C2:C5"},
        ],
        title="Scatter",
    )

    assert result["details"]["type"] == "scatter"

    charts = list_charts(chart_workbook, sheet_name="Sales")
    created_chart = next(chart for chart in charts if chart["anchor"] == "J1")
    assert created_chart["series"][0]["x_values"].endswith("$B$2:$B$5")
    assert created_chart["series"][0]["y_values"].endswith("$C$2:$C$5")


def test_list_charts_returns_created_chart_metadata(chart_workbook):
    create_chart_in_sheet(
        chart_workbook, "Sales", "A1:C5", "bar", "E1", title="Revenue", x_axis="Month", y_axis="EUR"
    )

    charts = list_charts(chart_workbook)

    assert len(charts) == 1
    assert charts[0]["sheet_name"] == "Sales"
    assert charts[0]["chart_type"] == "bar"
    assert charts[0]["title"] == "Revenue"
    assert charts[0]["x_axis_title"] == "Month"
    assert charts[0]["y_axis_title"] == "EUR"
    assert charts[0]["anchor"] == "E1"
    assert charts[0]["width"] == 15
    assert charts[0]["height"] == 7.5


def test_rename_sheet_updates_embedded_chart_series_references(chart_workbook):
    create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "bar", "E1", title="Revenue")

    result = rename_sheet(chart_workbook, "Sales", "Revenue Data")
    assert result["chart_reference_updates"] == 3

    charts = list_charts(chart_workbook, sheet_name="Revenue Data")
    created_chart = next(chart for chart in charts if chart["anchor"] == "E1")
    assert created_chart["series"][0]["title"] == "'Revenue Data'!B1"
    assert created_chart["series"][0]["categories"] == "'Revenue Data'!$A$2:$A$5"
    assert created_chart["series"][0]["values"] == "'Revenue Data'!$B$2:$B$5"


def test_rename_sheet_updates_chartsheet_chart_series_references(tmp_path):
    filepath = str(tmp_path / "chartsheet-rename.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Month", "Sales"), ("Jan", 10), ("Feb", 20), ("Mar", 15)]:
        ws.append(row)

    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb.save(filepath)
    wb.close()

    result = rename_sheet(filepath, "Data", "Revenue Data")
    assert result["chart_reference_updates"] == 3

    charts = list_charts(filepath, sheet_name="Charts")
    assert charts[0]["series"][0]["title"] == "'Revenue Data'!B1"
    assert charts[0]["series"][0]["categories"] == "'Revenue Data'!$A$2:$A$4"
    assert charts[0]["series"][0]["values"] == "'Revenue Data'!$B$2:$B$4"
    assert "occupied_range" not in charts[0]


def test_copy_sheet_duplicates_embedded_charts_with_rewritten_references(chart_workbook):
    create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "bar", "E1", title="Revenue")

    result = copy_sheet(chart_workbook, "Sales", "Sales Copy")

    assert result["copied_charts"] == 1

    source_charts = list_charts(chart_workbook, sheet_name="Sales")
    copied_charts = list_charts(chart_workbook, sheet_name="Sales Copy")

    assert len(source_charts) == 1
    assert len(copied_charts) == 1
    assert copied_charts[0]["anchor"] == "E1"
    assert copied_charts[0]["series"][0]["title"] == "'Sales Copy'!B1"
    assert copied_charts[0]["series"][0]["categories"] == "'Sales Copy'!$A$2:$A$5"
    assert copied_charts[0]["series"][0]["values"] == "'Sales Copy'!$B$2:$B$5"
    assert source_charts[0]["series"][0]["values"] == "'Sales'!$B$2:$B$5"


def test_list_charts_can_filter_by_sheet(chart_workbook):
    wb = load_workbook(chart_workbook)
    ws = wb.create_sheet("Inventory")
    ws["A1"] = "Item"
    ws["B1"] = "Count"
    ws["A2"] = "Widget"
    ws["B2"] = 10
    ws["A3"] = "Gadget"
    ws["B3"] = 5
    wb.save(chart_workbook)
    wb.close()

    create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "line", "E1", title="Sales Revenue")
    create_chart_in_sheet(chart_workbook, "Inventory", "A1:B3", "bar", "E1", title="Inventory Count")

    charts = list_charts(chart_workbook, sheet_name="Inventory")

    assert len(charts) == 1
    assert charts[0]["sheet_name"] == "Inventory"
    assert charts[0]["title"] == "Inventory Count"


def test_list_charts_tool_returns_json_envelope(chart_workbook):
    create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "bar", "E1", title="Revenue")

    payload = _load_tool_payload(list_charts_tool(chart_workbook))

    assert payload["operation"] == "list_charts"
    assert payload["data"]["charts"][0]["title"] == "Revenue"


def test_find_free_canvas_tool_returns_json_envelope(chart_workbook):
    create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "bar", "E1", width=10, height=6)

    payload = _load_tool_payload(
        find_free_canvas_tool(chart_workbook, "Sales", width=10, height=6, limit=2)
    )

    assert payload["operation"] == "find_free_canvas"
    assert len(payload["data"]["suggestions"]) == 2


def test_create_chart_from_series_tool_returns_json_envelope(chart_workbook):
    payload = _load_tool_payload(
        create_chart_from_series_tool(
            chart_workbook,
            "Sales",
            "scatter",
            "J1",
            [{"title": "Revenue vs Cost", "x_range": "B2:B5", "y_range": "C2:C5"}],
            title="Scatter",
        )
    )

    assert payload["operation"] == "create_chart_from_series"
    assert payload["data"]["details"]["series_count"] == 1


def test_create_chart_tool_accepts_explicit_series(chart_workbook):
    payload = _load_tool_payload(
        create_chart_tool(
            chart_workbook,
            "Sales",
            "scatter",
            "L1",
            series=[{"title": "Revenue vs Cost", "x_range": "B2:B5", "y_range": "C2:C5"}],
            title="Unified Scatter",
        )
    )

    assert payload["operation"] == "create_chart"
    assert payload["data"]["details"]["series_count"] == 1
    assert payload["data"]["details"]["type"] == "scatter"


def test_create_chart_tool_accepts_top_level_dimensions(chart_workbook):
    payload = _load_tool_payload(
        create_chart_tool(
            chart_workbook,
            "Sales",
            "bar",
            "L1",
            data_range="A1:B5",
            width=13,
            height=9,
        )
    )

    assert payload["operation"] == "create_chart"
    assert payload["data"]["details"]["width"] == 13
    assert payload["data"]["details"]["height"] == 9


def test_create_chart_tool_accepts_placement_without_target_cell(chart_workbook):
    payload = _load_tool_payload(
        create_chart_tool(
            chart_workbook,
            "Sales",
            "bar",
            data_range="A1:B5",
            placement={"relative_to": "data_range", "direction": "right", "padding_columns": 2},
        )
    )

    assert payload["operation"] == "create_chart"
    assert payload["data"]["details"]["location"] == "E1"
    assert payload["data"]["details"]["placement"]["relative_to"] == "data_range"


# --- Error cases ---

def test_chart_invalid_sheet(chart_workbook):
    with pytest.raises(ValidationError, match="not found"):
        create_chart_in_sheet(chart_workbook, "NoSheet", "A1:B5", "bar", "E1")


def test_chart_unsupported_type(chart_workbook):
    with pytest.raises(ValidationError, match="Unsupported chart type"):
        create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "radar", "E1")


def test_chart_invalid_data_range(chart_workbook):
    with pytest.raises(ValidationError, match="Invalid data range"):
        create_chart_in_sheet(chart_workbook, "Sales", "ZZZ", "bar", "E1")


def test_chart_rejects_invalid_dimensions(chart_workbook):
    with pytest.raises(ValidationError, match="width must be a positive number"):
        create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "bar", "E1", width=0)


def test_chart_rejects_boolean_dimensions(chart_workbook):
    with pytest.raises(ValidationError, match="width must be a positive number"):
        create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "bar", "E1", width=True)


def test_find_free_canvas_rejects_boolean_dimensions(chart_workbook):
    with pytest.raises(ValidationError, match="width must be a positive number"):
        find_free_canvas_slots(chart_workbook, "Sales", width=True, height=5, limit=1)


def test_chart_rejects_both_data_range_and_series(chart_workbook):
    with pytest.raises(ValidationError, match="either data_range or series, not both"):
        create_chart_in_sheet(
            chart_workbook,
            "Sales",
            "A1:B5",
            "bar",
            "E1",
            series=[{"title": "Revenue", "values_range": "B2:B5"}],
        )


def test_chart_requires_data_range_or_series(chart_workbook):
    with pytest.raises(ValidationError, match="Either data_range or series is required"):
        create_chart_in_sheet(chart_workbook, "Sales", None, "bar", "E1")


def test_chart_invalid_target_cell(chart_workbook):
    with pytest.raises((ValidationError, ChartError)):
        create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "bar", "")


def test_chart_rejects_target_cell_and_placement_together(chart_workbook):
    with pytest.raises(ValidationError, match="either target_cell or placement, not both"):
        create_chart_in_sheet(
            chart_workbook,
            "Sales",
            "A1:B5",
            "bar",
            "E1",
            placement={"relative_to": "data_range", "direction": "right"},
        )


def test_chart_cross_sheet_reference_invalid(chart_workbook):
    with pytest.raises(ValidationError, match="not found"):
        create_chart_in_sheet(chart_workbook, "Sales", "Missing!A1:B5", "bar", "E1")


def test_chart_accepts_quoted_sheet_name_with_apostrophe(tmp_path):
    from openpyxl import Workbook

    filepath = str(tmp_path / "quoted-sheet-chart.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Bob's Data"
    ws.append(["Label", "Value"])
    ws.append(["A", 10])
    ws.append(["B", 20])
    wb.save(filepath)
    wb.close()

    result = create_chart_in_sheet(
        filepath,
        "Bob's Data",
        "'Bob''s Data'!A1:B3",
        "bar",
        "D1",
    )

    assert result["details"]["location"] == "D1"


@pytest.mark.parametrize(
    ("chart_type", "data_range", "message"),
    [
        ("bar", "A1:A5", "category column and at least one value column"),
        ("scatter", "A1:A5", "X column and at least one Y series column"),
        ("bar", "A1:B1", "header row and at least one data row"),
        ("scatter", "A1:B1", "header row and at least one data row"),
    ],
)
def test_chart_rejects_contiguous_ranges_that_cannot_form_a_series(
    chart_workbook,
    chart_type,
    data_range,
    message,
):
    with pytest.raises(ValidationError, match=message):
        create_chart_in_sheet(chart_workbook, "Sales", data_range, chart_type, "E1")


def test_create_chart_from_series_rejects_missing_scatter_axis(chart_workbook):
    with pytest.raises(ValidationError, match="requires both x_range and y_range"):
        create_chart_from_series(
            chart_workbook,
            "Sales",
            "scatter",
            "J1",
            [{"title": "Broken", "x_range": "B2:B5"}],
        )


def test_create_chart_from_series_rejects_multiple_pie_series(chart_workbook):
    with pytest.raises(ValidationError, match="Pie charts require exactly one series"):
        create_chart_from_series(
            chart_workbook,
            "Sales",
            "pie",
            "J1",
            [
                {"title": "Revenue", "values_range": "B2:B5"},
                {"title": "Cost", "values_range": "C2:C5"},
            ],
            categories_range="A2:A5",
        )

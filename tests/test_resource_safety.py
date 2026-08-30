import os
from multiprocessing import get_context
from pathlib import Path
from queue import Empty
from threading import Event, Thread

import pytest
from openpyxl import load_workbook

import excel_mcp.workbook as workbook_module
from excel_mcp.exceptions import WorkbookError
from excel_mcp.workbook import create_workbook_snapshot, safe_workbook


def _process_workbook_writer(
    filepath: str,
    cell: str,
    value: str,
    entered,
    release,
    errors,
) -> None:
    try:
        with safe_workbook(filepath, save=True) as wb:
            wb["Sheet1"][cell] = value
            entered.set()
            if release is not None and not release.wait(timeout=10):
                raise TimeoutError("writer release was not signaled")
    except BaseException as exc:  # pragma: no cover - surfaced in parent process
        errors.put(f"{type(exc).__name__}: {exc!s}")


def _assert_no_sheetforge_artifacts(workbook_path: Path) -> None:
    temp_leftovers = list(
        workbook_path.parent.glob(
            f".{workbook_path.stem}.sheetforge-*{workbook_path.suffix or '.xlsx'}"
        )
    )
    backup_leftovers = list(
        workbook_path.parent.glob(f".{workbook_path.name}.sheetforge-backup-*.bak")
    )
    assert temp_leftovers == []
    assert backup_leftovers == []


def test_safe_workbook_closes_on_success(tmp_workbook):
    """Workbook should be closed after successful context manager exit."""
    with safe_workbook(tmp_workbook) as wb:
        assert "Sheet1" in wb.sheetnames


def test_safe_workbook_closes_on_error(tmp_workbook):
    """Workbook should be closed even when an exception occurs."""
    try:
        with safe_workbook(tmp_workbook) as wb:
            raise ValueError("simulated error")
    except ValueError:
        pass
    with safe_workbook(tmp_workbook) as wb:
        assert "Sheet1" in wb.sheetnames


def test_snapshot_verification_failure_removes_destination_and_temp_artifacts(
    tmp_workbook,
    tmp_path,
    monkeypatch,
):
    snapshot_path = tmp_path / "failed-snapshot.xlsx"
    original_verify = workbook_module._verify_saved_workbook

    def fail_destination_verification(filepath: str) -> None:
        if Path(filepath) == snapshot_path:
            raise ValueError("simulated snapshot verification failure")
        original_verify(filepath)

    monkeypatch.setattr(
        workbook_module,
        "_verify_saved_workbook",
        fail_destination_verification,
    )

    with pytest.raises(WorkbookError, match="simulated snapshot verification failure"):
        create_workbook_snapshot(tmp_workbook, str(snapshot_path))

    assert not snapshot_path.exists()
    assert list(tmp_path.glob(".failed-snapshot.sheetforge-snapshot-*.xlsx")) == []


def test_safe_workbook_saves_when_requested(tmp_workbook):
    """Workbook should save changes when save=True."""
    with safe_workbook(tmp_workbook, save=True) as wb:
        ws = wb["Sheet1"]
        ws["D1"] = "NewColumn"

    with safe_workbook(tmp_workbook) as wb:
        assert wb["Sheet1"]["D1"].value == "NewColumn"


def test_safe_workbook_does_not_save_on_error(tmp_workbook):
    """save=True should only persist changes when the block exits successfully."""
    try:
        with safe_workbook(tmp_workbook, save=True) as wb:
            wb["Sheet1"]["D1"] = "UnsavedColumn"
            raise ValueError("simulated error")
    except ValueError:
        pass

    with safe_workbook(tmp_workbook) as wb:
        assert wb["Sheet1"]["D1"].value is None


def test_safe_workbook_atomic_save_leaves_no_temp_artifacts(tmp_workbook):
    workbook_path = Path(tmp_workbook)

    with safe_workbook(tmp_workbook, save=True) as wb:
        wb["Sheet1"]["D1"] = "NewColumn"

    _assert_no_sheetforge_artifacts(workbook_path)


def test_safe_workbook_raises_workbook_error_on_post_save_verify_failure(tmp_workbook, monkeypatch):
    original_verify = workbook_module._verify_saved_workbook
    verification_calls = 0

    def _fail_after_replace(filepath: str) -> None:
        nonlocal verification_calls
        verification_calls += 1
        if verification_calls == 2:
            raise OSError("verification failed")
        original_verify(filepath)

    monkeypatch.setattr(
        workbook_module,
        "_verify_saved_workbook",
        _fail_after_replace,
    )

    with pytest.raises(WorkbookError, match="verification failed"):
        with safe_workbook(tmp_workbook, save=True) as wb:
            wb["Sheet1"]["D1"] = "NewColumn"

    workbook_path = Path(tmp_workbook)
    _assert_no_sheetforge_artifacts(workbook_path)

    with safe_workbook(tmp_workbook) as wb:
        assert wb["Sheet1"]["D1"].value is None


def test_safe_workbook_retains_recovery_backup_when_rollback_fails(
    tmp_workbook,
    monkeypatch,
):
    workbook_path = Path(tmp_workbook)
    original_verify = workbook_module._verify_saved_workbook
    original_replace = workbook_module.os.replace
    verification_calls = 0
    replace_calls = 0

    def _fail_after_replace(filepath: str) -> None:
        nonlocal verification_calls
        verification_calls += 1
        if verification_calls == 2:
            raise OSError("verification failed")
        original_verify(filepath)

    def _fail_rollback(source, destination) -> None:
        nonlocal replace_calls
        replace_calls += 1
        if replace_calls == 2:
            raise OSError("rollback failed")
        original_replace(source, destination)

    monkeypatch.setattr(workbook_module, "_verify_saved_workbook", _fail_after_replace)
    monkeypatch.setattr(workbook_module.os, "replace", _fail_rollback)

    with pytest.raises(WorkbookError, match="Recovery backup retained at"):
        with safe_workbook(tmp_workbook, save=True) as wb:
            wb["Sheet1"]["D1"] = "NewColumn"

    backups = list(
        workbook_path.parent.glob(f".{workbook_path.name}.sheetforge-backup-*.bak")
    )
    assert len(backups) == 1
    with backups[0].open("rb") as backup_file:
        backup_wb = load_workbook(backup_file)
        assert backup_wb["Sheet1"]["D1"].value is None
        backup_wb.close()

    backups[0].unlink()
    _assert_no_sheetforge_artifacts(workbook_path)


def test_safe_workbook_cleanup_failure_does_not_report_committed_save_as_failed(
    tmp_workbook,
    monkeypatch,
):
    workbook_path = Path(tmp_workbook)
    original_unlink = Path.unlink

    def _fail_backup_cleanup(path, *args, **kwargs):
        if ".sheetforge-backup-" in path.name:
            raise PermissionError("simulated cleanup failure")
        return original_unlink(path, *args, **kwargs)

    monkeypatch.setattr(Path, "unlink", _fail_backup_cleanup)

    with safe_workbook(tmp_workbook, save=True) as wb:
        wb["Sheet1"]["D1"] = "Committed"

    monkeypatch.setattr(Path, "unlink", original_unlink)
    with safe_workbook(tmp_workbook) as wb:
        assert wb["Sheet1"]["D1"].value == "Committed"

    backups = list(
        workbook_path.parent.glob(f".{workbook_path.name}.sheetforge-backup-*.bak")
    )
    assert len(backups) == 1
    backups[0].unlink()
    _assert_no_sheetforge_artifacts(workbook_path)


def test_safe_workbook_serializes_concurrent_mutations(tmp_workbook):
    first_inside = Event()
    release_first = Event()
    second_inside = Event()
    errors: list[BaseException] = []

    def first_writer() -> None:
        try:
            with safe_workbook(tmp_workbook, save=True) as wb:
                wb["Sheet1"]["B1"] = "left"
                first_inside.set()
                assert release_first.wait(timeout=5)
        except BaseException as exc:  # pragma: no cover - surfaced below
            errors.append(exc)

    def second_writer() -> None:
        try:
            assert first_inside.wait(timeout=5)
            with safe_workbook(tmp_workbook, save=True) as wb:
                second_inside.set()
                wb["Sheet1"]["C1"] = "right"
        except BaseException as exc:  # pragma: no cover - surfaced below
            errors.append(exc)

    first_thread = Thread(target=first_writer)
    second_thread = Thread(target=second_writer)
    first_thread.start()
    second_thread.start()

    assert first_inside.wait(timeout=5)
    assert not second_inside.wait(timeout=0.2)
    release_first.set()
    first_thread.join(timeout=5)
    second_thread.join(timeout=5)

    assert not first_thread.is_alive()
    assert not second_thread.is_alive()
    assert errors == []
    assert second_inside.is_set()

    with safe_workbook(tmp_workbook) as wb:
        assert wb["Sheet1"]["B1"].value == "left"
        assert wb["Sheet1"]["C1"].value == "right"


def test_safe_workbook_serializes_mutations_across_processes(tmp_workbook):
    context = get_context("spawn")
    first_entered = context.Event()
    release_first = context.Event()
    second_entered = context.Event()
    errors = context.Queue()

    first_process = context.Process(
        target=_process_workbook_writer,
        args=(tmp_workbook, "B1", "left", first_entered, release_first, errors),
    )
    second_process = context.Process(
        target=_process_workbook_writer,
        args=(tmp_workbook, "C1", "right", second_entered, None, errors),
    )

    first_process.start()
    assert first_entered.wait(timeout=10)
    second_process.start()
    assert not second_entered.wait(timeout=0.3)
    release_first.set()

    first_process.join(timeout=15)
    second_process.join(timeout=15)
    assert not first_process.is_alive()
    assert not second_process.is_alive()
    assert first_process.exitcode == 0
    assert second_process.exitcode == 0

    process_errors: list[str] = []
    while True:
        try:
            process_errors.append(errors.get_nowait())
        except Empty:
            break
    assert process_errors == []

    with safe_workbook(tmp_workbook) as wb:
        assert wb["Sheet1"]["B1"].value == "left"
        assert wb["Sheet1"]["C1"].value == "right"


def test_same_process_workbook_lock_honors_timeout(tmp_workbook):
    workbook_path = Path(tmp_workbook).resolve()
    timed_out = Event()
    errors: list[BaseException] = []

    def competing_writer() -> None:
        try:
            with workbook_module._exclusive_workbook_lock(workbook_path, timeout=0.05):
                errors.append(AssertionError("competing writer unexpectedly acquired lock"))
        except WorkbookError as exc:
            if "Timed out waiting" in str(exc):
                timed_out.set()
            else:  # pragma: no cover - surfaced below
                errors.append(exc)

    with workbook_module._exclusive_workbook_lock(workbook_path):
        thread = Thread(target=competing_writer)
        thread.start()
        thread.join(timeout=2)

    assert not thread.is_alive()
    assert timed_out.is_set()
    assert errors == []


def test_workbook_lock_directory_rejects_symlinks(tmp_path, monkeypatch):
    if not hasattr(os, "getuid"):
        pytest.skip("POSIX ownership checks are not available")

    real_directory = tmp_path / "redirected-locks"
    real_directory.mkdir()
    lock_root = tmp_path / f"sheetforge-workbook-locks-{os.getuid()}"
    try:
        lock_root.symlink_to(real_directory, target_is_directory=True)
    except OSError:
        pytest.skip("Symlinks are not available on this platform")

    monkeypatch.setattr(workbook_module.tempfile, "gettempdir", lambda: str(tmp_path))

    with pytest.raises(WorkbookError, match="must be a real directory"):
        workbook_module._workbook_lock_path(tmp_path / "report.xlsx")


def test_safe_workbook_preserves_symlink_when_saving(tmp_workbook, tmp_path):
    workbook_path = Path(tmp_workbook)
    alias_path = tmp_path / "alias.xlsx"
    try:
        alias_path.symlink_to(workbook_path)
    except OSError:
        pytest.skip("Symlinks are not available on this platform")

    with safe_workbook(str(alias_path), save=True) as wb:
        wb["Sheet1"]["D1"] = "via alias"

    assert alias_path.is_symlink()
    with safe_workbook(tmp_workbook) as wb:
        assert wb["Sheet1"]["D1"].value == "via alias"

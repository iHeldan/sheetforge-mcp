import json

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_mcp.server import (
    list_tables as list_tables_tool,
    read_excel_table as read_excel_table_tool,
    upsert_excel_table_rows as upsert_excel_table_rows_tool,
)
from excel_mcp.tables import (
    create_excel_table,
    list_excel_tables,
    read_excel_table,
    upsert_excel_table_rows,
)
from excel_mcp.exceptions import DataError


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    assert "operation" in payload
    return payload


def _create_workbook_with_table_and_chartsheet(tmp_path) -> str:
    filepath = str(tmp_path / "table-and-chart-sheet.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Name", "Value"), ("A", 1), ("B", 2)]:
        ws.append(row)

    table = Table(displayName="Metrics", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb.save(filepath)
    wb.close()
    return filepath


def test_create_table_with_auto_name(tmp_workbook):
    result = create_excel_table(tmp_workbook, "Sheet1", "A1:C6")
    assert "Successfully created table" in result["message"]
    assert result["range"] == "A1:C6"
    assert result["table_name"].startswith("Table_")


def test_create_table_with_custom_name(tmp_workbook):
    result = create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="MyTable")
    assert result["table_name"] == "MyTable"

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    table_names = [t.displayName for t in ws.tables.values()]
    assert "MyTable" in table_names
    wb.close()


def test_create_table_with_custom_style(tmp_workbook):
    result = create_excel_table(
        tmp_workbook, "Sheet1", "A1:C6", table_style="TableStyleLight1"
    )
    assert "Successfully created table" in result["message"]


def test_create_table_persists_to_file(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="PersistTest")

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert len(ws.tables) == 1
    wb.close()


def test_create_table_sheet_not_found(tmp_workbook):
    with pytest.raises(DataError, match="not found"):
        create_excel_table(tmp_workbook, "NoSheet", "A1:C6")


def test_create_table_duplicate_name(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C3", table_name="Dupl")
    with pytest.raises(DataError):
        create_excel_table(tmp_workbook, "Sheet1", "A1:C3", table_name="Dupl")


def test_list_tables_returns_created_table(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers", table_style="TableStyleLight1")

    result = list_excel_tables(tmp_workbook)

    assert result == [
        {
            "sheet_name": "Sheet1",
            "table_name": "Customers",
            "range": "A1:C6",
            "style": "TableStyleLight1",
            "headers": ["Name", "Age", "City"],
            "column_count": 3,
            "data_row_count": 5,
            "header_row_count": 1,
            "totals_row_count": 0,
            "totals_row_shown": False,
            "show_first_column": False,
            "show_last_column": False,
            "show_row_stripes": True,
            "show_column_stripes": False,
        }
    ]


def test_list_tables_can_filter_by_sheet(multi_sheet_workbook):
    create_excel_table(multi_sheet_workbook, "Sales", "A1:B2", table_name="SalesTable")
    create_excel_table(multi_sheet_workbook, "Inventory", "A1:B2", table_name="InventoryTable")

    result = list_excel_tables(multi_sheet_workbook, sheet_name="Inventory")

    assert result == [
        {
            "sheet_name": "Inventory",
            "table_name": "InventoryTable",
            "range": "A1:B2",
            "style": "TableStyleMedium9",
            "headers": ["Item", "Count"],
            "column_count": 2,
            "data_row_count": 1,
            "header_row_count": 1,
            "totals_row_count": 0,
            "totals_row_shown": False,
            "show_first_column": False,
            "show_last_column": False,
            "show_row_stripes": True,
            "show_column_stripes": False,
        }
    ]


def test_list_tables_tool_returns_json_envelope(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    payload = _load_tool_payload(list_tables_tool(tmp_workbook))

    assert payload["operation"] == "list_tables"
    assert payload["data"]["tables"][0]["table_name"] == "Customers"


def test_list_excel_tables_skips_chart_sheets_when_scanning_workbook(tmp_path):
    filepath = _create_workbook_with_table_and_chartsheet(tmp_path)

    result = list_excel_tables(filepath)

    assert len(result) == 1
    assert result[0]["sheet_name"] == "Data"
    assert result[0]["table_name"] == "Metrics"


def test_list_excel_tables_returns_empty_for_chart_sheet(tmp_path):
    filepath = _create_workbook_with_table_and_chartsheet(tmp_path)

    result = list_excel_tables(filepath, sheet_name="Charts")

    assert result == []


def test_list_tables_tool_returns_empty_for_chart_sheet(tmp_path):
    filepath = _create_workbook_with_table_and_chartsheet(tmp_path)

    payload = _load_tool_payload(list_tables_tool(filepath, "Charts"))

    assert payload["operation"] == "list_tables"
    assert payload["data"] == {"sheet_name": "Charts", "tables": []}


def test_read_excel_table_returns_rows_by_table_name(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    result = read_excel_table(tmp_workbook, "Customers")

    assert result == {
        "sheet_name": "Sheet1",
        "table_name": "Customers",
        "range": "A1:C6",
        "style": "TableStyleMedium9",
        "headers": ["Name", "Age", "City"],
        "rows": [
            ["Alice", 30, "Helsinki"],
            ["Bob", 25, "Tampere"],
            ["Carol", 35, "Turku"],
            ["Dave", 28, "Oulu"],
            ["Eve", 32, "Espoo"],
        ],
        "total_rows": 5,
        "truncated": False,
        "header_row_count": 1,
        "totals_row_count": 0,
        "totals_row_shown": False,
    }


def test_read_excel_table_supports_max_rows_and_compact(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    result = read_excel_table(tmp_workbook, "Customers", max_rows=2, compact=True)

    assert result == {
        "sheet_name": "Sheet1",
        "table_name": "Customers",
        "range": "A1:C6",
        "headers": ["Name", "Age", "City"],
        "rows": [
            ["Alice", 30, "Helsinki"],
            ["Bob", 25, "Tampere"],
        ],
        "total_rows": 5,
        "truncated": True,
    }


def test_read_excel_table_can_return_records_and_schema(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    result = read_excel_table(
        tmp_workbook,
        "Customers",
        compact=True,
        row_mode="objects",
        infer_schema=True,
    )

    assert result["records"][0] == {
        "name": "Alice",
        "age": 30,
        "city": "Helsinki",
    }
    assert result["schema"][1] == {
        "field": "age",
        "header": "Age",
        "type": "integer",
        "nullable": False,
    }
    assert result["row_mode"] == "objects"
    assert "rows" not in result


def test_read_excel_table_can_filter_by_sheet(multi_sheet_workbook):
    create_excel_table(multi_sheet_workbook, "Sales", "A1:B2", table_name="SalesTable")
    create_excel_table(multi_sheet_workbook, "Inventory", "A1:B2", table_name="InventoryTable")

    result = read_excel_table(multi_sheet_workbook, "InventoryTable", sheet_name="Inventory")

    assert result["sheet_name"] == "Inventory"
    assert result["headers"] == ["Item", "Count"]
    assert result["rows"] == [["Widget", 42]]


def test_read_excel_table_raises_for_missing_table(tmp_workbook):
    with pytest.raises(DataError, match="Table 'Missing' not found"):
        read_excel_table(tmp_workbook, "Missing")


def test_read_excel_table_reports_missing_table_for_chart_sheet(tmp_path):
    filepath = _create_workbook_with_table_and_chartsheet(tmp_path)

    with pytest.raises(DataError, match="Table 'Missing' not found in sheet 'Charts'."):
        read_excel_table(filepath, "Missing", sheet_name="Charts")


def test_read_excel_table_tool_returns_json_envelope(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    payload = _load_tool_payload(read_excel_table_tool(tmp_workbook, "Customers", max_rows=2, compact=True))

    assert payload["operation"] == "read_excel_table"
    assert payload["data"]["table_name"] == "Customers"
    assert payload["data"]["truncated"] is True


def test_read_excel_table_tool_can_return_records_and_schema(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    payload = _load_tool_payload(
        read_excel_table_tool(
            tmp_workbook,
            "Customers",
            compact=True,
            row_mode="objects",
            infer_schema=True,
        )
    )

    assert payload["operation"] == "read_excel_table"
    assert payload["data"]["records"][0]["name"] == "Alice"
    assert payload["data"]["schema"][2]["field"] == "city"


def test_upsert_excel_table_rows_updates_and_appends(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    result = upsert_excel_table_rows(
        tmp_workbook,
        "Customers",
        key_column="Name",
        rows=[
            {"Name": "Alice", "Age": 31},
            {"Name": "Frank", "Age": 29, "City": "Lahti"},
        ],
    )

    assert result["updated_rows"] == 1
    assert result["appended_rows"] == 1
    assert result["updated_keys"] == ["Alice"]
    assert result["appended_keys"] == ["Frank"]
    assert result["previous_table_range"] == "A1:C6"
    assert result["table_range"] == "A1:C7"

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws["B2"].value == 31
    assert ws["A7"].value == "Frank"
    assert ws["B7"].value == 29
    assert ws["C7"].value == "Lahti"
    assert ws.tables["Customers"].ref == "A1:C7"
    wb.close()


def test_upsert_excel_table_rows_rejects_occupied_space_below_table(tmp_workbook):
    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    ws["A7"] = "Occupied"
    wb.save(tmp_workbook)
    wb.close()

    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    with pytest.raises(DataError, match="Cannot expand table into occupied cells"):
        upsert_excel_table_rows(
            tmp_workbook,
            "Customers",
            key_column="Name",
            rows=[{"Name": "Frank", "Age": 29, "City": "Lahti"}],
        )


def test_upsert_excel_table_rows_rejects_appends_when_totals_row_is_enabled(tmp_path):
    filepath = str(tmp_path / "totals.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Qty"])
    ws.append(["Alice", 1])
    ws.append(["Bob", 2])
    ws.append(["Total", 3])

    table = Table(displayName="Sales", ref="A1:B4")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.totalsRowShown = True
    table.totalsRowCount = 1
    ws.add_table(table)
    ws["A5"] = "Below"
    ws["B5"] = 99
    wb.save(filepath)
    wb.close()

    with pytest.raises(DataError, match="totals row is enabled"):
        upsert_excel_table_rows(
            filepath,
            "Sales",
            key_column="Name",
            rows=[{"Name": "Carol", "Qty": 5}],
        )

    wb = load_workbook(filepath)
    ws = wb["Sheet1"]
    assert ws.tables["Sales"].ref == "A1:B4"
    assert ws["A4"].value == "Total"
    assert ws["A5"].value == "Below"
    wb.close()


def test_upsert_excel_table_rows_dry_run_rejects_totals_row_appends(tmp_path):
    filepath = str(tmp_path / "totals-dry-run.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Qty"])
    ws.append(["Alice", 1])
    ws.append(["Total", 1])

    table = Table(displayName="Sales", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.totalsRowShown = True
    table.totalsRowCount = 1
    ws.add_table(table)
    wb.save(filepath)
    wb.close()

    with pytest.raises(DataError, match="totals row is enabled"):
        upsert_excel_table_rows(
            filepath,
            "Sales",
            key_column="Name",
            rows=[{"Name": "Bob", "Qty": 2}],
            dry_run=True,
            include_changes=True,
        )


def test_upsert_excel_table_rows_tool_returns_json_envelope(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")

    payload = _load_tool_payload(
        upsert_excel_table_rows_tool(
            tmp_workbook,
            "Customers",
            "Name",
            [{"Name": "Frank", "Age": 29, "City": "Lahti"}],
        )
    )

    assert payload["operation"] == "upsert_excel_table_rows"
    assert payload["data"]["appended_rows"] == 1
    assert payload["data"]["table_range"] == "A1:C7"

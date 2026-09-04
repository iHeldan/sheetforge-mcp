import hashlib
import json
from pathlib import Path
import shutil

from openpyxl import load_workbook
import pytest

import excel_mcp.changeset as changeset_module
import excel_mcp.server as server_module
from excel_mcp.changeset import apply_workbook_changeset
from excel_mcp.exceptions import (
    DataError,
    PreconditionFailedError,
    ResponseTooLargeError,
    ValidationError,
    WorkbookError,
)


def _sha256(path: str | Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def _assert_no_changeset_artifacts(path: str | Path) -> None:
    workbook_path = Path(path)
    leftovers = list(
        workbook_path.parent.glob(
            f".{workbook_path.stem}.sheetforge-changeset-*{workbook_path.suffix}"
        )
    )
    assert leftovers == []


def test_changeset_preview_is_non_mutating_and_returns_commit_credentials(tmp_workbook):
    original_bytes = Path(tmp_workbook).read_bytes()
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {
                "sheet_name": "Sheet1",
                "start_cell": "D1",
                "data": [["Score"], [42]],
            },
        }
    ]
    assertions = [
        {
            "type": "cell_equals",
            "sheet_name": "Sheet1",
            "cell": "D2",
            "expected": 42,
        },
        {
            "type": "range_values_unchanged",
            "sheet_name": "Sheet1",
            "range_ref": "A1:C6",
        },
        {"type": "no_cell_ref_errors"},
    ]

    result = apply_workbook_changeset(
        tmp_workbook,
        operations,
        assertions,
        create_snapshot=False,
    )

    assert result["mode"] == "preview"
    assert result["persisted"] is False
    assert result["ready_to_commit"] is True
    assert result["expected_workbook_sha256"] == hashlib.sha256(original_bytes).hexdigest()
    assert result["changeset_token"].startswith("changeset_v1_")
    assert result["assertions_passed"] == 3
    assert result["diff"]["cell_changes"]["count"] == 2
    assert Path(tmp_workbook).read_bytes() == original_bytes
    _assert_no_changeset_artifacts(tmp_workbook)


def test_changeset_mcp_tool_defaults_to_preview(tmp_workbook):
    original_bytes = Path(tmp_workbook).read_bytes()
    payload = json.loads(
        server_module.apply_workbook_changeset(
            tmp_workbook,
            [
                {
                    "tool": "write_data_to_excel",
                    "args": {
                        "sheet_name": "Sheet1",
                        "start_cell": "D1",
                        "data": [["preview only"]],
                    },
                }
            ],
            create_snapshot=False,
        )
    )

    assert payload["ok"] is True
    assert payload["operation"] == "apply_workbook_changeset"
    assert payload["data"]["mode"] == "preview"
    assert payload["data"]["ready_to_commit"] is True
    assert Path(tmp_workbook).read_bytes() == original_bytes


def test_changeset_preview_supports_explicit_noncontiguous_chart_series(tmp_workbook):
    original_bytes = Path(tmp_workbook).read_bytes()
    result = apply_workbook_changeset(
        tmp_workbook,
        [
            {
                "tool": "create_chart",
                "args": {
                    "sheet_name": "Sheet1",
                    "chart_type": "bar",
                    "series": [
                        {"title": "Age", "values_range": "B2:B6"},
                        {"title": "City", "values_range": "C2:C6"},
                    ],
                    "categories_range": "A2:A6",
                    "target_cell": "E1",
                },
            }
        ],
        create_snapshot=False,
    )

    assert result["ready_to_commit"] is True
    assert result["diff"]["summary"]["chart_change_count"] == 1
    assert Path(tmp_workbook).read_bytes() == original_bytes


def test_changeset_commits_full_report_workflow_with_verified_snapshot(tmp_workbook):
    original_bytes = Path(tmp_workbook).read_bytes()
    operations = [
        {"tool": "create_worksheet", "args": {"sheet_name": "Report"}},
        {
            "tool": "write_data_to_excel",
            "args": {
                "sheet_name": "Report",
                "start_cell": "A1",
                "data": [["Region", "Revenue"], ["North", 120], ["South", 90]],
            },
        },
        {
            "tool": "format_ranges",
            "args": {
                "sheet_name": "Report",
                "ranges": [
                    {
                        "start_cell": "A1",
                        "end_cell": "B1",
                        "bold": True,
                        "bg_color": "#1F4E78",
                        "font_color": "#FFFFFF",
                    },
                    {
                        "start_cell": "B2",
                        "end_cell": "B3",
                        "number_format": "#,##0",
                    },
                ],
            },
        },
        {"tool": "freeze_panes", "args": {"sheet_name": "Report", "cell": "A2"}},
        {
            "tool": "set_autofilter",
            "args": {"sheet_name": "Report", "range_ref": "A1:B3"},
        },
        {
            "tool": "set_column_widths",
            "args": {"sheet_name": "Report", "widths": {"A": 18, "B": 14}},
        },
        {
            "tool": "set_row_heights",
            "args": {"sheet_name": "Report", "heights": {"1": 24}},
        },
        {
            "tool": "autofit_columns",
            "args": {
                "sheet_name": "Report",
                "columns": ["A", "B"],
                "min_width": 10,
                "max_width": 22,
            },
        },
        {
            "tool": "create_table",
            "args": {
                "sheet_name": "Report",
                "data_range": "A1:B3",
                "table_name": "RevenueReport",
            },
        },
        {
            "tool": "create_chart",
            "args": {
                "sheet_name": "Report",
                "data_range": "A1:B3",
                "chart_type": "bar",
                "target_cell": "D2",
                "title": "Revenue by region",
                "width": 12,
                "height": 7,
            },
        },
    ]
    assertions = [
        {
            "type": "sheet_exists",
            "sheet_name": "Report",
            "sheet_type": "worksheet",
        },
        {
            "type": "range_equals",
            "sheet_name": "Report",
            "range_ref": "A1:B3",
            "expected": [["Region", "Revenue"], ["North", 120], ["South", 90]],
        },
        {
            "type": "table_exists",
            "sheet_name": "Report",
            "table_name": "RevenueReport",
            "range_ref": "A1:B3",
        },
        {
            "type": "range_values_unchanged",
            "sheet_name": "Sheet1",
            "range_ref": "A1:C6",
        },
        {"type": "no_cell_ref_errors", "sheet_name": "Report"},
    ]

    preview = apply_workbook_changeset(tmp_workbook, operations, assertions)
    assert preview["ready_to_commit"] is True
    snapshot_path = Path(preview["snapshot"]["path"])
    assert preview["snapshot"]["status"] == "available"
    assert not snapshot_path.exists()

    committed = apply_workbook_changeset(
        tmp_workbook,
        operations,
        assertions,
        mode="commit",
        expected_workbook_sha256=preview["expected_workbook_sha256"],
        changeset_token=preview["changeset_token"],
    )

    assert committed["persisted"] is True
    assert committed["receipt"]["before_sha256"] == preview["expected_workbook_sha256"]
    assert committed["receipt"]["after_sha256"] == _sha256(tmp_workbook)
    assert committed["receipt"]["snapshot"]["status"] == "created"
    assert snapshot_path.read_bytes() == original_bytes
    assert committed["diff"]["cell_changes"]["count"] == 6

    wb = load_workbook(tmp_workbook)
    try:
        ws = wb["Report"]
        assert ws["B3"].value == 90
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == "A1:B3"
        assert "RevenueReport" in ws.tables
        assert len(ws._charts) == 1
        assert ws.row_dimensions[1].height == 24
    finally:
        wb.close()

    committed_bytes = Path(tmp_workbook).read_bytes()
    with pytest.raises(PreconditionFailedError) as exc_info:
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            assertions,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
        )
    assert exc_info.value.code == "stale_workbook"
    assert Path(tmp_workbook).read_bytes() == committed_bytes
    _assert_no_changeset_artifacts(tmp_workbook)


def test_changeset_rejects_stale_source_without_overwriting_intervening_edit(tmp_workbook):
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [["planned"]], "start_cell": "D1"},
        }
    ]
    preview = apply_workbook_changeset(
        tmp_workbook, operations, create_snapshot=False
    )

    wb = load_workbook(tmp_workbook)
    wb["Sheet1"]["E1"] = "human edit"
    wb.save(tmp_workbook)
    wb.close()
    intervening_bytes = Path(tmp_workbook).read_bytes()

    with pytest.raises(PreconditionFailedError) as exc_info:
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            create_snapshot=False,
        )

    assert exc_info.value.code == "stale_workbook"
    assert Path(tmp_workbook).read_bytes() == intervening_bytes
    _assert_no_changeset_artifacts(tmp_workbook)


def test_changeset_token_binds_exact_plan_and_target_path(tmp_workbook, tmp_path):
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [[1]], "start_cell": "D1"},
        }
    ]
    preview = apply_workbook_changeset(
        tmp_workbook, operations, create_snapshot=False
    )
    original_bytes = Path(tmp_workbook).read_bytes()

    changed_plan = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [[2]], "start_cell": "D1"},
        }
    ]
    with pytest.raises(PreconditionFailedError) as exc_info:
        apply_workbook_changeset(
            tmp_workbook,
            changed_plan,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            create_snapshot=False,
        )
    assert exc_info.value.code == "changeset_token_mismatch"
    assert Path(tmp_workbook).read_bytes() == original_bytes

    identical_path = tmp_path / "identical.xlsx"
    shutil.copy2(tmp_workbook, identical_path)
    with pytest.raises(PreconditionFailedError) as exc_info:
        apply_workbook_changeset(
            str(identical_path),
            operations,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            create_snapshot=False,
        )
    assert exc_info.value.code == "changeset_token_mismatch"
    assert identical_path.read_bytes() == original_bytes
    _assert_no_changeset_artifacts(tmp_workbook)
    _assert_no_changeset_artifacts(identical_path)


def test_changeset_failed_assertion_never_creates_snapshot_or_mutates_source(tmp_workbook):
    original_bytes = Path(tmp_workbook).read_bytes()
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [[10]], "start_cell": "D1"},
        }
    ]
    assertions = [
        {
            "type": "cell_equals",
            "sheet_name": "Sheet1",
            "cell": "D1",
            "expected": 99,
        }
    ]
    preview = apply_workbook_changeset(tmp_workbook, operations, assertions)
    snapshot_path = Path(preview["snapshot"]["path"])
    assert preview["ready_to_commit"] is False
    assert preview["assertions"][0]["passed"] is False

    with pytest.raises(PreconditionFailedError) as exc_info:
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            assertions,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
        )
    assert exc_info.value.code == "changeset_assertion_failed"
    assert Path(tmp_workbook).read_bytes() == original_bytes
    assert not snapshot_path.exists()
    _assert_no_changeset_artifacts(tmp_workbook)


def test_changeset_promotes_partial_batch_format_failure_to_transaction_failure(tmp_workbook):
    original_bytes = Path(tmp_workbook).read_bytes()
    operations = [
        {
            "tool": "format_ranges",
            "args": {
                "sheet_name": "Sheet1",
                "ranges": [
                    {"start_cell": "A1", "bold": True},
                    {"start_cell": "B1", "bg_color": "not-a-color"},
                ],
            },
        }
    ]

    with pytest.raises(ValidationError, match="failed atomically"):
        apply_workbook_changeset(
            tmp_workbook, operations, create_snapshot=False
        )

    assert Path(tmp_workbook).read_bytes() == original_bytes
    _assert_no_changeset_artifacts(tmp_workbook)


def test_changeset_snapshot_collision_is_reported_before_commit(tmp_workbook, tmp_path):
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [[1]], "start_cell": "D1"},
        }
    ]
    snapshot_path = tmp_path / "reserved.xlsx"
    snapshot_path.write_bytes(b"not the baseline")
    original_bytes = Path(tmp_workbook).read_bytes()

    preview = apply_workbook_changeset(
        tmp_workbook,
        operations,
        snapshot_filepath=str(snapshot_path),
    )
    assert preview["ready_to_commit"] is False
    assert preview["snapshot"]["status"] == "conflict"

    with pytest.raises(PreconditionFailedError) as exc_info:
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            snapshot_filepath=str(snapshot_path),
        )
    assert exc_info.value.code == "snapshot_conflict"
    assert Path(tmp_workbook).read_bytes() == original_bytes


def test_changeset_reuses_only_an_exact_matching_snapshot(tmp_workbook, tmp_path):
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [[1]], "start_cell": "D1"},
        }
    ]
    snapshot_path = tmp_path / "baseline.xlsx"
    original_bytes = Path(tmp_workbook).read_bytes()
    preview = apply_workbook_changeset(
        tmp_workbook,
        operations,
        snapshot_filepath=str(snapshot_path),
    )
    shutil.copy2(tmp_workbook, snapshot_path)

    committed = apply_workbook_changeset(
        tmp_workbook,
        operations,
        mode="commit",
        expected_workbook_sha256=preview["expected_workbook_sha256"],
        changeset_token=preview["changeset_token"],
        snapshot_filepath=str(snapshot_path),
    )

    assert committed["receipt"]["snapshot"]["status"] == "existing_matching"
    assert committed["receipt"]["snapshot"]["reused"] is True
    assert snapshot_path.read_bytes() == original_bytes


def test_changeset_rechecks_reused_snapshot_after_candidate_operations(
    tmp_workbook,
    tmp_path,
    monkeypatch,
):
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [[1]], "start_cell": "D1"},
        }
    ]
    snapshot_path = tmp_path / "baseline.xlsx"
    original_bytes = Path(tmp_workbook).read_bytes()
    preview = apply_workbook_changeset(
        tmp_workbook,
        operations,
        snapshot_filepath=str(snapshot_path),
    )
    shutil.copy2(tmp_workbook, snapshot_path)
    original_apply_operations = changeset_module._apply_operations

    def replace_snapshot_after_operations(*args, **kwargs):
        result = original_apply_operations(*args, **kwargs)
        snapshot_path.write_bytes(b"changed after initial snapshot check")
        return result

    monkeypatch.setattr(
        changeset_module,
        "_apply_operations",
        replace_snapshot_after_operations,
    )

    with pytest.raises(PreconditionFailedError) as exc_info:
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            snapshot_filepath=str(snapshot_path),
        )

    assert exc_info.value.code == "snapshot_conflict"
    assert Path(tmp_workbook).read_bytes() == original_bytes
    _assert_no_changeset_artifacts(tmp_workbook)


def test_changeset_preserves_source_symlink(tmp_workbook, tmp_path):
    alias = tmp_path / "alias.xlsx"
    try:
        alias.symlink_to(Path(tmp_workbook).name)
    except OSError:
        pytest.skip("Symlinks are not available on this platform")
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [["linked"]], "start_cell": "D1"},
        }
    ]
    preview = apply_workbook_changeset(
        str(alias), operations, create_snapshot=False
    )

    apply_workbook_changeset(
        str(alias),
        operations,
        mode="commit",
        expected_workbook_sha256=preview["expected_workbook_sha256"],
        changeset_token=preview["changeset_token"],
        create_snapshot=False,
    )

    assert alias.is_symlink()
    wb = load_workbook(tmp_workbook)
    try:
        assert wb["Sheet1"]["D1"].value == "linked"
    finally:
        wb.close()
    _assert_no_changeset_artifacts(tmp_workbook)


def test_changeset_rolls_back_exact_bytes_when_post_replace_verification_fails(
    tmp_workbook,
    monkeypatch,
):
    source = Path(tmp_workbook).resolve()
    original_bytes = source.read_bytes()
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [["new"]], "start_cell": "D1"},
        }
    ]
    preview = apply_workbook_changeset(
        tmp_workbook, operations, create_snapshot=False
    )
    original_verify = changeset_module.workbook_module._verify_saved_workbook
    failed_live_verification = False

    def fail_first_live_verification(filepath: str) -> None:
        nonlocal failed_live_verification
        if Path(filepath).resolve() == source and not failed_live_verification:
            failed_live_verification = True
            raise OSError("simulated post-replace verification failure")
        original_verify(filepath)

    monkeypatch.setattr(
        changeset_module.workbook_module,
        "_verify_saved_workbook",
        fail_first_live_verification,
    )

    with pytest.raises(WorkbookError, match="original workbook was restored"):
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            create_snapshot=False,
        )

    assert failed_live_verification is True
    assert source.read_bytes() == original_bytes
    _assert_no_changeset_artifacts(source)


def test_changeset_rolls_back_when_final_digest_cannot_be_read(
    tmp_workbook,
    monkeypatch,
):
    source = Path(tmp_workbook).resolve()
    original_bytes = source.read_bytes()
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [["new"]], "start_cell": "D1"},
        }
    ]
    preview = apply_workbook_changeset(
        tmp_workbook, operations, create_snapshot=False
    )
    original_sha256_file = changeset_module._sha256_file
    source_digest_calls = 0

    def fail_post_replace_digest(path: Path) -> str:
        nonlocal source_digest_calls
        if path.resolve() == source:
            source_digest_calls += 1
            if source_digest_calls == 3:
                raise OSError("simulated final digest failure")
        return original_sha256_file(path)

    monkeypatch.setattr(changeset_module, "_sha256_file", fail_post_replace_digest)

    with pytest.raises(WorkbookError, match="original workbook was restored"):
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            create_snapshot=False,
        )

    assert source_digest_calls == 4
    assert source.read_bytes() == original_bytes
    _assert_no_changeset_artifacts(source)


def test_changeset_enforces_response_budget_before_replacing_source(
    tmp_workbook,
    monkeypatch,
):
    original_bytes = Path(tmp_workbook).read_bytes()
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [["new"]], "start_cell": "D1"},
        }
    ]
    preview = apply_workbook_changeset(
        tmp_workbook, operations, create_snapshot=False
    )
    monkeypatch.setattr(changeset_module, "MAX_CHANGESET_RESPONSE_BYTES", 1)

    with pytest.raises(ResponseTooLargeError):
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            create_snapshot=False,
        )

    assert Path(tmp_workbook).read_bytes() == original_bytes
    _assert_no_changeset_artifacts(tmp_workbook)


def test_changeset_cleans_baseline_if_candidate_temp_creation_fails(
    tmp_workbook,
    monkeypatch,
):
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Sheet1", "data": [["new"]], "start_cell": "D1"},
        }
    ]
    preview = apply_workbook_changeset(
        tmp_workbook, operations, create_snapshot=False
    )
    original_new_temp = changeset_module._new_temp_workbook
    temp_calls = 0

    def fail_second_temp(*args, **kwargs):
        nonlocal temp_calls
        temp_calls += 1
        if temp_calls == 2:
            raise OSError("simulated candidate temp creation failure")
        return original_new_temp(*args, **kwargs)

    monkeypatch.setattr(changeset_module, "_new_temp_workbook", fail_second_temp)

    with pytest.raises(OSError, match="candidate temp"):
        apply_workbook_changeset(
            tmp_workbook,
            operations,
            mode="commit",
            expected_workbook_sha256=preview["expected_workbook_sha256"],
            changeset_token=preview["changeset_token"],
            create_snapshot=False,
        )

    _assert_no_changeset_artifacts(tmp_workbook)


@pytest.mark.parametrize(
    "operations,match",
    [
        (
            [{"tool": "delete_worksheet", "args": {"sheet_name": "Sheet1"}}],
            "Unsupported ChangeSet tool",
        ),
        (
            [
                {
                    "tool": "write_data_to_excel",
                    "args": {
                        "sheet_name": "Sheet1",
                        "data": [[1]],
                        "filepath": "other.xlsx",
                    },
                }
            ],
            "unsupported fields",
        ),
        (
            [
                {
                    "tool": "format_range",
                    "args": {
                        "sheet_name": "Sheet1",
                        "start_cell": "A1",
                        "end_cell": "B2",
                        "merge_cells": True,
                    },
                }
            ],
            "does not support merge_cells",
        ),
        (
            [
                {
                    "tool": "create_table",
                    "args": {
                        "sheet_name": "Sheet1",
                        "data_range": "A1:C6",
                        "table_name": "",
                    },
                }
            ],
            "table_name must be a non-empty string",
        ),
        (
            [
                {
                    "tool": "write_data_to_excel",
                    "args": {
                        "sheet_name": "Sheet1",
                        "data": [[1]],
                        "start_cell": "XFE1",
                    },
                }
            ],
            "must stay inside Excel's",
        ),
    ],
)
def test_changeset_rejects_unsafe_or_nondeterministic_operations(
    tmp_workbook,
    operations,
    match,
):
    original_bytes = Path(tmp_workbook).read_bytes()
    with pytest.raises(ValidationError, match=match):
        apply_workbook_changeset(
            tmp_workbook, operations, create_snapshot=False
        )
    assert Path(tmp_workbook).read_bytes() == original_bytes


def test_changeset_rejects_chart_sheet_for_cell_operations(complex_workbook):
    original_bytes = Path(complex_workbook).read_bytes()
    operations = [
        {
            "tool": "write_data_to_excel",
            "args": {"sheet_name": "Charts", "data": [[1]], "start_cell": "A1"},
        }
    ]

    with pytest.raises(DataError, match="chartsheet"):
        apply_workbook_changeset(
            complex_workbook, operations, create_snapshot=False
        )

    assert Path(complex_workbook).read_bytes() == original_bytes
    _assert_no_changeset_artifacts(complex_workbook)

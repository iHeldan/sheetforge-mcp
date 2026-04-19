import json

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName

from excel_mcp.server import freeze_panes as freeze_panes_tool
from excel_mcp.server import list_named_ranges as list_named_ranges_tool
from excel_mcp.server import merge_cells as merge_cells_tool
from excel_mcp.server import autofit_columns as autofit_columns_tool
from excel_mcp.server import describe_sheet_layout as describe_sheet_layout_tool
from excel_mcp.server import get_worksheet_protection as get_worksheet_protection_tool
from excel_mcp.server import set_print_area as set_print_area_tool
from excel_mcp.server import set_print_titles as set_print_titles_tool
from excel_mcp.server import read_range_formatting as read_range_formatting_tool
from excel_mcp.server import set_column_widths as set_column_widths_tool
from excel_mcp.server import set_row_heights as set_row_heights_tool
from excel_mcp.server import set_worksheet_protection as set_worksheet_protection_tool
from excel_mcp.server import set_worksheet_visibility as set_worksheet_visibility_tool
from excel_mcp.server import set_autofilter as set_autofilter_tool
from excel_mcp.server import unmerge_cells as unmerge_cells_tool
from excel_mcp.formatting import format_range as apply_range_formatting
from excel_mcp.formatting import read_range_formatting as read_range_formatting_impl
from excel_mcp.sheet import (
    delete_cols,
    delete_rows,
    delete_range_operation,
    autofit_columns,
    copy_range_operation,
    get_sheet_protection,
    insert_cols,
    insert_row,
    merge_range,
    set_auto_filter,
    set_column_widths,
    set_freeze_panes,
    set_print_area,
    set_print_titles,
    set_row_heights,
    set_sheet_protection,
    set_sheet_visibility,
    unmerge_range,
)
from excel_mcp.tables import create_excel_table
from excel_mcp.exceptions import ValidationError, WorkbookError
from excel_mcp.workbook import list_named_ranges
from excel_mcp.workbook import describe_sheet_layout as describe_sheet_layout_impl


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    assert "operation" in payload
    assert "message" in payload
    return payload


@pytest.fixture
def named_range_workbook(tmp_path):
    filepath = tmp_path / "named-ranges.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["A2"] = "Alice"
    ws["B2"] = 30
    wb.defined_names["PeopleTable"] = DefinedName("PeopleTable", attr_text="Sheet1!$A$1:$B$2")
    wb.save(filepath)
    wb.close()
    return str(filepath)


@pytest.fixture
def table_guard_workbook(tmp_path):
    filepath = str(tmp_path / "table-guard.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Value"])
    ws.append(["A", 1])
    ws.append(["B", 2])
    ws.append(["outside", "kept"])
    wb.save(filepath)
    wb.close()

    create_excel_table(filepath, "Data", "A1:B3", table_name="DataTable")
    return filepath


def test_set_freeze_panes_persists_value(tmp_workbook):
    result = set_freeze_panes(tmp_workbook, "Sheet1", "B2")
    assert result["freeze_panes"] == "B2"
    assert "changes" not in result

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].freeze_panes == "B2"
    wb.close()


def test_set_freeze_panes_dry_run_does_not_persist(tmp_workbook):
    result = set_freeze_panes(tmp_workbook, "Sheet1", "B2", dry_run=True)
    assert result["dry_run"] is True
    assert result["changes"][0]["new_value"] == "B2"

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].freeze_panes is None
    wb.close()


def test_merge_range_defaults_to_summary_without_changes(tmp_workbook):
    result = merge_range(tmp_workbook, "Sheet1", "A1", "B1")
    assert result["range"] == "A1:B1"
    assert "changes" not in result

    wb = load_workbook(tmp_workbook)
    assert any("A1:B1" in str(r) for r in wb["Sheet1"].merged_cells.ranges)
    wb.close()


def test_unmerge_range_can_include_changes_explicitly(tmp_workbook):
    merge_range(tmp_workbook, "Sheet1", "A1", "B1")

    result = unmerge_range(tmp_workbook, "Sheet1", "A1", "B1", include_changes=True)

    assert result["changes"][0]["range"] == "A1:B1"


def test_read_range_formatting_groups_styles_and_rules(tmp_workbook):
    apply_range_formatting(
        tmp_workbook,
        "Sheet1",
        "A1",
        "C1",
        bold=True,
        bg_color="FFF2CC",
    )
    apply_range_formatting(
        tmp_workbook,
        "Sheet1",
        "B2",
        "B3",
        number_format="0.0%",
        alignment="center",
        wrap_text=True,
        protection={"locked": False},
    )
    apply_range_formatting(
        tmp_workbook,
        "Sheet1",
        "C2",
        "C4",
        conditional_format={
            "type": "formula",
            "formula": ["$B2>20"],
        },
    )
    merge_range(tmp_workbook, "Sheet1", "A5", "B5")

    result = read_range_formatting_impl(tmp_workbook, "Sheet1", "A1:C5")

    assert result["range"] == "A1:C5"
    assert result["summary"]["style_group_count"] >= 2
    assert result["summary"]["has_conditional_formatting"] is True
    assert result["summary"]["has_merged_ranges"] is True
    assert result["merged_ranges"]["count"] == 1
    assert result["conditional_formats"]["count"] == 1
    assert any(group["font"]["bold"] is True for group in result["style_groups"])
    assert any(group.get("number_format") == "0.0%" for group in result["style_groups"])


def test_read_range_formatting_rejects_chartsheets(complex_workbook):
    with pytest.raises(ValidationError, match="chartsheet"):
        read_range_formatting_impl(complex_workbook, "Charts", "A1:B2")


def test_read_range_formatting_tool_returns_envelope(tmp_workbook):
    raw = read_range_formatting_tool(tmp_workbook, "Sheet1", "A1:B2", sample_limit=3)
    payload = _load_tool_payload(raw)

    assert payload["operation"] == "read_range_formatting"
    assert payload["data"]["range"] == "A1:B2"
    assert "style_groups" in payload["data"]


def test_describe_sheet_layout_summarizes_dashboard_structure(complex_workbook):
    result = describe_sheet_layout_impl(
        complex_workbook,
        "Dashboard",
        sample_limit=5,
        free_canvas_rows=4,
        free_canvas_cols=4,
        free_canvas_limit=2,
    )

    assert result["sheet_name"] == "Dashboard"
    assert result["freeze_panes"] == "A2"
    assert result["summary"]["chart_count"] == 1
    assert result["summary"]["merged_range_count"] == 1
    assert result["summary"]["data_validation_rule_count"] == 1
    assert result["summary"]["conditional_format_rule_count"] == 1
    assert result["charts"]["sample"][0]["anchor"] == "E2"
    assert result["free_canvas_preview"]["requested_block"] == {"rows": 4, "columns": 4}
    assert len(result["free_canvas_preview"]["suggestions"]) <= 2


def test_describe_sheet_layout_reports_custom_dimensions(tmp_workbook):
    set_column_widths(tmp_workbook, "Sheet1", {"A": 18})
    set_row_heights(tmp_workbook, "Sheet1", {"1": 24})

    result = describe_sheet_layout_impl(tmp_workbook, "Sheet1", sample_limit=5)

    assert result["custom_column_widths"]["count"] >= 1
    assert result["custom_row_heights"]["count"] >= 1


def test_describe_sheet_layout_rejects_chartsheets(complex_workbook):
    with pytest.raises(WorkbookError, match="chartsheet"):
        describe_sheet_layout_impl(complex_workbook, "Charts")


def test_describe_sheet_layout_tool_returns_envelope(complex_workbook):
    raw = describe_sheet_layout_tool(complex_workbook, "Dashboard", sample_limit=3)
    payload = _load_tool_payload(raw)

    assert payload["operation"] == "describe_sheet_layout"
    assert payload["data"]["sheet_name"] == "Dashboard"
    assert payload["data"]["summary"]["chart_count"] == 1


@pytest.mark.parametrize(
    ("operation", "args", "error_message"),
    [
        (insert_row, ("Sheet1", True), "start_row must be a positive integer"),
        (insert_cols, ("Sheet1", True), "start_col must be a positive integer"),
        (delete_rows, ("Sheet1", True), "start_row must be a positive integer"),
        (delete_cols, ("Sheet1", True), "start_col must be a positive integer"),
    ],
)
def test_row_and_column_operations_reject_boolean_indexes(
    tmp_workbook,
    operation,
    args,
    error_message,
):
    with pytest.raises(ValidationError, match=error_message):
        operation(tmp_workbook, *args, dry_run=True)


def test_set_autofilter_infers_used_range(tmp_workbook):
    result = set_auto_filter(tmp_workbook, "Sheet1")
    assert result["range"] == "A1:C6"
    assert "changes" not in result

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].auto_filter.ref == "A1:C6"
    wb.close()


def test_set_autofilter_dry_run_does_not_persist(tmp_workbook):
    result = set_auto_filter(tmp_workbook, "Sheet1", "A1:C3", dry_run=True)
    assert result["dry_run"] is True
    assert result["changes"][0]["new_value"] == "A1:C3"

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].auto_filter.ref is None
    wb.close()


def test_set_worksheet_visibility_persists_value(multi_sheet_workbook):
    result = set_sheet_visibility(multi_sheet_workbook, "Inventory", "hidden")
    assert result["visibility"] == "hidden"

    wb = load_workbook(multi_sheet_workbook)
    assert wb["Inventory"].sheet_state == "hidden"
    wb.close()


def test_set_worksheet_visibility_dry_run_does_not_persist(multi_sheet_workbook):
    result = set_sheet_visibility(multi_sheet_workbook, "Inventory", "veryHidden", dry_run=True)
    assert result["dry_run"] is True
    assert result["changes"][0]["new_value"] == "veryHidden"

    wb = load_workbook(multi_sheet_workbook)
    assert wb["Inventory"].sheet_state == "visible"
    wb.close()


def test_set_worksheet_visibility_rejects_hiding_only_visible_sheet(tmp_workbook):
    with pytest.raises(Exception, match="only visible sheet"):
        set_sheet_visibility(tmp_workbook, "Sheet1", "hidden")


def test_set_worksheet_visibility_counts_visible_chartsheets_when_hiding_worksheet(tmp_path):
    filepath = str(tmp_path / "visibility-with-chartsheet.xlsx")
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data.append(["Name", "Value"])
    data.append(["Alice", 10])
    data.append(["Bob", 12])

    chart = BarChart()
    values = Reference(data, min_col=2, min_row=1, max_row=3)
    categories = Reference(data, min_col=1, min_row=2, max_row=3)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    charts = wb.create_chartsheet("Charts")
    charts.add_chart(chart)
    wb.save(filepath)
    wb.close()

    result = set_sheet_visibility(filepath, "Data", "hidden")
    assert result["visibility"] == "hidden"

    wb = load_workbook(filepath)
    assert wb["Data"].sheet_state == "hidden"
    assert wb["Charts"].sheet_state == "visible"
    wb.close()


def test_set_worksheet_visibility_rejects_hiding_only_visible_chartsheet(tmp_path):
    filepath = str(tmp_path / "visibility-only-chartsheet.xlsx")
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data.append(["Name", "Value"])
    data.append(["Alice", 10])
    data.sheet_state = "hidden"

    chart = BarChart()
    values = Reference(data, min_col=2, min_row=1, max_row=2)
    categories = Reference(data, min_col=1, min_row=2, max_row=2)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    charts = wb.create_chartsheet("Charts")
    charts.add_chart(chart)
    wb.save(filepath)
    wb.close()

    with pytest.raises(Exception, match="only visible sheet"):
        set_sheet_visibility(filepath, "Charts", "hidden")


def test_copy_range_operation_uses_source_snapshot_for_overlapping_same_sheet_copy(tmp_path):
    filepath = str(tmp_path / "overlapping-copy.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "left"
    ws["B1"] = "middle"
    ws["C1"] = "right"
    wb.save(filepath)
    wb.close()

    result = copy_range_operation(filepath, "Sheet1", "A1", "B1", "B1")

    assert result["changes"] == [
        {
            "sheet_name": "Sheet1",
            "cell": "B1",
            "row": 1,
            "column": 2,
            "old_value": "middle",
            "new_value": "left",
            "source_cell": "A1",
        },
        {
            "sheet_name": "Sheet1",
            "cell": "C1",
            "row": 1,
            "column": 3,
            "old_value": "right",
            "new_value": "middle",
            "source_cell": "B1",
        },
    ]

    wb = load_workbook(filepath)
    ws = wb["Sheet1"]
    assert ws["A1"].value == "left"
    assert ws["B1"].value == "left"
    assert ws["C1"].value == "middle"
    wb.close()


def test_copy_range_operation_dry_run_preview_matches_overlapping_copy(tmp_path):
    filepath = str(tmp_path / "overlapping-copy-dry-run.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "left"
    ws["B1"] = "middle"
    ws["C1"] = "right"
    wb.save(filepath)
    wb.close()

    result = copy_range_operation(filepath, "Sheet1", "A1", "B1", "B1", dry_run=True)

    assert result["dry_run"] is True
    assert result["changes"][1]["new_value"] == "middle"

    wb = load_workbook(filepath)
    ws = wb["Sheet1"]
    assert ws["A1"].value == "left"
    assert ws["B1"].value == "middle"
    assert ws["C1"].value == "right"
    wb.close()


def test_delete_range_operation_shifts_only_selected_columns_up(tmp_path):
    filepath = str(tmp_path / "delete-range-up.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 5):
        ws[f"A{row}"] = f"A{row}"
        ws[f"B{row}"] = f"B{row}"
        ws[f"C{row}"] = f"C{row}"
        ws[f"D{row}"] = f"D{row}"
    wb.save(filepath)
    wb.close()

    result = delete_range_operation(filepath, "Sheet1", "B2", "C2", shift_direction="up")

    assert result["changes"] == [
        {
            "sheet_name": "Sheet1",
            "cell": "B2",
            "row": 2,
            "column": 2,
            "old_value": "B2",
            "new_value": "B3",
        },
        {
            "sheet_name": "Sheet1",
            "cell": "C2",
            "row": 2,
            "column": 3,
            "old_value": "C2",
            "new_value": "C3",
        },
        {
            "sheet_name": "Sheet1",
            "cell": "B3",
            "row": 3,
            "column": 2,
            "old_value": "B3",
            "new_value": "B4",
        },
        {
            "sheet_name": "Sheet1",
            "cell": "C3",
            "row": 3,
            "column": 3,
            "old_value": "C3",
            "new_value": "C4",
        },
        {
            "sheet_name": "Sheet1",
            "cell": "B4",
            "row": 4,
            "column": 2,
            "old_value": "B4",
            "new_value": None,
        },
        {
            "sheet_name": "Sheet1",
            "cell": "C4",
            "row": 4,
            "column": 3,
            "old_value": "C4",
            "new_value": None,
        },
    ]

    wb = load_workbook(filepath)
    ws = wb["Sheet1"]
    assert ws["A2"].value == "A2"
    assert ws["D2"].value == "D2"
    assert ws["B2"].value == "B3"
    assert ws["C2"].value == "C3"
    assert ws["B3"].value == "B4"
    assert ws["C3"].value == "C4"
    assert ws["B4"].value is None
    assert ws["C4"].value is None
    wb.close()


def test_delete_range_operation_dry_run_preview_matches_column_scoped_shift(tmp_path):
    filepath = str(tmp_path / "delete-range-up-dry-run.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 5):
        ws[f"A{row}"] = f"A{row}"
        ws[f"B{row}"] = f"B{row}"
        ws[f"C{row}"] = f"C{row}"
        ws[f"D{row}"] = f"D{row}"
    wb.save(filepath)
    wb.close()

    result = delete_range_operation(filepath, "Sheet1", "B2", "C2", shift_direction="up", dry_run=True)

    assert result["dry_run"] is True
    assert result["changes"][0]["cell"] == "B2"
    assert result["changes"][0]["new_value"] == "B3"
    assert result["changes"][-1] == {
        "sheet_name": "Sheet1",
        "cell": "C4",
        "row": 4,
        "column": 3,
        "old_value": "C4",
        "new_value": None,
    }

    wb = load_workbook(filepath)
    ws = wb["Sheet1"]
    assert ws["A2"].value == "A2"
    assert ws["B2"].value == "B2"
    assert ws["C2"].value == "C2"
    assert ws["D2"].value == "D2"
    wb.close()


def test_get_worksheet_protection_reports_defaults(tmp_workbook):
    result = get_sheet_protection(tmp_workbook, "Sheet1")
    assert result["enabled"] is False
    assert result["password_protected"] is False
    assert "formatCells" in result["options"]


def test_set_worksheet_protection_persists_state(tmp_workbook):
    result = set_sheet_protection(
        tmp_workbook,
        "Sheet1",
        enabled=True,
        password="secret",
        options={"selectUnlockedCells": True, "formatCells": False},
    )
    assert result["enabled"] is True
    assert result["password_protected"] is True
    assert result["options"]["selectUnlockedCells"] is True
    assert result["options"]["formatCells"] is False

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.protection.sheet is True
    assert bool(ws.protection.password) is True
    assert ws.protection.selectUnlockedCells is True
    assert ws.protection.formatCells is False
    wb.close()


def test_set_worksheet_protection_dry_run_does_not_persist(tmp_workbook):
    result = set_sheet_protection(
        tmp_workbook,
        "Sheet1",
        enabled=True,
        options={"selectUnlockedCells": True},
        dry_run=True,
    )
    assert result["dry_run"] is True
    assert result["changes"][0]["new_value"]["enabled"] is True

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.protection.sheet is False
    assert ws.protection.selectUnlockedCells is False
    wb.close()


def test_set_print_area_persists_value(tmp_workbook):
    result = set_print_area(tmp_workbook, "Sheet1", "A1:C4")
    assert result["print_area"] == "A1:C4"

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].print_area == "'Sheet1'!$A$1:$C$4"
    wb.close()


def test_set_print_area_can_clear_value(tmp_workbook):
    set_print_area(tmp_workbook, "Sheet1", "A1:C4")
    result = set_print_area(tmp_workbook, "Sheet1", None)
    assert result["print_area"] is None

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].print_area == ""
    wb.close()


def test_set_print_titles_persists_rows_and_columns(tmp_workbook):
    result = set_print_titles(tmp_workbook, "Sheet1", rows="1:2", columns="A:B")
    assert result["print_title_rows"] == "1:2"
    assert result["print_title_columns"] == "A:B"

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.print_title_rows == "$1:$2"
    assert ws.print_title_cols == "$A:$B"
    wb.close()


def test_set_print_titles_can_clear_rows_or_columns(tmp_workbook):
    set_print_titles(tmp_workbook, "Sheet1", rows="1:2", columns="A:B")
    result = set_print_titles(tmp_workbook, "Sheet1", rows="", columns=None)
    assert result["print_title_rows"] is None
    assert result["print_title_columns"] == "A:B"

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.print_title_rows is None
    assert ws.print_title_cols == "$A:$B"
    wb.close()


@pytest.mark.parametrize(
    ("operation", "args"),
    [
        (insert_row, ("Data", 2)),
        (delete_rows, ("Data", 2)),
        (insert_cols, ("Data", 2)),
        (delete_cols, ("Data", 2)),
    ],
)
def test_structural_row_and_column_operations_reject_native_table_impacts(
    table_guard_workbook,
    operation,
    args,
):
    with pytest.raises(Exception, match="native Excel table"):
        operation(table_guard_workbook, *args)


def test_insert_row_below_native_table_still_succeeds(table_guard_workbook):
    result = insert_row(table_guard_workbook, "Data", 5)

    assert result["start_row"] == 5

    wb = load_workbook(table_guard_workbook)
    ws = wb["Data"]
    assert ws.tables["DataTable"].ref == "A1:B3"
    assert ws["A4"].value == "outside"
    wb.close()


def test_set_column_widths_persists_values(tmp_workbook):
    result = set_column_widths(tmp_workbook, "Sheet1", {"A": 24, "c": 18.5})
    assert result["widths"] == {"A": 24.0, "C": 18.5}
    assert "changes" not in result

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.column_dimensions["A"].width == 24.0
    assert ws.column_dimensions["C"].width == 18.5
    wb.close()


def test_set_column_widths_can_include_changes_explicitly(tmp_workbook):
    result = set_column_widths(tmp_workbook, "Sheet1", {"B": 30}, include_changes=True)

    assert result["changes"][0]["new_value"] == 30.0


def test_set_column_widths_dry_run_does_not_persist(tmp_workbook):
    result = set_column_widths(tmp_workbook, "Sheet1", {"B": 30}, dry_run=True)
    assert result["dry_run"] is True
    assert result["changes"][0]["new_value"] == 30.0

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.column_dimensions["B"].width != 30.0
    wb.close()


def test_autofit_columns_persists_computed_width(tmp_workbook):
    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    ws["A2"] = "Extraordinarily long customer name"
    wb.save(tmp_workbook)
    wb.close()

    result = autofit_columns(tmp_workbook, "Sheet1", columns=["A"])
    assert result["columns_fitted"] == 1
    assert result["widths"]["A"] > 20

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.column_dimensions["A"].width == result["widths"]["A"]
    wb.close()


def test_autofit_columns_dry_run_does_not_persist(tmp_workbook):
    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    ws["A2"] = "Extraordinarily long customer name"
    wb.save(tmp_workbook)
    original_width = ws.column_dimensions["A"].width
    wb.close()

    result = autofit_columns(tmp_workbook, "Sheet1", columns=["A"], dry_run=True)
    assert result["dry_run"] is True
    assert result["widths"]["A"] > 20

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.column_dimensions["A"].width == original_width
    wb.close()


def test_set_row_heights_persists_values(tmp_workbook):
    result = set_row_heights(tmp_workbook, "Sheet1", {"1": 22, "3": 28.5})
    assert result["heights"] == {1: 22.0, 3: 28.5}
    assert "changes" not in result

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.row_dimensions[1].height == 22.0
    assert ws.row_dimensions[3].height == 28.5
    wb.close()


def test_set_row_heights_dry_run_does_not_persist(tmp_workbook):
    result = set_row_heights(tmp_workbook, "Sheet1", {"2": 31}, dry_run=True)
    assert result["dry_run"] is True
    assert result["changes"][0]["new_value"] == 31.0

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert ws.row_dimensions[2].height != 31.0
    wb.close()


def test_list_named_ranges_returns_destinations(named_range_workbook):
    result = list_named_ranges(named_range_workbook)
    assert result == [
        {
            "name": "PeopleTable",
            "type": "RANGE",
            "value": "Sheet1!$A$1:$B$2",
            "destinations": [{"sheet_name": "Sheet1", "range": "$A$1:$B$2"}],
            "local_sheet": None,
            "hidden": False,
            "broken_reference": False,
            "missing_sheets": [],
        }
    ]


def test_freeze_panes_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(freeze_panes_tool(tmp_workbook, "Sheet1", "B2", dry_run=True))
    assert payload["operation"] == "freeze_panes"
    assert payload["dry_run"] is True
    assert payload["data"]["freeze_panes"] == "B2"


def test_freeze_panes_tool_defaults_to_compact_committed_response(tmp_workbook):
    payload = _load_tool_payload(freeze_panes_tool(tmp_workbook, "Sheet1", "B2"))

    assert payload["operation"] == "freeze_panes"
    assert "changes" not in payload


def test_merge_cells_tool_defaults_to_compact_committed_response(tmp_workbook):
    payload = _load_tool_payload(merge_cells_tool(tmp_workbook, "Sheet1", "A1", "B1"))

    assert payload["operation"] == "merge_cells"
    assert "changes" not in payload


def test_unmerge_cells_tool_can_include_changes_explicitly(tmp_workbook):
    merge_cells_tool(tmp_workbook, "Sheet1", "A1", "B1")
    payload = _load_tool_payload(
        unmerge_cells_tool(tmp_workbook, "Sheet1", "A1", "B1", include_changes=True)
    )

    assert payload["operation"] == "unmerge_cells"
    assert payload["changes"][0]["range"] == "A1:B1"


def test_set_autofilter_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(set_autofilter_tool(tmp_workbook, "Sheet1", dry_run=True))
    assert payload["operation"] == "set_autofilter"
    assert payload["dry_run"] is True
    assert payload["data"]["range"] == "A1:C6"


def test_list_named_ranges_tool_returns_json_envelope(named_range_workbook):
    payload = _load_tool_payload(list_named_ranges_tool(named_range_workbook))
    assert payload["operation"] == "list_named_ranges"
    assert payload["data"]["named_ranges"][0]["name"] == "PeopleTable"


def test_set_worksheet_visibility_tool_returns_json_envelope(multi_sheet_workbook):
    payload = _load_tool_payload(
        set_worksheet_visibility_tool(multi_sheet_workbook, "Inventory", "hidden", dry_run=True)
    )
    assert payload["operation"] == "set_worksheet_visibility"
    assert payload["dry_run"] is True
    assert payload["data"]["visibility"] == "hidden"


def test_get_worksheet_protection_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(get_worksheet_protection_tool(tmp_workbook, "Sheet1"))
    assert payload["operation"] == "get_worksheet_protection"
    assert payload["data"]["enabled"] is False


def test_set_worksheet_protection_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(
        set_worksheet_protection_tool(
            tmp_workbook,
            "Sheet1",
            enabled=True,
            options={"selectUnlockedCells": True},
            dry_run=True,
        )
    )
    assert payload["operation"] == "set_worksheet_protection"
    assert payload["dry_run"] is True
    assert payload["data"]["enabled"] is True


def test_set_print_area_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(set_print_area_tool(tmp_workbook, "Sheet1", "A1:C4", dry_run=True))
    assert payload["operation"] == "set_print_area"
    assert payload["dry_run"] is True
    assert payload["data"]["print_area"] == "A1:C4"


def test_set_print_titles_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(
        set_print_titles_tool(tmp_workbook, "Sheet1", rows="1:2", columns="A:B", dry_run=True)
    )
    assert payload["operation"] == "set_print_titles"
    assert payload["dry_run"] is True
    assert payload["data"]["print_title_rows"] == "1:2"
    assert payload["data"]["print_title_columns"] == "A:B"


def test_set_column_widths_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(set_column_widths_tool(tmp_workbook, "Sheet1", {"A": 20}, dry_run=True))
    assert payload["operation"] == "set_column_widths"
    assert payload["dry_run"] is True
    assert payload["data"]["widths"]["A"] == 20.0


def test_autofit_columns_tool_returns_json_envelope(tmp_workbook):
    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    ws["A2"] = "Extraordinarily long customer name"
    wb.save(tmp_workbook)
    wb.close()

    payload = _load_tool_payload(autofit_columns_tool(tmp_workbook, "Sheet1", ["A"], dry_run=True))
    assert payload["operation"] == "autofit_columns"
    assert payload["dry_run"] is True
    assert payload["data"]["widths"]["A"] > 20


def test_set_row_heights_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(set_row_heights_tool(tmp_workbook, "Sheet1", {"1": 24}, dry_run=True))
    assert payload["operation"] == "set_row_heights"
    assert payload["dry_run"] is True
    assert payload["data"]["heights"]["1"] == 24.0

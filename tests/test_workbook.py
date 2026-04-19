import json

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from excel_mcp.chart import create_chart_in_sheet
from excel_mcp.server import (
    apply_workbook_repairs as apply_workbook_repairs_tool,
    analyze_range_impact as analyze_range_impact_tool,
    audit_workbook as audit_workbook_tool,
    delete_named_range as delete_named_range_tool,
    diff_workbooks as diff_workbooks_tool,
    explain_formula_cell as explain_formula_cell_tool,
    get_workbook_metadata as get_workbook_metadata_tool,
    inspect_conditional_format_rules as inspect_conditional_format_rules_tool,
    inspect_data_validation_rules as inspect_data_validation_rules_tool,
    inspect_named_range as inspect_named_range_tool,
    list_all_sheets as list_all_sheets_tool,
    remove_conditional_format_rules as remove_conditional_format_rules_tool,
    remove_data_validation_rules as remove_data_validation_rules_tool,
    plan_workbook_repairs as plan_workbook_repairs_tool,
    profile_workbook as profile_workbook_tool,
)
from excel_mcp.tables import create_excel_table
from excel_mcp.workbook import (
    apply_workbook_repairs,
    analyze_range_impact,
    audit_workbook,
    delete_named_range,
    diff_workbooks,
    explain_formula_cell,
    get_or_create_workbook,
    get_workbook_info,
    inspect_conditional_format_rules,
    inspect_data_validation_rules,
    inspect_named_range,
    list_named_ranges,
    list_sheets,
    plan_workbook_repairs,
    profile_workbook,
    remove_conditional_format_rules,
    remove_data_validation_rules,
)


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    assert "operation" in payload
    assert "message" in payload
    return payload


def test_get_or_create_raises_on_missing_file(tmp_path):
    """get_or_create_workbook should raise when file doesn't exist."""
    missing = str(tmp_path / "nonexistent.xlsx")
    with pytest.raises(FileNotFoundError):
        get_or_create_workbook(missing)


def test_get_or_create_loads_existing_file(tmp_workbook):
    """get_or_create_workbook should load existing files normally."""
    wb = get_or_create_workbook(tmp_workbook)
    assert "Sheet1" in wb.sheetnames
    wb.close()

def test_list_sheets_returns_names(multi_sheet_workbook):
    result = list_sheets(multi_sheet_workbook)
    assert len(result) == 2
    assert result[0]["name"] == "Sales"
    assert result[1]["name"] == "Inventory"
    assert result[0]["rows"] >= 2
    assert result[0]["columns"] >= 2
    assert result[0]["is_empty"] is False


def test_list_sheets_marks_empty_sheet(empty_workbook):
    result = list_sheets(empty_workbook)
    assert result == [
        {
            "name": "Sheet",
            "sheet_type": "worksheet",
            "rows": 0,
            "columns": 0,
            "column_range": None,
            "is_empty": True,
        }
    ]


def test_list_sheets_handles_chart_sheets(tmp_path):
    filepath = str(tmp_path / "chartsheet-list.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Name", "Value"), ("A", 1), ("B", 2)]:
        ws.append(row)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb.save(filepath)
    wb.close()

    result = list_sheets(filepath)

    assert result == [
        {
            "name": "Data",
            "sheet_type": "worksheet",
            "rows": 3,
            "columns": 2,
            "column_range": "A-B",
            "is_empty": False,
        },
        {
            "name": "Charts",
            "sheet_type": "chartsheet",
            "rows": 0,
            "columns": 0,
            "column_range": None,
            "is_empty": False,
        },
    ]


def test_list_all_sheets_tool_handles_chart_sheets(tmp_path):
    filepath = str(tmp_path / "chartsheet-tool.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Name", "Value"), ("A", 1), ("B", 2)]:
        ws.append(row)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb.save(filepath)
    wb.close()

    payload = json.loads(list_all_sheets_tool(filepath))

    assert payload["operation"] == "list_all_sheets"
    assert payload["data"]["sheets"][1]["sheet_type"] == "chartsheet"


def test_profile_workbook_summarizes_tables_and_charts(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")
    create_chart_in_sheet(
        filepath=tmp_workbook,
        sheet_name="Sheet1",
        chart_type="bar",
        target_cell="E1",
        data_range="A1:B6",
        title="Customers by Age",
    )

    result = profile_workbook(tmp_workbook)

    assert result["sheet_count"] == 1
    assert result["table_count"] == 1
    assert result["chart_count"] == 1
    assert result["named_range_count"] == 0

    sheet = result["sheets"][0]
    assert sheet["name"] == "Sheet1"
    assert sheet["used_range"] == "A1:C6"
    assert sheet["table_count"] == 1
    assert sheet["chart_count"] == 1
    assert sheet["tables"][0]["table_name"] == "Customers"
    assert sheet["charts"][0]["chart_type"] == "bar"
    assert sheet["charts"][0]["anchor"] == "E1"
    assert sheet["charts"][0]["occupied_range"].startswith("E1:")


def test_profile_workbook_tool_returns_json_envelope(tmp_workbook):
    payload = json.loads(profile_workbook_tool(tmp_workbook))

    assert payload["operation"] == "profile_workbook"
    assert payload["data"]["sheet_count"] == 1


def test_audit_workbook_reports_clean_structured_workbook(tmp_workbook):
    result = audit_workbook(tmp_workbook)

    assert result["summary"]["risk_level"] == "low"
    assert result["summary"]["finding_count"] == 0
    assert result["summary"]["worksheet_count"] == 1
    assert result["summary"]["chartsheet_count"] == 0
    assert result["findings"]["count"] == 0
    assert result["findings"]["sample"] == []
    assert result["sheet_assessments"] == [
        {
            "sheet_name": "Sheet1",
            "sheet_type": "worksheet",
            "visibility": "visible",
            "used_range": "A1:C6",
            "rows": 6,
            "columns": 3,
            "dataset_kind": "worksheet_table",
            "recommended_read_tool": "quick_read",
            "dominant_table_name": None,
            "table_count": 0,
            "chart_count": 0,
            "finding_count": 0,
            "highest_severity": None,
        }
    ]
    assert result["recommended_actions"] == []


def test_audit_workbook_reports_high_signal_findings(tmp_path):
    filepath = str(tmp_path / "audit.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Value", "Status"])
    ws.append(["Alice", 10, "ok"])
    ws["D2"] = "=SUM(#REF!)"
    ws["E2"] = "#DIV/0!"
    ws["E2"].data_type = "e"

    raw = wb.create_sheet("Raw")
    raw.append(["ID", "ID", None])
    raw.append([1, 2, 3])

    hidden = wb.create_sheet("Hidden")
    hidden.sheet_state = "veryHidden"

    wb.defined_names["BrokenRange"] = DefinedName(
        "BrokenRange",
        attr_text="MissingSheet!$A$1:$A$2",
    )
    wb.save(filepath)
    wb.close()

    result = audit_workbook(filepath)

    assert result["summary"]["risk_level"] == "high"
    assert result["summary"]["high_count"] >= 3
    assert result["summary"]["medium_count"] >= 2
    assert result["summary"]["hidden_sheet_count"] == 1
    assert result["summary"]["worksheet_count"] == 3

    by_code = {item["code"]: item["count"] for item in result["findings"]["by_code"]}
    assert by_code["broken_formula_reference"] == 1
    assert by_code["error_cells_present"] == 1
    assert by_code["duplicate_headers"] == 1
    assert by_code["blank_headers"] == 2
    assert by_code["hidden_sheet"] == 1
    assert by_code["named_range_missing_sheet"] == 1

    sample_by_code = {item["code"]: item for item in result["findings"]["sample"]}
    assert sample_by_code["broken_formula_reference"]["sheet_name"] == "Data"
    assert sample_by_code["broken_formula_reference"]["details"]["sample"] == ["D2"]
    assert sample_by_code["error_cells_present"]["details"]["sample"] == ["E2"]
    assert sample_by_code["duplicate_headers"]["sheet_name"] == "Raw"
    assert sample_by_code["blank_headers"]["sheet_name"] == "Raw"
    assert sample_by_code["hidden_sheet"]["details"]["visibility"] == "veryHidden"
    assert sample_by_code["named_range_missing_sheet"]["details"]["missing_sheets"] == ["MissingSheet"]

    assessments = {item["sheet_name"]: item for item in result["sheet_assessments"]}
    assert assessments["Data"]["highest_severity"] == "high"
    assert assessments["Raw"]["highest_severity"] == "medium"
    assert assessments["Hidden"]["highest_severity"] == "medium"


def test_audit_workbook_flags_broken_structured_reference_formulas(tmp_path):
    filepath = str(tmp_path / "audit-broken-structured-reference.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Value"])
    ws.append(["Alice", 10])
    ws.append(["Bob", 20])
    ws["D2"] = "=SUM(People[Value])"

    table = Table(displayName="People", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    del ws.tables["People"]
    wb.save(filepath)
    wb.close()

    result = audit_workbook(filepath)

    by_code = {item["code"]: item["count"] for item in result["findings"]["by_code"]}
    assert by_code["broken_formula_reference"] == 1
    sample_by_code = {item["code"]: item for item in result["findings"]["sample"]}
    assert sample_by_code["broken_formula_reference"]["sheet_name"] == "Data"
    assert sample_by_code["broken_formula_reference"]["details"]["sample"] == ["D2"]


def test_audit_workbook_handles_complex_workbook_orientation(complex_workbook):
    result = audit_workbook(complex_workbook)

    assert result["summary"]["chartsheet_count"] == 1
    assert result["summary"]["layout_like_sheet_count"] == 1

    assessments = {item["sheet_name"]: item for item in result["sheet_assessments"]}
    assert assessments["Charts"]["dataset_kind"] == "chartsheet"
    assert assessments["Charts"]["recommended_read_tool"] == "list_charts"
    assert assessments["Dashboard"]["dataset_kind"] == "layout_like_sheet"
    assert assessments["Dashboard"]["recommended_read_tool"] == "profile_workbook"
    assert assessments["Data"]["recommended_read_tool"] == "read_excel_table"


def test_audit_workbook_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(audit_workbook_tool(tmp_workbook))

    assert payload["operation"] == "audit_workbook"
    assert payload["data"]["summary"]["finding_count"] == 0


def test_audit_workbook_rejects_boolean_sample_limit(tmp_workbook):
    with pytest.raises(Exception, match="sample_limit must be a positive integer"):
        audit_workbook(tmp_workbook, sample_limit=True)


def test_plan_workbook_repairs_rejects_boolean_sample_limit(tmp_workbook):
    with pytest.raises(Exception, match="sample_limit must be a positive integer"):
        plan_workbook_repairs(tmp_workbook, sample_limit=True)


def test_plan_workbook_repairs_returns_empty_plan_for_clean_workbook(tmp_workbook):
    result = plan_workbook_repairs(tmp_workbook)

    assert result["audit_summary"]["finding_count"] == 0
    assert result["step_count"] == 0
    assert result["steps"] == []
    assert result["quick_wins"] == []


def test_plan_workbook_repairs_prioritizes_high_signal_actions(tmp_path):
    filepath = str(tmp_path / "repair-plan.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Value", "Status"])
    ws.append(["Alice", 10, "ok"])
    ws["D2"] = "=SUM(#REF!)"
    ws["E2"] = "#DIV/0!"
    ws["E2"].data_type = "e"

    raw = wb.create_sheet("Raw")
    raw.append(["ID", "ID", None])
    raw.append([1, 2, 3])

    hidden = wb.create_sheet("Hidden")
    hidden.sheet_state = "veryHidden"

    wb.defined_names["BrokenRange"] = DefinedName(
        "BrokenRange",
        attr_text="MissingSheet!$A$1:$A$2",
    )
    wb.save(filepath)
    wb.close()

    result = plan_workbook_repairs(filepath)

    assert result["step_count"] >= 5
    assert result["steps"][0]["priority"] == "high"

    steps_by_title = {step["title"]: step for step in result["steps"]}
    formula_step = steps_by_title["Repair broken formulas on 'Data'"]
    assert formula_step["can_execute_fully_in_sheetforge"] is False
    assert formula_step["suggested_tools"][0]["tool"] == "read_data_from_excel"
    assert formula_step["suggested_tools"][0]["args"]["sheet_name"] == "Data"
    assert formula_step["suggested_tools"][0]["args"]["start_cell"] == "D2"

    hidden_step = steps_by_title["Review hidden sheet 'Hidden' before workbook-wide automation"]
    assert hidden_step["can_execute_fully_in_sheetforge"] is True
    assert hidden_step["suggested_tools"][0]["tool"] == "set_worksheet_visibility"
    assert hidden_step["suggested_tools"][0]["args"]["dry_run"] is True

    header_step = steps_by_title["Normalize headers on 'Raw'"]
    assert sorted(header_step["finding_codes"]) == ["blank_headers", "duplicate_headers"]
    assert header_step["suggested_tools"][0]["tool"] == "quick_read"

    named_range_step = steps_by_title["Inspect and repair workbook named ranges"]
    assert named_range_step["suggested_tools"][0] == {
        "tool": "list_named_ranges",
        "args": {"filepath": filepath},
    }
    assert any(tool["tool"] == "inspect_named_range" for tool in named_range_step["suggested_tools"])
    assert any(tool["tool"] == "delete_named_range" for tool in named_range_step["suggested_tools"])

    assert "Review hidden sheet 'Hidden' before workbook-wide automation" in result["quick_wins"]


def test_plan_workbook_repairs_handles_layout_like_sheet(complex_workbook):
    result = plan_workbook_repairs(complex_workbook)

    steps_by_title = {step["title"]: step for step in result["steps"]}
    dashboard_step = steps_by_title["Treat 'Dashboard' as a layout-oriented sheet"]
    assert dashboard_step["priority"] == "low"
    assert dashboard_step["suggested_tools"][0]["tool"] == "profile_workbook"
    assert dashboard_step["suggested_tools"][1]["tool"] == "read_data_from_excel"
    assert dashboard_step["suggested_tools"][1]["args"]["sheet_name"] == "Dashboard"


def test_plan_workbook_repairs_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(plan_workbook_repairs_tool(tmp_workbook))

    assert payload["operation"] == "plan_workbook_repairs"
    assert payload["data"]["step_count"] == 0


def test_inspect_named_range_reports_scope_and_breakage(tmp_path):
    filepath = str(tmp_path / "named-range-inspect.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    wb.defined_names["BrokenRange"] = DefinedName(
        "BrokenRange",
        attr_text="MissingSheet!$A$1:$A$2",
    )
    ws.defined_names.add(
        DefinedName(
            "LocalRange",
            attr_text="Data!$B$2:$B$4",
        )
    )
    wb.save(filepath)
    wb.close()

    result = inspect_named_range(filepath, "BrokenRange")
    local_result = inspect_named_range(filepath, "LocalRange", scope_sheet="Data")

    assert result["match_count"] == 1
    assert result["matches"][0]["broken_reference"] is False
    assert result["matches"][0]["missing_sheets"] == ["MissingSheet"]
    assert local_result["matches"][0]["local_sheet"] == "Data"

    payload = _load_tool_payload(inspect_named_range_tool(filepath, "BrokenRange"))
    assert payload["operation"] == "inspect_named_range"
    assert payload["data"]["matches"][0]["name"] == "BrokenRange"


def test_delete_named_range_supports_dry_run_and_apply(tmp_path):
    filepath = str(tmp_path / "named-range-delete.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    wb.defined_names["BrokenRange"] = DefinedName(
        "BrokenRange",
        attr_text="MissingSheet!$A$1:$A$2",
    )
    wb.save(filepath)
    wb.close()

    payload = _load_tool_payload(delete_named_range_tool(filepath, "BrokenRange", dry_run=True))
    assert payload["operation"] == "delete_named_range"
    assert payload["dry_run"] is True

    preview = delete_named_range(filepath, "BrokenRange", dry_run=True)
    assert preview["removed_count"] == 1
    assert list_named_ranges(filepath)[0]["name"] == "BrokenRange"

    applied = delete_named_range(filepath, "BrokenRange", dry_run=False)
    assert applied["removed_count"] == 1
    assert list_named_ranges(filepath) == []


def test_inspect_and_remove_broken_validation_rules(tmp_path):
    filepath = str(tmp_path / "validation-rules.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    valid_rule = DataValidation(type="whole", formula1="1")
    valid_rule.add("A2:A4")
    broken_rule = DataValidation(type="list", formula1="=MissingSheet!$A$1:$A$3")
    broken_rule.add("B2:B4")
    ws.add_data_validation(valid_rule)
    ws.add_data_validation(broken_rule)
    wb.save(filepath)
    wb.close()

    inspected = inspect_data_validation_rules(filepath, "Data", broken_only=True)
    assert inspected["rule_count"] == 1
    assert inspected["rules"][0]["rule_index"] == 2
    assert inspected["rules"][0]["broken_reference"] is True

    preview = remove_data_validation_rules(filepath, "Data", broken_only=True, dry_run=True)
    assert preview["removed_count"] == 1
    assert inspect_data_validation_rules(filepath, "Data")["rule_count"] == 2

    applied = remove_data_validation_rules(filepath, "Data", broken_only=True, dry_run=False)
    assert applied["removed_count"] == 1
    remaining = inspect_data_validation_rules(filepath, "Data")
    assert remaining["rule_count"] == 1
    assert remaining["rules"][0]["applies_to"] == "A2:A4"

    payload = _load_tool_payload(inspect_data_validation_rules_tool(filepath, "Data", True))
    assert payload["operation"] == "inspect_data_validation_rules"
    payload = _load_tool_payload(remove_data_validation_rules_tool(filepath, "Data", None, True, True))
    assert payload["operation"] == "remove_data_validation_rules"
    assert payload["dry_run"] is True


def test_inspect_and_remove_broken_conditional_format_rules(tmp_path):
    filepath = str(tmp_path / "conditional-rules.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["A2"] = 2
    ws.conditional_formatting.add("A1:A2", FormulaRule(formula=["A1>0"]))
    ws.conditional_formatting.add("B1:B2", FormulaRule(formula=["MissingSheet!A1>0", "#REF!>0"]))
    wb.save(filepath)
    wb.close()

    inspected = inspect_conditional_format_rules(filepath, "Data", broken_only=True)
    assert inspected["rule_count"] == 1
    assert inspected["rules"][0]["rule_index"] == 2
    assert inspected["rules"][0]["broken_reference"] is True

    preview = remove_conditional_format_rules(filepath, "Data", broken_only=True, dry_run=True)
    assert preview["removed_count"] == 1
    assert inspect_conditional_format_rules(filepath, "Data")["rule_count"] == 2

    applied = remove_conditional_format_rules(filepath, "Data", broken_only=True, dry_run=False)
    assert applied["removed_count"] == 1
    remaining = inspect_conditional_format_rules(filepath, "Data")
    assert remaining["rule_count"] == 1
    assert remaining["rules"][0]["applies_to"] == "A1:A2"

    payload = _load_tool_payload(inspect_conditional_format_rules_tool(filepath, "Data", True))
    assert payload["operation"] == "inspect_conditional_format_rules"
    payload = _load_tool_payload(
        remove_conditional_format_rules_tool(filepath, "Data", None, True, True)
    )
    assert payload["operation"] == "remove_conditional_format_rules"
    assert payload["dry_run"] is True


def test_plan_workbook_repairs_points_to_repair_primitives(tmp_path):
    filepath = str(tmp_path / "repair-primitives-plan.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Value"])
    dv = DataValidation(type="list", formula1="=MissingSheet!$A$1:$A$2")
    dv.add("A2:A4")
    ws.add_data_validation(dv)
    ws.conditional_formatting.add("B2:B4", FormulaRule(formula=["#REF!>0"]))
    wb.defined_names["BrokenRange"] = DefinedName(
        "BrokenRange",
        attr_text="MissingSheet!$A$1:$A$2",
    )
    wb.save(filepath)
    wb.close()

    result = plan_workbook_repairs(filepath)
    steps_by_title = {step["title"]: step for step in result["steps"]}

    named_range_step = steps_by_title["Inspect and repair workbook named ranges"]
    assert named_range_step["can_execute_fully_in_sheetforge"] is True
    assert any(tool["tool"] == "inspect_named_range" for tool in named_range_step["suggested_tools"])
    assert any(tool["tool"] == "delete_named_range" for tool in named_range_step["suggested_tools"])

    validation_step = steps_by_title["Repair broken data validation rules on 'Data'"]
    assert validation_step["can_execute_fully_in_sheetforge"] is True
    assert validation_step["suggested_tools"][0]["tool"] == "inspect_data_validation_rules"
    assert validation_step["suggested_tools"][1]["tool"] == "remove_data_validation_rules"

    conditional_step = steps_by_title["Review broken conditional formatting rules on 'Data'"]
    assert conditional_step["can_execute_fully_in_sheetforge"] is True
    assert conditional_step["suggested_tools"][0]["tool"] == "inspect_conditional_format_rules"
    assert conditional_step["suggested_tools"][1]["tool"] == "remove_conditional_format_rules"


def test_apply_workbook_repairs_dry_run_and_apply(tmp_path):
    filepath = str(tmp_path / "apply-repairs.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Value"])

    broken_validation = DataValidation(type="list", formula1="=MissingSheet!$A$1:$A$2")
    broken_validation.add("A2:A4")
    ws.add_data_validation(broken_validation)
    ws.conditional_formatting.add("B2:B4", FormulaRule(formula=["#REF!>0"]))

    hidden = wb.create_sheet("Hidden")
    hidden.sheet_state = "veryHidden"

    wb.defined_names["BrokenRange"] = DefinedName(
        "BrokenRange",
        attr_text="MissingSheet!$A$1:$A$2",
    )
    wb.save(filepath)
    wb.close()

    repair_types = [
        "remove_broken_named_ranges",
        "remove_broken_validations",
        "remove_broken_conditional_formats",
        "reveal_hidden_sheets",
    ]

    preview = apply_workbook_repairs(
        filepath,
        repair_types=repair_types,
        dry_run=True,
    )
    assert preview["action_count"] == 4
    assert preview["audit_before"]["hidden_sheet_count"] == 1
    assert preview["audit_after"]["hidden_sheet_count"] == 1
    assert preview["diff"]["summary"]["named_range_change_count"] == 0

    applied = apply_workbook_repairs(
        filepath,
        repair_types=repair_types,
        dry_run=False,
    )
    assert applied["action_count"] == 4
    assert applied["audit_after"]["hidden_sheet_count"] == 0
    assert applied["audit_after"]["named_range_count"] == 0
    assert applied["audit_after"]["risk_level"] == "low"
    assert applied["audit_after"]["high_count"] == 0
    assert applied["audit_after"]["medium_count"] == 0
    assert applied["diff"]["summary"]["named_range_change_count"] == 1
    assert applied["diff"]["summary"]["validation_rule_change_count"] == 1
    assert applied["diff"]["summary"]["conditional_format_rule_change_count"] == 1
    assert applied["diff"]["summary"]["sheet_property_change_count"] == 1

    assert list_named_ranges(filepath) == []
    assert inspect_data_validation_rules(filepath, "Data")["rule_count"] == 0
    assert inspect_conditional_format_rules(filepath, "Data")["rule_count"] == 0
    audit_summary = audit_workbook(filepath)["summary"]
    assert audit_summary["risk_level"] == "low"
    assert audit_summary["high_count"] == 0
    assert audit_summary["medium_count"] == 0

    payload = _load_tool_payload(
        apply_workbook_repairs_tool(filepath, repair_types, None, 1, 25, True)
    )
    assert payload["operation"] == "apply_workbook_repairs"
    assert payload["dry_run"] is True


def test_diff_workbooks_reports_structural_and_cell_changes(tmp_path):
    before_path = str(tmp_path / "before.xlsx")
    after_path = str(tmp_path / "after.xlsx")

    before_wb = Workbook()
    before_ws = before_wb.active
    before_ws.title = "Data"
    before_ws["A1"] = "Value"
    before_ws["A2"] = 10
    before_wb.defined_names["InputRange"] = DefinedName(
        "InputRange",
        attr_text="Data!$A$2:$A$2",
    )
    before_wb.save(before_path)
    before_wb.close()

    after_wb = Workbook()
    after_ws = after_wb.active
    after_ws.title = "Data"
    after_ws["A1"] = "Value"
    after_ws["A2"] = 20
    after_ws["B2"] = "=A2*2"
    extra = after_wb.create_sheet("Summary")
    extra["A1"] = "Done"
    after_wb.save(after_path)
    after_wb.close()

    result = diff_workbooks(before_path, after_path)

    assert result["summary"]["sheet_count_before"] == 1
    assert result["summary"]["sheet_count_after"] == 2
    assert result["summary"]["named_range_change_count"] == 1
    assert result["cell_changes"]["count"] >= 2
    assert any(change["sheet_name"] == "Data" and change["cell"] == "A2" for change in result["cell_changes"]["sample"])
    assert result["sheet_changes"]["added"] == ["Summary"]

    payload = _load_tool_payload(diff_workbooks_tool(before_path, after_path, 25, True))
    assert payload["operation"] == "diff_workbooks"
    assert payload["data"]["summary"]["sheet_count_after"] == 2


def test_explain_formula_cell_reports_named_ranges_and_dependents(tmp_path):
    filepath = str(tmp_path / "formula-explain.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Input"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["B2"] = "=SUM(InputRange)"
    ws["C2"] = "=B2*2"
    wb.defined_names["InputRange"] = DefinedName(
        "InputRange",
        attr_text="Data!$A$2:$A$3",
    )
    wb.save(filepath)
    wb.close()

    result = explain_formula_cell(filepath, "Data", "B2")

    assert result["cell"] == "B2"
    assert result["formula"] == "=SUM(InputRange)"
    assert result["direct_reference_count"] == 1
    assert result["direct_references"][0]["reference_type"] == "named_range"
    assert result["direct_references"][0]["targets"][0]["reference"] == "Data!A2:A3"
    assert result["formula_chain"]["precedent_formula_count"] == 0
    assert result["formula_chain"]["max_depth_reached"] == 0
    assert result["formula_chain"]["truncated"] is False
    assert result["formula_chain"]["layer_summary"] == [
        {
            "depth": 0,
            "count": 1,
            "sample": [{"sheet_name": "Data", "cell": "B2"}],
        }
    ]
    assert result["dependent_formulas"]["count"] == 1
    assert result["dependent_formulas"]["sample"][0]["cell"] == "C2"
    assert "Formula depends on named ranges." in result["hints"]

    payload = _load_tool_payload(explain_formula_cell_tool(filepath, "Data", "B2", 3))
    assert payload["operation"] == "explain_formula_cell"
    assert payload["data"]["cell"] == "B2"


def test_explain_formula_cell_reports_transitive_formula_precedents(tmp_path):
    filepath = str(tmp_path / "formula-transitive.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["B2"] = "=SUM(A2:A3)"
    ws["C2"] = "=B2*2"
    ws["D2"] = "=C2+5"
    wb.save(filepath)
    wb.close()

    result = explain_formula_cell(filepath, "Data", "D2", max_depth=3)

    assert result["direct_formula_precedent_count"] == 1
    assert result["direct_formula_precedents"][0]["cell"] == "C2"
    assert result["transitive_formula_precedent_count"] == 1
    assert result["transitive_formula_precedents"][0]["cell"] == "B2"
    assert result["transitive_formula_precedents"][0]["depth"] == 2
    assert result["formula_chain"]["precedent_formula_count"] == 2
    assert result["formula_chain"]["max_depth_reached"] == 2
    assert result["formula_chain"]["truncated"] is False
    assert result["formula_chain"]["edge_count"] == 2
    assert result["formula_chain"]["leaf_formula_precedent_count"] == 1
    assert result["formula_chain"]["leaf_formula_precedents"] == [
        {"sheet_name": "Data", "cell": "B2", "depth": 2}
    ]
    assert result["formula_chain"]["path_sample"] == [
        [
            {"sheet_name": "Data", "cell": "D2"},
            {"sheet_name": "Data", "cell": "C2"},
            {"sheet_name": "Data", "cell": "B2"},
        ]
    ]


def test_explain_formula_cell_reports_branching_formula_chain_summary(tmp_path):
    filepath = str(tmp_path / "formula-branching.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["B2"] = "=SUM(A2:A3)"
    ws["B3"] = "=A2*2"
    ws["C2"] = "=B2+B3"
    ws["D2"] = "=C2+5"
    wb.save(filepath)
    wb.close()

    result = explain_formula_cell(filepath, "Data", "D2", max_depth=3)

    assert result["formula_chain"]["precedent_formula_count"] == 3
    assert result["formula_chain"]["max_depth_reached"] == 2
    assert result["formula_chain"]["edge_count"] == 3
    assert result["formula_chain"]["leaf_formula_precedent_count"] == 2
    assert result["formula_chain"]["leaf_formula_precedents"] == [
        {"sheet_name": "Data", "cell": "B2", "depth": 2},
        {"sheet_name": "Data", "cell": "B3", "depth": 2},
    ]
    assert result["formula_chain"]["layer_summary"] == [
        {
            "depth": 0,
            "count": 1,
            "sample": [{"sheet_name": "Data", "cell": "D2"}],
        },
        {
            "depth": 1,
            "count": 1,
            "sample": [{"sheet_name": "Data", "cell": "C2"}],
        },
        {
            "depth": 2,
            "count": 2,
            "sample": [
                {"sheet_name": "Data", "cell": "B2"},
                {"sheet_name": "Data", "cell": "B3"},
            ],
        },
    ]
    assert result["formula_chain"]["path_sample"] == [
        [
            {"sheet_name": "Data", "cell": "D2"},
            {"sheet_name": "Data", "cell": "C2"},
            {"sheet_name": "Data", "cell": "B2"},
        ],
        [
            {"sheet_name": "Data", "cell": "D2"},
            {"sheet_name": "Data", "cell": "C2"},
            {"sheet_name": "Data", "cell": "B3"},
        ],
    ]


def test_explain_formula_cell_marks_chain_as_truncated_when_depth_cap_hits(tmp_path):
    filepath = str(tmp_path / "formula-truncated.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A2"] = 10
    ws["B2"] = "=A2*2"
    ws["C2"] = "=B2+5"
    ws["D2"] = "=C2+7"
    wb.save(filepath)
    wb.close()

    result = explain_formula_cell(filepath, "Data", "D2", max_depth=1)

    assert result["direct_formula_precedent_count"] == 1
    assert result["transitive_formula_precedent_count"] == 0
    assert result["formula_chain"]["max_depth_reached"] == 1
    assert result["formula_chain"]["truncated"] is True
    assert result["formula_chain"]["leaf_formula_precedent_count"] == 0
    assert result["formula_chain"]["path_sample"] == [
        [
            {"sheet_name": "Data", "cell": "D2"},
            {"sheet_name": "Data", "cell": "C2"},
        ]
    ]


def test_analyze_range_impact_reports_overlapping_structures(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="Customers")
    create_chart_in_sheet(
        filepath=tmp_workbook,
        sheet_name="Sheet1",
        chart_type="bar",
        target_cell="E1",
        data_range="A1:B6",
        title="Customers by Age",
    )

    workbook = load_workbook(tmp_workbook)
    ws = workbook["Sheet1"]
    ws.merge_cells("B2:C2")
    ws.auto_filter.ref = "A1:C6"
    ws.print_area = "A1:F10"
    ws["D3"] = "=SUM(B2:C2)"
    ws["H2"] = "=SUM(B2:C3)"
    dependent_sheet = workbook.create_sheet("Dependent")
    dependent_sheet["A1"] = "=SUM(Sheet1!B2:C3)"
    dependent_sheet["B1"] = "=SUM(ImpactArea)"
    workbook.defined_names["ImpactArea"] = DefinedName(
        "ImpactArea",
        attr_text="Sheet1!$B$2:$F$4",
    )
    workbook.save(tmp_workbook)
    workbook.close()

    result = analyze_range_impact(tmp_workbook, "Sheet1", "A1:F4")

    assert result["summary"]["risk_level"] == "high"
    assert result["summary"]["table_count"] == 1
    assert result["summary"]["chart_count"] == 1
    assert result["summary"]["merged_range_count"] == 1
    assert result["summary"]["named_range_count"] == 1
    assert result["summary"]["formula_cell_count"] == 1
    assert result["summary"]["dependent_formula_count"] == 3
    assert result["summary"]["autofilter_overlap"] is True
    assert result["summary"]["print_area_overlap"] is True
    assert result["tables"][0]["covers_header"] is True
    assert result["charts"][0]["anchor"] == "E1"
    assert result["merged_ranges"][0]["range"] == "B2:C2"
    assert result["named_ranges"][0]["name"] == "ImpactArea"
    assert result["formula_cells"]["sample"] == ["D3"]
    assert result["dependent_formulas"]["count"] == 3
    dependent_cells = {
        (item["sheet_name"], item["cell"]) for item in result["dependent_formulas"]["sample"]
    }
    assert dependent_cells == {("Sheet1", "H2"), ("Dependent", "A1"), ("Dependent", "B1")}
    named_range_reference = next(
        reference
        for item in result["dependent_formulas"]["sample"]
        if item["cell"] == "B1"
        for reference in item["references"]
        if reference.get("via_named_range") == "ImpactArea"
    )
    assert named_range_reference["intersection_range"] == "B2:F4"


def test_analyze_range_impact_reports_low_risk_for_empty_area(tmp_workbook):
    result = analyze_range_impact(tmp_workbook, "Sheet1", "H20:I21")

    assert result["summary"]["risk_level"] == "low"
    assert result["summary"]["table_count"] == 0
    assert result["summary"]["chart_count"] == 0
    assert result["summary"]["dependent_formula_count"] == 0
    assert result["hints"] == ["No overlapping workbook structures detected for this range."]


def test_analyze_range_impact_accepts_quoted_sheet_name_with_apostrophe(tmp_path):
    filepath = str(tmp_path / "apostrophe-impact.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Bob's Data"
    ws["A1"] = "Label"
    ws["B1"] = "Value"
    ws["A2"] = "North"
    ws["B2"] = 10
    wb.save(filepath)
    wb.close()

    result = analyze_range_impact(filepath, "Bob's Data", "'Bob''s Data'!A1:B2")

    assert result["sheet_name"] == "Bob's Data"
    assert result["range"] == "A1:B2"


def test_analyze_range_impact_tracks_local_named_range_dependencies(tmp_workbook):
    workbook = load_workbook(tmp_workbook)
    ws = workbook["Sheet1"]
    ws["D1"] = "=SUM(LocalImpact)"
    ws.defined_names.add(
        DefinedName(
            "LocalImpact",
            attr_text="Sheet1!$B$2:$B$4",
        )
    )
    workbook.save(tmp_workbook)
    workbook.close()

    result = analyze_range_impact(tmp_workbook, "Sheet1", "B2:B4")

    assert result["summary"]["dependent_formula_count"] == 1
    dependency = result["dependent_formulas"]["sample"][0]
    assert dependency["cell"] == "D1"
    assert dependency["sheet_name"] == "Sheet1"
    assert dependency["references"][0]["via_named_range"] == "LocalImpact"
    assert dependency["references"][0]["intersection_range"] == "B2:B4"


def test_analyze_range_impact_tracks_transitive_formula_dependencies(tmp_workbook):
    workbook = load_workbook(tmp_workbook)
    ws = workbook["Sheet1"]
    for row in range(2, 6):
        ws[f"D{row}"] = f"=B{row}+1"
    ws["D6"] = "=SUM(D2:D5)"
    ws["E1"] = "=D6*2"
    workbook.save(tmp_workbook)
    workbook.close()

    result = analyze_range_impact(tmp_workbook, "Sheet1", "B2:C5")

    assert result["summary"]["dependent_formula_count"] == 6
    assert result["summary"]["direct_formula_count"] == 4
    assert result["summary"]["transitive_formula_count"] == 2
    assert result["dependent_formulas"]["count"] == 6
    assert result["dependent_formulas"]["direct_count"] == 4
    assert result["dependent_formulas"]["transitive_count"] == 2

    dependencies = {
        item["cell"]: item for item in result["dependent_formulas"]["sample"]
    }
    assert dependencies["D2"]["dependency_depth"] == 1
    assert dependencies["D2"]["dependency_type"] == "direct"
    assert dependencies["D6"]["dependency_depth"] == 2
    assert dependencies["D6"]["dependency_type"] == "transitive"
    assert {
        predecessor["cell"] for predecessor in dependencies["D6"]["transitive_via"]
    } == {"D2", "D3", "D4", "D5"}
    assert {
        reference["intersection_range"] for reference in dependencies["D6"]["references"]
    } == {"D2:D2", "D3:D3", "D4:D4", "D5:D5"}
    assert dependencies["E1"]["dependency_depth"] == 3
    assert dependencies["E1"]["dependency_type"] == "transitive"
    assert dependencies["E1"]["transitive_via"] == [{"sheet_name": "Sheet1", "cell": "D6"}]
    assert dependencies["E1"]["references"][0]["intersection_range"] == "D6:D6"


def test_analyze_range_impact_tracks_table_structured_reference_dependencies(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="People")

    workbook = load_workbook(tmp_workbook)
    ws = workbook["Sheet1"]
    ws["H2"] = "=SUM(People[Age])"
    ws["H3"] = "=COUNTA(People[#Headers])"
    ws["H4"] = "=COUNTA(People[[#All],[Name]])"
    ws["H5"] = "=SUM(People[[Age]:[City]])"
    workbook.save(tmp_workbook)
    workbook.close()

    data_result = analyze_range_impact(tmp_workbook, "Sheet1", "B2:C6")
    header_result = analyze_range_impact(tmp_workbook, "Sheet1", "A1:C1")

    assert data_result["summary"]["dependent_formula_count"] == 2
    assert {
        item["cell"] for item in data_result["dependent_formulas"]["sample"]
    } == {"H2", "H5"}
    assert any(
        reference.get("via_table") == "People" and reference.get("structured_reference") == "People[Age]"
        for item in data_result["dependent_formulas"]["sample"]
        for reference in item["references"]
    )
    assert any(
        reference.get("intersection_range") == "B2:C6"
        for item in data_result["dependent_formulas"]["sample"]
        for reference in item["references"]
    )

    assert header_result["summary"]["dependent_formula_count"] == 2
    assert {item["cell"] for item in header_result["dependent_formulas"]["sample"]} == {"H3", "H4"}
    assert any(
        reference.get("structured_reference") == "People[#Headers]"
        for item in header_result["dependent_formulas"]["sample"]
        for reference in item["references"]
    )
    assert any(
        reference.get("structured_reference") == "People[[#All],[Name]]"
        and reference.get("intersection_range") == "A1:A1"
        for item in header_result["dependent_formulas"]["sample"]
        for reference in item["references"]
    )


def test_analyze_range_impact_tracks_this_row_structured_references(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="People")

    workbook = load_workbook(tmp_workbook)
    ws = workbook["Sheet1"]
    ws["H3"] = "=SUM(People[@Age])"
    workbook.save(tmp_workbook)
    workbook.close()

    matching_row = analyze_range_impact(tmp_workbook, "Sheet1", "B3:B3")
    other_row = analyze_range_impact(tmp_workbook, "Sheet1", "B2:B2")

    assert matching_row["summary"]["dependent_formula_count"] == 1
    dependency = matching_row["dependent_formulas"]["sample"][0]
    assert dependency["cell"] == "H3"
    assert dependency["references"][0]["via_table"] == "People"
    assert dependency["references"][0]["structured_reference"] == "People[@Age]"
    assert dependency["references"][0]["intersection_range"] == "B3:B3"
    assert other_row["summary"]["dependent_formula_count"] == 0


def test_analyze_range_impact_reports_validation_and_conditional_format_overlaps(tmp_workbook):
    workbook = load_workbook(tmp_workbook)
    ws = workbook["Sheet1"]
    validation = DataValidation(type="whole", operator="between", formula1="18", formula2="65")
    validation.add("B2:B6")
    ws.add_data_validation(validation)
    ws.conditional_formatting.add(
        "C2:C6",
        CellIsRule(operator="equal", formula=["\"Turku\""], fill=None),
    )
    workbook.save(tmp_workbook)
    workbook.close()

    result = analyze_range_impact(tmp_workbook, "Sheet1", "B3:C4")

    assert result["summary"]["data_validation_count"] == 1
    assert result["summary"]["conditional_format_count"] == 1
    assert result["data_validations"]["count"] == 1
    assert result["conditional_formats"]["count"] == 1
    assert result["data_validations"]["sample"][0]["applies_to"] == "B2:B6"
    assert result["data_validations"]["sample"][0]["intersection_ranges"] == ["B3:B4"]
    assert result["conditional_formats"]["sample"][0]["applies_to"] == "C2:C6"
    assert result["conditional_formats"]["sample"][0]["intersection_ranges"] == ["C3:C4"]
    assert "Selected range overlaps worksheet data validation rules." in result["hints"]
    assert "Selected range overlaps conditional formatting rules." in result["hints"]


def test_analyze_range_impact_tracks_validation_and_conditional_format_dependencies(tmp_workbook):
    workbook = load_workbook(tmp_workbook)
    dependent_sheet = workbook.create_sheet("Checks")

    validation = DataValidation(
        type="whole",
        operator="between",
        formula1="Sheet1!$B$2",
        formula2="Sheet1!$B$4",
    )
    validation.add("A1:A3")
    dependent_sheet.add_data_validation(validation)
    dependent_sheet.conditional_formatting.add(
        "B1:B3",
        FormulaRule(formula=["COUNTIF(Sheet1!$B$2:$B$4,B1)>0"]),
    )
    workbook.save(tmp_workbook)
    workbook.close()

    result = analyze_range_impact(tmp_workbook, "Sheet1", "B2:B4")

    assert result["summary"]["dependent_validation_count"] == 1
    assert result["summary"]["dependent_conditional_format_count"] == 1
    assert result["dependent_validations"]["count"] == 1
    assert result["dependent_conditional_formats"]["count"] == 1

    dependent_validation = result["dependent_validations"]["sample"][0]
    assert dependent_validation["sheet_name"] == "Checks"
    assert dependent_validation["applies_to"] == "A1:A3"
    assert {
        reference["intersection_range"] for reference in dependent_validation["references"]
    } == {"B2:B2", "B4:B4"}

    dependent_cf = result["dependent_conditional_formats"]["sample"][0]
    assert dependent_cf["sheet_name"] == "Checks"
    assert dependent_cf["applies_to"] == "B1:B3"
    assert dependent_cf["formula"] == ["COUNTIF(Sheet1!$B$2:$B$4,B1)>0"]
    assert dependent_cf["references"][0]["intersection_range"] == "B2:B4"
    assert "Validation rules elsewhere in the workbook reference the selected range." in result["hints"]
    assert "Conditional formatting rules reference the selected range." in result["hints"]


def test_list_named_ranges_includes_local_and_workbook_scope(tmp_workbook):
    workbook = load_workbook(tmp_workbook)
    ws = workbook["Sheet1"]
    ws.defined_names.add(DefinedName("LocalImpact", attr_text="Sheet1!$B$2:$B$4"))
    workbook.defined_names["GlobalImpact"] = DefinedName(
        "GlobalImpact",
        attr_text="Sheet1!$A$1:$A$2",
    )
    workbook.save(tmp_workbook)
    workbook.close()

    result = list_named_ranges(tmp_workbook)
    scopes = {(item["name"], item["local_sheet"]) for item in result}

    assert ("LocalImpact", "Sheet1") in scopes
    assert ("GlobalImpact", None) in scopes


def test_analyze_range_impact_prefers_same_sheet_local_named_range_over_workbook_scope(tmp_workbook):
    workbook = load_workbook(tmp_workbook)
    ws = workbook["Sheet1"]
    ws["D1"] = "=SUM(ScopedImpact)"
    ws.defined_names.add(DefinedName("ScopedImpact", attr_text="Sheet1!$B$2:$B$4"))
    workbook.defined_names["ScopedImpact"] = DefinedName(
        "ScopedImpact",
        attr_text="Sheet1!$A$1:$A$2",
    )
    workbook.save(tmp_workbook)
    workbook.close()

    local_result = analyze_range_impact(tmp_workbook, "Sheet1", "B2:B4")
    global_result = analyze_range_impact(tmp_workbook, "Sheet1", "A1:A2")

    assert local_result["summary"]["dependent_formula_count"] == 1
    assert global_result["summary"]["dependent_formula_count"] == 0
    local_reference = local_result["dependent_formulas"]["sample"][0]["references"][0]
    assert local_reference["via_named_range"] == "ScopedImpact"
    assert local_reference["intersection_range"] == "B2:B4"


def test_analyze_range_impact_tool_returns_json_envelope(tmp_workbook):
    payload = _load_tool_payload(analyze_range_impact_tool(tmp_workbook, "Sheet1", "A1:C3"))

    assert payload["operation"] == "analyze_range_impact"
    assert payload["data"]["range"] == "A1:C3"


def test_profile_workbook_handles_chart_sheets(tmp_path):
    filepath = str(tmp_path / "chartsheet.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Name", "Value"), ("A", 1), ("B", 2), ("C", 3)]:
        ws.append(row)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    categories = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb.save(filepath)
    wb.close()

    result = profile_workbook(filepath)

    assert result["sheet_count"] == 2
    assert result["chart_count"] == 1
    assert result["table_count"] == 0
    assert result["sheets"][0]["sheet_type"] == "worksheet"
    chart_sheet = result["sheets"][1]
    assert chart_sheet["name"] == "Charts"
    assert chart_sheet["sheet_type"] == "chartsheet"
    assert chart_sheet["visibility"] == "visible"
    assert chart_sheet["table_count"] == 0
    assert chart_sheet["chart_count"] == 1
    assert chart_sheet["tables"] == []
    assert chart_sheet["charts"][0]["chart_index"] == 1
    assert chart_sheet["charts"][0]["chart_type"] == "bar"
    assert chart_sheet["charts"][0]["series_count"] == 1
    assert "occupied_range" not in chart_sheet["charts"][0]


def test_get_workbook_info_include_ranges_skips_chart_sheets(tmp_path):
    filepath = str(tmp_path / "chartsheet-ranges.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Name", "Value"), ("A", 1), ("B", 2)]:
        ws.append(row)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb.save(filepath)
    wb.close()

    result = get_workbook_info(filepath, include_ranges=True)

    assert result["sheets"] == ["Data", "Charts"]
    assert result["used_ranges"] == {"Data": "A1:B3"}


def test_get_workbook_metadata_tool_handles_chart_sheets_with_ranges(tmp_path):
    filepath = str(tmp_path / "chartsheet-metadata.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in [("Name", "Value"), ("A", 1), ("B", 2)]:
        ws.append(row)

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet = wb.create_chartsheet("Charts")
    chart_sheet.add_chart(chart)
    wb.save(filepath)
    wb.close()

    payload = json.loads(get_workbook_metadata_tool(filepath, include_ranges=True))

    assert payload["ok"] is True
    assert payload["operation"] == "get_workbook_metadata"
    assert payload["data"]["sheets"] == ["Data", "Charts"]
    assert payload["data"]["used_ranges"] == {"Data": "A1:B3"}

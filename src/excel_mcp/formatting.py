import json
import logging
import re
from typing import Any, Dict, List, Optional

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, Protection, Font,
    Color
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    FormulaRule, CellIsRule
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .workbook import (
    _extract_conditional_format_overlaps,
    _parse_range_reference,
    require_worksheet,
    safe_workbook,
)
from .cell_utils import parse_cell_range, validate_cell_reference
from .exceptions import ValidationError, FormattingError

logger = logging.getLogger(__name__)
HEX_COLOR_RE = re.compile(r"^[0-9A-Fa-f]{6}([0-9A-Fa-f]{2})?$")


def _should_include_changes(dry_run: bool, include_changes: Optional[bool]) -> bool:
    if include_changes is None:
        return dry_run
    return include_changes


def _validate_positive_integer(value: Any, *, argument_name: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise ValidationError(f"{argument_name} must be a positive integer")
    return value


def _normalize_argb_color(value: Any, *, label: str) -> str:
    if not isinstance(value, str):
        raise FormattingError(
            f"{label} must be a hex color like '1F4E78', '#1F4E78', or 'FF1F4E78'"
        )

    normalized = value.strip()
    if normalized.startswith("#"):
        normalized = normalized[1:]
    if not HEX_COLOR_RE.fullmatch(normalized):
        raise FormattingError(
            f"Invalid {label}: '{value}'. Use 6-hex RGB like '1F4E78' or '#1F4E78', "
            "or 8-hex ARGB like 'FF1F4E78'"
        )
    if len(normalized) == 6:
        normalized = f"FF{normalized}"
    return normalized.upper()


def _serialize_color_token(color: Any) -> Optional[str]:
    if color is None:
        return None

    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        rgb = getattr(color, "rgb", None)
        return str(rgb) if rgb else None
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        if theme is None:
            return None
        tint = getattr(color, "tint", None)
        return (
            f"theme:{theme},tint:{tint}"
            if tint not in (None, 0, 0.0)
            else f"theme:{theme}"
        )
    if color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        return None if indexed is None else f"indexed:{indexed}"
    if color_type == "auto" and getattr(color, "auto", False):
        return "auto"
    return None


def _serialize_font(font: Any) -> dict[str, Any]:
    return {
        "name": getattr(font, "name", None),
        "size": getattr(font, "size", None),
        "bold": bool(getattr(font, "bold", False)),
        "italic": bool(getattr(font, "italic", False)),
        "underline": getattr(font, "underline", None),
        "color": _serialize_color_token(getattr(font, "color", None)),
    }


def _serialize_fill(fill: Any) -> Optional[dict[str, Any]]:
    fill_type = getattr(fill, "fill_type", None) or getattr(fill, "patternType", None)
    fg_color = _serialize_color_token(getattr(fill, "fgColor", None))
    bg_color = _serialize_color_token(getattr(fill, "bgColor", None))
    if fill_type is None and fg_color is None and bg_color is None:
        return None
    return {
        "fill_type": fill_type,
        "fg_color": fg_color,
        "bg_color": bg_color,
    }


def _serialize_side(side: Any) -> Optional[dict[str, Any]]:
    style = getattr(side, "style", None)
    color = _serialize_color_token(getattr(side, "color", None))
    if style is None and color is None:
        return None
    return {
        "style": style,
        "color": color,
    }


def _serialize_border(border: Any) -> Optional[dict[str, Any]]:
    serialized = {
        "left": _serialize_side(getattr(border, "left", None)),
        "right": _serialize_side(getattr(border, "right", None)),
        "top": _serialize_side(getattr(border, "top", None)),
        "bottom": _serialize_side(getattr(border, "bottom", None)),
    }
    if not any(value is not None for value in serialized.values()):
        return None
    return serialized


def _serialize_alignment(alignment: Any) -> Optional[dict[str, Any]]:
    if alignment is None:
        return None

    serialized = {
        "horizontal": getattr(alignment, "horizontal", None),
        "vertical": getattr(alignment, "vertical", None),
        "text_rotation": getattr(alignment, "text_rotation", 0) or 0,
        "wrap_text": bool(getattr(alignment, "wrap_text", False)),
        "shrink_to_fit": bool(getattr(alignment, "shrink_to_fit", False)),
    }
    if not any(
        [
            serialized["horizontal"],
            serialized["vertical"],
            serialized["text_rotation"],
            serialized["wrap_text"],
            serialized["shrink_to_fit"],
        ]
    ):
        return None
    return serialized


def _serialize_protection(protection: Any) -> Optional[dict[str, Any]]:
    if protection is None:
        return None

    locked = getattr(protection, "locked", None)
    hidden = getattr(protection, "hidden", None)
    if locked in (None, True) and hidden in (None, False):
        return None
    return {
        "locked": None if locked is None else bool(locked),
        "hidden": None if hidden is None else bool(hidden),
    }


def _serialize_cell_style(cell: Any) -> dict[str, Any]:
    style: dict[str, Any] = {
        "font": _serialize_font(cell.font),
    }

    fill = _serialize_fill(cell.fill)
    if fill is not None:
        style["fill"] = fill

    border = _serialize_border(cell.border)
    if border is not None:
        style["border"] = border

    if cell.number_format not in (None, "", "General"):
        style["number_format"] = cell.number_format

    alignment = _serialize_alignment(cell.alignment)
    if alignment is not None:
        style["alignment"] = alignment

    protection = _serialize_protection(cell.protection)
    if protection is not None:
        style["protection"] = protection

    return style


def _style_signature(style: dict[str, Any]) -> str:
    return json.dumps(style, sort_keys=True, separators=(",", ":"))


def _range_size(bounds: tuple[int, int, int, int]) -> int:
    min_row, min_col, max_row, max_col = bounds
    return (max_row - min_row + 1) * (max_col - min_col + 1)


def _merged_range_overlaps(
    sheet: Worksheet,
    *,
    target_bounds: tuple[int, int, int, int],
) -> list[str]:
    min_row, min_col, max_row, max_col = target_bounds
    overlaps: list[str] = []
    for merged_range in sheet.merged_cells.ranges:
        merged_min_col, merged_min_row, merged_max_col, merged_max_row = (
            merged_range.bounds
        )
        if (
            merged_max_row < min_row
            or merged_min_row > max_row
            or merged_max_col < min_col
            or merged_min_col > max_col
        ):
            continue
        overlaps.append(str(merged_range))
    return overlaps


def read_range_formatting(
    filepath: str,
    sheet_name: str,
    range_ref: str,
    sample_limit: int = 10,
) -> Dict[str, Any]:
    """Read a compact formatting summary for a worksheet range."""
    try:
        _validate_positive_integer(sample_limit, argument_name="sample_limit")

        with safe_workbook(filepath) as wb:
            sheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="formatting inspection",
            )
            bounds, normalized_range = _parse_range_reference(
                range_ref,
                worksheet=sheet,
                expected_sheet=sheet_name,
                error_cls=ValidationError,
            )

            min_row, min_col, max_row, max_col = bounds
            style_groups: dict[str, dict[str, Any]] = {}
            inspected_cell_count = 0
            merged_placeholder_count = 0

            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        merged_placeholder_count += 1
                        continue

                    inspected_cell_count += 1
                    style = _serialize_cell_style(cell)
                    signature = _style_signature(style)
                    group = style_groups.setdefault(
                        signature,
                        {
                            "style": style,
                            "cell_count": 0,
                            "sample_cells": [],
                        },
                    )
                    group["cell_count"] += 1
                    if len(group["sample_cells"]) < 5:
                        group["sample_cells"].append(
                            f"{get_column_letter(cell.column)}{cell.row}"
                        )

            sorted_groups = sorted(
                style_groups.values(),
                key=lambda group: (-group["cell_count"], group["sample_cells"][0]),
            )
            style_group_sample = []
            for style_index, group in enumerate(sorted_groups[:sample_limit], start=1):
                style_group_sample.append(
                    {
                        "style_index": style_index,
                        "cell_count": group["cell_count"],
                        "sample_cells": group["sample_cells"],
                        **group["style"],
                    }
                )

            conditional_formats = _extract_conditional_format_overlaps(
                sheet,
                sheet_name=sheet_name,
                target_bounds=bounds,
            )
            merged_ranges = _merged_range_overlaps(sheet, target_bounds=bounds)

        warnings: list[str] = []
        if len(sorted_groups) > sample_limit:
            warnings.append(
                f"Sampled {sample_limit} of {len(sorted_groups)} style groups; "
                "increase sample_limit for more detail"
            )
        if len(conditional_formats) > sample_limit:
            warnings.append(
                f"Sampled {sample_limit} of {len(conditional_formats)} overlapping conditional-format rules"
            )
        if len(merged_ranges) > sample_limit:
            warnings.append(
                f"Sampled {sample_limit} of {len(merged_ranges)} overlapping merged ranges"
            )

        result = {
            "sheet_name": sheet_name,
            "range": normalized_range,
            "cell_count": _range_size(bounds),
            "inspected_cell_count": inspected_cell_count,
            "merged_placeholder_count": merged_placeholder_count,
            "summary": {
                "style_group_count": len(sorted_groups),
                "uniform_formatting": len(sorted_groups) <= 1,
                "has_conditional_formatting": bool(conditional_formats),
                "has_merged_ranges": bool(merged_ranges),
            },
            "style_groups": style_group_sample,
            "merged_ranges": {
                "count": len(merged_ranges),
                "sample": merged_ranges[:sample_limit],
            },
            "conditional_formats": {
                "count": len(conditional_formats),
                "sample": conditional_formats[:sample_limit],
            },
        }
        if warnings:
            result["warnings"] = warnings
        return result
    except (ValidationError, FormattingError):
        raise
    except Exception as e:
        logger.error(f"Failed to read range formatting: {e}")
        raise FormattingError(str(e))


def _build_format_preview(
    *,
    sheet_name: str,
    range_str: str,
    bold: bool,
    italic: bool,
    underline: bool,
    font_size: Optional[int],
    font_color: Optional[str],
    bg_color: Optional[str],
    border_style: Optional[str],
    border_color: Optional[str],
    number_format: Optional[str],
    alignment: Optional[str],
    wrap_text: bool,
    merge_cells: bool,
    protection: Optional[Dict[str, Any]],
    conditional_format: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    return {
        "sheet_name": sheet_name,
        "range": range_str,
        "font": {
            "bold": bold,
            "italic": italic,
            "underline": underline,
            "font_size": font_size,
            "font_color": font_color,
        },
        "fill": {"bg_color": bg_color} if bg_color is not None else None,
        "border": {
            "border_style": border_style,
            "border_color": border_color,
        }
        if border_style is not None or border_color is not None
        else None,
        "number_format": number_format,
        "alignment": alignment,
        "wrap_text": wrap_text,
        "merge_cells": merge_cells,
        "protection": protection,
        "conditional_format": conditional_format,
    }


def _apply_conditional_format(
    sheet: Worksheet,
    range_str: str,
    conditional_format: Dict[str, Any],
) -> None:
    rule_type = conditional_format.get("type")
    if not rule_type:
        raise FormattingError("Conditional format type not specified")

    nested_params = conditional_format.get("params", {})
    if nested_params is None:
        nested_params = {}
    if not isinstance(nested_params, dict):
        raise FormattingError("Conditional format params must be an object")

    top_level_params = {
        key: value
        for key, value in conditional_format.items()
        if key not in {"type", "params"}
    }
    params = dict(top_level_params)
    params.update(nested_params)

    if rule_type == "cell_is" and "fill" in params:
        fill_params = params["fill"]
        if isinstance(fill_params, dict):
            try:
                fill_color = fill_params.get("fgColor", "FFC7CE")
                fill_color = _normalize_argb_color(
                    fill_color,
                    label="conditional format fill color",
                )
                params["fill"] = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid",
                )
            except FormattingError:
                raise
            except ValueError as e:
                raise FormattingError(
                    f"Invalid conditional format fill color: {str(e)}"
                ) from e

    try:
        if rule_type == "color_scale":
            rule = ColorScaleRule(**params)
        elif rule_type == "data_bar":
            rule = DataBarRule(**params)
        elif rule_type == "icon_set":
            rule = IconSetRule(**params)
        elif rule_type == "formula":
            rule = FormulaRule(**params)
        elif rule_type == "cell_is":
            rule = CellIsRule(**params)
        else:
            raise FormattingError(f"Invalid conditional format type: {rule_type}")

        sheet.conditional_formatting.add(range_str, rule)
    except FormattingError:
        raise
    except Exception as e:
        hint = ""
        if not params:
            hint = (
                " Provide rule parameters under conditional_format.params "
                "or as top-level keys."
            )
        raise FormattingError(
            f"Failed to apply conditional formatting: {str(e)}{hint}"
        ) from e


def _apply_format_to_sheet(
    sheet: Worksheet,
    *,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    if not validate_cell_reference(start_cell):
        raise ValidationError(f"Invalid start cell reference: {start_cell}")

    if end_cell and not validate_cell_reference(end_cell):
        raise ValidationError(f"Invalid end cell reference: {end_cell}")

    try:
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
    except ValueError as e:
        raise ValidationError(f"Invalid cell range: {str(e)}") from e

    if end_row is None:
        end_row = start_row
    if end_col is None:
        end_col = start_col

    range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
    preview = _build_format_preview(
        sheet_name=sheet_name,
        range_str=range_str,
        bold=bold,
        italic=italic,
        underline=underline,
        font_size=font_size,
        font_color=font_color,
        bg_color=bg_color,
        border_style=border_style,
        border_color=border_color,
        number_format=number_format,
        alignment=alignment,
        wrap_text=wrap_text,
        merge_cells=merge_cells,
        protection=protection,
        conditional_format=conditional_format,
    )

    font_args = {
        "bold": bold,
        "italic": italic,
        "underline": "single" if underline else None,
    }
    if font_size is not None:
        font_args["size"] = font_size
    if font_color is not None:
        try:
            font_color = _normalize_argb_color(font_color, label="font color")
            font_args["color"] = Color(rgb=font_color)
        except FormattingError:
            raise
        except ValueError as e:
            raise FormattingError(f"Invalid font color: {str(e)}") from e
    font = Font(**font_args)

    fill = None
    if bg_color is not None:
        try:
            bg_color = _normalize_argb_color(bg_color, label="background color")
            fill = PatternFill(
                start_color=Color(rgb=bg_color),
                end_color=Color(rgb=bg_color),
                fill_type="solid",
            )
        except FormattingError:
            raise
        except ValueError as e:
            raise FormattingError(f"Invalid background color: {str(e)}") from e

    border = None
    if border_style is not None:
        try:
            normalized_border_color = border_color if border_color else "000000"
            normalized_border_color = _normalize_argb_color(
                normalized_border_color,
                label="border color",
            )
            side = Side(style=border_style, color=Color(rgb=normalized_border_color))
            border = Border(left=side, right=side, top=side, bottom=side)
        except FormattingError:
            raise
        except ValueError as e:
            raise FormattingError(f"Invalid border settings: {str(e)}") from e

    align = None
    if alignment is not None or wrap_text:
        try:
            align = Alignment(
                horizontal=alignment,
                vertical="center",
                wrap_text=wrap_text,
            )
        except ValueError as e:
            raise FormattingError(f"Invalid alignment settings: {str(e)}") from e

    protect = None
    if protection is not None:
        try:
            protect = Protection(**protection)
        except ValueError as e:
            raise FormattingError(f"Invalid protection settings: {str(e)}") from e

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = font
            if fill is not None:
                cell.fill = fill
            if border is not None:
                cell.border = border
            if align is not None:
                cell.alignment = align
            if protect is not None:
                cell.protection = protect
            if number_format is not None:
                cell.number_format = number_format

    if merge_cells and end_cell:
        try:
            sheet.merge_cells(range_str)
        except ValueError as e:
            raise FormattingError(f"Failed to merge cells: {str(e)}") from e

    if conditional_format is not None:
        _apply_conditional_format(sheet, range_str, conditional_format)

    return {"range": range_str, "preview": preview}

def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Apply formatting to a range of cells.
    
    This function handles all Excel formatting operations including:
    - Font properties (bold, italic, size, color, etc.)
    - Cell fill/background color
    - Borders (style and color)
    - Number formatting
    - Alignment and text wrapping
    - Cell merging
    - Protection
    - Conditional formatting
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell reference
        end_cell: Optional ending cell reference
        bold: Whether to make text bold
        italic: Whether to make text italic
        underline: Whether to underline text
        font_size: Font size in points
        font_color: Font color (hex code)
        bg_color: Background color (hex code)
        border_style: Border style (thin, medium, thick, double)
        border_color: Border color (hex code)
        number_format: Excel number format string
        alignment: Text alignment (left, center, right, justify)
        wrap_text: Whether to wrap text
        merge_cells: Whether to merge the range
        protection: Cell protection settings
        conditional_format: Conditional formatting rules
        
    Returns:
        Dictionary with operation status
    """
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            sheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="formatting cells",
            )
            applied = _apply_format_to_sheet(
                sheet,
                sheet_name=sheet_name,
                start_cell=start_cell,
                end_cell=end_cell,
                bold=bold,
                italic=italic,
                underline=underline,
                font_size=font_size,
                font_color=font_color,
                bg_color=bg_color,
                border_style=border_style,
                border_color=border_color,
                number_format=number_format,
                alignment=alignment,
                wrap_text=wrap_text,
                merge_cells=merge_cells,
                protection=protection,
                conditional_format=conditional_format,
            )

        result = {
            "message": f"{'Previewed' if dry_run else 'Applied'} formatting to range {applied['range']}",
            "range": applied["range"],
            "dry_run": dry_run,
        }
        if _should_include_changes(dry_run, include_changes):
            result["changes"] = [applied["preview"]]
        return result

    except (ValidationError, FormattingError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to apply formatting: {e}")
        raise FormattingError(str(e))


def format_ranges(
    filepath: str,
    sheet_name: str,
    ranges: List[Dict[str, Any]],
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Apply formatting to multiple ranges in a single workbook pass."""
    try:
        if not ranges:
            raise FormattingError("At least one format operation must be provided")

        with safe_workbook(filepath, save=not dry_run) as wb:
            sheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="formatting cells",
            )
            applied_ranges: List[str] = []
            previews: List[Dict[str, Any]] = []
            errors: List[Dict[str, Any]] = []

            for index, operation in enumerate(ranges, start=1):
                if not isinstance(operation, dict):
                    raise FormattingError(
                        f"Format operation at index {index} must be an object"
                    )

                operation_data = dict(operation)
                operation_data.pop("dry_run", None)
                operation_data.pop("include_changes", None)
                range_label = operation_data.get("start_cell")
                if operation_data.get("end_cell"):
                    range_label = f"{range_label}:{operation_data['end_cell']}"

                try:
                    applied = _apply_format_to_sheet(
                        sheet,
                        sheet_name=sheet_name,
                        **operation_data,
                    )
                except (ValidationError, FormattingError) as e:
                    errors.append(
                        {
                            "index": index,
                            "range": range_label,
                            "error": str(e),
                        }
                    )
                    continue
                except TypeError as e:
                    errors.append(
                        {
                            "index": index,
                            "range": range_label,
                            "error": f"Invalid format operation at index {index}: {str(e)}",
                        }
                    )
                    continue

                applied_ranges.append(applied["range"])
                previews.append(applied["preview"])

        ranges_failed = len(errors)
        if applied_ranges and not errors:
            message = (
                f"{'Previewed' if dry_run else 'Applied'} formatting to "
                f"{len(applied_ranges)} range(s) in sheet '{sheet_name}'"
            )
        elif applied_ranges:
            message = (
                f"{'Previewed' if dry_run else 'Applied'} formatting to "
                f"{len(applied_ranges)} range(s) in sheet '{sheet_name}'; "
                f"{ranges_failed} range(s) failed"
            )
        else:
            message = (
                f"No formatting operations were applied in sheet '{sheet_name}'; "
                f"{ranges_failed} range(s) failed"
            )

        result = {
            "message": message,
            "sheet_name": sheet_name,
            "ranges_formatted": len(applied_ranges),
            "ranges_failed": ranges_failed,
            "ranges": applied_ranges,
            "dry_run": dry_run,
        }
        if errors:
            result["errors"] = errors
            result["warnings"] = [
                f"{ranges_failed} range(s) failed during batch formatting"
            ]
        if _should_include_changes(dry_run, include_changes):
            result["changes"] = previews
        return result

    except (ValidationError, FormattingError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to apply batch formatting: {e}")
        raise FormattingError(str(e))

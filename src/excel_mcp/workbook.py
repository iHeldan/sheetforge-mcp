import logging
import os
import re
import shutil
import tempfile
from collections import Counter, OrderedDict, deque
from contextlib import contextmanager, suppress
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.utils.cell import quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import WorkbookError

logger = logging.getLogger(__name__)

AUDIT_SAMPLE_ROWS = 25
AUDIT_LARGE_DATASET_THRESHOLD = 1000
AUDIT_SEVERITY_RANK = {"low": 1, "medium": 2, "high": 3}
SUPPORTED_REPAIR_TYPES = {
    "remove_broken_named_ranges",
    "remove_broken_validations",
    "remove_broken_conditional_formats",
    "reveal_hidden_sheets",
}
_STRUCTURED_REFERENCE_FLAG_RANGE_RE = re.compile(
    r"^\[\[(?P<flag>#[^,\]]+)\],\[(?P<start>[^\]]+)\]:\[(?P<end>[^\]]+)\]\]$"
)
_STRUCTURED_REFERENCE_FLAG_COLUMN_RE = re.compile(
    r"^\[\[(?P<flag>#[^,\]]+)\],\[(?P<column>[^\]]+)\]\]$"
)
_STRUCTURED_REFERENCE_COLUMN_RANGE_RE = re.compile(
    r"^\[\[(?P<start>[^\]]+)\]:\[(?P<end>[^\]]+)\]\]$"
)
_STRUCTURED_REFERENCE_FLAG_ONLY_RE = re.compile(r"^\[(?P<flag>#[^\]]+)\]$")
_STRUCTURED_REFERENCE_THIS_ROW_COLUMN_RE = re.compile(r"^\[@(?P<column>[^\]]+)\]$")
_STRUCTURED_REFERENCE_COLUMN_RE = re.compile(r"^\[(?P<column>[^\]]+)\]$")
_SHEET_REFERENCE_TOKEN_RE = re.compile(r"(?<!\])(?P<sheet>'(?:[^']|'')+'|[A-Za-z0-9_.]+)!")


def _get_sheet_usage(ws) -> tuple[int, int, str | None, bool]:
    """Return user-facing sheet dimensions instead of openpyxl's default 1x1 empty sheet."""
    is_empty = ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None
    if is_empty:
        return 0, 0, None, True
    return ws.max_row, ws.max_column, f"A-{get_column_letter(ws.max_column)}", False


def _get_used_range(ws) -> str | None:
    rows, columns, _, is_empty = _get_sheet_usage(ws)
    if is_empty or rows == 0 or columns == 0:
        return None
    return f"A1:{get_column_letter(columns)}{rows}"


def _bounds_to_range(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _normalize_sheet_reference_name(raw_sheet_name: Any) -> str:
    cleaned = str(raw_sheet_name).strip()
    if len(cleaned) >= 2 and cleaned[0] == "'" and cleaned[-1] == "'":
        cleaned = cleaned[1:-1]
    return cleaned.replace("''", "'")


def _rewrite_sheet_references_in_text(
    text: Any,
    *,
    old_sheet_name: str,
    new_sheet_name: str,
) -> Any:
    if not isinstance(text, str) or "!" not in text:
        return text

    quoted_new_sheet = quote_sheetname(new_sheet_name)

    def _replace(match: re.Match[str]) -> str:
        if _normalize_sheet_reference_name(match.group("sheet")) != old_sheet_name:
            return match.group(0)
        return f"{quoted_new_sheet}!"

    return _SHEET_REFERENCE_TOKEN_RE.sub(_replace, text)


def _update_defined_name_sheet_references(
    wb: Any,
    *,
    old_sheet_name: str,
    new_sheet_name: str,
) -> int:
    updated_count = 0
    seen_ids: set[int] = set()

    for _, defined_name, _ in _iter_defined_name_entries(wb):
        defined_name_id = id(defined_name)
        if defined_name_id in seen_ids:
            continue
        seen_ids.add(defined_name_id)

        original_text = getattr(defined_name, "attr_text", None)
        if original_text is None:
            original_text = str(getattr(defined_name, "value", "") or "")

        updated_text = _rewrite_sheet_references_in_text(
            original_text,
            old_sheet_name=old_sheet_name,
            new_sheet_name=new_sheet_name,
        )
        if updated_text == original_text:
            continue
        defined_name.attr_text = updated_text
        updated_count += 1

    return updated_count


def _bounds_intersect(
    first: tuple[int, int, int, int],
    second: tuple[int, int, int, int],
) -> bool:
    return not (
        first[2] < second[0]
        or second[2] < first[0]
        or first[3] < second[1]
        or second[3] < first[1]
    )


def _intersection_bounds(
    first: tuple[int, int, int, int],
    second: tuple[int, int, int, int],
) -> tuple[int, int, int, int] | None:
    if not _bounds_intersect(first, second):
        return None
    return (
        max(first[0], second[0]),
        max(first[1], second[1]),
        min(first[2], second[2]),
        min(first[3], second[3]),
    )


def _parse_range_reference(
    range_ref: str,
    *,
    worksheet: Worksheet | None = None,
    expected_sheet: str | None = None,
    error_cls: type[Exception] = WorkbookError,
) -> tuple[tuple[int, int, int, int], str]:
    cleaned = str(range_ref).strip()
    if not cleaned:
        raise error_cls("Range reference is required")

    if "!" in cleaned:
        range_sheet, cleaned = cleaned.rsplit("!", 1)
        normalized_sheet = _normalize_sheet_reference_name(range_sheet)
        if expected_sheet is not None and normalized_sheet != expected_sheet:
            raise error_cls(
                f"Range '{range_ref}' refers to sheet '{normalized_sheet}', expected '{expected_sheet}'"
            )

    cleaned = cleaned.replace("$", "")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cleaned)
    except ValueError as exc:
        raise error_cls(f"Invalid range reference: {range_ref}") from exc

    if None in (min_col, min_row, max_col, max_row):
        if worksheet is None:
            raise error_cls(f"Range reference requires worksheet bounds: {range_ref}")
        min_col = 1 if min_col is None else min_col
        min_row = 1 if min_row is None else min_row
        max_col = worksheet.max_column if max_col is None else max_col
        max_row = worksheet.max_row if max_row is None else max_row

    return (min_row, min_col, max_row, max_col), _bounds_to_range(min_row, min_col, max_row, max_col)


def _iter_range_references(
    range_ref: Any,
    *,
    worksheet: Worksheet | None = None,
    expected_sheet: str | None = None,
) -> list[tuple[tuple[int, int, int, int], str]]:
    if not range_ref:
        return []

    if hasattr(range_ref, "ranges"):
        parts = [str(item) for item in range_ref.ranges]
    elif isinstance(range_ref, (list, tuple)):
        parts = range_ref
    else:
        parts = str(range_ref).split(",")
    references: list[tuple[tuple[int, int, int, int], str]] = []
    for part in parts:
        cleaned = str(part).strip()
        if not cleaned:
            continue

        if "!" in cleaned:
            range_sheet, local_range = cleaned.rsplit("!", 1)
            normalized_sheet = _normalize_sheet_reference_name(range_sheet)
            if expected_sheet is not None and normalized_sheet != expected_sheet:
                continue
        else:
            local_range = cleaned

        local_range = local_range.replace("$", "")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(local_range)
        except ValueError:
            continue
        if None in (min_col, min_row, max_col, max_row):
            if worksheet is None:
                continue
            min_col = 1 if min_col is None else min_col
            min_row = 1 if min_row is None else min_row
            max_col = worksheet.max_column if max_col is None else max_col
            max_row = worksheet.max_row if max_row is None else max_row
        references.append(
            ((min_row, min_col, max_row, max_col), _bounds_to_range(min_row, min_col, max_row, max_col))
        )
    return references


def _cell_is_within_bounds(
    sheet_name: str,
    row_index: int,
    column_index: int,
    target_sheet: str,
    target_bounds: tuple[int, int, int, int],
) -> bool:
    return (
        sheet_name == target_sheet
        and target_bounds[0] <= row_index <= target_bounds[2]
        and target_bounds[1] <= column_index <= target_bounds[3]
    )


def _defined_name_local_sheet(
    wb: Any,
    defined_name: Any,
) -> str | None:
    local_sheet_id = getattr(defined_name, "localSheetId", None)
    if local_sheet_id is None:
        return None
    try:
        return wb.sheetnames[local_sheet_id]
    except Exception:
        return None


def _iter_defined_name_entries(wb: Any):
    seen: set[tuple[str, str | None]] = set()

    for name, defined_name in wb.defined_names.items():
        local_sheet = _defined_name_local_sheet(wb, defined_name)
        entry_key = (name, local_sheet)
        if entry_key in seen:
            continue
        seen.add(entry_key)
        yield name, defined_name, local_sheet

    for ws in getattr(wb, "worksheets", []):
        for name, defined_name in ws.defined_names.items():
            entry_key = (name, ws.title)
            if entry_key in seen:
                continue
            seen.add(entry_key)
            yield name, defined_name, ws.title


def _resolve_defined_name(
    wb: Any,
    *,
    name: str,
    formula_sheet_name: str,
    scope_sheet_name: str | None = None,
) -> tuple[Any, str | None] | None:
    if scope_sheet_name is not None and scope_sheet_name in wb.sheetnames:
        scoped_ws = wb[scope_sheet_name]
        scoped_name = scoped_ws.defined_names.get(name)
        if scoped_name is not None:
            return scoped_name, scope_sheet_name

    if scope_sheet_name is not None:
        scoped_name = wb.defined_names.get(name)
        if scoped_name is None:
            return None
        local_sheet = _defined_name_local_sheet(wb, scoped_name)
        if local_sheet != scope_sheet_name:
            return None
        return scoped_name, local_sheet

    if formula_sheet_name in wb.sheetnames:
        local_ws = wb[formula_sheet_name]
        local_name = local_ws.defined_names.get(name)
        if local_name is not None:
            return local_name, formula_sheet_name

    workbook_name = wb.defined_names.get(name)
    if workbook_name is None:
        return None

    local_sheet = _defined_name_local_sheet(wb, workbook_name)
    if local_sheet is not None and local_sheet != formula_sheet_name:
        return None

    return workbook_name, local_sheet


def _resolve_named_range_references(
    wb: Any,
    *,
    name: str,
    formula_sheet_name: str,
    target_sheet: str,
    target_bounds: tuple[int, int, int, int],
    scope_sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    resolved = _resolve_defined_name(
        wb,
        name=name,
        formula_sheet_name=formula_sheet_name,
        scope_sheet_name=scope_sheet_name,
    )
    if resolved is None:
        return []

    defined_name, _ = resolved

    matches: list[dict[str, Any]] = []
    try:
        destinations = list(defined_name.destinations)
    except Exception:
        destinations = []

    for destination_sheet_name, destination_range in destinations:
        if destination_sheet_name != target_sheet or destination_sheet_name not in wb.sheetnames:
            continue

        destination_ws = wb[destination_sheet_name]
        if _sheet_type(destination_ws) == "chartsheet":
            continue

        for destination_bounds, destination_ref in _iter_range_references(
            destination_range,
            worksheet=destination_ws,
            expected_sheet=target_sheet,
        ):
            intersection = _intersection_bounds(target_bounds, destination_bounds)
            if intersection is None:
                continue
            matches.append(
                {
                    "reference": f"{destination_sheet_name}!{destination_ref}",
                    "intersection_range": _bounds_to_range(*intersection),
                    "via_named_range": defined_name.name,
                }
            )

    return matches


def _find_table_reference(
    wb: Any,
    *,
    table_name: str,
    sheet_name: str | None = None,
) -> tuple[str, Worksheet, Any] | None:
    sheet_names = [sheet_name] if sheet_name is not None else list(wb.sheetnames)
    for current_sheet_name in sheet_names:
        if current_sheet_name not in wb.sheetnames:
            continue
        ws = wb[current_sheet_name]
        if _sheet_type(ws) == "chartsheet":
            continue
        for table in ws.tables.values():
            if table.displayName == table_name:
                return current_sheet_name, ws, table
    return None


def _table_column_lookup(ws: Worksheet, table: Any) -> dict[str, int]:
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    column_names = list(getattr(table, "column_names", []) or [])
    if len(column_names) == (max_col - min_col + 1):
        return {
            str(column_name): min_col + offset
            for offset, column_name in enumerate(column_names)
        }

    return {
        str(ws.cell(row=min_row, column=column_index).value): column_index
        for column_index in range(min_col, max_col + 1)
    }


def _normalize_structured_flag(flag: str | None) -> str:
    if not flag:
        return "data"
    normalized = flag.strip().lower()
    if normalized.startswith("#"):
        normalized = normalized[1:]
    return normalized.replace(" ", "")


def _parse_structured_reference(
    local_reference: str,
) -> tuple[str, dict[str, Any]] | None:
    if "[" not in local_reference or not local_reference.endswith("]"):
        return None

    table_name, spec = local_reference.split("[", 1)
    table_name = table_name.strip()
    if not table_name:
        return None
    spec = f"[{spec}"

    match = _STRUCTURED_REFERENCE_FLAG_RANGE_RE.fullmatch(spec)
    if match:
        return table_name, {
            "row_selector": _normalize_structured_flag(match.group("flag")),
            "start_column": match.group("start"),
            "end_column": match.group("end"),
        }

    match = _STRUCTURED_REFERENCE_FLAG_COLUMN_RE.fullmatch(spec)
    if match:
        return table_name, {
            "row_selector": _normalize_structured_flag(match.group("flag")),
            "start_column": match.group("column"),
            "end_column": match.group("column"),
        }

    match = _STRUCTURED_REFERENCE_COLUMN_RANGE_RE.fullmatch(spec)
    if match:
        return table_name, {
            "row_selector": "data",
            "start_column": match.group("start"),
            "end_column": match.group("end"),
        }

    match = _STRUCTURED_REFERENCE_FLAG_ONLY_RE.fullmatch(spec)
    if match:
        return table_name, {
            "row_selector": _normalize_structured_flag(match.group("flag")),
            "start_column": None,
            "end_column": None,
        }

    match = _STRUCTURED_REFERENCE_THIS_ROW_COLUMN_RE.fullmatch(spec)
    if match:
        return table_name, {
            "row_selector": "thisrow",
            "start_column": match.group("column"),
            "end_column": match.group("column"),
        }

    match = _STRUCTURED_REFERENCE_COLUMN_RE.fullmatch(spec)
    if match:
        return table_name, {
            "row_selector": "data",
            "start_column": match.group("column"),
            "end_column": match.group("column"),
        }

    return None


def _resolve_table_structured_reference(
    wb: Any,
    *,
    local_reference: str,
    formula_sheet_name: str,
    formula_row: int,
    scope_sheet_name: str | None,
    target_sheet: str,
    target_bounds: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    parsed_reference = _parse_structured_reference(local_reference)
    if parsed_reference is None:
        return []

    table_name, reference_parts = parsed_reference
    table_match = _find_table_reference(wb, table_name=table_name, sheet_name=scope_sheet_name)
    if table_match is None:
        return []

    table_sheet_name, table_ws, table = table_match
    if table_sheet_name != target_sheet:
        return []

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    header_row_count = int(table.headerRowCount or 0)
    totals_row_count = int(table.totalsRowCount or 0)
    data_start_row = min_row + header_row_count
    data_end_row = max_row - totals_row_count

    row_selector = reference_parts["row_selector"]
    if row_selector == "all":
        row_start, row_end = min_row, max_row
    elif row_selector == "data":
        row_start, row_end = data_start_row, data_end_row
    elif row_selector == "headers":
        if header_row_count < 1:
            return []
        row_start, row_end = min_row, min_row + header_row_count - 1
    elif row_selector == "totals":
        if totals_row_count < 1:
            return []
        row_start, row_end = max_row - totals_row_count + 1, max_row
    elif row_selector == "thisrow":
        if table_sheet_name != formula_sheet_name or not (data_start_row <= formula_row <= data_end_row):
            return []
        row_start = row_end = formula_row
    else:
        return []

    if row_start > row_end:
        return []

    start_column_name = reference_parts["start_column"]
    end_column_name = reference_parts["end_column"]
    if start_column_name is None:
        column_start, column_end = min_col, max_col
    else:
        column_lookup = _table_column_lookup(table_ws, table)
        if start_column_name not in column_lookup or end_column_name not in column_lookup:
            return []
        column_start = column_lookup[start_column_name]
        column_end = column_lookup[end_column_name]
        if column_start > column_end:
            column_start, column_end = column_end, column_start

    reference_bounds = (row_start, column_start, row_end, column_end)
    intersection = _intersection_bounds(target_bounds, reference_bounds)
    if intersection is None:
        return []

    return [
        {
            "reference": f"{table_sheet_name}!{_bounds_to_range(*reference_bounds)}",
            "intersection_range": _bounds_to_range(*intersection),
            "via_table": table.displayName,
            "structured_reference": local_reference,
        }
    ]


def _resolve_table_structured_reference_targets(
    wb: Any,
    *,
    local_reference: str,
    formula_sheet_name: str,
    formula_row: int,
    scope_sheet_name: str | None,
) -> list[dict[str, Any]]:
    parsed_reference = _parse_structured_reference(local_reference)
    if parsed_reference is None:
        return []

    table_name, reference_parts = parsed_reference
    table_match = _find_table_reference(wb, table_name=table_name, sheet_name=scope_sheet_name)
    if table_match is None:
        return []

    table_sheet_name, table_ws, table = table_match
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    header_row_count = int(table.headerRowCount or 0)
    totals_row_count = int(table.totalsRowCount or 0)
    data_start_row = min_row + header_row_count
    data_end_row = max_row - totals_row_count

    row_selector = reference_parts["row_selector"]
    if row_selector == "all":
        row_start, row_end = min_row, max_row
    elif row_selector == "data":
        row_start, row_end = data_start_row, data_end_row
    elif row_selector == "headers":
        if header_row_count < 1:
            return []
        row_start, row_end = min_row, min_row + header_row_count - 1
    elif row_selector == "totals":
        if totals_row_count < 1:
            return []
        row_start, row_end = max_row - totals_row_count + 1, max_row
    elif row_selector == "thisrow":
        if table_sheet_name != formula_sheet_name or not (data_start_row <= formula_row <= data_end_row):
            return []
        row_start = row_end = formula_row
    else:
        return []

    if row_start > row_end:
        return []

    start_column_name = reference_parts["start_column"]
    end_column_name = reference_parts["end_column"]
    if start_column_name is None:
        column_start, column_end = min_col, max_col
    else:
        column_lookup = _table_column_lookup(table_ws, table)
        if start_column_name not in column_lookup or end_column_name not in column_lookup:
            return []
        column_start = column_lookup[start_column_name]
        column_end = column_lookup[end_column_name]
        if column_start > column_end:
            column_start, column_end = column_end, column_start

    resolved_range = _bounds_to_range(row_start, column_start, row_end, column_end)
    return [
        {
            "sheet_name": table_sheet_name,
            "range": resolved_range,
            "reference": f"{table_sheet_name}!{resolved_range}",
            "reference_type": "structured_reference",
            "table_name": table.displayName,
            "structured_reference": local_reference,
        }
    ]


def _resolve_formula_reference_targets(
    wb: Any,
    *,
    token_value: str,
    formula_sheet_name: str,
    formula_row: int,
) -> list[dict[str, Any]]:
    local_reference = token_value
    reference_scope_sheet = None
    if "!" in token_value:
        range_sheet, local_reference = token_value.rsplit("!", 1)
        reference_scope_sheet = _normalize_sheet_reference_name(range_sheet)

    table_targets = _resolve_table_structured_reference_targets(
        wb,
        local_reference=local_reference,
        formula_sheet_name=formula_sheet_name,
        formula_row=formula_row,
        scope_sheet_name=reference_scope_sheet,
    )
    if table_targets:
        return table_targets

    reference_sheet_name = formula_sheet_name if reference_scope_sheet is None else reference_scope_sheet
    if reference_scope_sheet is not None and reference_sheet_name not in wb.sheetnames:
        return [
            {
                "reference": token_value,
                "reference_type": "worksheet_range",
                "broken_reference": True,
                "missing_sheet": reference_sheet_name,
            }
        ]

    if reference_sheet_name in wb.sheetnames:
        reference_ws = wb[reference_sheet_name]
        if _sheet_type(reference_ws) != "chartsheet":
            try:
                _, normalized_reference = _parse_range_reference(
                    local_reference,
                    worksheet=reference_ws,
                    error_cls=WorkbookError,
                )
            except WorkbookError:
                pass
            else:
                return [
                    {
                        "sheet_name": reference_sheet_name,
                        "range": normalized_reference,
                        "reference": f"{reference_sheet_name}!{normalized_reference}",
                        "reference_type": "worksheet_range",
                    }
                ]

    if "[" in local_reference:
        return [
            {
                "reference": token_value,
                "reference_type": "structured_reference",
                "broken_reference": True,
            }
        ]

    resolved_named_range = _resolve_defined_name(
        wb,
        name=local_reference,
        formula_sheet_name=formula_sheet_name,
        scope_sheet_name=reference_scope_sheet,
    )
    if resolved_named_range is None:
        return []

    defined_name, local_sheet = resolved_named_range
    try:
        destinations = list(defined_name.destinations)
    except Exception:
        destinations = []

    if not destinations:
        return [
            {
                "reference": local_reference,
                "reference_type": "named_range",
                "named_range": local_reference,
                "local_sheet": local_sheet,
                "broken_reference": True,
            }
        ]

    resolved_targets: list[dict[str, Any]] = []
    for destination_sheet_name, destination_range in destinations:
        if destination_sheet_name not in wb.sheetnames:
            resolved_targets.append(
                {
                    "reference": f"{destination_sheet_name}!{destination_range}",
                    "reference_type": "named_range",
                    "named_range": local_reference,
                    "local_sheet": local_sheet,
                    "broken_reference": True,
                    "missing_sheet": destination_sheet_name,
                }
            )
            continue

        destination_ws = wb[destination_sheet_name]
        if _sheet_type(destination_ws) == "chartsheet":
            resolved_targets.append(
                {
                    "reference": f"{destination_sheet_name}!{destination_range}",
                    "reference_type": "named_range",
                    "named_range": local_reference,
                    "local_sheet": local_sheet,
                    "broken_reference": True,
                }
            )
            continue

        for _, normalized_reference in _iter_range_references(
            destination_range,
            worksheet=destination_ws,
            expected_sheet=destination_sheet_name,
        ):
            resolved_targets.append(
                {
                    "sheet_name": destination_sheet_name,
                    "range": normalized_reference,
                    "reference": f"{destination_sheet_name}!{normalized_reference}",
                    "reference_type": "named_range",
                    "named_range": local_reference,
                    "local_sheet": local_sheet,
                }
            )

    return resolved_targets


def _resolve_formula_token_references(
    wb: Any,
    *,
    token_value: str,
    formula_sheet_name: str,
    formula_row: int,
    target_sheet: str,
    target_bounds: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    reference_scope_sheet = None
    local_reference = token_value
    if "!" in token_value:
        range_sheet, local_reference = token_value.rsplit("!", 1)
        reference_scope_sheet = _normalize_sheet_reference_name(range_sheet)

    table_matches = _resolve_table_structured_reference(
        wb,
        local_reference=local_reference,
        formula_sheet_name=formula_sheet_name,
        formula_row=formula_row,
        scope_sheet_name=reference_scope_sheet,
        target_sheet=target_sheet,
        target_bounds=target_bounds,
    )
    if table_matches:
        return table_matches

    reference_sheet_name = formula_sheet_name if reference_scope_sheet is None else reference_scope_sheet
    if reference_sheet_name == target_sheet and reference_sheet_name in wb.sheetnames:
        reference_ws = wb[reference_sheet_name]
        if _sheet_type(reference_ws) != "chartsheet":
            try:
                reference_bounds, normalized_reference = _parse_range_reference(
                    local_reference,
                    worksheet=reference_ws,
                    error_cls=WorkbookError,
                )
            except WorkbookError:
                pass
            else:
                intersection = _intersection_bounds(target_bounds, reference_bounds)
                if intersection is not None:
                    return [
                        {
                            "reference": f"{reference_sheet_name}!{normalized_reference}",
                            "intersection_range": _bounds_to_range(*intersection),
                        }
                    ]

    if "[" in local_reference:
        return []

    return _resolve_named_range_references(
        wb,
        name=local_reference,
        formula_sheet_name=formula_sheet_name,
        target_sheet=target_sheet,
        target_bounds=target_bounds,
        scope_sheet_name=reference_scope_sheet,
    )


def _formula_text_has_broken_reference(
    wb: Any,
    *,
    formula_text: Any,
    formula_sheet_name: str,
) -> bool:
    if formula_text is None:
        return False

    normalized_formula = str(formula_text).strip()
    if not normalized_formula:
        return False
    if "#REF!" in normalized_formula.upper():
        return True
    if not normalized_formula.startswith("="):
        normalized_formula = f"={normalized_formula}"

    try:
        tokenizer = Tokenizer(normalized_formula)
    except Exception:
        return False

    for token in tokenizer.items:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue

        token_value = str(token.value).strip()
        if not token_value:
            continue
        if "#REF!" in token_value.upper():
            return True

        local_reference = token_value
        reference_scope_sheet = None
        if "!" in token_value:
            range_sheet, local_reference = token_value.rsplit("!", 1)
            reference_scope_sheet = _normalize_sheet_reference_name(range_sheet)
            if reference_scope_sheet and reference_scope_sheet not in wb.sheetnames:
                return True

        resolved_targets = _resolve_formula_reference_targets(
            wb,
            token_value=token_value,
            formula_sheet_name=formula_sheet_name,
            formula_row=1,
        )
        if any(target.get("broken_reference") for target in resolved_targets):
            return True

        if "[" in local_reference:
            continue

        resolved_named_range = _resolve_defined_name(
            wb,
            name=local_reference,
            formula_sheet_name=formula_sheet_name,
            scope_sheet_name=reference_scope_sheet,
        )
        if resolved_named_range is None:
            continue

        defined_name, _ = resolved_named_range
        try:
            destinations = list(defined_name.destinations)
        except Exception:
            destinations = []

        if any(sheet_name not in wb.sheetnames for sheet_name, _ in destinations):
            return True

    return False


def _extract_formula_dependencies(
    wb: Any,
    *,
    target_sheet: str,
    target_bounds: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    def _dedupe_reference_matches(references: list[dict[str, Any]]) -> list[dict[str, Any]]:
        deduped: list[dict[str, Any]] = []
        seen: set[tuple[tuple[str, str], ...]] = set()
        for reference in references:
            key = tuple(sorted((str(item_key), str(item_value)) for item_key, item_value in reference.items()))
            if key in seen:
                continue
            seen.add(key)
            deduped.append(reference)
        return deduped

    def _dedupe_formula_cells(cells: list[dict[str, Any]]) -> list[dict[str, Any]]:
        deduped: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        for cell in cells:
            key = (cell["sheet_name"], cell["cell"])
            if key in seen:
                continue
            seen.add(key)
            deduped.append(cell)
        return deduped

    formula_entries: list[dict[str, Any]] = []

    for formula_sheet_name in wb.sheetnames:
        formula_ws = wb[formula_sheet_name]
        if _sheet_type(formula_ws) == "chartsheet":
            continue

        for row in formula_ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                if _cell_is_within_bounds(
                    formula_sheet_name,
                    cell.row,
                    cell.column,
                    target_sheet,
                    target_bounds,
                ):
                    continue

                matched_references: list[dict[str, Any]] = []
                try:
                    tokenizer = Tokenizer(cell.value)
                except Exception:
                    continue

                token_values: list[str] = []
                for token in tokenizer.items:
                    if token.type != "OPERAND" or token.subtype != "RANGE":
                        continue

                    token_value = str(token.value).strip()
                    if not token_value:
                        continue

                    token_values.append(token_value)
                    matched_references.extend(
                        _resolve_formula_token_references(
                            wb,
                            token_value=token_value,
                            formula_sheet_name=formula_sheet_name,
                            formula_row=cell.row,
                            target_sheet=target_sheet,
                            target_bounds=target_bounds,
                        )
                    )

                formula_entries.append(
                    {
                        "sheet_name": formula_sheet_name,
                        "cell": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "formula": cell.value,
                        "token_values": token_values,
                        "direct_references": _dedupe_reference_matches(matched_references),
                    }
                )

    dependencies: list[dict[str, Any]] = []
    formula_entry_lookup = {
        (entry["sheet_name"], entry["cell"]): entry for entry in formula_entries
    }

    direct_frontier: set[tuple[str, str]] = set()
    for entry in formula_entries:
        if not entry["direct_references"]:
            continue
        direct_frontier.add((entry["sheet_name"], entry["cell"]))
        dependencies.append(
            {
                "sheet_name": entry["sheet_name"],
                "cell": entry["cell"],
                "formula": entry["formula"],
                "references": entry["direct_references"],
                "dependency_depth": 1,
                "dependency_type": "direct",
            }
        )

    discovered = set(direct_frontier)
    frontier = set(direct_frontier)
    depth = 2

    while frontier:
        frontier_entries = [formula_entry_lookup[key] for key in frontier]
        next_frontier: set[tuple[str, str]] = set()

        for entry in formula_entries:
            entry_key = (entry["sheet_name"], entry["cell"])
            if entry_key in discovered:
                continue

            matched_references: list[dict[str, Any]] = []
            transitive_via: list[dict[str, Any]] = []
            for predecessor in frontier_entries:
                predecessor_bounds = (
                    predecessor["row"],
                    predecessor["column"],
                    predecessor["row"],
                    predecessor["column"],
                )
                predecessor_matches: list[dict[str, Any]] = []
                for token_value in entry["token_values"]:
                    predecessor_matches.extend(
                        _resolve_formula_token_references(
                            wb,
                            token_value=token_value,
                            formula_sheet_name=entry["sheet_name"],
                            formula_row=entry["row"],
                            target_sheet=predecessor["sheet_name"],
                            target_bounds=predecessor_bounds,
                        )
                    )

                if predecessor_matches:
                    matched_references.extend(predecessor_matches)
                    transitive_via.append(
                        {
                            "sheet_name": predecessor["sheet_name"],
                            "cell": predecessor["cell"],
                        }
                    )

            matched_references = _dedupe_reference_matches(matched_references)
            transitive_via = _dedupe_formula_cells(transitive_via)
            if not matched_references:
                continue

            dependencies.append(
                {
                    "sheet_name": entry["sheet_name"],
                    "cell": entry["cell"],
                    "formula": entry["formula"],
                    "references": matched_references,
                    "dependency_depth": depth,
                    "dependency_type": "transitive",
                    "transitive_via": transitive_via,
                }
            )
            discovered.add(entry_key)
            next_frontier.add(entry_key)

        frontier = next_frontier
        depth += 1

    return dependencies


def _extract_reference_matches_from_formula_text(
    wb: Any,
    *,
    formula_text: Any,
    formula_sheet_name: str,
    formula_row: int,
    target_sheet: str,
    target_bounds: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    if formula_text is None:
        return []

    normalized_formula = str(formula_text).strip()
    if not normalized_formula:
        return []
    if not normalized_formula.startswith("="):
        normalized_formula = f"={normalized_formula}"

    try:
        tokenizer = Tokenizer(normalized_formula)
    except Exception:
        return []

    matched_references: list[dict[str, Any]] = []
    for token in tokenizer.items:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue

        token_value = str(token.value).strip()
        if not token_value:
            continue
        matched_references.extend(
            _resolve_formula_token_references(
                wb,
                token_value=token_value,
                formula_sheet_name=formula_sheet_name,
                formula_row=formula_row,
                target_sheet=target_sheet,
                target_bounds=target_bounds,
            )
        )

    return matched_references


def _formula_token_values(formula_text: Any) -> list[str]:
    if formula_text is None:
        return []

    normalized_formula = str(formula_text).strip()
    if not normalized_formula:
        return []
    if not normalized_formula.startswith("="):
        normalized_formula = f"={normalized_formula}"

    tokenizer = Tokenizer(normalized_formula)
    token_values: list[str] = []
    for token in tokenizer.items:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        token_value = str(token.value).strip()
        if token_value:
            token_values.append(token_value)
    return token_values


def _iter_formula_cells(wb: Any) -> list[dict[str, Any]]:
    formula_entries: list[dict[str, Any]] = []
    for formula_sheet_name in wb.sheetnames:
        formula_ws = wb[formula_sheet_name]
        if _sheet_type(formula_ws) == "chartsheet":
            continue

        for row in formula_ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                try:
                    token_values = _formula_token_values(cell.value)
                except Exception:
                    token_values = []
                formula_entries.append(
                    {
                        "sheet_name": formula_sheet_name,
                        "cell": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "formula": cell.value,
                        "token_values": token_values,
                    }
                )
    return formula_entries


def _formula_dependency_graph(
    wb: Any,
    formula_entries: list[dict[str, Any]],
) -> dict[tuple[str, str], set[tuple[str, str]]]:
    graph: dict[tuple[str, str], set[tuple[str, str]]] = {
        (entry["sheet_name"], entry["cell"]): set() for entry in formula_entries
    }

    for entry in formula_entries:
        entry_key = (entry["sheet_name"], entry["cell"])
        for token_value in entry["token_values"]:
            resolved_targets = _resolve_formula_reference_targets(
                wb,
                token_value=token_value,
                formula_sheet_name=entry["sheet_name"],
                formula_row=entry["row"],
            )
            for target in resolved_targets:
                target_sheet_name = target.get("sheet_name")
                target_range = target.get("range")
                if (
                    not target_sheet_name
                    or not target_range
                    or target_sheet_name not in wb.sheetnames
                ):
                    continue

                target_ws = wb[target_sheet_name]
                if _sheet_type(target_ws) == "chartsheet":
                    continue

                try:
                    target_bounds, _ = _parse_range_reference(
                        target_range,
                        worksheet=target_ws,
                        expected_sheet=target_sheet_name,
                        error_cls=WorkbookError,
                    )
                except WorkbookError:
                    continue

                for row in target_ws.iter_rows(
                    min_row=target_bounds[0],
                    max_row=target_bounds[2],
                    min_col=target_bounds[1],
                    max_col=target_bounds[3],
                ):
                    for precedent_cell in row:
                        if not isinstance(precedent_cell.value, str) or not precedent_cell.value.startswith("="):
                            continue
                        graph[entry_key].add(
                            (target_sheet_name, precedent_cell.coordinate)
                        )

    return graph


def _collect_formula_precedent_entries(
    wb: Any,
    *,
    formula_sheet_name: str,
    formula_cell: Any,
) -> list[dict[str, Any]]:
    try:
        tokenizer = Tokenizer(formula_cell.value)
    except Exception:
        return []

    formula_precedents: list[dict[str, Any]] = []
    seen_formula_cells: set[tuple[str, str]] = set()

    for token in tokenizer.items:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue

        token_value = str(token.value).strip()
        if not token_value:
            continue

        resolved_targets = _resolve_formula_reference_targets(
            wb,
            token_value=token_value,
            formula_sheet_name=formula_sheet_name,
            formula_row=formula_cell.row,
        )
        for target in resolved_targets:
            target_sheet_name = target.get("sheet_name")
            target_range = target.get("range")
            if not target_sheet_name or not target_range or target_sheet_name not in wb.sheetnames:
                continue

            target_ws = wb[target_sheet_name]
            if _sheet_type(target_ws) == "chartsheet":
                continue

            target_bounds, _ = _parse_range_reference(
                target_range,
                worksheet=target_ws,
                expected_sheet=target_sheet_name,
                error_cls=WorkbookError,
            )
            for row in target_ws.iter_rows(
                min_row=target_bounds[0],
                max_row=target_bounds[2],
                min_col=target_bounds[1],
                max_col=target_bounds[3],
            ):
                for precedent_cell in row:
                    if not isinstance(precedent_cell.value, str) or not precedent_cell.value.startswith("="):
                        continue
                    precedent_key = (target_sheet_name, precedent_cell.coordinate)
                    if precedent_key in seen_formula_cells:
                        continue
                    seen_formula_cells.add(precedent_key)
                    formula_precedents.append(
                        {
                            "sheet_name": target_sheet_name,
                            "cell": precedent_cell.coordinate,
                            "formula": precedent_cell.value,
                            "reached_via": target["reference"],
                        }
                    )

    return formula_precedents


def _sample_formula_chain_paths(
    *,
    root_key: tuple[str, str],
    child_map: dict[tuple[str, str], list[tuple[str, str]]],
    path_limit: int,
) -> list[list[dict[str, Any]]]:
    sampled_paths: list[list[dict[str, Any]]] = []
    frontier: deque[tuple[tuple[str, str], list[tuple[str, str]]]] = deque([(root_key, [root_key])])

    while frontier and len(sampled_paths) < path_limit:
        current_key, path = frontier.popleft()
        child_keys = child_map.get(current_key, [])
        if not child_keys:
            if len(path) > 1:
                sampled_paths.append(
                    [
                        {
                            "sheet_name": sheet_name,
                            "cell": cell,
                        }
                        for sheet_name, cell in path
                    ]
                )
            continue

        for child_key in child_keys:
            if child_key in path:
                continue
            frontier.append((child_key, path + [child_key]))

    return sampled_paths


def _tarjan_strongly_connected_components(
    graph: dict[tuple[str, str], set[tuple[str, str]]]
) -> list[list[tuple[str, str]]]:
    index = 0
    stack: list[tuple[str, str]] = []
    on_stack: set[tuple[str, str]] = set()
    indices: dict[tuple[str, str], int] = {}
    lowlinks: dict[tuple[str, str], int] = {}
    components: list[list[tuple[str, str]]] = []

    def strongconnect(node: tuple[str, str]) -> None:
        nonlocal index
        indices[node] = index
        lowlinks[node] = index
        index += 1
        stack.append(node)
        on_stack.add(node)

        for neighbor in graph.get(node, set()):
            if neighbor not in indices:
                strongconnect(neighbor)
                lowlinks[node] = min(lowlinks[node], lowlinks[neighbor])
            elif neighbor in on_stack:
                lowlinks[node] = min(lowlinks[node], indices[neighbor])

        if lowlinks[node] == indices[node]:
            component: list[tuple[str, str]] = []
            while stack:
                current = stack.pop()
                on_stack.remove(current)
                component.append(current)
                if current == node:
                    break
            components.append(component)

    for node in graph:
        if node not in indices:
            strongconnect(node)

    return components


def _extract_validation_overlaps(
    ws: Worksheet,
    *,
    sheet_name: str,
    target_bounds: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    overlaps: list[dict[str, Any]] = []

    for validation in getattr(getattr(ws, "data_validations", None), "dataValidation", []):
        intersection_ranges: list[str] = []
        for validation_bounds, validation_ref in _iter_range_references(
            validation.sqref,
            worksheet=ws,
            expected_sheet=sheet_name,
        ):
            intersection = _intersection_bounds(target_bounds, validation_bounds)
            if intersection is None:
                continue
            intersection_ranges.append(_bounds_to_range(*intersection))

        if not intersection_ranges:
            continue

        overlaps.append(
            {
                "applies_to": str(validation.sqref),
                "intersection_ranges": intersection_ranges,
                "validation_type": validation.type,
                "operator": validation.operator or None,
                "formula1": validation.formula1 or None,
                "formula2": validation.formula2 or None,
            }
        )

    return overlaps


def _extract_validation_dependencies(
    wb: Any,
    *,
    target_sheet: str,
    target_bounds: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    dependencies: list[dict[str, Any]] = []

    for validation_sheet_name in wb.sheetnames:
        validation_ws = wb[validation_sheet_name]
        if _sheet_type(validation_ws) == "chartsheet":
            continue

        for validation in getattr(getattr(validation_ws, "data_validations", None), "dataValidation", []):
            applies_to_refs = _iter_range_references(
                validation.sqref,
                worksheet=validation_ws,
                expected_sheet=validation_sheet_name,
            )
            formula_row = applies_to_refs[0][0][0] if applies_to_refs else 1

            matched_references: list[dict[str, Any]] = []
            matched_references.extend(
                _extract_reference_matches_from_formula_text(
                    wb,
                    formula_text=validation.formula1,
                    formula_sheet_name=validation_sheet_name,
                    formula_row=formula_row,
                    target_sheet=target_sheet,
                    target_bounds=target_bounds,
                )
            )
            matched_references.extend(
                _extract_reference_matches_from_formula_text(
                    wb,
                    formula_text=validation.formula2,
                    formula_sheet_name=validation_sheet_name,
                    formula_row=formula_row,
                    target_sheet=target_sheet,
                    target_bounds=target_bounds,
                )
            )

            if not matched_references:
                continue

            dependencies.append(
                {
                    "sheet_name": validation_sheet_name,
                    "applies_to": str(validation.sqref),
                    "validation_type": validation.type,
                    "operator": validation.operator or None,
                    "formula1": validation.formula1 or None,
                    "formula2": validation.formula2 or None,
                    "references": matched_references,
                }
            )

    return dependencies


def _extract_conditional_format_overlaps(
    ws: Worksheet,
    *,
    sheet_name: str,
    target_bounds: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    overlaps: list[dict[str, Any]] = []

    for conditional_format, rules in getattr(ws.conditional_formatting, "_cf_rules", {}).items():
        intersection_ranges: list[str] = []
        for format_bounds, _ in _iter_range_references(
            conditional_format.sqref,
            worksheet=ws,
            expected_sheet=sheet_name,
        ):
            intersection = _intersection_bounds(target_bounds, format_bounds)
            if intersection is None:
                continue
            intersection_ranges.append(_bounds_to_range(*intersection))

        if not intersection_ranges:
            continue

        for rule in rules:
            overlaps.append(
                {
                    "applies_to": str(conditional_format.sqref),
                    "intersection_ranges": intersection_ranges,
                    "rule_type": getattr(rule, "type", None),
                    "operator": getattr(rule, "operator", None),
                    "formula": list(getattr(rule, "formula", None) or []),
                }
            )

    return overlaps


def _extract_conditional_format_dependencies(
    wb: Any,
    *,
    target_sheet: str,
    target_bounds: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    dependencies: list[dict[str, Any]] = []

    for format_sheet_name in wb.sheetnames:
        format_ws = wb[format_sheet_name]
        if _sheet_type(format_ws) == "chartsheet":
            continue

        for conditional_format, rules in getattr(format_ws.conditional_formatting, "_cf_rules", {}).items():
            applies_to_refs = _iter_range_references(
                conditional_format.sqref,
                worksheet=format_ws,
                expected_sheet=format_sheet_name,
            )
            formula_row = applies_to_refs[0][0][0] if applies_to_refs else 1

            for rule in rules:
                formulas = list(getattr(rule, "formula", None) or [])
                matched_references: list[dict[str, Any]] = []
                for formula_text in formulas:
                    matched_references.extend(
                        _extract_reference_matches_from_formula_text(
                            wb,
                            formula_text=formula_text,
                            formula_sheet_name=format_sheet_name,
                            formula_row=formula_row,
                            target_sheet=target_sheet,
                            target_bounds=target_bounds,
                        )
                    )

                if not matched_references:
                    continue

                dependencies.append(
                    {
                        "sheet_name": format_sheet_name,
                        "applies_to": str(conditional_format.sqref),
                        "rule_type": getattr(rule, "type", None),
                        "operator": getattr(rule, "operator", None),
                        "formula": formulas,
                        "references": matched_references,
                    }
                )

    return dependencies


def _freeze_panes_value(ws) -> str | None:
    value = ws.freeze_panes
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return getattr(value, "coordinate", None)


def _sheet_type(ws: Any) -> str:
    return "chartsheet" if ws.__class__.__name__ == "Chartsheet" else "worksheet"


def _audit_finding(
    severity: str,
    code: str,
    message: str,
    *,
    sheet_name: str | None = None,
    recommendation: str | None = None,
    details: dict[str, Any] | None = None,
) -> dict[str, Any]:
    finding = {
        "severity": severity,
        "code": code,
        "message": message,
    }
    if sheet_name is not None:
        finding["sheet_name"] = sheet_name
    if recommendation is not None:
        finding["recommendation"] = recommendation
    if details:
        finding["details"] = details
    return finding


def _highest_severity(findings: list[dict[str, Any]]) -> str | None:
    if not findings:
        return None
    return max(
        (finding["severity"] for finding in findings),
        key=lambda severity: AUDIT_SEVERITY_RANK.get(severity, 0),
    )


def _cells_with_broken_formula_references(
    wb: Any,
    *,
    ws: Worksheet,
    sheet_name: str,
) -> list[str]:
    cells: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue
            if _formula_text_has_broken_reference(
                wb,
                formula_text=cell.value,
                formula_sheet_name=sheet_name,
            ):
                cells.append(cell.coordinate)
    return cells


def _cells_with_error_values(ws: Worksheet) -> list[str]:
    cells: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "e":
                cells.append(cell.coordinate)
    return cells


def _broken_validation_rules(
    wb: Any,
    *,
    ws: Worksheet,
    sheet_name: str,
) -> list[dict[str, Any]]:
    return [
        {
            "applies_to": rule["applies_to"],
            "validation_type": rule["validation_type"],
            "formula1": rule["formula1"],
            "formula2": rule["formula2"],
            "rule_index": rule["rule_index"],
        }
        for rule in _inspect_validation_rules(
            wb,
            ws=ws,
            sheet_name=sheet_name,
        )
        if rule["broken_reference"]
    ]


def _broken_conditional_format_rules(
    wb: Any,
    *,
    ws: Worksheet,
    sheet_name: str,
) -> list[dict[str, Any]]:
    return [
        {
            "applies_to": rule["applies_to"],
            "rule_type": rule["rule_type"],
            "formula": rule["formula"],
            "rule_index": rule["rule_index"],
        }
        for rule in _inspect_conditional_format_rules(
            wb,
            ws=ws,
            sheet_name=sheet_name,
        )
        if rule["broken_reference"]
    ]


def _worksheet_audit_assessment(
    ws: Worksheet,
    *,
    sheet_name: str,
    header_row: int,
) -> dict[str, Any]:
    from .data import (
        _header_profile,
        _read_table_from_worksheet,
        _table_dominates_sheet,
        _worksheet_dataset_kind,
    )
    from .tables import _build_table_metadata

    rows, columns, column_range, is_empty = _get_sheet_usage(ws)
    used_range = _get_used_range(ws)
    chart_count = len(getattr(ws, "_charts", []))
    merged_range_count = len(ws.merged_cells.ranges)
    native_tables = [
        _build_table_metadata(sheet_name, ws, table)
        for table in ws.tables.values()
    ]

    if is_empty:
        return {
            "sheet_name": sheet_name,
            "sheet_type": "worksheet",
            "rows": rows,
            "columns": columns,
            "column_range": column_range,
            "used_range": used_range,
            "dataset_kind": "empty_sheet",
            "recommended_read_tool": "quick_read",
            "header_profile": {
                "total_headers": 0,
                "non_empty_headers": 0,
                "blank_headers": 0,
                "string_headers": 0,
                "duplicate_headers": 0,
                "score": 0.0,
                "confidence": "low",
            },
            "total_rows": 0,
            "chart_count": chart_count,
            "merged_range_count": merged_range_count,
            "native_tables": native_tables,
            "dominant_table": None,
            "header_quality_profile": {
                "total_headers": 0,
                "non_empty_headers": 0,
                "blank_headers": 0,
                "string_headers": 0,
                "duplicate_headers": 0,
                "score": 0.0,
                "confidence": "low",
            },
            "header_quality_scope": "worksheet",
        }

    sample = _read_table_from_worksheet(
        ws,
        sheet_name,
        header_row=header_row,
        max_rows=AUDIT_SAMPLE_ROWS,
        include_headers=True,
        row_mode="arrays",
        infer_schema=False,
    )
    header_profile = _header_profile(sample["headers"])
    dataset_kind = _worksheet_dataset_kind(
        used_range=used_range,
        total_rows=sample["total_rows"],
        header_confidence=header_profile["confidence"],
        chart_count=chart_count,
        merged_range_count=merged_range_count,
    )
    dominant_table = _table_dominates_sheet(
        {
            "native_tables": [
                {
                    "table_name": table["table_name"],
                    "data_row_count": table["data_row_count"],
                    "column_count": table["column_count"],
                }
                for table in native_tables
            ],
            "total_rows": sample["total_rows"],
            "column_count": len(sample["headers"]),
        }
    )
    preferred_table = dominant_table
    if preferred_table is None and len(native_tables) == 1 and len(sample["headers"]) > 0:
        sole_table = native_tables[0]
        row_coverage = sole_table["data_row_count"] / sample["total_rows"] if sample["total_rows"] else 0.0
        col_coverage = sole_table["column_count"] / len(sample["headers"])
        if row_coverage >= 0.8 and col_coverage >= 0.5:
            preferred_table = sole_table

    if preferred_table is not None:
        recommended_read_tool = "read_excel_table"
    elif dataset_kind == "layout_like_sheet":
        recommended_read_tool = "profile_workbook"
    else:
        recommended_read_tool = "quick_read"

    header_quality_profile = (
        _header_profile(preferred_table["headers"])
        if preferred_table is not None
        else header_profile
    )
    header_quality_scope = "dominant_table" if preferred_table is not None else "worksheet"

    return {
        "sheet_name": sheet_name,
        "sheet_type": "worksheet",
        "rows": rows,
        "columns": columns,
        "column_range": column_range,
        "used_range": used_range,
        "dataset_kind": dataset_kind,
        "recommended_read_tool": recommended_read_tool,
        "header_profile": header_profile,
        "total_rows": sample["total_rows"],
        "chart_count": chart_count,
        "merged_range_count": merged_range_count,
        "native_tables": native_tables,
        "dominant_table": preferred_table,
        "header_quality_profile": header_quality_profile,
        "header_quality_scope": header_quality_scope,
    }


def _workbook_named_range_findings(
    wb: Any,
    *,
    named_ranges: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    for named_range in named_ranges:
        value = str(named_range.get("value") or "")
        if "#REF!" in value.upper():
            findings.append(
                _audit_finding(
                    "high",
                    "broken_named_range_reference",
                    f"Named range '{named_range['name']}' contains a broken #REF! reference.",
                    sheet_name=named_range.get("local_sheet"),
                    recommendation="Repair or remove broken defined names before relying on workbook formulas or automation.",
                    details={
                        "name": named_range["name"],
                        "local_sheet": named_range.get("local_sheet"),
                        "value": named_range.get("value"),
                    },
                )
            )

        missing_sheets = sorted(
            {
                destination["sheet_name"]
                for destination in named_range.get("destinations", [])
                if destination["sheet_name"] not in wb.sheetnames
            }
        )
        if missing_sheets:
            findings.append(
                _audit_finding(
                    "high",
                    "named_range_missing_sheet",
                    f"Named range '{named_range['name']}' points to missing sheet destinations.",
                    sheet_name=named_range.get("local_sheet"),
                    recommendation="Repair or remove defined names that reference deleted or renamed sheets.",
                    details={
                        "name": named_range["name"],
                        "local_sheet": named_range.get("local_sheet"),
                        "missing_sheets": missing_sheets,
                    },
                )
            )
    return findings


def _unique_step_tools(tools: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for tool in tools:
        tool_name = tool.get("tool")
        args = tool.get("args", {})
        key = (str(tool_name), str(sorted(args.items())))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(tool)
    return deduped


def _sort_steps(steps: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        steps,
        key=lambda step: (
            -AUDIT_SEVERITY_RANK.get(step["priority"], 0),
            str(step.get("sheet_name") or ""),
            step["title"],
        ),
    )


def require_worksheet(
    wb: Any,
    sheet_name: str,
    *,
    error_cls: type[Exception] = WorkbookError,
    operation: str = "worksheet operations",
) -> Worksheet:
    if sheet_name not in wb.sheetnames:
        raise error_cls(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]
    if not isinstance(ws, Worksheet):
        raise error_cls(
            f"Sheet '{sheet_name}' is a chartsheet and cannot be used for {operation}"
        )
    return ws


def first_worksheet(
    wb: Any,
    *,
    error_cls: type[Exception] = WorkbookError,
) -> tuple[str, Worksheet]:
    if not wb.sheetnames:
        raise error_cls("Workbook contains no sheets")

    worksheets = list(getattr(wb, "worksheets", []))
    if not worksheets:
        raise error_cls("Workbook contains no worksheets")

    worksheet = worksheets[0]
    return worksheet.title, worksheet


def _serialize_named_ranges(wb: Any) -> list[dict[str, Any]]:
    ranges = []
    for name, defined_name, local_sheet in _iter_defined_name_entries(wb):
        destinations = []
        try:
            destinations = [
                {
                    "sheet_name": sheet_name,
                    "range": cell_range,
                }
                for sheet_name, cell_range in defined_name.destinations
            ]
        except Exception:
            destinations = []

        ranges.append(
            {
                "name": name,
                "type": defined_name.type,
                "value": defined_name.value,
                "destinations": destinations,
                "local_sheet": local_sheet,
                "hidden": bool(getattr(defined_name, "hidden", False)),
                "broken_reference": "#REF!" in str(defined_name.value or "").upper(),
                "missing_sheets": sorted(
                    {
                        destination["sheet_name"]
                        for destination in destinations
                        if destination["sheet_name"] not in wb.sheetnames
                    }
                ),
            }
        )

    return sorted(ranges, key=lambda item: item["name"].lower())


def _named_range_sources(
    wb: Any,
    *,
    name: str | None = None,
    scope_sheet: str | None = None,
) -> list[dict[str, Any]]:
    sources: list[dict[str, Any]] = []

    for defined_name, local_sheet in (
        (
            defined_name,
            _defined_name_local_sheet(wb, defined_name),
        )
        for defined_name in wb.defined_names.values()
    ):
        current_name = getattr(defined_name, "name", None)
        if current_name is None:
            continue
        if name is not None and current_name != name:
            continue
        if scope_sheet is not None and local_sheet != scope_sheet:
            continue
        sources.append(
            {
                "name": current_name,
                "local_sheet": local_sheet,
                "container": "workbook",
                "defined_name": defined_name,
            }
        )

    for ws in getattr(wb, "worksheets", []):
        if scope_sheet is not None and ws.title != scope_sheet:
            continue
        for current_name, defined_name in ws.defined_names.items():
            if name is not None and current_name != name:
                continue
            sources.append(
                {
                    "name": current_name,
                    "local_sheet": ws.title,
                    "container": "worksheet",
                    "defined_name": defined_name,
                }
            )

    return sources


def _serialize_named_range_source(wb: Any, source: dict[str, Any]) -> dict[str, Any]:
    defined_name = source["defined_name"]
    destinations = []
    try:
        destinations = [
            {
                "sheet_name": sheet_name,
                "range": cell_range,
            }
            for sheet_name, cell_range in defined_name.destinations
        ]
    except Exception:
        destinations = []

    return {
        "name": source["name"],
        "type": defined_name.type,
        "value": defined_name.value,
        "destinations": destinations,
        "local_sheet": source["local_sheet"],
        "hidden": bool(getattr(defined_name, "hidden", False)),
        "container": source["container"],
        "broken_reference": "#REF!" in str(defined_name.value or "").upper(),
        "missing_sheets": sorted(
            {
                destination["sheet_name"]
                for destination in destinations
                if destination["sheet_name"] not in wb.sheetnames
            }
        ),
    }


def _inspect_validation_rules(
    wb: Any,
    *,
    ws: Worksheet,
    sheet_name: str,
) -> list[dict[str, Any]]:
    rules: list[dict[str, Any]] = []
    for rule_index, validation in enumerate(
        getattr(getattr(ws, "data_validations", None), "dataValidation", []),
        start=1,
    ):
        formulas = [validation.formula1, validation.formula2]
        rules.append(
            {
                "rule_index": rule_index,
                "applies_to": str(validation.sqref),
                "validation_type": validation.type,
                "operator": validation.operator or None,
                "formula1": validation.formula1 or None,
                "formula2": validation.formula2 or None,
                "allow_blank": bool(validation.allowBlank),
                "broken_reference": any(
                    _formula_text_has_broken_reference(
                        wb,
                        formula_text=formula,
                        formula_sheet_name=sheet_name,
                    )
                    for formula in formulas
                ),
            }
        )
    return rules


def _inspect_conditional_format_rules(
    wb: Any,
    *,
    ws: Worksheet,
    sheet_name: str,
) -> list[dict[str, Any]]:
    rules: list[dict[str, Any]] = []
    rule_index = 1
    for conditional_format, format_rules in getattr(ws.conditional_formatting, "_cf_rules", {}).items():
        for rule in format_rules:
            formulas = list(getattr(rule, "formula", None) or [])
            rules.append(
                {
                    "rule_index": rule_index,
                    "applies_to": str(conditional_format.sqref),
                    "rule_type": getattr(rule, "type", None),
                    "operator": getattr(rule, "operator", None),
                    "priority": getattr(rule, "priority", None),
                    "stop_if_true": bool(getattr(rule, "stopIfTrue", False)),
                    "formula": formulas,
                    "broken_reference": any(
                        _formula_text_has_broken_reference(
                            wb,
                            formula_text=formula,
                            formula_sheet_name=sheet_name,
                        )
                        for formula in formulas
                    ),
                }
            )
            rule_index += 1
    return rules


def _normalize_rule_indexes(
    rule_indexes: list[int] | None,
    *,
    label: str,
) -> list[int]:
    if rule_indexes is None:
        return []
    if not isinstance(rule_indexes, list):
        raise WorkbookError(f"{label} must be a list of positive integers")

    normalized: list[int] = []
    for value in rule_indexes:
        if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
            raise WorkbookError(f"{label} must contain only positive integers")
        normalized.append(value)
    return normalized


def _remove_named_range_sources(
    wb: Any,
    *,
    name: str,
    scope_sheet: str | None = None,
) -> list[dict[str, Any]]:
    matches = _named_range_sources(wb, name=name, scope_sheet=scope_sheet)
    if not matches:
        raise WorkbookError(
            f"Named range '{name}'"
            + (f" scoped to '{scope_sheet}'" if scope_sheet else "")
            + " not found"
        )
    if scope_sheet is None and len(matches) > 1:
        raise WorkbookError(
            f"Named range '{name}' exists in multiple scopes; specify scope_sheet to remove the intended one"
        )

    removed: list[dict[str, Any]] = []
    for match in matches:
        removed.append(_serialize_named_range_source(wb, match))
        if match["container"] == "worksheet":
            ws = wb[match["local_sheet"]]
            ws.defined_names.pop(name, None)
        else:
            wb.defined_names.pop(name, None)
    return removed


def _format_absolute_named_range(
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> str:
    start = f"${get_column_letter(min_col)}${min_row}"
    end = f"${get_column_letter(max_col)}${max_row}"
    if start == end:
        return start
    return f"{start}:{end}"


def _resolve_named_range_target(
    wb: Any,
    *,
    range_ref: Any,
    sheet_name: str | None = None,
    scope_sheet: str | None = None,
) -> tuple[str, str, str]:
    normalized_range_ref = str(range_ref).strip()
    if not normalized_range_ref:
        raise WorkbookError("range_ref is required")

    normalized_sheet_name = sheet_name.strip() if isinstance(sheet_name, str) else None
    normalized_scope_sheet = scope_sheet.strip() if isinstance(scope_sheet, str) else None

    if "!" in normalized_range_ref:
        raw_sheet_name, local_range_ref = normalized_range_ref.rsplit("!", 1)
        target_sheet_name = _normalize_sheet_reference_name(raw_sheet_name)
        if normalized_sheet_name and target_sheet_name != normalized_sheet_name:
            raise WorkbookError(
                "sheet_name does not match the sheet embedded in range_ref"
            )
    else:
        target_sheet_name = normalized_sheet_name or normalized_scope_sheet
        local_range_ref = normalized_range_ref
        if not target_sheet_name:
            raise WorkbookError(
                "sheet_name is required when range_ref is not sheet-qualified"
            )

    require_worksheet(
        wb,
        target_sheet_name,
        error_cls=WorkbookError,
        operation="named range creation",
    )

    cleaned_local_range = str(local_range_ref).strip()
    if not cleaned_local_range:
        raise WorkbookError("range_ref is required")

    try:
        min_col, min_row, max_col, max_row = range_boundaries(
            cleaned_local_range.replace("$", "")
        )
    except ValueError as exc:
        raise WorkbookError(
            "range_ref must be a valid A1-style cell or range reference"
        ) from exc

    if None in (min_col, min_row, max_col, max_row):
        raise WorkbookError(
            "range_ref must be a valid A1-style cell or range reference"
        )

    absolute_range = _format_absolute_named_range(
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
    )
    full_reference = f"{quote_sheetname(target_sheet_name)}!{absolute_range}"
    return target_sheet_name, absolute_range, full_reference


def _matching_named_range_sources_for_scope(
    wb: Any,
    *,
    name: str,
    scope_sheet: str | None,
) -> list[dict[str, Any]]:
    candidates = _named_range_sources(wb, name=name, scope_sheet=scope_sheet)
    if scope_sheet is None:
        return [candidate for candidate in candidates if candidate["local_sheet"] is None]
    return [
        candidate
        for candidate in candidates
        if candidate["local_sheet"] == scope_sheet
    ]


def _remove_named_range_scope_matches(
    wb: Any,
    *,
    matches: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    removed: list[dict[str, Any]] = []
    for match in matches:
        removed.append(_serialize_named_range_source(wb, match))
        if match["container"] == "worksheet":
            wb[match["local_sheet"]].defined_names.pop(match["name"], None)
        else:
            wb.defined_names.pop(match["name"], None)
    return removed


def create_named_range(
    filepath: str,
    name: str,
    range_ref: str,
    sheet_name: str | None = None,
    scope_sheet: str | None = None,
    hidden: bool = False,
    replace: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Create a workbook-level or sheet-scoped named range."""
    try:
        normalized_name = str(name).strip()
        normalized_scope_sheet = scope_sheet.strip() if isinstance(scope_sheet, str) else None
        normalized_sheet_name = sheet_name.strip() if isinstance(sheet_name, str) else None

        if not normalized_name:
            raise WorkbookError("name is required")
        if isinstance(hidden, bool) is False:
            raise WorkbookError("hidden must be a boolean")
        if isinstance(replace, bool) is False:
            raise WorkbookError("replace must be a boolean")
        if isinstance(dry_run, bool) is False:
            raise WorkbookError("dry_run must be a boolean")

        with safe_workbook(filepath, save=not dry_run) as wb:
            if normalized_scope_sheet is not None:
                scope_ws = require_worksheet(
                    wb,
                    normalized_scope_sheet,
                    error_cls=WorkbookError,
                    operation="named range creation",
                )
                normalized_scope_sheet = scope_ws.title

            target_sheet_name, absolute_range, full_reference = _resolve_named_range_target(
                wb,
                range_ref=range_ref,
                sheet_name=normalized_sheet_name,
                scope_sheet=normalized_scope_sheet,
            )

            existing_matches = _matching_named_range_sources_for_scope(
                wb,
                name=normalized_name,
                scope_sheet=normalized_scope_sheet,
            )

            if existing_matches and not replace:
                scope_label = (
                    f"sheet scope '{normalized_scope_sheet}'"
                    if normalized_scope_sheet
                    else "workbook scope"
                )
                raise WorkbookError(
                    f"Named range '{normalized_name}' already exists in {scope_label}; "
                    "set replace=True to overwrite it"
                )

            replaced = _remove_named_range_scope_matches(
                wb,
                matches=existing_matches,
            ) if existing_matches else []

            if not dry_run:
                defined_name = DefinedName(
                    normalized_name,
                    attr_text=full_reference,
                    hidden=hidden,
                )
                if normalized_scope_sheet is None:
                    wb.defined_names[normalized_name] = defined_name
                else:
                    wb[normalized_scope_sheet].defined_names.add(defined_name)

            created = {
                "name": normalized_name,
                "type": "RANGE",
                "value": full_reference,
                "destinations": [
                    {
                        "sheet_name": target_sheet_name,
                        "range": absolute_range,
                    }
                ],
                "local_sheet": normalized_scope_sheet,
                "hidden": hidden,
                "broken_reference": False,
                "missing_sheets": [],
            }

            return {
                "message": (
                    f"{'Previewed' if dry_run else 'Created'} named range '{normalized_name}'"
                ),
                "name": normalized_name,
                "sheet_name": target_sheet_name,
                "scope_sheet": normalized_scope_sheet,
                "dry_run": dry_run,
                "replaced_count": len(replaced),
                "replaced": replaced,
                "named_range": created,
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to create named range '{name}': {e}")
        raise WorkbookError(str(e))


def _remove_validation_rules(
    wb: Any,
    ws: Worksheet,
    *,
    rule_indexes: list[int] | None = None,
    broken_only: bool = False,
) -> list[dict[str, Any]]:
    normalized_indexes = set(_normalize_rule_indexes(rule_indexes, label="rule_indexes"))
    if not normalized_indexes and not broken_only:
        raise WorkbookError("Specify rule_indexes or set broken_only=True")

    current_rules = _inspect_validation_rules(
        wb,
        ws=ws,
        sheet_name=ws.title,
    )
    removed = [
        rule
        for rule in current_rules
        if (rule["rule_index"] in normalized_indexes)
        or (broken_only and rule["broken_reference"])
    ]
    if not removed:
        return []

    removed_indexes = {rule["rule_index"] for rule in removed}
    kept = [
        validation
        for rule_index, validation in enumerate(
            getattr(getattr(ws, "data_validations", None), "dataValidation", []),
            start=1,
        )
        if rule_index not in removed_indexes
    ]
    ws.data_validations.dataValidation = kept
    return removed


def _remove_conditional_format_rules(
    wb: Any,
    ws: Worksheet,
    *,
    rule_indexes: list[int] | None = None,
    broken_only: bool = False,
) -> list[dict[str, Any]]:
    normalized_indexes = set(_normalize_rule_indexes(rule_indexes, label="rule_indexes"))
    if not normalized_indexes and not broken_only:
        raise WorkbookError("Specify rule_indexes or set broken_only=True")

    current_rules = _inspect_conditional_format_rules(
        wb,
        ws=ws,
        sheet_name=ws.title,
    )
    removed = [
        rule
        for rule in current_rules
        if (rule["rule_index"] in normalized_indexes)
        or (broken_only and rule["broken_reference"])
    ]
    if not removed:
        return []

    removed_indexes = {rule["rule_index"] for rule in removed}
    rebuilt_rules: OrderedDict[Any, list[Any]] = OrderedDict()
    current_index = 1
    for conditional_format, format_rules in getattr(ws.conditional_formatting, "_cf_rules", {}).items():
        kept_rules: list[Any] = []
        for rule in format_rules:
            if current_index not in removed_indexes:
                kept_rules.append(rule)
            current_index += 1
        if kept_rules:
            rebuilt_rules[conditional_format] = kept_rules

    ws.conditional_formatting._cf_rules = rebuilt_rules
    return removed


@contextmanager
def safe_workbook(filepath: str, save: bool = False, read_only: bool = False):
    """Context manager that ensures workbook is always closed.

    Args:
        filepath: Path to Excel file.
        save: If True, save the workbook before closing.
        read_only: If True, open in read-only mode.
    """
    wb = load_workbook(filepath, read_only=read_only)
    try:
        yield wb
    except Exception:
        raise
    else:
        if save and not read_only:
            _persist_workbook_atomically(wb, filepath)
    finally:
        wb.close()


def _verify_saved_workbook(filepath: str) -> None:
    """Reopen a saved workbook to verify the persisted file is readable."""
    verification_wb = load_workbook(filepath, read_only=True)
    try:
        _ = verification_wb.sheetnames
    finally:
        verification_wb.close()


def _fsync_file(path: Path) -> None:
    with path.open("rb") as handle:
        os.fsync(handle.fileno())


def _fsync_directory(path: Path) -> None:
    try:
        dir_fd = os.open(path, os.O_RDONLY)
    except OSError:
        return

    try:
        os.fsync(dir_fd)
    except OSError:
        pass
    finally:
        os.close(dir_fd)


def _persist_workbook_atomically(wb: Workbook, filepath: str) -> None:
    """Persist workbook changes via temp file + atomic replace + reopen verify."""
    destination = Path(filepath)
    destination.parent.mkdir(parents=True, exist_ok=True)
    backup_path: Path | None = None
    temp_suffix = destination.suffix or ".xlsx"
    fd, temp_name = tempfile.mkstemp(
        prefix=f".{destination.stem}.sheetforge-",
        suffix=temp_suffix,
        dir=str(destination.parent),
    )
    os.close(fd)
    temp_path = Path(temp_name)

    try:
        wb.save(temp_path)
        if destination.exists():
            with suppress(OSError):
                os.chmod(temp_path, destination.stat().st_mode)
        _fsync_file(temp_path)
        _verify_saved_workbook(str(temp_path))

        if destination.exists():
            backup_fd, backup_name = tempfile.mkstemp(
                prefix=f".{destination.name}.sheetforge-backup-",
                suffix=".bak",
                dir=str(destination.parent),
            )
            os.close(backup_fd)
            backup_path = Path(backup_name)
            shutil.copy2(destination, backup_path)
            _fsync_file(backup_path)

        os.replace(temp_path, destination)
        _fsync_directory(destination.parent)
        try:
            _verify_saved_workbook(str(destination))
        except Exception:
            if backup_path is not None and backup_path.exists():
                os.replace(backup_path, destination)
                _fsync_directory(destination.parent)
            else:
                with suppress(FileNotFoundError):
                    destination.unlink()
                _fsync_directory(destination.parent)
            raise
    except Exception as exc:
        raise WorkbookError(f"Failed to save workbook atomically: {exc!s}") from exc
    finally:
        with suppress(FileNotFoundError):
            temp_path.unlink()
        if backup_path is not None:
            with suppress(FileNotFoundError):
                backup_path.unlink()

def create_workbook(filepath: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Create a new Excel workbook with optional custom sheet name"""
    try:
        wb = Workbook()
        # Rename default sheet
        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            sheet.title = sheet_name
        else:
            wb.create_sheet(sheet_name)

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return {
            "message": f"Created workbook: {filepath}",
            "active_sheet": sheet_name,
            "workbook": wb
        }
    except Exception as e:
        logger.error(f"Failed to create workbook: {e}")
        raise WorkbookError(f"Failed to create workbook: {e!s}")

def get_or_create_workbook(filepath: str) -> Workbook:
    """Load an existing workbook. Raises FileNotFoundError if it doesn't exist."""
    return load_workbook(filepath)

def create_sheet(filepath: str, sheet_name: str) -> dict:
    """Create a new worksheet in the workbook if it doesn't exist."""
    try:
        with safe_workbook(filepath, save=True) as wb:
            # Check if sheet already exists
            if sheet_name in wb.sheetnames:
                raise WorkbookError(f"Sheet {sheet_name} already exists")

            # Create new sheet
            wb.create_sheet(sheet_name)
        return {"message": f"Sheet {sheet_name} created successfully"}
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create sheet: {e}")
        raise WorkbookError(str(e))

def list_sheets(filepath: str) -> list[dict[str, Any]]:
    """List all sheets with basic metadata."""
    try:
        with safe_workbook(filepath) as wb:
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                if _sheet_type(ws) == "chartsheet":
                    sheets.append(
                        {
                            "name": name,
                            "sheet_type": "chartsheet",
                            "rows": 0,
                            "columns": 0,
                            "column_range": None,
                            "is_empty": len(getattr(ws, "_charts", [])) == 0,
                        }
                    )
                    continue

                rows, columns, column_range, is_empty = _get_sheet_usage(ws)
                sheets.append({
                    "name": name,
                    "sheet_type": "worksheet",
                    "rows": rows,
                    "columns": columns,
                    "column_range": column_range,
                    "is_empty": is_empty,
                })
            return sheets
    except Exception as e:
        logger.error(f"Failed to list sheets: {e}")
        raise WorkbookError(str(e))


def list_named_ranges(filepath: str) -> list[dict[str, Any]]:
    """List workbook-level and local defined names."""
    try:
        with safe_workbook(filepath) as wb:
            return _serialize_named_ranges(wb)
    except Exception as e:
        logger.error(f"Failed to list named ranges: {e}")
        raise WorkbookError(str(e))


def inspect_named_range(
    filepath: str,
    name: str,
    scope_sheet: str | None = None,
) -> dict[str, Any]:
    """Inspect a named range and report scope, destinations, and breakage signals."""
    try:
        if not str(name).strip():
            raise WorkbookError("name is required")

        with safe_workbook(filepath) as wb:
            matches = _named_range_sources(
                wb,
                name=str(name).strip(),
                scope_sheet=scope_sheet,
            )
            if not matches:
                raise WorkbookError(
                    f"Named range '{name}'"
                    + (f" scoped to '{scope_sheet}'" if scope_sheet else "")
                    + " not found"
                )

            serialized = [
                _serialize_named_range_source(wb, match)
                for match in matches
            ]
            return {
                "name": str(name).strip(),
                "scope_sheet": scope_sheet,
                "match_count": len(serialized),
                "matches": serialized,
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to inspect named range '{name}': {e}")
        raise WorkbookError(str(e))

def get_workbook_info(filepath: str, include_ranges: bool = False) -> dict[str, Any]:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")

        with safe_workbook(filepath) as wb:
            info = {
                "filename": path.name,
                "sheets": wb.sheetnames,
                "size": path.stat().st_size,
                "modified": path.stat().st_mtime
            }

            if include_ranges:
                # Add used ranges for each sheet
                ranges = {}
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if _sheet_type(ws) == "chartsheet":
                        continue

                    used_range = _get_used_range(ws)
                    if used_range is not None:
                        ranges[sheet_name] = used_range
                info["used_ranges"] = ranges

        return info

    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get workbook info: {e}")
        raise WorkbookError(str(e))


def profile_workbook(filepath: str) -> dict[str, Any]:
    """Return a workbook-level inventory tuned for agent orientation."""
    try:
        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")

        with safe_workbook(filepath) as wb:
            from .chart import (
                _chart_occupied_range,
                _chart_type_name,
                _extract_chart_anchor,
                _extract_chart_dimensions,
                _extract_title_text,
            )
            from .sheet import _sheet_protection_state
            from .tables import _build_table_metadata

            named_ranges = _serialize_named_ranges(wb)
            sheets: list[dict[str, Any]] = []
            total_tables = 0
            total_charts = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                charts = []
                for chart_index, chart in enumerate(getattr(ws, "_charts", []), start=1):
                    width, height = _extract_chart_dimensions(chart)
                    anchor = _extract_chart_anchor(chart)
                    series = getattr(chart, "ser", None) or list(getattr(chart, "series", []))
                    chart_info = {
                        "chart_index": chart_index,
                        "chart_type": _chart_type_name(chart),
                        "anchor": anchor,
                        "title": _extract_title_text(getattr(chart, "title", None)),
                        "width": width,
                        "height": height,
                        "series_count": len(series),
                    }
                    if anchor and width and height:
                        chart_info["occupied_range"] = _chart_occupied_range(
                            ws,
                            anchor,
                            width=width,
                            height=height,
                        )
                    charts.append(chart_info)

                if _sheet_type(ws) == "chartsheet":
                    total_charts += len(charts)
                    sheets.append(
                        {
                            "name": sheet_name,
                            "sheet_type": "chartsheet",
                            "visibility": ws.sheet_state,
                            "table_count": 0,
                            "chart_count": len(charts),
                            "tables": [],
                            "charts": charts,
                        }
                    )
                    continue

                rows, columns, column_range, is_empty = _get_sheet_usage(ws)
                tables = [
                    _build_table_metadata(sheet_name, ws, table)
                    for table in ws.tables.values()
                ]

                total_tables += len(tables)
                total_charts += len(charts)

                sheets.append(
                    {
                        "name": sheet_name,
                        "sheet_type": "worksheet",
                        "rows": rows,
                        "columns": columns,
                        "column_range": column_range,
                        "used_range": _get_used_range(ws),
                        "is_empty": is_empty,
                        "visibility": ws.sheet_state,
                        "freeze_panes": _freeze_panes_value(ws),
                        "has_autofilter": bool(ws.auto_filter.ref),
                        "autofilter_range": ws.auto_filter.ref or None,
                        "print_area": ws.print_area or None,
                        "print_title_rows": ws.print_title_rows or None,
                        "print_title_columns": ws.print_title_cols or None,
                        "merged_range_count": len(ws.merged_cells.ranges),
                        "table_count": len(tables),
                        "chart_count": len(charts),
                        "protection": {
                            "enabled": _sheet_protection_state(ws)["enabled"],
                            "password_protected": _sheet_protection_state(ws)["password_protected"],
                        },
                        "tables": tables,
                        "charts": charts,
                    }
                )

            return {
                "filename": path.name,
                "size": path.stat().st_size,
                "modified": path.stat().st_mtime,
                "sheet_count": len(wb.sheetnames),
                "named_range_count": len(named_ranges),
                "table_count": total_tables,
                "chart_count": total_charts,
                "sheets": sheets,
                "named_ranges": named_ranges,
            }

    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to profile workbook: {e}")
        raise WorkbookError(str(e))


def _validate_positive_int(value: Any, *, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise WorkbookError(f"{label} must be a positive integer")
    return value


def _sample_payload(
    items: list[Any],
    *,
    sample_limit: int,
) -> dict[str, Any]:
    return {
        "count": len(items),
        "sample": items[:sample_limit],
    }


def _custom_column_layout(ws: Worksheet) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for column_key, dimension in ws.column_dimensions.items():
        width = getattr(dimension, "width", None)
        hidden = bool(getattr(dimension, "hidden", False))
        if width is None and not hidden:
            continue

        column_letter = str(getattr(dimension, "index", None) or column_key).upper()
        try:
            sort_index = column_index_from_string(column_letter)
        except ValueError:
            continue

        entries.append(
            {
                "column": column_letter,
                "width": float(width) if width is not None else None,
                "hidden": hidden,
                "outline_level": int(getattr(dimension, "outlineLevel", 0) or 0),
                "_sort_index": sort_index,
            }
        )

    entries.sort(key=lambda entry: entry["_sort_index"])
    for entry in entries:
        entry.pop("_sort_index", None)
    return entries


def _custom_row_layout(ws: Worksheet) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for row_key, dimension in ws.row_dimensions.items():
        height = getattr(dimension, "height", None)
        hidden = bool(getattr(dimension, "hidden", False))
        if height is None and not hidden:
            continue
        try:
            row_index = int(row_key)
        except (TypeError, ValueError):
            continue

        entries.append(
            {
                "row": row_index,
                "height": float(height) if height is not None else None,
                "hidden": hidden,
                "outline_level": int(getattr(dimension, "outlineLevel", 0) or 0),
            }
        )

    entries.sort(key=lambda entry: entry["row"])
    return entries


def describe_sheet_layout(
    filepath: str,
    sheet_name: str,
    sample_limit: int = 10,
    free_canvas_rows: int = 8,
    free_canvas_cols: int = 6,
    free_canvas_limit: int = 3,
) -> dict[str, Any]:
    """Return a compact structural summary for one worksheet."""
    try:
        _validate_positive_int(sample_limit, label="sample_limit")
        _validate_positive_int(free_canvas_rows, label="free_canvas_rows")
        _validate_positive_int(free_canvas_cols, label="free_canvas_cols")
        _validate_positive_int(free_canvas_limit, label="free_canvas_limit")

        with safe_workbook(filepath) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=WorkbookError,
                operation="sheet-layout inspection",
            )
            from .chart import (
                _chart_occupied_range,
                _chart_type_name,
                _extract_chart_anchor,
                _extract_chart_dimensions,
                _extract_title_text,
                _find_free_canvas_slots_in_worksheet,
            )
            from .sheet import (
                _display_print_title_columns,
                _display_print_title_rows,
                _sheet_protection_state,
            )
            from .tables import _build_table_metadata

            rows, columns, column_range, is_empty = _get_sheet_usage(ws)
            merged_ranges = [str(merged_range) for merged_range in ws.merged_cells.ranges]
            tables = [
                _build_table_metadata(sheet_name, ws, table)
                for table in ws.tables.values()
            ]
            charts: list[dict[str, Any]] = []
            for chart_index, chart in enumerate(getattr(ws, "_charts", []), start=1):
                width, height = _extract_chart_dimensions(chart)
                anchor = _extract_chart_anchor(chart)
                series = getattr(chart, "ser", None) or list(getattr(chart, "series", []))
                chart_info = {
                    "chart_index": chart_index,
                    "chart_type": _chart_type_name(chart),
                    "anchor": anchor,
                    "title": _extract_title_text(getattr(chart, "title", None)),
                    "width": width,
                    "height": height,
                    "series_count": len(series),
                }
                if anchor and width and height:
                    chart_info["occupied_range"] = _chart_occupied_range(
                        ws,
                        anchor,
                        width=width,
                        height=height,
                    )
                charts.append({key: value for key, value in chart_info.items() if value is not None})

            validation_rules = _inspect_validation_rules(
                wb,
                ws=ws,
                sheet_name=sheet_name,
            )
            conditional_rules = _inspect_conditional_format_rules(
                wb,
                ws=ws,
                sheet_name=sheet_name,
            )
            free_canvas_preview = _find_free_canvas_slots_in_worksheet(
                ws,
                min_rows=free_canvas_rows,
                min_cols=free_canvas_cols,
                limit=free_canvas_limit,
            )
            free_canvas_preview["requested_block"] = {
                "rows": free_canvas_rows,
                "columns": free_canvas_cols,
            }

            custom_column_widths = _custom_column_layout(ws)
            custom_row_heights = _custom_row_layout(ws)
            visibility = ws.sheet_state
            used_range = _get_used_range(ws)
            freeze_panes = _freeze_panes_value(ws)
            autofilter_range = ws.auto_filter.ref or None
            print_area = ws.print_area or None
            print_title_rows = _display_print_title_rows(ws)
            print_title_columns = _display_print_title_columns(ws)
            protection = _sheet_protection_state(ws)

        warnings: list[str] = []
        sampled_groups = {
            "merged_ranges": len(merged_ranges),
            "tables": len(tables),
            "charts": len(charts),
            "data_validation_rules": len(validation_rules),
            "conditional_format_rules": len(conditional_rules),
            "custom_column_widths": len(custom_column_widths),
            "custom_row_heights": len(custom_row_heights),
        }
        for label, count in sampled_groups.items():
            if count > sample_limit:
                warnings.append(
                    f"Sampled {sample_limit} of {count} {label.replace('_', ' ')}"
                )

        result = {
            "sheet_name": sheet_name,
            "sheet_type": "worksheet",
            "visibility": visibility,
            "rows": rows,
            "columns": columns,
            "column_range": column_range,
            "used_range": used_range,
            "is_empty": is_empty,
            "freeze_panes": freeze_panes,
            "autofilter_range": autofilter_range,
            "print_area": print_area,
            "print_title_rows": print_title_rows,
            "print_title_columns": print_title_columns,
            "protection": protection,
            "summary": {
                "merged_range_count": len(merged_ranges),
                "table_count": len(tables),
                "chart_count": len(charts),
                "data_validation_rule_count": len(validation_rules),
                "conditional_format_rule_count": len(conditional_rules),
                "custom_column_width_count": len(custom_column_widths),
                "custom_row_height_count": len(custom_row_heights),
            },
            "merged_ranges": _sample_payload(merged_ranges, sample_limit=sample_limit),
            "tables": _sample_payload(tables, sample_limit=sample_limit),
            "charts": _sample_payload(charts, sample_limit=sample_limit),
            "data_validation_rules": _sample_payload(validation_rules, sample_limit=sample_limit),
            "conditional_format_rules": _sample_payload(conditional_rules, sample_limit=sample_limit),
            "custom_column_widths": _sample_payload(custom_column_widths, sample_limit=sample_limit),
            "custom_row_heights": _sample_payload(custom_row_heights, sample_limit=sample_limit),
            "free_canvas_preview": free_canvas_preview,
        }
        if warnings:
            result["warnings"] = warnings
        return result
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to describe sheet layout: {e}")
        raise WorkbookError(str(e))


def audit_workbook(
    filepath: str,
    header_row: int = 1,
    sample_limit: int = 25,
) -> dict[str, Any]:
    """Audit workbook structure for high-signal issues that affect agent workflows."""
    try:
        if not isinstance(header_row, int) or isinstance(header_row, bool) or header_row <= 0:
            raise WorkbookError("header_row must be a positive integer")
        if not isinstance(sample_limit, int) or isinstance(sample_limit, bool) or sample_limit <= 0:
            raise WorkbookError("sample_limit must be a positive integer")

        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")

        with safe_workbook(filepath) as wb:
            named_ranges = _serialize_named_ranges(wb)
            findings: list[dict[str, Any]] = []
            sheet_assessments: list[dict[str, Any]] = []
            worksheet_count = 0
            chartsheet_count = 0
            hidden_sheet_count = 0
            empty_sheet_count = 0
            layout_like_sheet_count = 0
            total_tables = 0
            total_charts = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_findings: list[dict[str, Any]] = []
                visibility = getattr(ws, "sheet_state", "visible")
                if visibility != "visible":
                    hidden_sheet_count += 1
                    sheet_findings.append(
                        _audit_finding(
                            "medium" if visibility == "veryHidden" else "low",
                            "hidden_sheet",
                            f"Sheet '{sheet_name}' is {visibility}.",
                            sheet_name=sheet_name,
                            recommendation="Confirm whether hidden sheets should be included before running broad workbook automation.",
                            details={"visibility": visibility},
                        )
                    )

                if _sheet_type(ws) == "chartsheet":
                    chartsheet_count += 1
                    chart_count = len(getattr(ws, "_charts", []))
                    total_charts += chart_count
                    if chart_count == 0:
                        sheet_findings.append(
                            _audit_finding(
                                "low",
                                "empty_chartsheet",
                                f"Chartsheet '{sheet_name}' does not currently contain any charts.",
                                sheet_name=sheet_name,
                                recommendation="Remove empty chart sheets or populate them before sharing the workbook.",
                            )
                        )

                    highest_severity = _highest_severity(sheet_findings)
                    findings.extend(sheet_findings)
                    sheet_assessments.append(
                        {
                            "sheet_name": sheet_name,
                            "sheet_type": "chartsheet",
                            "visibility": visibility,
                            "dataset_kind": "chartsheet",
                            "recommended_read_tool": "list_charts",
                            "chart_count": chart_count,
                            "table_count": 0,
                            "finding_count": len(sheet_findings),
                            "highest_severity": highest_severity,
                        }
                    )
                    continue

                worksheet_count += 1
                assessment = _worksheet_audit_assessment(
                    ws,
                    sheet_name=sheet_name,
                    header_row=header_row,
                )
                total_tables += len(assessment["native_tables"])
                total_charts += assessment["chart_count"]

                if assessment["dataset_kind"] == "empty_sheet":
                    empty_sheet_count += 1
                    sheet_findings.append(
                        _audit_finding(
                            "low",
                            "empty_sheet",
                            f"Worksheet '{sheet_name}' is empty.",
                            sheet_name=sheet_name,
                            recommendation="Skip empty sheets in automation flows unless you plan to populate them.",
                        )
                    )
                elif assessment["dataset_kind"] == "layout_like_sheet":
                    layout_like_sheet_count += 1
                    sheet_findings.append(
                        _audit_finding(
                            "low",
                            "layout_like_sheet",
                            f"Worksheet '{sheet_name}' looks layout-heavy rather than cleanly tabular.",
                            sheet_name=sheet_name,
                            recommendation="Prefer profile_workbook or read_data_from_excel before assuming a header-based table read.",
                            details={
                                "chart_count": assessment["chart_count"],
                                "merged_range_count": assessment["merged_range_count"],
                                "header_confidence": assessment["header_profile"]["confidence"],
                            },
                        )
                    )

                header_profile = assessment["header_quality_profile"]
                if assessment["dataset_kind"] != "layout_like_sheet" and header_profile["blank_headers"] > 0:
                    details = {"blank_headers": header_profile["blank_headers"]}
                    if assessment["header_quality_scope"] == "dominant_table":
                        details["scope"] = "dominant_table"
                        if assessment["dominant_table"] is not None:
                            details["table_name"] = assessment["dominant_table"]["table_name"]
                    sheet_findings.append(
                        _audit_finding(
                            "medium",
                            "blank_headers",
                            f"Worksheet '{sheet_name}' has blank cells in the configured header row.",
                            sheet_name=sheet_name,
                            recommendation="Fill blank headers before relying on object-mode reads, queries, or row updates by field name.",
                            details=details,
                        )
                    )

                if assessment["dataset_kind"] != "layout_like_sheet" and header_profile["duplicate_headers"] > 0:
                    details = {"duplicate_headers": header_profile["duplicate_headers"]}
                    if assessment["header_quality_scope"] == "dominant_table":
                        details["scope"] = "dominant_table"
                        if assessment["dominant_table"] is not None:
                            details["table_name"] = assessment["dominant_table"]["table_name"]
                    sheet_findings.append(
                        _audit_finding(
                            "medium",
                            "duplicate_headers",
                            f"Worksheet '{sheet_name}' has duplicate header labels in the configured header row.",
                            sheet_name=sheet_name,
                            recommendation="Deduplicate headers before relying on object-mode reads, queries, or row updates by field name.",
                            details=details,
                        )
                    )

                largest_table_rows = max(
                    [table["data_row_count"] for table in assessment["native_tables"]],
                    default=0,
                )
                if max(assessment["total_rows"], largest_table_rows) >= AUDIT_LARGE_DATASET_THRESHOLD:
                    details: dict[str, Any] = {
                        "row_count": max(assessment["total_rows"], largest_table_rows),
                    }
                    if assessment["dominant_table"] is not None:
                        details["table_name"] = assessment["dominant_table"]["table_name"]
                    sheet_findings.append(
                        _audit_finding(
                            "low",
                            "large_tabular_dataset",
                            f"Worksheet '{sheet_name}' contains a large structured dataset.",
                            sheet_name=sheet_name,
                            recommendation="Prefer query_table, aggregate_table, or paginated reads for large datasets instead of full workbook dumps.",
                            details=details,
                        )
                    )

                broken_formula_cells = _cells_with_broken_formula_references(
                    wb,
                    ws=ws,
                    sheet_name=sheet_name,
                )
                if broken_formula_cells:
                    sheet_findings.append(
                        _audit_finding(
                            "high",
                            "broken_formula_reference",
                            f"Worksheet '{sheet_name}' contains formulas with broken #REF! references.",
                            sheet_name=sheet_name,
                            recommendation="Repair broken formulas before trusting workbook calculations or mutation side effects.",
                            details={
                                "count": len(broken_formula_cells),
                                "sample": broken_formula_cells[:sample_limit],
                            },
                        )
                    )

                error_cells = _cells_with_error_values(ws)
                if error_cells:
                    sheet_findings.append(
                        _audit_finding(
                            "high",
                            "error_cells_present",
                            f"Worksheet '{sheet_name}' contains Excel error cells.",
                            sheet_name=sheet_name,
                            recommendation="Inspect and resolve workbook error cells before depending on the affected dataset.",
                            details={
                                "count": len(error_cells),
                                "sample": error_cells[:sample_limit],
                            },
                        )
                    )

                broken_validations = _broken_validation_rules(
                    wb,
                    ws=ws,
                    sheet_name=sheet_name,
                )
                if broken_validations:
                    sheet_findings.append(
                        _audit_finding(
                            "high",
                            "broken_validation_reference",
                            f"Worksheet '{sheet_name}' contains data validation rules with broken #REF! references.",
                            sheet_name=sheet_name,
                            recommendation="Repair data validation formulas before applying validation-aware writes or relying on workbook integrity.",
                            details={
                                "count": len(broken_validations),
                                "sample": broken_validations[:sample_limit],
                            },
                        )
                    )

                broken_conditional_formats = _broken_conditional_format_rules(
                    wb,
                    ws=ws,
                    sheet_name=sheet_name,
                )
                if broken_conditional_formats:
                    sheet_findings.append(
                        _audit_finding(
                            "high",
                            "broken_conditional_format_reference",
                            f"Worksheet '{sheet_name}' contains conditional formatting rules with broken #REF! references.",
                            sheet_name=sheet_name,
                            recommendation="Repair conditional formatting formulas before relying on dashboard semantics or visual QA.",
                            details={
                                "count": len(broken_conditional_formats),
                                "sample": broken_conditional_formats[:sample_limit],
                            },
                        )
                    )

                highest_severity = _highest_severity(sheet_findings)
                findings.extend(sheet_findings)
                sheet_assessments.append(
                    {
                        "sheet_name": sheet_name,
                        "sheet_type": "worksheet",
                        "visibility": visibility,
                        "used_range": assessment["used_range"],
                        "rows": assessment["rows"],
                        "columns": assessment["columns"],
                        "dataset_kind": assessment["dataset_kind"],
                        "recommended_read_tool": assessment["recommended_read_tool"],
                        "dominant_table_name": (
                            assessment["dominant_table"]["table_name"]
                            if assessment["dominant_table"] is not None
                            else None
                        ),
                        "table_count": len(assessment["native_tables"]),
                        "chart_count": assessment["chart_count"],
                        "finding_count": len(sheet_findings),
                        "highest_severity": highest_severity,
                    }
                )

            findings.extend(
                _workbook_named_range_findings(
                    wb,
                    named_ranges=named_ranges,
                )
            )

            severity_counter = Counter(finding["severity"] for finding in findings)
            code_counter = Counter(finding["code"] for finding in findings)
            high_count = severity_counter.get("high", 0)
            medium_count = severity_counter.get("medium", 0)
            low_count = severity_counter.get("low", 0)

            if high_count > 0 or medium_count >= 4:
                risk_level = "high"
            elif medium_count > 0:
                risk_level = "medium"
            else:
                risk_level = "low"

            recommendations = [
                finding["recommendation"]
                for finding in findings
                if finding.get("recommendation")
            ]
            deduped_recommendations: list[str] = []
            seen_recommendations: set[str] = set()
            for recommendation in recommendations:
                if recommendation in seen_recommendations:
                    continue
                seen_recommendations.add(recommendation)
                deduped_recommendations.append(recommendation)

            sheets_needing_attention = [
                {
                    "sheet_name": assessment["sheet_name"],
                    "highest_severity": assessment["highest_severity"],
                    "finding_count": assessment["finding_count"],
                }
                for assessment in sheet_assessments
                if assessment["finding_count"] > 0
            ]
            sheets_needing_attention.sort(
                key=lambda item: (
                    AUDIT_SEVERITY_RANK.get(item["highest_severity"] or "low", 0),
                    item["finding_count"],
                    item["sheet_name"],
                ),
                reverse=True,
            )

            return {
                "filename": path.name,
                "size": path.stat().st_size,
                "modified": path.stat().st_mtime,
                "summary": {
                    "risk_level": risk_level,
                    "finding_count": len(findings),
                    "high_count": high_count,
                    "medium_count": medium_count,
                    "low_count": low_count,
                    "sheet_count": len(wb.sheetnames),
                    "worksheet_count": worksheet_count,
                    "chartsheet_count": chartsheet_count,
                    "table_count": total_tables,
                    "chart_count": total_charts,
                    "named_range_count": len(named_ranges),
                    "hidden_sheet_count": hidden_sheet_count,
                    "empty_sheet_count": empty_sheet_count,
                    "layout_like_sheet_count": layout_like_sheet_count,
                    "sheets_needing_attention": sheets_needing_attention[:sample_limit],
                },
                "sheet_assessments": sheet_assessments,
                "findings": {
                    "count": len(findings),
                    "high_count": high_count,
                    "medium_count": medium_count,
                    "low_count": low_count,
                    "by_code": [
                        {"code": code, "count": count}
                        for code, count in sorted(
                            code_counter.items(),
                            key=lambda item: (-item[1], item[0]),
                        )
                    ],
                    "sample": findings[:sample_limit],
                    "truncated": len(findings) > sample_limit,
                },
                "recommended_actions": deduped_recommendations[:sample_limit],
            }

    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to audit workbook: {e}")
        raise WorkbookError(str(e))


def plan_workbook_repairs(
    filepath: str,
    header_row: int = 1,
    sample_limit: int = 25,
) -> dict[str, Any]:
    """Translate workbook audit findings into prioritized next steps for SheetForge users."""
    try:
        audit = audit_workbook(
            filepath,
            header_row=header_row,
            sample_limit=sample_limit,
        )

        audit_for_plan = audit
        findings_count = audit["findings"]["count"]
        if findings_count > sample_limit:
            audit_for_plan = audit_workbook(
                filepath,
                header_row=header_row,
                sample_limit=findings_count,
            )

        findings_sample = audit_for_plan["findings"]["sample"]
        sheet_assessments = {
            item["sheet_name"]: item for item in audit_for_plan["sheet_assessments"]
        }

        def findings_for(
            code: str,
            *,
            sheet_name: str | None = None,
        ) -> list[dict[str, Any]]:
            return [
                finding
                for finding in findings_sample
                if finding["code"] == code and finding.get("sheet_name") == sheet_name
            ]

        steps: list[dict[str, Any]] = []

        workbook_named_range_findings = [
            finding
            for finding in findings_sample
            if finding["code"] in {"broken_named_range_reference", "named_range_missing_sheet"}
        ]
        if workbook_named_range_findings:
            named_range_tools = []
            for finding in workbook_named_range_findings:
                details = finding.get("details", {})
                named_range_name = details.get("name")
                local_sheet = details.get("local_sheet")
                if not named_range_name:
                    continue
                named_range_tools.extend(
                    [
                        {
                            "tool": "inspect_named_range",
                            "args": {
                                "filepath": filepath,
                                "name": named_range_name,
                                "scope_sheet": local_sheet,
                            },
                        },
                        {
                            "tool": "delete_named_range",
                            "args": {
                                "filepath": filepath,
                                "name": named_range_name,
                                "scope_sheet": local_sheet,
                                "dry_run": True,
                            },
                        },
                    ]
                )
            steps.append(
                {
                    "priority": "high",
                    "title": "Inspect and repair workbook named ranges",
                    "finding_codes": sorted(
                        {
                            finding["code"]
                            for finding in workbook_named_range_findings
                        }
                    ),
                    "reason": "Broken or missing-sheet named ranges can silently break formulas, validations, and downstream automation.",
                    "can_execute_fully_in_sheetforge": True,
                    "suggested_tools": _unique_step_tools(
                        [
                            {
                                "tool": "list_named_ranges",
                                "args": {"filepath": filepath},
                            },
                            {
                                "tool": "apply_workbook_repairs",
                                "args": {
                                    "filepath": filepath,
                                    "repair_types": ["remove_broken_named_ranges"],
                                    "dry_run": True,
                                },
                            },
                            *named_range_tools,
                        ]
                    ),
                    "follow_up": "Inspect the broken scopes, then dry-run or apply named-range deletion for entries that are no longer valid.",
                }
            )

        for sheet_name, assessment in sheet_assessments.items():
            hidden_findings = findings_for("hidden_sheet", sheet_name=sheet_name)
            if hidden_findings:
                steps.append(
                    {
                        "priority": hidden_findings[0]["severity"],
                        "title": f"Review hidden sheet '{sheet_name}' before workbook-wide automation",
                        "sheet_name": sheet_name,
                        "finding_codes": ["hidden_sheet"],
                        "reason": "Hidden sheets can contain formulas, support data, or reporting logic that broad workbook automation might otherwise miss.",
                        "can_execute_fully_in_sheetforge": True,
                        "suggested_tools": _unique_step_tools(
                            [
                                {
                                    "tool": "set_worksheet_visibility",
                                    "args": {
                                        "filepath": filepath,
                                        "sheet_name": sheet_name,
                                        "visibility": "visible",
                                        "dry_run": True,
                                    },
                                },
                                {
                                    "tool": "quick_read",
                                    "args": {
                                        "filepath": filepath,
                                        "sheet_name": sheet_name,
                                        "max_rows": 5,
                                    },
                                },
                                {
                                    "tool": "apply_workbook_repairs",
                                    "args": {
                                        "filepath": filepath,
                                        "repair_types": ["reveal_hidden_sheets"],
                                        "sheet_names": [sheet_name],
                                        "dry_run": True,
                                    },
                                },
                            ]
                        ),
                        "follow_up": "If the hidden sheet matters to the workflow, rerun the planned automation after reviewing it.",
                    }
                )

            broken_formula_findings = findings_for("broken_formula_reference", sheet_name=sheet_name)
            if broken_formula_findings:
                sample_cells = broken_formula_findings[0].get("details", {}).get("sample", [])
                suggested_tools = []
                if sample_cells:
                    suggested_tools.append(
                        {
                            "tool": "read_data_from_excel",
                            "args": {
                                "filepath": filepath,
                                "sheet_name": sheet_name,
                                "start_cell": sample_cells[0],
                                "end_cell": sample_cells[0],
                            },
                        }
                    )
                steps.append(
                    {
                        "priority": "high",
                        "title": f"Repair broken formulas on '{sheet_name}'",
                        "sheet_name": sheet_name,
                        "finding_codes": ["broken_formula_reference"],
                        "reason": "Formulas with #REF! are already broken and can make downstream reads or edits unreliable.",
                        "can_execute_fully_in_sheetforge": False,
                        "suggested_tools": suggested_tools,
                        "follow_up": "After inspecting the broken cells, rewrite the formulas with apply_formula once the intended references are known.",
                    }
                )

            error_cell_findings = findings_for("error_cells_present", sheet_name=sheet_name)
            if error_cell_findings:
                sample_cells = error_cell_findings[0].get("details", {}).get("sample", [])
                suggested_tools = []
                if sample_cells:
                    suggested_tools.append(
                        {
                            "tool": "read_data_from_excel",
                            "args": {
                                "filepath": filepath,
                                "sheet_name": sheet_name,
                                "start_cell": sample_cells[0],
                                "end_cell": sample_cells[0],
                            },
                        }
                    )
                steps.append(
                    {
                        "priority": "high",
                        "title": f"Investigate Excel error cells on '{sheet_name}'",
                        "sheet_name": sheet_name,
                        "finding_codes": ["error_cells_present"],
                        "reason": "Workbook cells already evaluate to Excel errors, which can contaminate downstream calculations and exports.",
                        "can_execute_fully_in_sheetforge": False,
                        "suggested_tools": suggested_tools,
                        "follow_up": "Identify the upstream source of the error cells before mutating dependent workbook areas.",
                    }
                )

            validation_findings = findings_for("broken_validation_reference", sheet_name=sheet_name)
            if validation_findings:
                steps.append(
                    {
                        "priority": "high",
                        "title": f"Repair broken data validation rules on '{sheet_name}'",
                        "sheet_name": sheet_name,
                        "finding_codes": ["broken_validation_reference"],
                        "reason": "Validation formulas with #REF! can mislead users and break validation-aware automation.",
                        "can_execute_fully_in_sheetforge": True,
                        "suggested_tools": [
                            {
                                "tool": "inspect_data_validation_rules",
                                "args": {
                                    "filepath": filepath,
                                    "sheet_name": sheet_name,
                                    "broken_only": True,
                                },
                            },
                            {
                                "tool": "remove_data_validation_rules",
                                "args": {
                                    "filepath": filepath,
                                    "sheet_name": sheet_name,
                                    "broken_only": True,
                                    "dry_run": True,
                                },
                            },
                            {
                                "tool": "apply_workbook_repairs",
                                "args": {
                                    "filepath": filepath,
                                    "repair_types": ["remove_broken_validations"],
                                    "sheet_names": [sheet_name],
                                    "dry_run": True,
                                },
                            },
                        ],
                        "follow_up": "Inspect the broken validation rules, then dry-run or apply removal for the invalid entries before recreating them if needed.",
                    }
                )

            conditional_format_findings = findings_for(
                "broken_conditional_format_reference",
                sheet_name=sheet_name,
            )
            if conditional_format_findings:
                steps.append(
                    {
                        "priority": "high",
                        "title": f"Review broken conditional formatting rules on '{sheet_name}'",
                        "sheet_name": sheet_name,
                        "finding_codes": ["broken_conditional_format_reference"],
                        "reason": "Broken conditional formatting rules can make dashboards and visual QA misleading even when cell values still exist.",
                        "can_execute_fully_in_sheetforge": True,
                        "suggested_tools": [
                            {
                                "tool": "inspect_conditional_format_rules",
                                "args": {
                                    "filepath": filepath,
                                    "sheet_name": sheet_name,
                                    "broken_only": True,
                                },
                            },
                            {
                                "tool": "remove_conditional_format_rules",
                                "args": {
                                    "filepath": filepath,
                                    "sheet_name": sheet_name,
                                    "broken_only": True,
                                    "dry_run": True,
                                },
                            },
                            {
                                "tool": "apply_workbook_repairs",
                                "args": {
                                    "filepath": filepath,
                                    "repair_types": ["remove_broken_conditional_formats"],
                                    "sheet_names": [sheet_name],
                                    "dry_run": True,
                                },
                            },
                        ],
                        "follow_up": "Inspect the broken conditional-format rules, then dry-run or apply removal for invalid rules before recreating them if needed.",
                    }
                )

            if (
                assessment["dataset_kind"] != "layout_like_sheet"
                and assessment.get("dominant_table_name") is None
            ):
                header_issue_findings = [
                    *findings_for("blank_headers", sheet_name=sheet_name),
                    *findings_for("duplicate_headers", sheet_name=sheet_name),
                ]
                if header_issue_findings:
                    steps.append(
                        {
                            "priority": "medium",
                            "title": f"Normalize headers on '{sheet_name}'",
                            "sheet_name": sheet_name,
                            "finding_codes": sorted(
                                {
                                    finding["code"] for finding in header_issue_findings
                                }
                            ),
                            "reason": "Blank or duplicate headers make object-mode reads, queries, and keyed updates harder to trust.",
                            "can_execute_fully_in_sheetforge": False,
                            "suggested_tools": [
                                {
                                    "tool": "quick_read",
                                    "args": {
                                        "filepath": filepath,
                                        "sheet_name": sheet_name,
                                        "header_row": header_row,
                                        "max_rows": 3,
                                        "row_mode": "arrays",
                                    },
                                }
                            ],
                            "follow_up": "After inspecting the current header row, rewrite the header cells with write_data_to_excel before relying on field-based reads.",
                        }
                    )

            if assessment["dataset_kind"] == "layout_like_sheet":
                used_range = assessment.get("used_range") or "A1"
                end_cell = used_range.split(":")[-1]
                steps.append(
                    {
                        "priority": "low",
                        "title": f"Treat '{sheet_name}' as a layout-oriented sheet",
                        "sheet_name": sheet_name,
                        "finding_codes": ["layout_like_sheet"],
                        "reason": "This sheet looks dashboard-like, so forcing a tabular workflow is more likely to produce noisy or misleading results.",
                        "can_execute_fully_in_sheetforge": True,
                        "suggested_tools": [
                            {
                                "tool": "profile_workbook",
                                "args": {"filepath": filepath},
                            },
                            {
                                "tool": "read_data_from_excel",
                                "args": {
                                    "filepath": filepath,
                                    "sheet_name": sheet_name,
                                    "start_cell": "A1",
                                    "end_cell": end_cell,
                                    "values_only": True,
                                    "preview_only": True,
                                },
                            },
                        ],
                        "follow_up": "Use chart, layout, or range-based tools on this sheet instead of assuming a row-based dataset.",
                    }
                )

        sorted_steps = _sort_steps(steps)
        quick_wins = [
            step["title"]
            for step in sorted_steps
            if step["priority"] in {"high", "medium"} and step["can_execute_fully_in_sheetforge"]
        ]

        return {
            "audit_summary": audit["summary"],
            "step_count": len(sorted_steps),
            "steps": sorted_steps,
            "quick_wins": quick_wins[:sample_limit],
        }

    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to plan workbook repairs: {e}")
        raise WorkbookError(str(e))


def analyze_range_impact(filepath: str, sheet_name: str, range_ref: str) -> dict[str, Any]:
    """Inspect workbook structures that overlap a worksheet range before mutation."""
    try:
        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")

        with safe_workbook(filepath) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=WorkbookError,
                operation="range impact analysis",
            )
            target_bounds, normalized_range = _parse_range_reference(
                range_ref,
                worksheet=ws,
                expected_sheet=sheet_name,
                error_cls=WorkbookError,
            )

            from .chart import (
                DEFAULT_CHART_HEIGHT,
                DEFAULT_CHART_WIDTH,
                _chart_occupied_range,
                _chart_type_name,
                _extract_chart_anchor,
                _extract_chart_dimensions,
                _extract_title_text,
            )
            from .tables import _build_table_metadata

            tables: list[dict[str, Any]] = []
            for table in ws.tables.values():
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                table_bounds = (min_row, min_col, max_row, max_col)
                intersection = _intersection_bounds(target_bounds, table_bounds)
                if intersection is None:
                    continue

                metadata = _build_table_metadata(sheet_name, ws, table)
                header_row_count = metadata["header_row_count"]
                totals_row_count = metadata["totals_row_count"]
                header_bounds = (
                    (min_row, min_col, min_row + header_row_count - 1, max_col)
                    if header_row_count > 0
                    else None
                )
                data_start_row = min_row + header_row_count
                data_end_row = max_row - totals_row_count
                data_bounds = (
                    (data_start_row, min_col, data_end_row, max_col)
                    if data_start_row <= data_end_row
                    else None
                )
                totals_bounds = (
                    (max_row - totals_row_count + 1, min_col, max_row, max_col)
                    if totals_row_count > 0
                    else None
                )
                tables.append(
                    {
                        "table_name": table.displayName,
                        "range": table.ref,
                        "intersection_range": _bounds_to_range(*intersection),
                        "covers_header": bool(
                            header_bounds and _bounds_intersect(target_bounds, header_bounds)
                        ),
                        "covers_data": bool(
                            data_bounds and _bounds_intersect(target_bounds, data_bounds)
                        ),
                        "covers_totals_row": bool(
                            totals_bounds and _bounds_intersect(target_bounds, totals_bounds)
                        ),
                    }
                )

            charts: list[dict[str, Any]] = []
            for chart_index, chart in enumerate(getattr(ws, "_charts", []), start=1):
                anchor = _extract_chart_anchor(chart)
                if not anchor:
                    continue
                width, height = _extract_chart_dimensions(chart)
                occupied_range = _chart_occupied_range(
                    ws,
                    anchor,
                    width=width or DEFAULT_CHART_WIDTH,
                    height=height or DEFAULT_CHART_HEIGHT,
                )
                chart_bounds, _ = _parse_range_reference(occupied_range, error_cls=WorkbookError)
                intersection = _intersection_bounds(target_bounds, chart_bounds)
                if intersection is None:
                    continue
                charts.append(
                    {
                        "chart_index": chart_index,
                        "chart_type": _chart_type_name(chart),
                        "title": _extract_title_text(getattr(chart, "title", None)),
                        "anchor": anchor,
                        "occupied_range": occupied_range,
                        "intersection_range": _bounds_to_range(*intersection),
                    }
                )

            merged_ranges: list[dict[str, Any]] = []
            for merged_range in ws.merged_cells.ranges:
                merged_bounds, merged_ref = _parse_range_reference(
                    str(merged_range),
                    worksheet=ws,
                    error_cls=WorkbookError,
                )
                intersection = _intersection_bounds(target_bounds, merged_bounds)
                if intersection is None:
                    continue
                merged_ranges.append(
                    {
                        "range": merged_ref,
                        "intersection_range": _bounds_to_range(*intersection),
                    }
                )

            named_ranges: list[dict[str, Any]] = []
            for named_range in _serialize_named_ranges(wb):
                matching_destinations = []
                for destination in named_range["destinations"]:
                    if destination["sheet_name"] != sheet_name:
                        continue
                    for destination_bounds, destination_ref in _iter_range_references(
                        destination["range"],
                        worksheet=ws,
                        expected_sheet=sheet_name,
                    ):
                        intersection = _intersection_bounds(target_bounds, destination_bounds)
                        if intersection is None:
                            continue
                        matching_destinations.append(
                            {
                                "range": destination_ref,
                                "intersection_range": _bounds_to_range(*intersection),
                            }
                        )

                if matching_destinations:
                    named_ranges.append(
                        {
                            "name": named_range["name"],
                            "local_sheet": named_range["local_sheet"],
                            "hidden": named_range["hidden"],
                            "destinations": matching_destinations,
                        }
                    )

            autofilter = None
            if ws.auto_filter.ref:
                autofilter_bounds, autofilter_ref = _parse_range_reference(
                    ws.auto_filter.ref,
                    worksheet=ws,
                    error_cls=WorkbookError,
                )
                intersection = _intersection_bounds(target_bounds, autofilter_bounds)
                if intersection is not None:
                    autofilter = {
                        "range": autofilter_ref,
                        "intersection_range": _bounds_to_range(*intersection),
                    }

            print_area_matches = []
            for print_bounds, print_ref in _iter_range_references(
                ws.print_area,
                worksheet=ws,
                expected_sheet=sheet_name,
            ):
                intersection = _intersection_bounds(target_bounds, print_bounds)
                if intersection is None:
                    continue
                print_area_matches.append(
                    {
                        "range": print_ref,
                        "intersection_range": _bounds_to_range(*intersection),
                    }
                )

            formula_cells = []
            for row in ws.iter_rows(
                min_row=target_bounds[0],
                max_row=target_bounds[2],
                min_col=target_bounds[1],
                max_col=target_bounds[3],
            ):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells.append(cell.coordinate)

            dependent_formulas = _extract_formula_dependencies(
                wb,
                target_sheet=sheet_name,
                target_bounds=target_bounds,
            )
            direct_formula_count = sum(
                1 for dependency in dependent_formulas if dependency.get("dependency_depth", 1) == 1
            )
            transitive_formula_count = len(dependent_formulas) - direct_formula_count
            data_validations = _extract_validation_overlaps(
                ws,
                sheet_name=sheet_name,
                target_bounds=target_bounds,
            )
            dependent_validations = _extract_validation_dependencies(
                wb,
                target_sheet=sheet_name,
                target_bounds=target_bounds,
            )
            conditional_formats = _extract_conditional_format_overlaps(
                ws,
                sheet_name=sheet_name,
                target_bounds=target_bounds,
            )
            dependent_conditional_formats = _extract_conditional_format_dependencies(
                wb,
                target_sheet=sheet_name,
                target_bounds=target_bounds,
            )

            impact_score = (
                len(tables) * 3
                + len(charts) * 3
                + len(merged_ranges) * 2
                + len(named_ranges) * 2
                + len(data_validations) * 2
                + len(conditional_formats) * 2
                + (1 if formula_cells else 0)
                + len(dependent_formulas) * 3
                + len(dependent_validations) * 2
                + len(dependent_conditional_formats) * 2
                + (1 if autofilter else 0)
                + (1 if print_area_matches else 0)
            )
            if impact_score >= 3:
                risk_level = "high"
            elif impact_score >= 1:
                risk_level = "medium"
            else:
                risk_level = "low"

            hints: list[str] = []
            if tables:
                hints.append("Selected range overlaps native Excel tables.")
                if any(table["covers_header"] for table in tables):
                    hints.append("Table header rows are inside the selected range.")
                if any(table["covers_totals_row"] for table in tables):
                    hints.append("Table totals rows are inside the selected range.")
            if charts:
                hints.append("Selected range overlaps embedded chart footprints.")
            if merged_ranges:
                hints.append("Selected range touches merged cells that may need to be unmerged first.")
            if named_ranges:
                hints.append("Named ranges point into the selected range.")
            if data_validations:
                hints.append("Selected range overlaps worksheet data validation rules.")
            if conditional_formats:
                hints.append("Selected range overlaps conditional formatting rules.")
            if autofilter:
                hints.append("Selected range overlaps the worksheet autofilter.")
            if print_area_matches:
                hints.append("Selected range overlaps the worksheet print area.")
            if formula_cells:
                hints.append("Selected range contains formula cells that may recalculate or break.")
            if dependent_formulas:
                hints.append("Formulas elsewhere in the workbook reference the selected range.")
            if dependent_validations:
                hints.append("Validation rules elsewhere in the workbook reference the selected range.")
            if dependent_conditional_formats:
                hints.append("Conditional formatting rules reference the selected range.")
            if not hints:
                hints.append("No overlapping workbook structures detected for this range.")

            return {
                "sheet_name": sheet_name,
                "range": normalized_range,
                "used_range": _get_used_range(ws),
                "summary": {
                    "risk_level": risk_level,
                    "table_count": len(tables),
                    "chart_count": len(charts),
                    "merged_range_count": len(merged_ranges),
                    "named_range_count": len(named_ranges),
                    "data_validation_count": len(data_validations),
                    "conditional_format_count": len(conditional_formats),
                    "formula_cell_count": len(formula_cells),
                    "dependent_formula_count": len(dependent_formulas),
                    "direct_formula_count": direct_formula_count,
                    "transitive_formula_count": transitive_formula_count,
                    "dependent_validation_count": len(dependent_validations),
                    "dependent_conditional_format_count": len(dependent_conditional_formats),
                    "autofilter_overlap": autofilter is not None,
                    "print_area_overlap": bool(print_area_matches),
                },
                "tables": tables,
                "charts": charts,
                "merged_ranges": merged_ranges,
                "named_ranges": named_ranges,
                "data_validations": {
                    "count": len(data_validations),
                    "sample": data_validations[:10],
                },
                "conditional_formats": {
                    "count": len(conditional_formats),
                    "sample": conditional_formats[:10],
                },
                "formula_cells": {
                    "count": len(formula_cells),
                    "sample": formula_cells[:10],
                },
                "dependent_formulas": {
                    "count": len(dependent_formulas),
                    "direct_count": direct_formula_count,
                    "transitive_count": transitive_formula_count,
                    "sample": dependent_formulas[:10],
                },
                "dependent_validations": {
                    "count": len(dependent_validations),
                    "sample": dependent_validations[:10],
                },
                "dependent_conditional_formats": {
                    "count": len(dependent_conditional_formats),
                    "sample": dependent_conditional_formats[:10],
                },
                "worksheet_features": {
                    "autofilter": autofilter,
                    "print_area": print_area_matches,
                },
                "hints": hints,
            }

    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to analyze range impact: {e}")
        raise WorkbookError(str(e))


def inspect_data_validation_rules(
    filepath: str,
    sheet_name: str,
    broken_only: bool = False,
) -> dict[str, Any]:
    """Inspect worksheet data validation rules with stable rule indexes."""
    try:
        if isinstance(broken_only, bool) is False:
            raise WorkbookError("broken_only must be a boolean")

        with safe_workbook(filepath) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=WorkbookError,
                operation="data validation inspection",
            )
            rules = _inspect_validation_rules(
                wb,
                ws=ws,
                sheet_name=sheet_name,
            )
            if broken_only:
                rules = [rule for rule in rules if rule["broken_reference"]]
            return {
                "sheet_name": sheet_name,
                "broken_only": broken_only,
                "rule_count": len(rules),
                "rules": rules,
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to inspect data validation rules: {e}")
        raise WorkbookError(str(e))


def inspect_conditional_format_rules(
    filepath: str,
    sheet_name: str,
    broken_only: bool = False,
) -> dict[str, Any]:
    """Inspect worksheet conditional formatting rules with stable rule indexes."""
    try:
        if isinstance(broken_only, bool) is False:
            raise WorkbookError("broken_only must be a boolean")

        with safe_workbook(filepath) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=WorkbookError,
                operation="conditional formatting inspection",
            )
            rules = _inspect_conditional_format_rules(
                wb,
                ws=ws,
                sheet_name=sheet_name,
            )
            if broken_only:
                rules = [rule for rule in rules if rule["broken_reference"]]
            return {
                "sheet_name": sheet_name,
                "broken_only": broken_only,
                "rule_count": len(rules),
                "rules": rules,
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to inspect conditional formatting rules: {e}")
        raise WorkbookError(str(e))


def delete_named_range(
    filepath: str,
    name: str,
    scope_sheet: str | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Delete a workbook-level or sheet-scoped named range."""
    try:
        if not str(name).strip():
            raise WorkbookError("name is required")
        if isinstance(dry_run, bool) is False:
            raise WorkbookError("dry_run must be a boolean")

        with safe_workbook(filepath, save=not dry_run) as wb:
            removed = _remove_named_range_sources(
                wb,
                name=str(name).strip(),
                scope_sheet=scope_sheet,
            )
            return {
                "message": (
                    f"{'Previewed deletion of' if dry_run else 'Deleted'} "
                    f"{len(removed)} named range(s)"
                ),
                "name": str(name).strip(),
                "scope_sheet": scope_sheet,
                "removed_count": len(removed),
                "removed": removed,
                "dry_run": dry_run,
                "changes": [
                    {
                        "type": "delete_named_range",
                        "name": item["name"],
                        "local_sheet": item["local_sheet"],
                        "scope": item["container"],
                    }
                    for item in removed
                ],
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to delete named range '{name}': {e}")
        raise WorkbookError(str(e))


def remove_data_validation_rules(
    filepath: str,
    sheet_name: str,
    rule_indexes: list[int] | None = None,
    broken_only: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Remove selected worksheet data validation rules."""
    try:
        if isinstance(broken_only, bool) is False:
            raise WorkbookError("broken_only must be a boolean")
        if isinstance(dry_run, bool) is False:
            raise WorkbookError("dry_run must be a boolean")
        normalized_indexes = _normalize_rule_indexes(rule_indexes, label="rule_indexes")
        if not normalized_indexes and not broken_only:
            raise WorkbookError("Specify rule_indexes or set broken_only=True")

        with safe_workbook(filepath, save=not dry_run) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=WorkbookError,
                operation="data validation removal",
            )
            before_rules = _inspect_validation_rules(
                wb,
                ws=ws,
                sheet_name=sheet_name,
            )
            if dry_run:
                removed = [
                    rule
                    for rule in before_rules
                    if (rule["rule_index"] in normalized_indexes)
                    or (broken_only and rule["broken_reference"])
                ]
            else:
                removed = _remove_validation_rules(
                    wb,
                    ws,
                    rule_indexes=normalized_indexes,
                    broken_only=broken_only,
                )
            return {
                "message": (
                    f"{'Previewed removal of' if dry_run else 'Removed'} "
                    f"{len(removed)} data validation rule(s) from '{sheet_name}'"
                ),
                "sheet_name": sheet_name,
                "broken_only": broken_only,
                "rule_indexes": normalized_indexes,
                "removed_count": len(removed),
                "removed_rules": removed,
                "remaining_rule_count": len(before_rules) - len(removed),
                "dry_run": dry_run,
                "changes": [
                    {
                        "type": "remove_data_validation_rule",
                        "sheet_name": sheet_name,
                        "rule_index": rule["rule_index"],
                        "applies_to": rule["applies_to"],
                    }
                    for rule in removed
                ],
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to remove data validation rules: {e}")
        raise WorkbookError(str(e))


def remove_conditional_format_rules(
    filepath: str,
    sheet_name: str,
    rule_indexes: list[int] | None = None,
    broken_only: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Remove selected worksheet conditional formatting rules."""
    try:
        if isinstance(broken_only, bool) is False:
            raise WorkbookError("broken_only must be a boolean")
        if isinstance(dry_run, bool) is False:
            raise WorkbookError("dry_run must be a boolean")
        normalized_indexes = _normalize_rule_indexes(rule_indexes, label="rule_indexes")
        if not normalized_indexes and not broken_only:
            raise WorkbookError("Specify rule_indexes or set broken_only=True")

        with safe_workbook(filepath, save=not dry_run) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=WorkbookError,
                operation="conditional formatting removal",
            )
            before_rules = _inspect_conditional_format_rules(
                wb,
                ws=ws,
                sheet_name=sheet_name,
            )
            if dry_run:
                removed = [
                    rule
                    for rule in before_rules
                    if (rule["rule_index"] in normalized_indexes)
                    or (broken_only and rule["broken_reference"])
                ]
            else:
                removed = _remove_conditional_format_rules(
                    wb,
                    ws,
                    rule_indexes=normalized_indexes,
                    broken_only=broken_only,
                )
            return {
                "message": (
                    f"{'Previewed removal of' if dry_run else 'Removed'} "
                    f"{len(removed)} conditional formatting rule(s) from '{sheet_name}'"
                ),
                "sheet_name": sheet_name,
                "broken_only": broken_only,
                "rule_indexes": normalized_indexes,
                "removed_count": len(removed),
                "removed_rules": removed,
                "remaining_rule_count": len(before_rules) - len(removed),
                "dry_run": dry_run,
                "changes": [
                    {
                        "type": "remove_conditional_format_rule",
                        "sheet_name": sheet_name,
                        "rule_index": rule["rule_index"],
                        "applies_to": rule["applies_to"],
                    }
                    for rule in removed
                ],
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to remove conditional formatting rules: {e}")
        raise WorkbookError(str(e))


def _normalize_string_list(
    values: list[str] | None,
    *,
    label: str,
) -> list[str]:
    if values is None:
        return []
    if not isinstance(values, list):
        raise WorkbookError(f"{label} must be a list of strings")

    normalized: list[str] = []
    for value in values:
        if not isinstance(value, str) or not value.strip():
            raise WorkbookError(f"{label} must contain only non-empty strings")
        normalized.append(value.strip())
    return normalized


def _normalize_repair_types(repair_types: list[str] | None) -> list[str]:
    normalized = _normalize_string_list(repair_types, label="repair_types")
    invalid = sorted(set(normalized) - SUPPORTED_REPAIR_TYPES)
    if invalid:
        raise WorkbookError(
            "Unsupported repair_types: " + ", ".join(invalid)
        )
    return normalized


def _snapshot_workbook_state(
    wb: Any,
) -> dict[str, Any]:
    from .chart import (
        DEFAULT_CHART_HEIGHT,
        DEFAULT_CHART_WIDTH,
        _chart_occupied_range,
        _chart_type_name,
        _extract_chart_anchor,
        _extract_chart_dimensions,
        _extract_title_text,
    )
    from .tables import _build_table_metadata

    sheets: dict[str, dict[str, Any]] = {}
    total_tables = 0
    total_charts = 0
    total_validation_rules = 0
    total_conditional_rules = 0
    hidden_sheet_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        visibility = getattr(ws, "sheet_state", "visible")
        if visibility != "visible":
            hidden_sheet_count += 1

        if _sheet_type(ws) == "chartsheet":
            charts = []
            for chart_index, chart in enumerate(getattr(ws, "_charts", []), start=1):
                width, height = _extract_chart_dimensions(chart)
                anchor = _extract_chart_anchor(chart)
                chart_info = {
                    "chart_index": chart_index,
                    "chart_type": _chart_type_name(chart),
                    "anchor": anchor,
                    "title": _extract_title_text(getattr(chart, "title", None)),
                    "width": width,
                    "height": height,
                }
                if anchor:
                    chart_info["occupied_range"] = _chart_occupied_range(
                        ws,
                        anchor,
                        width=width or DEFAULT_CHART_WIDTH,
                        height=height or DEFAULT_CHART_HEIGHT,
                    )
                charts.append(chart_info)

            total_charts += len(charts)
            sheets[sheet_name] = {
                "sheet_name": sheet_name,
                "sheet_type": "chartsheet",
                "visibility": visibility,
                "charts": charts,
            }
            continue

        rows, columns, _, _ = _get_sheet_usage(ws)
        tables = [
            {
                "table_name": metadata["table_name"],
                "range": metadata["range"],
                "data_row_count": metadata["data_row_count"],
                "column_count": metadata["column_count"],
            }
            for metadata in (
                _build_table_metadata(sheet_name, ws, table)
                for table in ws.tables.values()
            )
        ]
        charts = []
        for chart_index, chart in enumerate(getattr(ws, "_charts", []), start=1):
            width, height = _extract_chart_dimensions(chart)
            anchor = _extract_chart_anchor(chart)
            chart_info = {
                "chart_index": chart_index,
                "chart_type": _chart_type_name(chart),
                "anchor": anchor,
                "title": _extract_title_text(getattr(chart, "title", None)),
                "width": width,
                "height": height,
            }
            if anchor:
                chart_info["occupied_range"] = _chart_occupied_range(
                    ws,
                    anchor,
                    width=width or DEFAULT_CHART_WIDTH,
                    height=height or DEFAULT_CHART_HEIGHT,
                )
            charts.append(chart_info)

        validation_rules = _inspect_validation_rules(
            wb,
            ws=ws,
            sheet_name=sheet_name,
        )
        conditional_rules = _inspect_conditional_format_rules(
            wb,
            ws=ws,
            sheet_name=sheet_name,
        )

        total_tables += len(tables)
        total_charts += len(charts)
        total_validation_rules += len(validation_rules)
        total_conditional_rules += len(conditional_rules)

        sheets[sheet_name] = {
            "sheet_name": sheet_name,
            "sheet_type": "worksheet",
            "visibility": visibility,
            "rows": rows,
            "columns": columns,
            "used_range": _get_used_range(ws),
            "tables": tables,
            "charts": charts,
            "validation_rules": validation_rules,
            "conditional_format_rules": conditional_rules,
        }

    named_ranges = _serialize_named_ranges(wb)
    return {
        "summary": {
            "sheet_count": len(wb.sheetnames),
            "table_count": total_tables,
            "chart_count": total_charts,
            "named_range_count": len(named_ranges),
            "validation_rule_count": total_validation_rules,
            "conditional_format_rule_count": total_conditional_rules,
            "hidden_sheet_count": hidden_sheet_count,
        },
        "sheet_order": list(wb.sheetnames),
        "sheets": sheets,
        "named_ranges": named_ranges,
    }


def _signature_map(
    items: list[dict[str, Any]],
    *,
    key_fields: tuple[str, ...],
) -> dict[tuple[Any, ...], dict[str, Any]]:
    return {
        tuple(item.get(field) for field in key_fields): item
        for item in items
    }


def _strip_rule_index(rule: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in rule.items()
        if key != "rule_index"
    }


def _freeze_signature_value(value: Any) -> Any:
    if isinstance(value, list):
        return tuple(_freeze_signature_value(item) for item in value)
    if isinstance(value, dict):
        return tuple(
            (key, _freeze_signature_value(item))
            for key, item in sorted(value.items())
        )
    return value


def _diff_named_ranges(
    before_snapshot: dict[str, Any],
    after_snapshot: dict[str, Any],
) -> dict[str, list[dict[str, Any]]]:
    before_map = _signature_map(
        before_snapshot["named_ranges"],
        key_fields=("name", "local_sheet"),
    )
    after_map = _signature_map(
        after_snapshot["named_ranges"],
        key_fields=("name", "local_sheet"),
    )

    added = [
        item for key, item in after_map.items()
        if key not in before_map
    ]
    removed = [
        item for key, item in before_map.items()
        if key not in after_map
    ]
    changed = [
        {
            "name": key[0],
            "local_sheet": key[1],
            "before": before_map[key],
            "after": after_map[key],
        }
        for key in before_map.keys() & after_map.keys()
        if before_map[key] != after_map[key]
    ]
    return {
        "added": sorted(added, key=lambda item: (item["name"], str(item["local_sheet"]))),
        "removed": sorted(removed, key=lambda item: (item["name"], str(item["local_sheet"]))),
        "changed": sorted(changed, key=lambda item: (item["name"], str(item["local_sheet"]))),
    }


def _diff_workbook_snapshots(
    before_snapshot: dict[str, Any],
    after_snapshot: dict[str, Any],
    *,
    sample_limit: int,
) -> dict[str, Any]:
    before_sheets = before_snapshot["sheets"]
    after_sheets = after_snapshot["sheets"]

    added_sheet_names = sorted(set(after_sheets) - set(before_sheets))
    removed_sheet_names = sorted(set(before_sheets) - set(after_sheets))

    sheet_property_changes: list[dict[str, Any]] = []
    table_changes: list[dict[str, Any]] = []
    chart_changes: list[dict[str, Any]] = []
    validation_changes: list[dict[str, Any]] = []
    conditional_changes: list[dict[str, Any]] = []

    for sheet_name in sorted(set(before_sheets) & set(after_sheets)):
        before_sheet = before_sheets[sheet_name]
        after_sheet = after_sheets[sheet_name]

        for field in ("sheet_type", "visibility", "used_range", "rows", "columns"):
            if before_sheet.get(field) != after_sheet.get(field):
                sheet_property_changes.append(
                    {
                        "sheet_name": sheet_name,
                        "field": field,
                        "before": before_sheet.get(field),
                        "after": after_sheet.get(field),
                    }
                )

        before_tables = _signature_map(
            before_sheet.get("tables", []),
            key_fields=("table_name",),
        )
        after_tables = _signature_map(
            after_sheet.get("tables", []),
            key_fields=("table_name",),
        )
        for table_name in sorted(set(after_tables) - set(before_tables)):
            table_changes.append(
                {
                    "sheet_name": sheet_name,
                    "change_type": "added",
                    "table": after_tables[table_name],
                }
            )
        for table_name in sorted(set(before_tables) - set(after_tables)):
            table_changes.append(
                {
                    "sheet_name": sheet_name,
                    "change_type": "removed",
                    "table": before_tables[table_name],
                }
            )
        for table_name in sorted(set(before_tables) & set(after_tables)):
            if before_tables[table_name] != after_tables[table_name]:
                table_changes.append(
                    {
                        "sheet_name": sheet_name,
                        "change_type": "changed",
                        "before": before_tables[table_name],
                        "after": after_tables[table_name],
                    }
                )

        before_charts = _signature_map(
            before_sheet.get("charts", []),
            key_fields=("anchor", "chart_type", "title"),
        )
        after_charts = _signature_map(
            after_sheet.get("charts", []),
            key_fields=("anchor", "chart_type", "title"),
        )
        for chart_key in sorted(set(after_charts) - set(before_charts)):
            chart_changes.append(
                {
                    "sheet_name": sheet_name,
                    "change_type": "added",
                    "chart": after_charts[chart_key],
                }
            )
        for chart_key in sorted(set(before_charts) - set(after_charts)):
            chart_changes.append(
                {
                    "sheet_name": sheet_name,
                    "change_type": "removed",
                    "chart": before_charts[chart_key],
                }
            )
        for chart_key in sorted(set(before_charts) & set(after_charts)):
            if before_charts[chart_key] != after_charts[chart_key]:
                chart_changes.append(
                    {
                        "sheet_name": sheet_name,
                        "change_type": "changed",
                        "before": before_charts[chart_key],
                        "after": after_charts[chart_key],
                    }
                )

        before_validation = {
            tuple(
                (key, _freeze_signature_value(value))
                for key, value in _strip_rule_index(rule).items()
            ): _strip_rule_index(rule)
            for rule in before_sheet.get("validation_rules", [])
        }
        after_validation = {
            tuple(
                (key, _freeze_signature_value(value))
                for key, value in _strip_rule_index(rule).items()
            ): _strip_rule_index(rule)
            for rule in after_sheet.get("validation_rules", [])
        }
        for rule_key in sorted(set(after_validation) - set(before_validation)):
            validation_changes.append(
                {
                    "sheet_name": sheet_name,
                    "change_type": "added",
                    "rule": after_validation[rule_key],
                }
            )
        for rule_key in sorted(set(before_validation) - set(after_validation)):
            validation_changes.append(
                {
                    "sheet_name": sheet_name,
                    "change_type": "removed",
                    "rule": before_validation[rule_key],
                }
            )

        before_conditional = {
            tuple(
                (key, _freeze_signature_value(value))
                for key, value in _strip_rule_index(rule).items()
            ): _strip_rule_index(rule)
            for rule in before_sheet.get("conditional_format_rules", [])
        }
        after_conditional = {
            tuple(
                (key, _freeze_signature_value(value))
                for key, value in _strip_rule_index(rule).items()
            ): _strip_rule_index(rule)
            for rule in after_sheet.get("conditional_format_rules", [])
        }
        for rule_key in sorted(set(after_conditional) - set(before_conditional)):
            conditional_changes.append(
                {
                    "sheet_name": sheet_name,
                    "change_type": "added",
                    "rule": after_conditional[rule_key],
                }
            )
        for rule_key in sorted(set(before_conditional) - set(after_conditional)):
            conditional_changes.append(
                {
                    "sheet_name": sheet_name,
                    "change_type": "removed",
                    "rule": before_conditional[rule_key],
                }
            )

    named_range_diff = _diff_named_ranges(before_snapshot, after_snapshot)
    summary = {
        "sheet_count_before": before_snapshot["summary"]["sheet_count"],
        "sheet_count_after": after_snapshot["summary"]["sheet_count"],
        "named_range_count_before": before_snapshot["summary"]["named_range_count"],
        "named_range_count_after": after_snapshot["summary"]["named_range_count"],
        "validation_rule_count_before": before_snapshot["summary"]["validation_rule_count"],
        "validation_rule_count_after": after_snapshot["summary"]["validation_rule_count"],
        "conditional_format_rule_count_before": before_snapshot["summary"]["conditional_format_rule_count"],
        "conditional_format_rule_count_after": after_snapshot["summary"]["conditional_format_rule_count"],
        "hidden_sheet_count_before": before_snapshot["summary"]["hidden_sheet_count"],
        "hidden_sheet_count_after": after_snapshot["summary"]["hidden_sheet_count"],
        "sheet_property_change_count": len(sheet_property_changes),
        "named_range_change_count": (
            len(named_range_diff["added"])
            + len(named_range_diff["removed"])
            + len(named_range_diff["changed"])
        ),
        "table_change_count": len(table_changes),
        "chart_change_count": len(chart_changes),
        "validation_rule_change_count": len(validation_changes),
        "conditional_format_rule_change_count": len(conditional_changes),
    }

    return {
        "summary": summary,
        "sheet_changes": {
            "added": added_sheet_names[:sample_limit],
            "removed": removed_sheet_names[:sample_limit],
            "property_changes": sheet_property_changes[:sample_limit],
        },
        "named_range_changes": {
            "added": named_range_diff["added"][:sample_limit],
            "removed": named_range_diff["removed"][:sample_limit],
            "changed": named_range_diff["changed"][:sample_limit],
        },
        "table_changes": table_changes[:sample_limit],
        "chart_changes": chart_changes[:sample_limit],
        "validation_rule_changes": validation_changes[:sample_limit],
        "conditional_format_rule_changes": conditional_changes[:sample_limit],
    }


def _diff_workbook_cell_values(
    before_wb: Any,
    after_wb: Any,
    *,
    sample_limit: int,
) -> dict[str, Any]:
    cell_changes: list[dict[str, Any]] = []
    change_count = 0

    for sheet_name in sorted(set(before_wb.sheetnames) & set(after_wb.sheetnames)):
        before_ws = before_wb[sheet_name]
        after_ws = after_wb[sheet_name]
        if _sheet_type(before_ws) == "chartsheet" or _sheet_type(after_ws) == "chartsheet":
            continue

        before_rows, before_cols, _, _ = _get_sheet_usage(before_ws)
        after_rows, after_cols, _, _ = _get_sheet_usage(after_ws)
        max_row = max(before_rows, after_rows)
        max_col = max(before_cols, after_cols)
        if max_row == 0 or max_col == 0:
            continue

        for row_index in range(1, max_row + 1):
            for column_index in range(1, max_col + 1):
                before_value = (
                    before_ws.cell(row=row_index, column=column_index).value
                    if row_index <= before_ws.max_row and column_index <= before_ws.max_column
                    else None
                )
                after_value = (
                    after_ws.cell(row=row_index, column=column_index).value
                    if row_index <= after_ws.max_row and column_index <= after_ws.max_column
                    else None
                )
                if before_value == after_value:
                    continue
                change_count += 1
                if len(cell_changes) < sample_limit:
                    cell_changes.append(
                        {
                            "sheet_name": sheet_name,
                            "cell": f"{get_column_letter(column_index)}{row_index}",
                            "before": before_value,
                            "after": after_value,
                        }
                    )

    return {
        "count": change_count,
        "sample": cell_changes,
        "truncated": change_count > sample_limit,
    }


def apply_workbook_repairs(
    filepath: str,
    repair_types: list[str] | None = None,
    sheet_names: list[str] | None = None,
    header_row: int = 1,
    sample_limit: int = 25,
    dry_run: bool = True,
) -> dict[str, Any]:
    """Execute safe workbook repairs with an audit summary and before/after structural diff."""
    try:
        if not isinstance(header_row, int) or isinstance(header_row, bool) or header_row <= 0:
            raise WorkbookError("header_row must be a positive integer")
        if not isinstance(sample_limit, int) or isinstance(sample_limit, bool) or sample_limit <= 0:
            raise WorkbookError("sample_limit must be a positive integer")
        if not isinstance(dry_run, bool):
            raise WorkbookError("dry_run must be a boolean")

        normalized_repair_types = _normalize_repair_types(repair_types)
        if not normalized_repair_types:
            normalized_repair_types = [
                "remove_broken_named_ranges",
                "remove_broken_validations",
                "remove_broken_conditional_formats",
            ]
        normalized_sheet_names = _normalize_string_list(sheet_names, label="sheet_names")
        sheet_name_filter = set(normalized_sheet_names)

        audit_before = audit_workbook(
            filepath,
            header_row=header_row,
            sample_limit=sample_limit,
        )

        with safe_workbook(filepath, save=not dry_run) as wb:
            before_snapshot = _snapshot_workbook_state(wb)
            actions: list[dict[str, Any]] = []

            if "remove_broken_named_ranges" in normalized_repair_types:
                for named_range in _serialize_named_ranges(wb):
                    destinations = named_range.get("destinations", [])
                    destination_sheets = {
                        destination["sheet_name"]
                        for destination in destinations
                    }
                    if sheet_name_filter:
                        if (
                            named_range.get("local_sheet") not in sheet_name_filter
                            and not (destination_sheets & sheet_name_filter)
                        ):
                            continue
                    if not (
                        named_range["broken_reference"]
                        or named_range["missing_sheets"]
                    ):
                        continue

                    actions.append(
                        {
                            "repair_type": "remove_broken_named_ranges",
                            "name": named_range["name"],
                            "scope_sheet": named_range.get("local_sheet"),
                            "status": "planned" if dry_run else "applied",
                        }
                    )
                    if not dry_run:
                        _remove_named_range_sources(
                            wb,
                            name=named_range["name"],
                            scope_sheet=named_range.get("local_sheet"),
                        )

            for sheet_name in wb.sheetnames:
                if sheet_name_filter and sheet_name not in sheet_name_filter:
                    continue

                ws = wb[sheet_name]

                if "reveal_hidden_sheets" in normalized_repair_types:
                    visibility = getattr(ws, "sheet_state", "visible")
                    if visibility != "visible":
                        actions.append(
                            {
                                "repair_type": "reveal_hidden_sheets",
                                "sheet_name": sheet_name,
                                "old_visibility": visibility,
                                "new_visibility": "visible",
                                "status": "planned" if dry_run else "applied",
                            }
                        )
                        if not dry_run:
                            ws.sheet_state = "visible"

                if _sheet_type(ws) == "chartsheet":
                    continue

                if "remove_broken_validations" in normalized_repair_types:
                    broken_validations = _broken_validation_rules(
                        wb,
                        ws=ws,
                        sheet_name=sheet_name,
                    )
                    if broken_validations:
                        actions.append(
                            {
                                "repair_type": "remove_broken_validations",
                                "sheet_name": sheet_name,
                                "rule_count": len(broken_validations),
                                "rules": broken_validations[:sample_limit],
                                "status": "planned" if dry_run else "applied",
                            }
                        )
                        if not dry_run:
                            _remove_validation_rules(
                                wb,
                                ws,
                                broken_only=True,
                            )

                if "remove_broken_conditional_formats" in normalized_repair_types:
                    broken_conditional_formats = _broken_conditional_format_rules(
                        wb,
                        ws=ws,
                        sheet_name=sheet_name,
                    )
                    if broken_conditional_formats:
                        actions.append(
                            {
                                "repair_type": "remove_broken_conditional_formats",
                                "sheet_name": sheet_name,
                                "rule_count": len(broken_conditional_formats),
                                "rules": broken_conditional_formats[:sample_limit],
                                "status": "planned" if dry_run else "applied",
                            }
                        )
                        if not dry_run:
                            _remove_conditional_format_rules(
                                wb,
                                ws,
                                broken_only=True,
                            )

            after_snapshot = before_snapshot if dry_run else _snapshot_workbook_state(wb)

        audit_after = audit_before if dry_run else audit_workbook(
            filepath,
            header_row=header_row,
            sample_limit=sample_limit,
        )

        return {
            "repair_types": normalized_repair_types,
            "sheet_names": normalized_sheet_names or None,
            "dry_run": dry_run,
            "action_count": len(actions),
            "actions": actions[:sample_limit],
            "truncated": len(actions) > sample_limit,
            "audit_before": audit_before["summary"],
            "audit_after": audit_after["summary"],
            "diff": _diff_workbook_snapshots(
                before_snapshot,
                after_snapshot,
                sample_limit=sample_limit,
            ),
        }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to apply workbook repairs: {e}")
        raise WorkbookError(str(e))


def diff_workbooks(
    before_filepath: str,
    after_filepath: str,
    sample_limit: int = 25,
    include_cell_changes: bool = True,
) -> dict[str, Any]:
    """Diff two workbook files and report structural changes plus sampled cell-value changes."""
    try:
        if not isinstance(sample_limit, int) or isinstance(sample_limit, bool) or sample_limit <= 0:
            raise WorkbookError("sample_limit must be a positive integer")
        if not isinstance(include_cell_changes, bool):
            raise WorkbookError("include_cell_changes must be a boolean")

        before_path = Path(before_filepath)
        after_path = Path(after_filepath)
        if not before_path.exists():
            raise WorkbookError(f"File not found: {before_filepath}")
        if not after_path.exists():
            raise WorkbookError(f"File not found: {after_filepath}")

        with safe_workbook(before_filepath) as before_wb, safe_workbook(after_filepath) as after_wb:
            before_snapshot = _snapshot_workbook_state(before_wb)
            after_snapshot = _snapshot_workbook_state(after_wb)
            diff = _diff_workbook_snapshots(
                before_snapshot,
                after_snapshot,
                sample_limit=sample_limit,
            )
            if include_cell_changes:
                diff["cell_changes"] = _diff_workbook_cell_values(
                    before_wb,
                    after_wb,
                    sample_limit=sample_limit,
                )
            else:
                diff["cell_changes"] = {
                    "count": 0,
                    "sample": [],
                    "truncated": False,
                }
            return {
                "before_file": before_path.name,
                "after_file": after_path.name,
                "sample_limit": sample_limit,
                "include_cell_changes": include_cell_changes,
                **diff,
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to diff workbooks: {e}")
        raise WorkbookError(str(e))


def explain_formula_cell(
    filepath: str,
    sheet_name: str,
    cell: str,
    max_depth: int = 3,
) -> dict[str, Any]:
    """Explain a formula cell's direct references, upstream formula chain, and downstream dependents."""
    try:
        if not isinstance(max_depth, int) or isinstance(max_depth, bool) or max_depth <= 0:
            raise WorkbookError("max_depth must be a positive integer")

        with safe_workbook(filepath) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=WorkbookError,
                operation="formula explanation",
            )
            formula_cell = ws[cell]
            if not isinstance(formula_cell.value, str) or not formula_cell.value.startswith("="):
                raise WorkbookError(f"Cell '{cell}' does not contain a formula")

            tokenizer = Tokenizer(formula_cell.value)
            direct_references: list[dict[str, Any]] = []
            seen_reference_keys: set[tuple[str, str]] = set()
            for token in tokenizer.items:
                if token.type != "OPERAND" or token.subtype != "RANGE":
                    continue

                token_value = str(token.value).strip()
                if not token_value:
                    continue
                resolved_targets = _resolve_formula_reference_targets(
                    wb,
                    token_value=token_value,
                    formula_sheet_name=sheet_name,
                    formula_row=formula_cell.row,
                )
                if not resolved_targets:
                    direct_references.append(
                        {
                            "token": token_value,
                            "reference_type": "unresolved",
                            "targets": [],
                        }
                    )
                    continue

                deduped_targets: list[dict[str, Any]] = []
                for target in resolved_targets:
                    key = (target.get("reference", ""), target.get("reference_type", ""))
                    if key in seen_reference_keys:
                        continue
                    seen_reference_keys.add(key)
                    deduped_targets.append(target)

                direct_references.append(
                    {
                        "token": token_value,
                        "reference_type": deduped_targets[0].get("reference_type"),
                        "targets": deduped_targets,
                    }
                )

            direct_formula_precedents: list[dict[str, Any]] = []
            transitive_formula_precedents: list[dict[str, Any]] = []
            seen_formula_cells: set[tuple[str, str]] = set()
            root_key = (sheet_name, formula_cell.coordinate)
            chain_children: dict[tuple[str, str], list[tuple[str, str]]] = {}
            chain_edges: list[dict[str, Any]] = []
            non_leaf_formula_cells: set[tuple[str, str]] = set()
            frontier: deque[dict[str, Any]] = deque()

            direct_precedent_entries = _collect_formula_precedent_entries(
                wb,
                formula_sheet_name=sheet_name,
                formula_cell=formula_cell,
            )
            for precedent in direct_precedent_entries:
                precedent_key = (precedent["sheet_name"], precedent["cell"])
                if precedent_key in seen_formula_cells:
                    continue
                seen_formula_cells.add(precedent_key)
                precedent_entry = dict(precedent)
                precedent_entry["depth"] = 1
                direct_formula_precedents.append(precedent_entry)
                frontier.append(precedent_entry)
                chain_children.setdefault(root_key, []).append(precedent_key)
                chain_edges.append(
                    {
                        "from_sheet_name": sheet_name,
                        "from_cell": formula_cell.coordinate,
                        "to_sheet_name": precedent["sheet_name"],
                        "to_cell": precedent["cell"],
                        "via_reference": precedent["reached_via"],
                        "depth": 1,
                    }
                )

            chain_truncated = False
            while frontier:
                precedent_entry = frontier.popleft()
                depth = int(precedent_entry["depth"])
                if depth >= max_depth:
                    if _collect_formula_precedent_entries(
                        wb,
                        formula_sheet_name=precedent_entry["sheet_name"],
                        formula_cell=wb[precedent_entry["sheet_name"]][precedent_entry["cell"]],
                    ):
                        chain_truncated = True
                        non_leaf_formula_cells.add(
                            (precedent_entry["sheet_name"], precedent_entry["cell"])
                        )
                    continue

                precedent_ws = wb[precedent_entry["sheet_name"]]
                precedent_cell = precedent_ws[precedent_entry["cell"]]
                nested_precedents = _collect_formula_precedent_entries(
                    wb,
                    formula_sheet_name=precedent_entry["sheet_name"],
                    formula_cell=precedent_cell,
                )
                parent_key = (precedent_entry["sheet_name"], precedent_entry["cell"])
                for nested_precedent in nested_precedents:
                    nested_key = (nested_precedent["sheet_name"], nested_precedent["cell"])
                    chain_edges.append(
                        {
                            "from_sheet_name": precedent_entry["sheet_name"],
                            "from_cell": precedent_entry["cell"],
                            "to_sheet_name": nested_precedent["sheet_name"],
                            "to_cell": nested_precedent["cell"],
                            "via_reference": nested_precedent["reached_via"],
                            "depth": depth + 1,
                        }
                    )
                    chain_children.setdefault(parent_key, []).append(nested_key)
                    non_leaf_formula_cells.add(parent_key)
                    if nested_key in seen_formula_cells:
                        continue
                    seen_formula_cells.add(nested_key)
                    nested_entry = {
                        "sheet_name": nested_precedent["sheet_name"],
                        "cell": nested_precedent["cell"],
                        "formula": nested_precedent["formula"],
                        "depth": depth + 1,
                        "reached_via": nested_precedent["reached_via"],
                        "parent_formula_cell": {
                            "sheet_name": precedent_entry["sheet_name"],
                            "cell": precedent_entry["cell"],
                        },
                    }
                    transitive_formula_precedents.append(nested_entry)
                    frontier.append(nested_entry)

            target_bounds = (
                formula_cell.row,
                formula_cell.column,
                formula_cell.row,
                formula_cell.column,
            )
            dependents = _extract_formula_dependencies(
                wb,
                target_sheet=sheet_name,
                target_bounds=target_bounds,
            )

            hints: list[str] = []
            if any(
                target.get("reference_type") == "named_range"
                for reference in direct_references
                for target in reference["targets"]
            ):
                hints.append("Formula depends on named ranges.")
            if any(
                target.get("reference_type") == "structured_reference"
                for reference in direct_references
                for target in reference["targets"]
            ):
                hints.append("Formula depends on Excel table structured references.")
            if any(
                target.get("broken_reference")
                for reference in direct_references
                for target in reference["targets"]
            ):
                hints.append("At least one formula reference is broken or points to a missing sheet.")
            if dependents:
                hints.append("Other formulas in the workbook depend on this cell.")
            if not hints:
                hints.append("Formula references were resolved without broken workbook links.")

            all_formula_precedents = direct_formula_precedents + transitive_formula_precedents
            max_depth_reached = max(
                (int(entry["depth"]) for entry in all_formula_precedents),
                default=0,
            )
            layer_summary: list[dict[str, Any]] = [
                {
                    "depth": 0,
                    "count": 1,
                    "sample": [
                        {
                            "sheet_name": sheet_name,
                            "cell": formula_cell.coordinate,
                        }
                    ],
                }
            ]
            for layer_depth in range(1, max_depth_reached + 1):
                layer_entries = [
                    entry for entry in all_formula_precedents if int(entry["depth"]) == layer_depth
                ]
                layer_summary.append(
                    {
                        "depth": layer_depth,
                        "count": len(layer_entries),
                        "sample": [
                            {
                                "sheet_name": entry["sheet_name"],
                                "cell": entry["cell"],
                            }
                            for entry in layer_entries[:10]
                        ],
                    }
                )

            leaf_formula_precedents = [
                {
                    "sheet_name": entry["sheet_name"],
                    "cell": entry["cell"],
                    "depth": entry["depth"],
                }
                for entry in all_formula_precedents
                if (entry["sheet_name"], entry["cell"]) not in non_leaf_formula_cells
            ]
            path_sample = _sample_formula_chain_paths(
                root_key=root_key,
                child_map=chain_children,
                path_limit=10,
            )

            return {
                "sheet_name": sheet_name,
                "cell": formula_cell.coordinate,
                "formula": formula_cell.value,
                "max_depth": max_depth,
                "direct_reference_count": len(direct_references),
                "direct_references": direct_references,
                "direct_formula_precedent_count": len(direct_formula_precedents),
                "direct_formula_precedents": direct_formula_precedents,
                "transitive_formula_precedent_count": len(transitive_formula_precedents),
                "transitive_formula_precedents": transitive_formula_precedents,
                "formula_chain": {
                    "root": {
                        "sheet_name": sheet_name,
                        "cell": formula_cell.coordinate,
                    },
                    "precedent_formula_count": len(all_formula_precedents),
                    "max_depth_requested": max_depth,
                    "max_depth_reached": max_depth_reached,
                    "truncated": chain_truncated,
                    "layer_summary": layer_summary,
                    "edge_count": len(chain_edges),
                    "edge_sample": chain_edges[:20],
                    "leaf_formula_precedent_count": len(leaf_formula_precedents),
                    "leaf_formula_precedents": leaf_formula_precedents[:10],
                    "path_sample": path_sample,
                },
                "dependent_formulas": {
                    "count": len(dependents),
                    "sample": dependents[:10],
                },
                "hints": hints,
            }
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to explain formula cell: {e}")
        raise WorkbookError(str(e))


def detect_circular_dependencies(
    filepath: str,
    sample_limit: int = 25,
) -> dict[str, Any]:
    """Detect circular workbook formula dependencies, including self-references."""
    try:
        if not isinstance(sample_limit, int) or isinstance(sample_limit, bool) or sample_limit <= 0:
            raise WorkbookError("sample_limit must be a positive integer")

        with safe_workbook(filepath) as wb:
            formula_entries = _iter_formula_cells(wb)
            formula_lookup = {
                (entry["sheet_name"], entry["cell"]): entry
                for entry in formula_entries
            }
            graph = _formula_dependency_graph(wb, formula_entries)
            components = _tarjan_strongly_connected_components(graph)

            cycles: list[dict[str, Any]] = []
            for component in components:
                sorted_component = sorted(component)
                includes_self_reference = (
                    len(sorted_component) == 1
                    and sorted_component[0] in graph.get(sorted_component[0], set())
                )
                if len(sorted_component) <= 1 and not includes_self_reference:
                    continue

                cells = [
                    {
                        "sheet_name": sheet_name,
                        "cell": cell,
                        "formula": formula_lookup[(sheet_name, cell)]["formula"],
                    }
                    for sheet_name, cell in sorted_component
                ]
                cycles.append(
                    {
                        "size": len(sorted_component),
                        "includes_self_reference": includes_self_reference,
                        "cell_refs": [f"{sheet_name}!{cell}" for sheet_name, cell in sorted_component],
                        "sheets": sorted({sheet_name for sheet_name, _ in sorted_component}),
                        "cells": cells,
                    }
                )

        cycle_count = len(cycles)
        self_referential_count = sum(1 for cycle in cycles if cycle["includes_self_reference"])
        multi_cell_count = cycle_count - self_referential_count
        dependency_edge_count = sum(len(neighbors) for neighbors in graph.values())

        result = {
            "summary": {
                "formula_cell_count": len(formula_entries),
                "dependency_edge_count": dependency_edge_count,
                "cycle_count": cycle_count,
                "self_referential_cycle_count": self_referential_count,
                "multi_cell_cycle_count": multi_cell_count,
                "has_circular_dependencies": cycle_count > 0,
            },
            "cycles": {
                "count": cycle_count,
                "sample": cycles[:sample_limit],
            },
        }
        if cycle_count == 0:
            result["hints"] = ["No circular formula dependencies detected."]
        else:
            result["hints"] = [
                "Review circular formula groups before relying on workbook calculations or downstream automation."
            ]
            if cycle_count > sample_limit:
                result["warnings"] = [
                    f"Sampled {sample_limit} of {cycle_count} circular dependency groups."
                ]

        return result
    except WorkbookError:
        raise
    except Exception as e:
        logger.error(f"Failed to detect circular dependencies: {e}")
        raise WorkbookError(str(e))

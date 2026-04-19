import logging
from copy import copy
import re
from typing import Any, Dict, Optional

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.styles import Font, Border, PatternFill, Side
from openpyxl.formula.translate import Translator, TranslatorError

from .cell_utils import parse_cell_range, validate_cell_reference
from .exceptions import SheetError, ValidationError
from .workbook import (
    _rewrite_sheet_references_in_text,
    _update_defined_name_sheet_references,
    require_worksheet,
    safe_workbook,
)

logger = logging.getLogger(__name__)

PROTECTION_OPTION_FIELDS = (
    "selectLockedCells",
    "selectUnlockedCells",
    "formatCells",
    "formatColumns",
    "formatRows",
    "insertColumns",
    "insertRows",
    "insertHyperlinks",
    "deleteColumns",
    "deleteRows",
    "sort",
    "autoFilter",
    "pivotTables",
    "objects",
    "scenarios",
)

ROW_RANGE_PATTERN = re.compile(r"^\$?(\d+):\$?(\d+)$")
COLUMN_RANGE_PATTERN = re.compile(r"^\$?([A-Za-z]+):\$?([A-Za-z]+)$")


def _should_include_changes(dry_run: bool, include_changes: Optional[bool]) -> bool:
    if include_changes is None:
        return dry_run
    return include_changes


def _attach_changes(
    payload: Dict[str, Any],
    *,
    changes: list[dict[str, Any]],
    dry_run: bool,
    include_changes: Optional[bool],
) -> Dict[str, Any]:
    if _should_include_changes(dry_run, include_changes):
        payload["changes"] = changes
    return payload


def _validate_positive_integer(value: Any, *, argument_name: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise ValidationError(f"{argument_name} must be a positive integer")
    return value


def _translated_copy_value(
    value: Any,
    *,
    source_coordinate: str,
    target_coordinate: str,
) -> Any:
    if not isinstance(value, str) or not value.startswith("="):
        return value
    try:
        return Translator(value, origin=source_coordinate).translate_formula(target_coordinate)
    except TranslatorError:
        return value


def _rewrite_sheet_reference_formula(
    formula: Any,
    *,
    old_sheet_name: str,
    new_sheet_name: str,
) -> Any:
    return _rewrite_sheet_references_in_text(
        formula,
        old_sheet_name=old_sheet_name,
        new_sheet_name=new_sheet_name,
    )


def _update_formula_container_sheet_references(
    container: Any,
    *,
    old_sheet_name: str,
    new_sheet_name: str,
) -> int:
    if container is None:
        return 0

    updated_count = 0
    for attr_name in ("numRef", "strRef", "multiLvlStrRef"):
        reference = getattr(container, attr_name, None)
        formula = getattr(reference, "f", None)
        updated_formula = _rewrite_sheet_reference_formula(
            formula,
            old_sheet_name=old_sheet_name,
            new_sheet_name=new_sheet_name,
        )
        if updated_formula == formula:
            continue
        reference.f = updated_formula
        updated_count += 1

    return updated_count


def _update_chart_sheet_references(
    chart: Any,
    *,
    old_sheet_name: str,
    new_sheet_name: str,
) -> int:
    updated_count = 0

    title = getattr(chart, "title", None)
    updated_count += _update_formula_container_sheet_references(
        getattr(title, "tx", None),
        old_sheet_name=old_sheet_name,
        new_sheet_name=new_sheet_name,
    )

    series_items = getattr(chart, "ser", None) or list(getattr(chart, "series", []))
    for series in series_items:
        updated_count += _update_formula_container_sheet_references(
            getattr(series, "tx", None),
            old_sheet_name=old_sheet_name,
            new_sheet_name=new_sheet_name,
        )
        for attr_name in ("cat", "val", "xVal", "yVal"):
            updated_count += _update_formula_container_sheet_references(
                getattr(series, attr_name, None),
                old_sheet_name=old_sheet_name,
                new_sheet_name=new_sheet_name,
            )

    return updated_count


def _update_workbook_chart_sheet_references(
    workbook: Any,
    *,
    old_sheet_name: str,
    new_sheet_name: str,
) -> int:
    updated_count = 0
    for sheet in getattr(workbook, "_sheets", []):
        for chart in getattr(sheet, "_charts", []):
            updated_count += _update_chart_sheet_references(
                chart,
                old_sheet_name=old_sheet_name,
                new_sheet_name=new_sheet_name,
            )
    return updated_count


def _copy_local_named_ranges(
    source_sheet: Worksheet,
    target_sheet: Worksheet,
) -> int:
    copied_count = 0

    for _, defined_name in source_sheet.defined_names.items():
        cloned_name = copy(defined_name)
        cloned_name.attr_text = _rewrite_sheet_references_in_text(
            getattr(defined_name, "attr_text", None),
            old_sheet_name=source_sheet.title,
            new_sheet_name=target_sheet.title,
        )
        target_sheet.defined_names.add(cloned_name)
        copied_count += 1

    return copied_count

def copy_sheet(filepath: str, source_sheet: str, target_sheet: str) -> Dict[str, Any]:
    """Copy a worksheet within the same workbook."""
    try:
        with safe_workbook(filepath, save=True) as wb:
            if target_sheet in wb.sheetnames:
                raise SheetError(f"Target sheet '{target_sheet}' already exists")

            source = require_worksheet(
                wb,
                source_sheet,
                error_cls=SheetError,
                operation="copying worksheets",
            )
            target = wb.copy_worksheet(source)
            target.title = target_sheet
            copied_named_range_count = _copy_local_named_ranges(source, target)

        return {
            "message": f"Sheet '{source_sheet}' copied to '{target_sheet}'",
            "copied_local_named_ranges": copied_named_range_count,
        }
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to copy sheet: {e}")
        raise SheetError(str(e))

def delete_sheet(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """Delete a worksheet from the workbook."""
    try:
        with safe_workbook(filepath, save=True) as wb:
            if sheet_name not in wb.sheetnames:
                raise SheetError(f"Sheet '{sheet_name}' not found")

            if len(wb.sheetnames) == 1:
                raise SheetError("Cannot delete the only sheet in workbook")

            del wb[sheet_name]
        return {"message": f"Sheet '{sheet_name}' deleted"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete sheet: {e}")
        raise SheetError(str(e))

def rename_sheet(filepath: str, old_name: str, new_name: str) -> Dict[str, Any]:
    """Rename a worksheet."""
    try:
        with safe_workbook(filepath, save=True) as wb:
            if old_name not in wb.sheetnames:
                raise SheetError(f"Sheet '{old_name}' not found")

            if new_name in wb.sheetnames:
                raise SheetError(f"Sheet '{new_name}' already exists")

            sheet = wb[old_name]
            sheet.title = new_name
            updated_named_range_count = _update_defined_name_sheet_references(
                wb,
                old_sheet_name=old_name,
                new_sheet_name=new_name,
            )
            updated_chart_reference_count = _update_workbook_chart_sheet_references(
                wb,
                old_sheet_name=old_name,
                new_sheet_name=new_name,
            )
        return {
            "message": f"Sheet renamed from '{old_name}' to '{new_name}'",
            "named_range_reference_updates": updated_named_range_count,
            "chart_reference_updates": updated_chart_reference_count,
        }
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to rename sheet: {e}")
        raise SheetError(str(e))


def set_sheet_visibility(
    filepath: str,
    sheet_name: str,
    visibility: str,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Set worksheet visibility to visible, hidden, or veryHidden."""
    valid_visibilities = {"visible", "hidden", "veryHidden"}

    try:
        if visibility not in valid_visibilities:
            raise ValidationError(
                f"Invalid worksheet visibility: {visibility}. "
                "Must be one of: visible, hidden, veryHidden"
            )

        with safe_workbook(filepath, save=not dry_run) as wb:
            if sheet_name not in wb.sheetnames:
                raise SheetError(f"Sheet '{sheet_name}' not found")

            worksheet = wb[sheet_name]
            previous_visibility = worksheet.sheet_state

            visible_sheets = [
                current_sheet_name
                for current_sheet_name in wb.sheetnames
                if getattr(wb[current_sheet_name], "sheet_state", "visible") == "visible"
            ]
            if (
                visibility != "visible"
                and previous_visibility == "visible"
                and len(visible_sheets) == 1
            ):
                raise SheetError("Cannot hide the only visible sheet in workbook")

            worksheet.sheet_state = visibility

        return _attach_changes({
            "message": (
                f"{'Previewed' if dry_run else 'Set'} worksheet visibility for "
                f"'{sheet_name}' to '{visibility}'"
            ),
            "sheet_name": sheet_name,
            "visibility": visibility,
            "dry_run": dry_run,
        }, changes=[
            {
                "type": "set_worksheet_visibility",
                "sheet_name": sheet_name,
                "old_value": previous_visibility,
                "new_value": visibility,
            }
        ], dry_run=dry_run, include_changes=include_changes)
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to set worksheet visibility: {e}")
        raise SheetError(str(e))


def _validate_dimension_value(value: Any, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValidationError(f"{label} must be a number greater than 0")
    if value <= 0:
        raise ValidationError(f"{label} must be greater than 0")
    return float(value)


def _worksheet_has_data(worksheet: Worksheet) -> bool:
    return not (
        worksheet.max_row == 1
        and worksheet.max_column == 1
        and worksheet.cell(1, 1).value is None
    )


def _display_width(value: Any) -> int:
    if value is None:
        return 0
    text = str(value)
    return max((len(line) for line in text.splitlines()), default=0)


def _serialize_protection_options(worksheet: Worksheet) -> Dict[str, bool]:
    return {
        field: bool(getattr(worksheet.protection, field))
        for field in PROTECTION_OPTION_FIELDS
    }


def _sheet_protection_state(worksheet: Worksheet) -> Dict[str, Any]:
    return {
        "enabled": bool(worksheet.protection.sheet),
        "password_protected": bool(worksheet.protection.password),
        "options": _serialize_protection_options(worksheet),
    }


def _table_row_operation_impacts(
    worksheet: Worksheet,
    *,
    start_row: int,
) -> list[dict[str, Any]]:
    impacts: list[dict[str, Any]] = []
    for table in getattr(worksheet, "tables", {}).values():
        _, min_row, _, max_row = range_boundaries(table.ref)
        if start_row > max_row:
            continue
        impacts.append(
            {
                "table_name": table.displayName,
                "range": table.ref,
                "position": "before" if start_row < min_row else "inside",
            }
        )
    return impacts


def _table_column_operation_impacts(
    worksheet: Worksheet,
    *,
    start_col: int,
) -> list[dict[str, Any]]:
    impacts: list[dict[str, Any]] = []
    for table in getattr(worksheet, "tables", {}).values():
        min_col, _, max_col, _ = range_boundaries(table.ref)
        if start_col > max_col:
            continue
        impacts.append(
            {
                "table_name": table.displayName,
                "range": table.ref,
                "position": "before" if start_col < min_col else "inside",
            }
        )
    return impacts


def _raise_table_structure_guard(
    *,
    operation: str,
    impacts: list[dict[str, Any]],
) -> None:
    if not impacts:
        return

    impacted_tables = ", ".join(
        f"{impact['table_name']} ({impact['range']})"
        for impact in impacts
    )
    raise SheetError(
        f"Cannot {operation} because it would shift native Excel table(s) without a safe "
        f"table-aware update: {impacted_tables}. Use table-specific tools or choose a "
        "position after the affected tables."
    )


def get_sheet_protection(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """Return worksheet protection state and option flags."""
    try:
        with safe_workbook(filepath) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="worksheet protection operations",
            )
            return {
                "sheet_name": sheet_name,
                **_sheet_protection_state(worksheet),
            }
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get worksheet protection: {e}")
        raise SheetError(str(e))


def set_sheet_protection(
    filepath: str,
    sheet_name: str,
    enabled: bool = True,
    password: Optional[str] = None,
    options: Optional[Dict[str, bool]] = None,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Enable or disable worksheet protection with optional capability flags."""
    try:
        if options is not None:
            if not isinstance(options, dict):
                raise ValidationError("Protection options must be an object of boolean flags")
            unknown_options = sorted(
                key for key in options.keys() if key not in PROTECTION_OPTION_FIELDS
            )
            if unknown_options:
                raise ValidationError(
                    f"Unknown protection options: {', '.join(unknown_options)}"
                )
            for key, value in options.items():
                if not isinstance(value, bool):
                    raise ValidationError(f"Protection option '{key}' must be true or false")

        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="worksheet protection operations",
            )
            previous_state = _sheet_protection_state(worksheet)

            worksheet.protection.sheet = enabled
            if enabled and password is not None:
                worksheet.protection.set_password(password)

            if options:
                for key, value in options.items():
                    setattr(worksheet.protection, key, value)

            current_state = _sheet_protection_state(worksheet)

        return _attach_changes({
            "message": (
                f"{'Previewed' if dry_run else 'Set'} worksheet protection "
                f"for '{sheet_name}' to {'enabled' if enabled else 'disabled'}"
            ),
            "sheet_name": sheet_name,
            **current_state,
            "dry_run": dry_run,
        }, changes=[
            {
                "type": "set_worksheet_protection",
                "sheet_name": sheet_name,
                "old_value": previous_state,
                "new_value": current_state,
            }
        ], dry_run=dry_run, include_changes=include_changes)
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to set worksheet protection: {e}")
        raise SheetError(str(e))


def _normalize_print_area(range_ref: str) -> str:
    if not isinstance(range_ref, str) or not range_ref.strip():
        raise ValidationError("Print area must be a cell range like A1:C10")

    cleaned_range = range_ref.strip()
    if ":" in cleaned_range:
        start_cell, end_cell = cleaned_range.split(":", maxsplit=1)
        try:
            parse_cell_range(start_cell, end_cell)
        except ValueError as e:
            raise ValidationError(f"Invalid print area: {str(e)}") from e
        return f"{start_cell.upper()}:{end_cell.upper()}"

    if not validate_cell_reference(cleaned_range):
        raise ValidationError(f"Invalid print area: {range_ref}")
    return cleaned_range.upper()


def _display_print_area(worksheet: Worksheet) -> Optional[str]:
    value = worksheet.print_area
    if not value:
        return None
    if isinstance(value, str):
        raw_value = value
    else:
        raw_value = ",".join(str(item) for item in value)

    prefixes = [
        f"'{worksheet.title}'!",
        f"{worksheet.title}!",
    ]
    for prefix in prefixes:
        raw_value = raw_value.replace(prefix, "")
    return raw_value.replace("$", "")


def _normalize_print_title_rows(rows: str) -> str:
    match = ROW_RANGE_PATTERN.match(rows.strip())
    if not match:
        raise ValidationError("Print title rows must be in the form '1:3'")
    start_row = int(match.group(1))
    end_row = int(match.group(2))
    if start_row > end_row:
        raise ValidationError("Print title rows must be in ascending order")
    return f"{start_row}:{end_row}"


def _normalize_print_title_columns(columns: str) -> str:
    match = COLUMN_RANGE_PATTERN.match(columns.strip())
    if not match:
        raise ValidationError("Print title columns must be in the form 'A:C'")
    start_column = match.group(1).upper()
    end_column = match.group(2).upper()
    if column_index_from_string(start_column) > column_index_from_string(end_column):
        raise ValidationError("Print title columns must be in ascending order")
    return f"{start_column}:{end_column}"


def _display_print_title_rows(worksheet: Worksheet) -> Optional[str]:
    value = worksheet.print_title_rows
    return value.replace("$", "") if value else None


def _display_print_title_columns(worksheet: Worksheet) -> Optional[str]:
    value = worksheet.print_title_cols
    return value.replace("$", "") if value else None


def set_print_area(
    filepath: str,
    sheet_name: str,
    range_ref: Optional[str] = None,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Set or clear worksheet print area."""
    try:
        normalized_range = (
            _normalize_print_area(range_ref) if range_ref is not None else None
        )

        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="print layout operations",
            )
            previous_area = _display_print_area(worksheet)
            worksheet.print_area = normalized_range
            current_area = _display_print_area(worksheet)

        return _attach_changes({
            "message": (
                f"{'Previewed' if dry_run else 'Set'} print area for "
                f"'{sheet_name}' to '{current_area}'"
                if current_area is not None
                else f"{'Previewed clearing' if dry_run else 'Cleared'} print area for '{sheet_name}'"
            ),
            "sheet_name": sheet_name,
            "print_area": current_area,
            "dry_run": dry_run,
        }, changes=[
            {
                "type": "set_print_area",
                "sheet_name": sheet_name,
                "old_value": previous_area,
                "new_value": current_area,
            }
        ], dry_run=dry_run, include_changes=include_changes)
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to set print area: {e}")
        raise SheetError(str(e))


def set_print_titles(
    filepath: str,
    sheet_name: str,
    rows: Optional[str] = None,
    columns: Optional[str] = None,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Set, preserve, or clear worksheet repeating title rows and columns."""
    try:
        if rows is None and columns is None:
            raise ValidationError("Provide rows, columns, or both")

        normalized_rows = (
            _normalize_print_title_rows(rows) if rows not in (None, "") else rows
        )
        normalized_columns = (
            _normalize_print_title_columns(columns)
            if columns not in (None, "")
            else columns
        )

        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="print layout operations",
            )
            previous_rows = _display_print_title_rows(worksheet)
            previous_columns = _display_print_title_columns(worksheet)

            if rows is not None:
                if normalized_rows == "":
                    worksheet._print_rows = None
                else:
                    worksheet.print_title_rows = normalized_rows

            if columns is not None:
                if normalized_columns == "":
                    worksheet._print_cols = None
                else:
                    worksheet.print_title_cols = normalized_columns

            current_rows = _display_print_title_rows(worksheet)
            current_columns = _display_print_title_columns(worksheet)

        return _attach_changes({
            "message": (
                f"{'Previewed' if dry_run else 'Set'} print titles for '{sheet_name}'"
            ),
            "sheet_name": sheet_name,
            "print_title_rows": current_rows,
            "print_title_columns": current_columns,
            "dry_run": dry_run,
        }, changes=[
            {
                "type": "set_print_titles",
                "sheet_name": sheet_name,
                "old_value": {
                    "rows": previous_rows,
                    "columns": previous_columns,
                },
                "new_value": {
                    "rows": current_rows,
                    "columns": current_columns,
                },
            }
        ], dry_run=dry_run, include_changes=include_changes)
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to set print titles: {e}")
        raise SheetError(str(e))


def set_column_widths(
    filepath: str,
    sheet_name: str,
    widths: Dict[str, float],
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Set explicit widths for one or more worksheet columns."""
    try:
        if not widths:
            raise ValidationError("At least one column width must be provided")

        normalized_widths: Dict[str, float] = {}
        for column_key, width_value in widths.items():
            column_letter = str(column_key).strip().upper()
            try:
                column_index_from_string(column_letter)
            except ValueError as e:
                raise ValidationError(f"Invalid column letter: {column_key}") from e
            normalized_widths[column_letter] = _validate_dimension_value(
                width_value, f"Width for column {column_letter}"
            )

        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="column sizing operations",
            )
            changes = []
            for column_letter, width_value in normalized_widths.items():
                old_value = worksheet.column_dimensions[column_letter].width
                worksheet.column_dimensions[column_letter].width = width_value
                changes.append(
                    {
                        "type": "set_column_width",
                        "sheet_name": sheet_name,
                        "column": column_letter,
                        "old_value": old_value,
                        "new_value": width_value,
                    }
                )

        return _attach_changes({
            "message": (
                f"{'Previewed' if dry_run else 'Set'} widths for "
                f"{len(normalized_widths)} column(s) in sheet '{sheet_name}'"
            ),
            "sheet_name": sheet_name,
            "widths": normalized_widths,
            "dry_run": dry_run,
        }, changes=changes, dry_run=dry_run, include_changes=include_changes)
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to set column widths: {e}")
        raise SheetError(str(e))


def set_row_heights(
    filepath: str,
    sheet_name: str,
    heights: Dict[str, float],
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Set explicit heights for one or more worksheet rows."""
    try:
        if not heights:
            raise ValidationError("At least one row height must be provided")

        normalized_heights: Dict[int, float] = {}
        for row_key, height_value in heights.items():
            try:
                row_number = int(str(row_key).strip())
            except ValueError as e:
                raise ValidationError(f"Invalid row number: {row_key}") from e
            if row_number < 1:
                raise ValidationError(f"Invalid row number: {row_key}")
            normalized_heights[row_number] = _validate_dimension_value(
                height_value, f"Height for row {row_number}"
            )

        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="row sizing operations",
            )
            changes = []
            for row_number, height_value in normalized_heights.items():
                old_value = worksheet.row_dimensions[row_number].height
                worksheet.row_dimensions[row_number].height = height_value
                changes.append(
                    {
                        "type": "set_row_height",
                        "sheet_name": sheet_name,
                        "row": row_number,
                        "old_value": old_value,
                        "new_value": height_value,
                    }
                )

        return _attach_changes({
            "message": (
                f"{'Previewed' if dry_run else 'Set'} heights for "
                f"{len(normalized_heights)} row(s) in sheet '{sheet_name}'"
            ),
            "sheet_name": sheet_name,
            "heights": normalized_heights,
            "dry_run": dry_run,
        }, changes=changes, dry_run=dry_run, include_changes=include_changes)
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to set row heights: {e}")
        raise SheetError(str(e))


def autofit_columns(
    filepath: str,
    sheet_name: str,
    columns: Optional[list[str]] = None,
    min_width: float = 8.43,
    max_width: Optional[float] = None,
    padding: float = 2.0,
    dry_run: bool = False,
) -> Dict[str, Any]:
    """Auto-fit worksheet columns based on the longest displayed value."""
    try:
        min_width_value = _validate_dimension_value(min_width, "Minimum width")
        padding_value = _validate_dimension_value(padding, "Padding")
        max_width_value = None
        if max_width is not None:
            max_width_value = _validate_dimension_value(max_width, "Maximum width")
            if max_width_value < min_width_value:
                raise ValidationError("Maximum width must be greater than or equal to minimum width")

        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="auto-fitting columns",
            )
            if columns:
                normalized_columns = []
                for column_key in columns:
                    column_letter = str(column_key).strip().upper()
                    try:
                        column_index_from_string(column_letter)
                    except ValueError as e:
                        raise ValidationError(f"Invalid column letter: {column_key}") from e
                    normalized_columns.append(column_letter)
            else:
                if not _worksheet_has_data(worksheet):
                    raise ValidationError("Worksheet has no populated cells to auto-fit")
                normalized_columns = [
                    get_column_letter(column_index)
                    for column_index in range(1, worksheet.max_column + 1)
                ]

            fitted_widths: Dict[str, float] = {}
            for column_letter in normalized_columns:
                column_index = column_index_from_string(column_letter)
                longest_value = 0
                for row_index in range(1, worksheet.max_row + 1):
                    longest_value = max(
                        longest_value,
                        _display_width(worksheet.cell(row=row_index, column=column_index).value),
                    )

                computed_width = max(min_width_value, float(longest_value) + padding_value)
                if max_width_value is not None:
                    computed_width = min(computed_width, max_width_value)
                fitted_widths[column_letter] = computed_width
                worksheet.column_dimensions[column_letter].width = computed_width

        return {
            "message": (
                f"{'Previewed' if dry_run else 'Auto-fit'} "
                f"{len(fitted_widths)} column(s) in sheet '{sheet_name}'"
            ),
            "sheet_name": sheet_name,
            "columns_fitted": len(fitted_widths),
            "widths": fitted_widths,
            "dry_run": dry_run,
        }
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to auto-fit columns: {e}")
        raise SheetError(str(e))

def format_range_string(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    """Format range string from row and column indices."""
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

def copy_range(
    source_ws: Worksheet,
    target_ws: Worksheet,
    source_range: str,
    target_start: Optional[str] = None,
) -> None:
    """Copy range from source worksheet to target worksheet."""
    # Parse source range
    if ':' in source_range:
        source_start, source_end = source_range.split(':')
    else:
        source_start = source_range
        source_end = None
        
    src_start_row, src_start_col, src_end_row, src_end_col = parse_cell_range(
        source_start, source_end
    )

    if src_end_row is None:
        src_end_row = src_start_row
        src_end_col = src_start_col

    if target_start is None:
        target_start = source_start

    tgt_start_row, tgt_start_col, _, _ = parse_cell_range(target_start)

    for i, row in enumerate(range(src_start_row, src_end_row + 1)):
        for j, col in enumerate(range(src_start_col, src_end_col + 1)):
            source_cell = source_ws.cell(row=row, column=col)
            target_cell = target_ws.cell(row=tgt_start_row + i, column=tgt_start_col + j)

            target_cell.value = source_cell.value

            try:
                # Copy font
                font_kwargs = {}
                if hasattr(source_cell.font, 'name'):
                    font_kwargs['name'] = source_cell.font.name
                if hasattr(source_cell.font, 'size'):
                    font_kwargs['size'] = source_cell.font.size
                if hasattr(source_cell.font, 'bold'):
                    font_kwargs['bold'] = source_cell.font.bold
                if hasattr(source_cell.font, 'italic'):
                    font_kwargs['italic'] = source_cell.font.italic
                if hasattr(source_cell.font, 'color'):
                    font_color = None
                    if source_cell.font.color:
                        font_color = source_cell.font.color.rgb
                    font_kwargs['color'] = font_color
                target_cell.font = Font(**font_kwargs)

                # Copy border
                new_border = Border()
                for side in ['left', 'right', 'top', 'bottom']:
                    source_side = getattr(source_cell.border, side)
                    if source_side and source_side.style:
                        side_color = source_side.color.rgb if source_side.color else None
                        setattr(new_border, side, Side(
                            style=source_side.style,
                            color=side_color
                        ))
                target_cell.border = new_border

                # Copy fill
                if hasattr(source_cell, 'fill'):
                    fill_kwargs = {'patternType': source_cell.fill.patternType}
                    if hasattr(source_cell.fill, 'fgColor') and source_cell.fill.fgColor:
                        fg_color = None
                        if hasattr(source_cell.fill.fgColor, 'rgb'):
                            fg_color = source_cell.fill.fgColor.rgb
                        fill_kwargs['fgColor'] = fg_color
                    if hasattr(source_cell.fill, 'bgColor') and source_cell.fill.bgColor:
                        bg_color = None
                        if hasattr(source_cell.fill.bgColor, 'rgb'):
                            bg_color = source_cell.fill.bgColor.rgb
                        fill_kwargs['bgColor'] = bg_color
                    target_cell.fill = PatternFill(**fill_kwargs)

                # Copy number format and alignment
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.alignment:
                    target_cell.alignment = source_cell.alignment

            except Exception:
                continue

def delete_range(worksheet: Worksheet, start_cell: str, end_cell: Optional[str] = None) -> None:
    """Delete contents and formatting of a range."""
    start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

    if end_row is None:
        end_row = start_row
        end_col = start_col

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = None
            cell.font = Font()
            cell.border = Border()
            cell.fill = PatternFill()
            cell.number_format = "General"
            cell.alignment = None


def _snapshot_cell_values(
    worksheet: Worksheet,
    *,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> dict[tuple[int, int], Any]:
    snapshot: dict[tuple[int, int], Any] = {}
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            snapshot[(row, col)] = worksheet.cell(row=row, column=col).value
    return snapshot


def _cell_value_changes(
    *,
    before: dict[tuple[int, int], Any],
    after: dict[tuple[int, int], Any],
    sheet_name: str,
) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for row, col in sorted(before.keys()):
        old_value = before[(row, col)]
        new_value = after[(row, col)]
        if old_value == new_value:
            continue
        changes.append(
            {
                "sheet_name": sheet_name,
                "cell": f"{get_column_letter(col)}{row}",
                "row": row,
                "column": col,
                "old_value": old_value,
                "new_value": new_value,
            }
        )
    return changes

def merge_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Merge a range of cells."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

            if end_row is None or end_col is None:
                raise SheetError("Both start and end cells must be specified for merging")

            range_string = format_range_string(start_row, start_col, end_row, end_col)
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="merging cells",
            )
            worksheet.merge_cells(range_string)
        return _attach_changes({
            "message": f"Range '{range_string}' {'would be merged' if dry_run else 'merged'} in sheet '{sheet_name}'",
            "range": range_string,
            "sheet_name": sheet_name,
            "dry_run": dry_run,
        }, changes=[{
                "type": "merge_cells",
                "sheet_name": sheet_name,
                "range": range_string,
            }], dry_run=dry_run, include_changes=include_changes)
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to merge range: {e}")
        raise SheetError(str(e))

def unmerge_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Unmerge a range of cells."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="merging cells",
            )

            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

            if end_row is None or end_col is None:
                raise SheetError("Both start and end cells must be specified for unmerging")

            range_string = format_range_string(start_row, start_col, end_row, end_col)

            # Check if range is actually merged
            merged_ranges = worksheet.merged_cells.ranges
            target_range = range_string.upper()

            if not any(str(merged_range).upper() == target_range for merged_range in merged_ranges):
                raise SheetError(f"Range '{range_string}' is not merged")

            worksheet.unmerge_cells(range_string)
        return _attach_changes({
            "message": f"Range '{range_string}' {'would be unmerged' if dry_run else 'unmerged successfully'}",
            "range": range_string,
            "sheet_name": sheet_name,
            "dry_run": dry_run,
        }, changes=[{
                "type": "unmerge_cells",
                "sheet_name": sheet_name,
                "range": range_string,
            }], dry_run=dry_run, include_changes=include_changes)
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to unmerge range: {e}")
        raise SheetError(str(e))

def get_merged_ranges(filepath: str, sheet_name: str) -> list[str]:
    """Get merged cells in a worksheet."""
    try:
        with safe_workbook(filepath) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="merged-cell inspection",
            )
            return [str(merged_range) for merged_range in worksheet.merged_cells.ranges]
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get merged cells: {e}")
        raise SheetError(str(e))


def set_freeze_panes(
    filepath: str,
    sheet_name: str,
    cell: Optional[str],
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Set or clear worksheet freeze panes."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="freeze panes",
            )
            previous_value = worksheet.freeze_panes
            previous_cell = previous_value.coordinate if hasattr(previous_value, "coordinate") else previous_value

            normalized_cell = None if cell in (None, "", "A1") else cell
            if normalized_cell is not None and not validate_cell_reference(normalized_cell):
                raise ValidationError(f"Invalid freeze pane cell reference: {normalized_cell}")

            worksheet.freeze_panes = normalized_cell

        changes = [{
            "type": "freeze_panes",
            "sheet_name": sheet_name,
            "old_value": previous_cell,
            "new_value": normalized_cell,
        }]

        if normalized_cell is None:
            message = f"{'Previewed clearing' if dry_run else 'Cleared'} freeze panes in sheet '{sheet_name}'"
        else:
            message = f"{'Previewed' if dry_run else 'Set'} freeze panes at {normalized_cell} in sheet '{sheet_name}'"

        return _attach_changes({
            "message": message,
            "sheet_name": sheet_name,
            "freeze_panes": normalized_cell,
            "dry_run": dry_run,
        }, changes=changes, dry_run=dry_run, include_changes=include_changes)
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to set freeze panes: {e}")
        raise SheetError(str(e))


def set_auto_filter(
    filepath: str,
    sheet_name: str,
    range_ref: Optional[str] = None,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Set worksheet autofilter for an explicit or inferred range."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="auto filters",
            )
            previous_ref = worksheet.auto_filter.ref

            resolved_ref = range_ref
            if not resolved_ref:
                is_empty = worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(1, 1).value is None
                if is_empty:
                    raise SheetError("Cannot infer autofilter range from an empty sheet")
                resolved_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

            try:
                start_ref, end_ref = resolved_ref.split(":", 1)
            except ValueError:
                raise ValidationError("Autofilter range must be in A1:B2 format")

            parse_cell_range(start_ref, end_ref)
            worksheet.auto_filter.ref = resolved_ref

        return _attach_changes({
            "message": f"{'Previewed' if dry_run else 'Set'} autofilter range {resolved_ref} in sheet '{sheet_name}'",
            "sheet_name": sheet_name,
            "range": resolved_ref,
            "dry_run": dry_run,
        }, changes=[{
            "type": "set_autofilter",
            "sheet_name": sheet_name,
            "old_value": previous_ref,
            "new_value": resolved_ref,
        }], dry_run=dry_run, include_changes=include_changes)
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to set autofilter: {e}")
        raise SheetError(str(e))

def copy_range_operation(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None,
    dry_run: bool = False,
) -> Dict:
    """Copy a range of cells to another location."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            source_ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="copying cell ranges",
            )
            resolved_target_sheet = target_sheet or sheet_name
            target_ws = require_worksheet(
                wb,
                resolved_target_sheet,
                error_cls=ValidationError,
                operation="copying cell ranges",
            )

            # Parse source range
            try:
                start_row, start_col, end_row, end_col = parse_cell_range(source_start, source_end)
            except ValueError as e:
                logger.error(f"Invalid source range: {e}")
                raise ValidationError(f"Invalid source range: {str(e)}")

            # Parse target starting point
            try:
                target_row = int(''.join(filter(str.isdigit, target_start)))
                target_col = column_index_from_string(''.join(filter(str.isalpha, target_start)))
            except ValueError as e:
                logger.error(f"Invalid target cell: {e}")
                raise ValidationError(f"Invalid target cell: {str(e)}")

            # Copy the range
            row_offset = target_row - start_row
            col_offset = target_col - start_col
            changes = []
            source_cells = []

            for i in range(start_row, end_row + 1):
                for j in range(start_col, end_col + 1):
                    source_cell = source_ws.cell(row=i, column=j)
                    source_cells.append(
                        {
                            "source_row": i,
                            "source_col": j,
                            "value": source_cell.value,
                            "has_style": source_cell.has_style,
                            "style": copy(source_cell._style) if source_cell.has_style else None,
                        }
                    )

            for source_entry in source_cells:
                i = source_entry["source_row"]
                j = source_entry["source_col"]
                source_value = source_entry["value"]
                source_style = source_entry["style"]
                source_has_style = bool(source_entry["has_style"])
                source_cell = source_ws.cell(row=i, column=j)
                target_cell = target_ws.cell(row=i + row_offset, column=j + col_offset)
                translated_value = _translated_copy_value(
                    source_value,
                    source_coordinate=source_cell.coordinate,
                    target_coordinate=target_cell.coordinate,
                )
                if target_cell.value != translated_value:
                    changes.append({
                        "sheet_name": resolved_target_sheet,
                        "cell": f"{get_column_letter(j + col_offset)}{i + row_offset}",
                        "row": i + row_offset,
                        "column": j + col_offset,
                        "old_value": target_cell.value,
                        "new_value": translated_value,
                        "source_cell": f"{get_column_letter(j)}{i}",
                    })
                target_cell.value = translated_value
                if source_has_style and source_style is not None:
                    target_cell._style = copy(source_style)

        return {
            "message": f"{'Previewed' if dry_run else 'Copied'} range successfully",
            "source_range": f"{source_start}:{source_end}",
            "target_sheet": resolved_target_sheet,
            "target_start": target_start,
            "dry_run": dry_run,
            "changes": changes,
        }

    except (ValidationError, SheetError):
        raise
    except Exception as e:
        logger.error(f"Failed to copy range: {e}")
        raise SheetError(f"Failed to copy range: {str(e)}")

def delete_range_operation(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    shift_direction: str = "up",
    dry_run: bool = False,
) -> Dict[str, Any]:
    """Delete a range of cells and shift remaining cells."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="deleting cell ranges",
            )

            # Validate range
            try:
                start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
                if end_row and end_row > worksheet.max_row:
                    raise SheetError(f"End row {end_row} out of bounds (1-{worksheet.max_row})")
                if end_col and end_col > worksheet.max_column:
                    raise SheetError(f"End column {end_col} out of bounds (1-{worksheet.max_column})")
            except ValueError as e:
                raise SheetError(f"Invalid range: {str(e)}")

            # Validate shift direction
            if shift_direction not in ["up", "left"]:
                raise ValidationError(f"Invalid shift direction: {shift_direction}. Must be 'up' or 'left'")

            range_string = format_range_string(
                start_row, start_col,
                end_row or start_row,
                end_col or start_col
            )
            normalized_end_row = end_row or start_row
            normalized_end_col = end_col or start_col
            range_height = normalized_end_row - start_row + 1
            range_width = normalized_end_col - start_col + 1

            # Shift cells if needed
            if shift_direction == "up":
                _raise_table_structure_guard(
                    operation=f"delete range {range_string} with upward shift",
                    impacts=_table_row_operation_impacts(worksheet, start_row=start_row),
                )
                impact_bounds = {
                    "min_row": start_row,
                    "min_col": start_col,
                    "max_row": worksheet.max_row,
                    "max_col": normalized_end_col,
                }
                before_snapshot = _snapshot_cell_values(worksheet, **impact_bounds)

                if normalized_end_row < worksheet.max_row:
                    worksheet.move_range(
                        format_range_string(
                            normalized_end_row + 1,
                            start_col,
                            worksheet.max_row,
                            normalized_end_col,
                        ),
                        rows=-range_height,
                        cols=0,
                    )

                vacated_start_row = max(worksheet.max_row - range_height + 1, start_row)
                delete_range(
                    worksheet,
                    f"{get_column_letter(start_col)}{vacated_start_row}",
                    f"{get_column_letter(normalized_end_col)}{worksheet.max_row}",
                )
                after_snapshot = _snapshot_cell_values(worksheet, **impact_bounds)
            elif shift_direction == "left":
                _raise_table_structure_guard(
                    operation=f"delete range {range_string} with left shift",
                    impacts=_table_column_operation_impacts(worksheet, start_col=start_col),
                )
                impact_bounds = {
                    "min_row": start_row,
                    "min_col": start_col,
                    "max_row": normalized_end_row,
                    "max_col": worksheet.max_column,
                }
                before_snapshot = _snapshot_cell_values(worksheet, **impact_bounds)

                if normalized_end_col < worksheet.max_column:
                    worksheet.move_range(
                        format_range_string(
                            start_row,
                            normalized_end_col + 1,
                            normalized_end_row,
                            worksheet.max_column,
                        ),
                        rows=0,
                        cols=-range_width,
                    )

                vacated_start_col = max(worksheet.max_column - range_width + 1, start_col)
                delete_range(
                    worksheet,
                    f"{get_column_letter(vacated_start_col)}{start_row}",
                    f"{get_column_letter(worksheet.max_column)}{normalized_end_row}",
                )
                after_snapshot = _snapshot_cell_values(worksheet, **impact_bounds)

            changes = _cell_value_changes(
                before=before_snapshot,
                after=after_snapshot,
                sheet_name=sheet_name,
            )

        return {
            "message": f"{'Previewed' if dry_run else 'Deleted'} range {range_string} successfully",
            "range": range_string,
            "shift_direction": shift_direction,
            "dry_run": dry_run,
            "changes": changes,
        }
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete range: {e}")
        raise SheetError(str(e))

def insert_row(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1,
    dry_run: bool = False,
) -> Dict[str, Any]:
    """Insert one or more rows starting at the specified row."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="inserting rows",
            )

            # Validate parameters
            _validate_positive_integer(start_row, argument_name="start_row")
            _validate_positive_integer(count, argument_name="count")

            _raise_table_structure_guard(
                operation=f"insert rows starting at row {start_row}",
                impacts=_table_row_operation_impacts(worksheet, start_row=start_row),
            )
            worksheet.insert_rows(start_row, count)

        return {
            "message": f"{'Previewed' if dry_run else 'Inserted'} {count} row(s) starting at row {start_row} in sheet '{sheet_name}'",
            "sheet_name": sheet_name,
            "start_row": start_row,
            "count": count,
            "dry_run": dry_run,
            "changes": [{
                "type": "insert_rows",
                "sheet_name": sheet_name,
                "start_row": start_row,
                "count": count,
            }],
        }
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to insert rows: {e}")
        raise SheetError(str(e))

def insert_cols(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1,
    dry_run: bool = False,
) -> Dict[str, Any]:
    """Insert one or more columns starting at the specified column."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="inserting columns",
            )

            # Validate parameters
            _validate_positive_integer(start_col, argument_name="start_col")
            _validate_positive_integer(count, argument_name="count")

            _raise_table_structure_guard(
                operation=f"insert columns starting at column {start_col}",
                impacts=_table_column_operation_impacts(worksheet, start_col=start_col),
            )
            worksheet.insert_cols(start_col, count)

        return {
            "message": f"{'Previewed' if dry_run else 'Inserted'} {count} column(s) starting at column {start_col} in sheet '{sheet_name}'",
            "sheet_name": sheet_name,
            "start_col": start_col,
            "count": count,
            "dry_run": dry_run,
            "changes": [{
                "type": "insert_columns",
                "sheet_name": sheet_name,
                "start_col": start_col,
                "count": count,
            }],
        }
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to insert columns: {e}")
        raise SheetError(str(e))

def delete_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1,
    dry_run: bool = False,
) -> Dict[str, Any]:
    """Delete one or more rows starting at the specified row."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="deleting rows",
            )

            # Validate parameters
            _validate_positive_integer(start_row, argument_name="start_row")
            _validate_positive_integer(count, argument_name="count")
            if start_row > worksheet.max_row:
                raise ValidationError(f"Start row {start_row} exceeds worksheet bounds (max row: {worksheet.max_row})")

            _raise_table_structure_guard(
                operation=f"delete rows starting at row {start_row}",
                impacts=_table_row_operation_impacts(worksheet, start_row=start_row),
            )
            worksheet.delete_rows(start_row, count)

        return {
            "message": f"{'Previewed' if dry_run else 'Deleted'} {count} row(s) starting at row {start_row} in sheet '{sheet_name}'",
            "sheet_name": sheet_name,
            "start_row": start_row,
            "count": count,
            "dry_run": dry_run,
            "changes": [{
                "type": "delete_rows",
                "sheet_name": sheet_name,
                "start_row": start_row,
                "count": count,
            }],
        }
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete rows: {e}")
        raise SheetError(str(e))

def delete_cols(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1,
    dry_run: bool = False,
) -> Dict[str, Any]:
    """Delete one or more columns starting at the specified column."""
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=SheetError,
                operation="deleting columns",
            )

            # Validate parameters
            _validate_positive_integer(start_col, argument_name="start_col")
            _validate_positive_integer(count, argument_name="count")
            if start_col > worksheet.max_column:
                raise ValidationError(f"Start column {start_col} exceeds worksheet bounds (max column: {worksheet.max_column})")

            _raise_table_structure_guard(
                operation=f"delete columns starting at column {start_col}",
                impacts=_table_column_operation_impacts(worksheet, start_col=start_col),
            )
            worksheet.delete_cols(start_col, count)

        return {
            "message": f"{'Previewed' if dry_run else 'Deleted'} {count} column(s) starting at column {start_col} in sheet '{sheet_name}'",
            "sheet_name": sheet_name,
            "start_col": start_col,
            "count": count,
            "dry_run": dry_run,
            "changes": [{
                "type": "delete_columns",
                "sheet_name": sheet_name,
                "start_col": start_col,
                "count": count,
            }],
        }
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete columns: {e}")
        raise SheetError(str(e))

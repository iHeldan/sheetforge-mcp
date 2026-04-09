from enum import Enum
import logging
from typing import Any, Dict, List, Optional

from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.units import EMU_to_cm

from .cell_utils import parse_cell_range
from .exceptions import ValidationError, ChartError
from .workbook import safe_workbook

logger = logging.getLogger(__name__)
DEFAULT_CHART_WIDTH = 15.0
DEFAULT_CHART_HEIGHT = 7.5

class ChartType(str, Enum):
    """Supported chart types"""
    LINE = "line"
    BAR = "bar"
    PIE = "pie"
    SCATTER = "scatter"
    AREA = "area"

class ChartStyle:
    """Chart style configuration"""
    def __init__(
        self,
        title_size: int = 14,
        title_bold: bool = True,
        axis_label_size: int = 12,
        show_legend: bool = True,
        legend_position: str = "r",
        show_data_labels: bool = True,
        grid_lines: bool = False,
        style_id: int = 2
    ):
        self.title_size = title_size
        self.title_bold = title_bold
        self.axis_label_size = axis_label_size
        self.show_legend = show_legend
        self.legend_position = legend_position
        self.show_data_labels = show_data_labels
        self.grid_lines = grid_lines
        self.style_id = style_id


def _extract_text_runs(rich_text: Any) -> Optional[str]:
    parts: list[str] = []
    for paragraph in getattr(rich_text, "p", []) or []:
        for run in getattr(paragraph, "r", []) or []:
            text = getattr(run, "t", None)
            if text:
                parts.append(text)
        for field in getattr(paragraph, "fld", []) or []:
            text = getattr(field, "t", None)
            if text:
                parts.append(text)
    return "".join(parts) or None


def _extract_title_text(title: Any) -> Optional[str]:
    if title is None:
        return None
    if isinstance(title, str):
        return title or None

    tx = getattr(title, "tx", None)
    if tx is not None:
        rich_text = getattr(tx, "rich", None)
        if rich_text is not None:
            extracted = _extract_text_runs(rich_text)
            if extracted:
                return extracted

        str_ref = getattr(tx, "strRef", None)
        if str_ref is not None and getattr(str_ref, "f", None):
            return str_ref.f

    str_ref = getattr(title, "strRef", None)
    if str_ref is not None and getattr(str_ref, "f", None):
        return str_ref.f

    value = getattr(title, "v", None)
    if value is not None:
        value_text = str(value).strip()
        if not value_text or value_text == "None":
            return None
        return value_text

    return None


def _extract_reference_formula(data_source: Any) -> Optional[str]:
    if data_source is None:
        return None

    for attr_name in ("numRef", "strRef", "multiLvlStrRef"):
        reference = getattr(data_source, attr_name, None)
        if reference is not None and getattr(reference, "f", None):
            return reference.f

    return None


def _extract_chart_anchor(chart: Any) -> Optional[str]:
    anchor = getattr(chart, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"


def _extract_series_metadata(series: Any) -> Dict[str, Any]:
    metadata: Dict[str, Any] = {}

    title = _extract_title_text(getattr(series, "tx", None))
    if title:
        metadata["title"] = title

    categories = _extract_reference_formula(getattr(series, "cat", None))
    if categories:
        metadata["categories"] = categories

    x_values = _extract_reference_formula(getattr(series, "xVal", None))
    if x_values:
        metadata["x_values"] = x_values

    values = _extract_reference_formula(getattr(series, "val", None))
    if values:
        metadata["values"] = values

    y_values = _extract_reference_formula(getattr(series, "yVal", None))
    if y_values:
        metadata["y_values"] = y_values

    return metadata


def _extract_chart_dimensions(chart: Any) -> tuple[Optional[float], Optional[float]]:
    anchor = getattr(chart, "anchor", None)
    ext = getattr(anchor, "ext", None)
    if ext is not None and getattr(ext, "cx", None) is not None and getattr(ext, "cy", None) is not None:
        return EMU_to_cm(ext.cx), EMU_to_cm(ext.cy)
    return getattr(chart, "width", None), getattr(chart, "height", None)


def _chart_type_name(chart: Any) -> str:
    class_name = type(chart).__name__
    if class_name.endswith("Chart"):
        return class_name.removesuffix("Chart").lower()
    return class_name.lower()


def list_charts(
    filepath: str,
    sheet_name: Optional[str] = None,
) -> list[dict[str, Any]]:
    """List embedded charts for one worksheet or the whole workbook."""
    try:
        with safe_workbook(filepath) as wb:
            if sheet_name is not None and sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            sheet_names = [sheet_name] if sheet_name is not None else list(wb.sheetnames)
            charts: list[dict[str, Any]] = []

            for current_sheet_name in sheet_names:
                worksheet = wb[current_sheet_name]
                for chart_index, chart in enumerate(getattr(worksheet, "_charts", []), start=1):
                    series = getattr(chart, "ser", None) or list(getattr(chart, "series", []))
                    width, height = _extract_chart_dimensions(chart)
                    chart_info = {
                        "sheet_name": current_sheet_name,
                        "chart_index": chart_index,
                        "chart_type": _chart_type_name(chart),
                        "anchor": _extract_chart_anchor(chart),
                        "title": _extract_title_text(getattr(chart, "title", None)),
                        "x_axis_title": _extract_title_text(
                            getattr(getattr(chart, "x_axis", None), "title", None)
                        ),
                        "y_axis_title": _extract_title_text(
                            getattr(getattr(chart, "y_axis", None), "title", None)
                        ),
                        "legend_position": getattr(getattr(chart, "legend", None), "position", None),
                        "style": getattr(chart, "style", None),
                        "width": width,
                        "height": height,
                        "series": [_extract_series_metadata(item) for item in series],
                    }
                    charts.append({key: value for key, value in chart_info.items() if value is not None})

            return charts

    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Failed to list charts: {e}")
        raise ChartError(str(e))


def _validate_target_cell(target_cell: str) -> None:
    if not target_cell:
        raise ValidationError("Invalid target cell format: target cell is required")

    column_part = "".join(character for character in target_cell if character.isalpha())
    row_part = "".join(character for character in target_cell if character.isdigit())
    if not column_part or not row_part:
        raise ValidationError(f"Invalid target cell format: {target_cell}")

    try:
        column_index_from_string(column_part)
        row_index = int(row_part)
    except ValueError as e:
        raise ValidationError(f"Invalid target cell: {str(e)}") from e

    if row_index < 1:
        raise ValidationError(f"Invalid target cell: {target_cell}")


def _normalize_style(style: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    normalized_style = dict(style or {})
    normalized_style.setdefault("show_data_labels", True)
    return normalized_style


def _resolve_chart_class(chart_type: str) -> tuple[str, Any]:
    chart_classes = {
        "line": LineChart,
        "bar": BarChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart,
    }

    chart_type_lower = chart_type.lower()
    chart_class = chart_classes.get(chart_type_lower)
    if not chart_class:
        raise ValidationError(
            f"Unsupported chart type: {chart_type}. "
            f"Supported types: {', '.join(chart_classes.keys())}"
        )
    return chart_type_lower, chart_class


def _build_chart(
    chart_type: str,
    *,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
) -> Any:
    _, chart_class = _resolve_chart_class(chart_type)
    chart = chart_class()
    if title:
        chart.title = title
    if hasattr(chart, "x_axis") and x_axis:
        chart.x_axis.title = x_axis
    if hasattr(chart, "y_axis") and y_axis:
        chart.y_axis.title = y_axis
    return chart


def _resolve_chart_dimensions(
    style: Optional[Dict[str, Any]],
    width: Optional[float],
    height: Optional[float],
) -> tuple[float, float]:
    style = style or {}

    raw_width = width if width is not None else style.get("width")
    raw_height = height if height is not None else style.get("height")

    def _coerce_dimension(raw_value: Any, name: str, default: float) -> float:
        if raw_value is None:
            return default
        try:
            value = float(raw_value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(f"{name} must be a positive number") from exc
        if value <= 0:
            raise ValidationError(f"{name} must be a positive number")
        return value

    return (
        _coerce_dimension(raw_width, "width", DEFAULT_CHART_WIDTH),
        _coerce_dimension(raw_height, "height", DEFAULT_CHART_HEIGHT),
    )


def _resolve_range_source(
    workbook: Any,
    default_worksheet: Any,
    range_ref: str,
) -> tuple[Any, int, int, int, int]:
    if not range_ref:
        raise ValidationError("Range reference is required")

    if "!" in range_ref:
        range_sheet_name, cell_range = range_ref.rsplit("!", 1)
        range_sheet_name = range_sheet_name.strip("'")
        if range_sheet_name not in workbook.sheetnames:
            raise ValidationError(f"Sheet '{range_sheet_name}' referenced in range not found")
        source_worksheet = workbook[range_sheet_name]
    else:
        source_worksheet = default_worksheet
        cell_range = range_ref

    if ":" not in cell_range:
        raise ValidationError(f"Invalid data range format: {range_ref}")

    try:
        start_cell, end_cell = cell_range.split(":")
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
    except ValueError as e:
        raise ValidationError(f"Invalid data range format: {str(e)}") from e

    return source_worksheet, start_row, start_col, end_row, end_col


def _reference_from_range(
    workbook: Any,
    default_worksheet: Any,
    range_ref: str,
) -> Reference:
    source_worksheet, start_row, start_col, end_row, end_col = _resolve_range_source(
        workbook,
        default_worksheet,
        range_ref,
    )
    return Reference(
        source_worksheet,
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    )


def _apply_chart_style(chart: Any, style: Dict[str, Any]) -> None:
    try:
        if style.get("show_legend", True):
            chart.legend = Legend()
            chart.legend.position = style.get("legend_position", "r")
        else:
            chart.legend = None

        if style.get("show_data_labels", False):
            data_labels = DataLabelList()
            data_label_options = style.get("data_label_options", {})
            if not isinstance(data_label_options, dict):
                data_label_options = {}

            def _opt(name: str, default: bool) -> bool:
                return bool(data_label_options.get(name, default))

            data_labels.showVal = _opt("show_val", True)
            data_labels.showCatName = _opt("show_cat_name", False)
            data_labels.showSerName = _opt("show_ser_name", False)
            data_labels.showLegendKey = _opt("show_legend_key", False)
            data_labels.showPercent = _opt("show_percent", False)
            data_labels.showBubbleSize = _opt("show_bubble_size", False)
            chart.dataLabels = data_labels

        if style.get("grid_lines", False):
            if hasattr(chart, "x_axis"):
                chart.x_axis.majorGridlines = ChartLines()
            if hasattr(chart, "y_axis"):
                chart.y_axis.majorGridlines = ChartLines()
    except Exception as e:
        logger.error(f"Failed to apply chart style: {e}")
        raise ChartError(f"Failed to apply chart style: {str(e)}")


def _finalize_chart(
    worksheet: Any,
    chart: Any,
    target_cell: str,
    *,
    width: float = DEFAULT_CHART_WIDTH,
    height: float = DEFAULT_CHART_HEIGHT,
) -> None:
    try:
        _validate_target_cell(target_cell)
        chart.width = width
        chart.height = height
        worksheet.add_chart(chart, target_cell)
    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Failed to create chart drawing: {e}")
        raise ChartError(f"Failed to create chart drawing: {str(e)}")


def create_chart_in_sheet(
    filepath: str,
    sheet_name: str,
    data_range: Optional[str],
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
    style: Optional[Dict] = None,
    series: Optional[List[Dict[str, Any]]] = None,
    categories_range: Optional[str] = None,
    width: Optional[float] = None,
    height: Optional[float] = None,
) -> dict[str, Any]:
    """Create chart in sheet with either a contiguous data range or explicit series."""
    style = _normalize_style(style)
    resolved_width, resolved_height = _resolve_chart_dimensions(style, width, height)
    if data_range and series:
        raise ValidationError("Provide either data_range or series, not both")
    if not data_range and not series:
        raise ValidationError("Either data_range or series is required")
    if series is not None:
        return create_chart_from_series(
            filepath=filepath,
            sheet_name=sheet_name,
            chart_type=chart_type,
            target_cell=target_cell,
            series=series,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis,
            categories_range=categories_range,
            style=style,
            width=resolved_width,
            height=resolved_height,
        )

    try:
        with safe_workbook(filepath, save=True) as wb:
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found")
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            worksheet = wb[sheet_name]
            source_worksheet, start_row, start_col, end_row, end_col = _resolve_range_source(
                wb,
                worksheet,
                data_range,
            )
            chart_type_lower, _ = _resolve_chart_class(chart_type)
            chart = _build_chart(chart_type_lower, title=title, x_axis=x_axis, y_axis=y_axis)

            try:
                if chart_type_lower == "scatter":
                    for col in range(start_col + 1, end_col + 1):
                        series_title = source_worksheet.cell(row=start_row, column=col).value
                        x_values = Reference(
                            source_worksheet,
                            min_row=start_row + 1,
                            max_row=end_row,
                            min_col=start_col,
                        )
                        y_values = Reference(
                            source_worksheet,
                            min_row=start_row + 1,
                            max_row=end_row,
                            min_col=col,
                        )
                        series = Series(y_values, x_values, title=series_title)
                        chart.series.append(series)
                else:
                    data = Reference(
                        source_worksheet,
                        min_row=start_row,
                        max_row=end_row,
                        min_col=start_col + 1,
                        max_col=end_col,
                    )
                    cats = Reference(
                        source_worksheet,
                        min_row=start_row + 1,
                        max_row=end_row,
                        min_col=start_col,
                    )
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
            except Exception as e:
                logger.error(f"Failed to create chart data references: {e}")
                raise ChartError(f"Failed to create chart data references: {str(e)}")

            _apply_chart_style(chart, style)
            _finalize_chart(
                worksheet,
                chart,
                target_cell,
                width=resolved_width,
                height=resolved_height,
            )

        return {
            "message": f"{chart_type.capitalize()} chart created successfully",
            "details": {
                "type": chart_type,
                "location": target_cell,
                "data_range": data_range,
                "width": resolved_width,
                "height": resolved_height,
            }
        }

    except (ValidationError, ChartError):
        raise
    except Exception as e:
        logger.error(f"Unexpected error creating chart: {e}")
        raise ChartError(f"Unexpected error creating chart: {str(e)}")


def create_chart_from_series(
    filepath: str,
    sheet_name: str,
    chart_type: str,
    target_cell: str,
    series: List[Dict[str, Any]],
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
    categories_range: Optional[str] = None,
    style: Optional[Dict[str, Any]] = None,
    width: Optional[float] = None,
    height: Optional[float] = None,
) -> dict[str, Any]:
    """Create a chart from explicit series definitions."""
    normalized_style = _normalize_style(style)
    resolved_width, resolved_height = _resolve_chart_dimensions(normalized_style, width, height)
    if not isinstance(series, list) or not series:
        raise ValidationError("At least one series definition is required")

    try:
        with safe_workbook(filepath, save=True) as wb:
            if sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            worksheet = wb[sheet_name]
            chart_type_lower, _ = _resolve_chart_class(chart_type)
            chart = _build_chart(chart_type_lower, title=title, x_axis=x_axis, y_axis=y_axis)

            if chart_type_lower == "scatter" and categories_range is not None:
                raise ValidationError("categories_range is not supported for scatter charts")
            if chart_type_lower == "pie" and len(series) != 1:
                raise ValidationError("Pie charts require exactly one series definition")

            shared_categories = None
            if chart_type_lower != "scatter" and categories_range is not None:
                shared_categories = _reference_from_range(wb, worksheet, categories_range)

            try:
                for index, series_definition in enumerate(series, start=1):
                    if not isinstance(series_definition, dict):
                        raise ValidationError("Each series definition must be an object")

                    series_title = series_definition.get("title")
                    if chart_type_lower == "scatter":
                        x_range = series_definition.get("x_range")
                        y_range = series_definition.get("y_range")
                        if not x_range or not y_range:
                            raise ValidationError(
                                f"Scatter series {index} requires both x_range and y_range"
                            )

                        x_values = _reference_from_range(wb, worksheet, x_range)
                        y_values = _reference_from_range(wb, worksheet, y_range)
                        chart.series.append(Series(y_values, x_values, title=series_title))
                        continue

                    values_range = series_definition.get("values_range")
                    if not values_range:
                        raise ValidationError(
                            f"Series {index} requires values_range for {chart_type_lower} charts"
                        )

                    values = _reference_from_range(wb, worksheet, values_range)
                    chart.series.append(Series(values, title=series_title))

                if shared_categories is not None:
                    chart.set_categories(shared_categories)
            except ValidationError:
                raise
            except Exception as e:
                logger.error(f"Failed to create chart series: {e}")
                raise ChartError(f"Failed to create chart series: {str(e)}")

            _apply_chart_style(chart, normalized_style)
            _finalize_chart(
                worksheet,
                chart,
                target_cell,
                width=resolved_width,
                height=resolved_height,
            )

        return {
            "message": f"{chart_type.capitalize()} chart created successfully",
            "details": {
                "type": chart_type,
                "location": target_cell,
                "series_count": len(series),
                "categories_range": categories_range,
                "width": resolved_width,
                "height": resolved_height,
            },
        }

    except (ValidationError, ChartError):
        raise
    except Exception as e:
        logger.error(f"Unexpected error creating chart from series: {e}")
        raise ChartError(f"Unexpected error creating chart from series: {str(e)}")

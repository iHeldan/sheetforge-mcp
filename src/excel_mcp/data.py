import base64
from datetime import date, datetime, time, timezone
from decimal import Decimal
import hashlib
import json
from pathlib import Path
import logging
import re
import unicodedata
from typing import Any, Dict, List, Optional

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from .exceptions import DataError, PreconditionFailedError
from .cell_utils import parse_cell_range
from .cell_validation import get_data_validation_for_cell
from .workbook import first_worksheet, require_worksheet, safe_workbook

logger = logging.getLogger(__name__)
ROW_MODES = {"arrays", "objects"}
RANGE_READ_CURSOR_VERSION = 2
DEFAULT_DATASET_SAMPLE_ROWS = 25
KEY_CANDIDATE_SCAN_LIMIT = 100
TABULAR_BLANK_GAP_TOLERANCE = 5
DATASET_TOKEN_VERSION = 1
CONTENT_TOKEN_QUANTILES = (0.25, 0.5, 0.75)


def _cell_address(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _range_string(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    return f"{_cell_address(start_row, start_col)}:{_cell_address(end_row, end_col)}"


def _compact_table_payload(table_data: Dict[str, Any]) -> Dict[str, Any]:
    compact_data = {
        "headers": table_data["headers"],
        "rows": table_data["rows"],
    }
    if table_data["truncated"]:
        compact_data["total_rows"] = table_data["total_rows"]
        compact_data["truncated"] = True
    return compact_data


def _canonical_json(data: Any) -> str:
    return json.dumps(
        data,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
    )


def _token_digest(prefix: str, payload: Dict[str, Any]) -> str:
    digest = hashlib.sha256(_canonical_json(payload).encode("utf-8")).hexdigest()[:20]
    return f"{prefix}_v{DATASET_TOKEN_VERSION}_{digest}"


def _snapshot_metadata(filepath: str) -> Dict[str, Any]:
    stat = Path(filepath).stat()
    return {
        "calculated_at": datetime.now(timezone.utc).isoformat(),
        "file_mtime": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
        "file_size": stat.st_size,
        "token_basis": "live_workbook_snapshot",
    }


def _normalize_decimal_string(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return str(normalized.quantize(Decimal("1")))
    return format(normalized, "f")


def _normalize_token_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return {"type": "boolean", "value": value}
    if isinstance(value, int):
        return {"type": "integer", "value": str(value)}
    if isinstance(value, (float, Decimal)):
        return {
            "type": "number",
            "value": _normalize_decimal_string(Decimal(str(value))),
        }
    if isinstance(value, datetime):
        return {"type": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"type": "date", "value": value.isoformat()}
    if isinstance(value, time):
        return {"type": "time", "value": value.isoformat()}
    if isinstance(value, str):
        normalized = value.strip()
        return {
            "type": "formula" if normalized.startswith("=") else "string",
            "value": normalized,
        }
    return {
        "type": type(value).__name__.lower(),
        "value": str(value).strip(),
    }


def _dataset_anchor_indexes(row_count: int) -> List[int]:
    if row_count <= 0:
        return []

    indexes = {0, row_count - 1}
    if row_count > 1:
        indexes.add(1)
        indexes.add(row_count - 2)

    for quantile in CONTENT_TOKEN_QUANTILES:
        indexes.add(int(round((row_count - 1) * quantile)))

    return sorted(index for index in indexes if 0 <= index < row_count)


def _token_header_columns(ws: Worksheet, header_row: int) -> List[int]:
    non_empty_header_columns = [
        col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=col).value not in (None, "")
    ]
    if not non_empty_header_columns:
        return _selected_columns(1, ws.max_column)
    return _selected_columns(non_empty_header_columns[0], non_empty_header_columns[-1])


def _read_rows_for_columns(
    ws: Worksheet,
    columns: List[int],
    *,
    first_row: int,
    last_row: int,
    limit: Optional[int] = None,
) -> List[List[Any]]:
    if first_row > last_row:
        return []

    rows: List[List[Any]] = []
    remaining = None if limit is None else limit
    for row_idx in range(first_row, last_row + 1):
        if remaining is not None and remaining <= 0:
            break
        rows.append([ws.cell(row=row_idx, column=col).value for col in columns])
        if remaining is not None:
            remaining -= 1
    return rows


def _worksheet_structure_summary(
    ws: Worksheet,
    *,
    sheet_name: str,
    header_row: int,
) -> Dict[str, Any]:
    columns = _token_header_columns(ws, header_row)
    headers = [ws.cell(row=header_row, column=col).value for col in columns]
    data_extent = _detect_tabular_data_extent(ws, header_row, columns)
    return {
        "target_kind": "worksheet",
        "sheet_name": sheet_name,
        "header_row": header_row,
        "headers": headers,
        "column_count": len(headers),
        "first_data_row": data_extent["first_data_row"],
        "data_end_row": data_extent["last_data_row"],
        "data_row_count": max(data_extent["last_data_row"] - header_row, 0),
        "ignored_trailing_row_count": data_extent["ignored_trailing_row_count"],
        "header_columns": columns,
    }


def _worksheet_dataset_tokens(
    ws: Worksheet,
    *,
    sheet_name: str,
    header_row: int,
) -> Dict[str, Any]:
    summary = _worksheet_structure_summary(ws, sheet_name=sheet_name, header_row=header_row)
    headers = summary["headers"]
    columns = summary["header_columns"]
    first_data_row = summary["first_data_row"]
    data_end_row = summary["data_end_row"]
    data_row_count = summary["data_row_count"]

    structure_payload = {
        "target_kind": summary["target_kind"],
        "sheet_name": sheet_name,
        "header_row": header_row,
        "headers": headers,
        "column_count": summary["column_count"],
        "first_data_row": first_data_row,
        "data_end_row": data_end_row,
        "ignored_trailing_row_count": summary["ignored_trailing_row_count"],
    }

    key_field = None
    anchor_keys: List[Any] = []
    if first_data_row is not None and data_end_row >= first_data_row:
        candidate_rows = _read_rows_for_columns(
            ws,
            columns,
            first_row=first_data_row,
            last_row=data_end_row,
            limit=KEY_CANDIDATE_SCAN_LIMIT,
        )
        if candidate_rows:
            schema = _build_schema(headers, candidate_rows)
            key_candidates = _infer_key_candidates(headers, candidate_rows, schema)
            for candidate in key_candidates:
                if candidate["confidence"] == "high":
                    key_field = candidate["field"]
                    break
            if key_field is not None:
                key_index = next(
                    (
                        index
                        for index, column in enumerate(schema)
                        if column["field"] == key_field
                    ),
                    None,
                )
                if key_index is not None:
                    for anchor_index in _dataset_anchor_indexes(data_row_count):
                        row_index = first_data_row + anchor_index
                        anchor_keys.append(
                            _normalize_token_value(
                                ws.cell(row=row_index, column=columns[key_index]).value
                            )
                        )

    anchor_rows: List[Dict[str, Any]] = []
    if first_data_row is not None and data_end_row >= first_data_row:
        for anchor_index in _dataset_anchor_indexes(data_row_count):
            row_index = first_data_row + anchor_index
            anchor_rows.append(
                {
                    "row": row_index,
                    "values": [
                        _normalize_token_value(ws.cell(row=row_index, column=col).value)
                        for col in columns
                    ],
                }
            )

    content_payload = {
        "target_kind": summary["target_kind"],
        "sheet_name": sheet_name,
        "header_row": header_row,
        "data_row_count": data_row_count,
        "data_end_row": data_end_row,
        "anchor_rows": anchor_rows,
        "key_field": key_field,
        "anchor_keys": anchor_keys,
    }

    return {
        "structure_token": _token_digest("sf_struct", structure_payload),
        "content_token": _token_digest("sf_content", content_payload),
        "structure_summary": {
            key: value
            for key, value in summary.items()
            if key != "header_columns"
        },
    }


def _table_dataset_tokens(
    ws: Worksheet,
    *,
    sheet_name: str,
    table_name: str,
    table_range: str,
    headers: List[Any],
    header_row_count: int,
    totals_row_count: int,
    totals_row_shown: bool,
) -> Dict[str, Any]:
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    data_start_row = min_row + header_row_count
    data_end_row = max_row - totals_row_count
    data_row_count = max(data_end_row - data_start_row + 1, 0)
    columns = _selected_columns(min_col, max_col)

    key_field = None
    anchor_keys: List[Any] = []
    if data_row_count > 0:
        candidate_rows = _read_rows_for_columns(
            ws,
            columns,
            first_row=data_start_row,
            last_row=data_end_row,
            limit=KEY_CANDIDATE_SCAN_LIMIT,
        )
        if candidate_rows:
            schema = _build_schema(headers, candidate_rows)
            key_candidates = _infer_key_candidates(headers, candidate_rows, schema)
            for candidate in key_candidates:
                if candidate["confidence"] == "high":
                    key_field = candidate["field"]
                    break
            if key_field is not None:
                key_index = next(
                    (
                        index
                        for index, column in enumerate(schema)
                        if column["field"] == key_field
                    ),
                    None,
                )
                if key_index is not None:
                    for anchor_index in _dataset_anchor_indexes(data_row_count):
                        row_index = data_start_row + anchor_index
                        anchor_keys.append(
                            _normalize_token_value(
                                ws.cell(row=row_index, column=columns[key_index]).value
                            )
                        )

    structure_summary = {
        "target_kind": "excel_table",
        "sheet_name": sheet_name,
        "table_name": table_name,
        "range": table_range,
        "headers": headers,
        "column_count": len(headers),
        "header_row_count": header_row_count,
        "totals_row_count": totals_row_count,
        "totals_row_shown": totals_row_shown,
        "data_row_count": data_row_count,
    }
    structure_payload = dict(structure_summary)
    anchor_rows: List[Dict[str, Any]] = []
    if data_row_count > 0:
        for anchor_index in _dataset_anchor_indexes(data_row_count):
            row_index = data_start_row + anchor_index
            anchor_rows.append(
                {
                    "row": row_index,
                    "values": [
                        _normalize_token_value(ws.cell(row=row_index, column=col).value)
                        for col in columns
                    ],
                }
            )

    content_payload = {
        "target_kind": "excel_table",
        "sheet_name": sheet_name,
        "table_name": table_name,
        "data_row_count": data_row_count,
        "anchor_rows": anchor_rows,
        "key_field": key_field,
        "anchor_keys": anchor_keys,
    }

    return {
        "structure_token": _token_digest("sf_struct", structure_payload),
        "content_token": _token_digest("sf_content", content_payload),
        "structure_summary": structure_summary,
    }


def _attach_dataset_identity(
    payload: Dict[str, Any],
    *,
    dataset_tokens: Dict[str, Any],
    filepath: str,
) -> Dict[str, Any]:
    result = dict(payload)
    result["structure_token"] = dataset_tokens["structure_token"]
    result["content_token"] = dataset_tokens["content_token"]
    result["snapshot_metadata"] = _snapshot_metadata(filepath)
    return result


def _raise_structure_token_mismatch(
    *,
    expected_structure_token: str,
    dataset_tokens: Dict[str, Any],
) -> None:
    summary = dataset_tokens["structure_summary"]
    details = {
        "expected_structure_token": expected_structure_token,
        "actual_structure_token": dataset_tokens["structure_token"],
        "current_target_kind": summary["target_kind"],
        "sheet_name": summary.get("sheet_name"),
        "table_name": summary.get("table_name"),
        "headers": summary.get("headers"),
        "column_count": summary.get("column_count"),
        "data_end_row": summary.get("data_end_row"),
        "range": summary.get("range"),
    }
    raise PreconditionFailedError(
        "Dataset structure changed since the read that produced expected_structure_token.",
        code="stale_structure_token",
        details=details,
        suggested_next_tool="describe_dataset",
    )


def _assert_expected_structure_token(
    *,
    expected_structure_token: Optional[str],
    dataset_tokens: Dict[str, Any],
) -> None:
    if expected_structure_token is None:
        return
    if expected_structure_token != dataset_tokens["structure_token"]:
        _raise_structure_token_mismatch(
            expected_structure_token=expected_structure_token,
            dataset_tokens=dataset_tokens,
        )


def _require_structure_change_intent(
    *,
    expected_structure_token: Optional[str],
    allow_structure_change: bool,
    dataset_tokens: Dict[str, Any],
    reason: str,
) -> None:
    if expected_structure_token is None or allow_structure_change:
        return

    summary = dataset_tokens["structure_summary"]
    raise PreconditionFailedError(
        reason,
        code="structure_change_intent_required",
        details={
            "actual_structure_token": dataset_tokens["structure_token"],
            "current_target_kind": summary["target_kind"],
            "sheet_name": summary.get("sheet_name"),
            "table_name": summary.get("table_name"),
        },
        suggested_next_tool="describe_dataset",
    )


def _validate_row_mode(row_mode: str) -> None:
    if row_mode not in ROW_MODES:
        raise DataError("row_mode must be 'arrays' or 'objects'")


def _validate_positive_integer(
    value: Optional[int],
    *,
    argument_name: str,
    allow_none: bool = False,
) -> Optional[int]:
    if value is None:
        if allow_none:
            return None
        raise DataError(f"{argument_name} must be a positive integer")
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise DataError(f"{argument_name} must be a positive integer")
    return value


def _field_name_from_header(header: Any, index: int) -> str:
    raw_header = "" if header is None else str(header).strip()
    if not raw_header:
        return f"column_{index}"

    ascii_header = (
        unicodedata.normalize("NFKD", raw_header).encode("ascii", "ignore").decode("ascii")
    )
    normalized = re.sub(r"[^0-9A-Za-z]+", "_", ascii_header).strip("_").lower()
    if not normalized:
        return f"column_{index}"
    if normalized[0].isdigit():
        return f"column_{normalized}"
    return normalized


def _dedupe_field_name(field_name: str, seen_fields: set[str]) -> str:
    if field_name not in seen_fields:
        seen_fields.add(field_name)
        return field_name

    suffix = 2
    candidate = f"{field_name}_{suffix}"
    while candidate in seen_fields:
        suffix += 1
        candidate = f"{field_name}_{suffix}"
    seen_fields.add(candidate)
    return candidate


def _infer_value_type(value: Any) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, (float, Decimal)):
        return "number"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    if isinstance(value, str):
        if value.startswith("="):
            return "formula"
        return "string"
    return type(value).__name__.lower()


def _infer_column_type(values: List[Any]) -> str:
    non_null_types = {_infer_value_type(value) for value in values if value is not None}
    if not non_null_types:
        return "unknown"
    if len(non_null_types) == 1:
        return next(iter(non_null_types))
    if "formula" in non_null_types:
        return "formula"
    if non_null_types <= {"integer", "number"}:
        return "number"
    if non_null_types <= {"date", "datetime"}:
        return "datetime"
    return "mixed"


def _build_schema(headers: List[Any], rows: List[List[Any]]) -> List[Dict[str, Any]]:
    schema: List[Dict[str, Any]] = []
    seen_fields: set[str] = set()

    for index, header in enumerate(headers, start=1):
        field_name = _dedupe_field_name(_field_name_from_header(header, index), seen_fields)
        column_values = [row[index - 1] if index - 1 < len(row) else None for row in rows]
        schema.append(
            {
                "field": field_name,
                "header": header,
                "type": _infer_column_type(column_values),
                "nullable": any(value is None for value in column_values),
            }
        )

    return schema


def _rows_to_records(rows: List[List[Any]], schema: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for row in rows:
        record: Dict[str, Any] = {}
        for index, column in enumerate(schema):
            record[column["field"]] = row[index] if index < len(row) else None
        records.append(record)
    return records


def augment_tabular_payload(
    payload: Dict[str, Any],
    *,
    headers: List[Any],
    rows: List[List[Any]],
    row_mode: str = "arrays",
    infer_schema: bool = False,
    include_headers: bool = True,
    next_start_row: Optional[int] = None,
) -> Dict[str, Any]:
    _validate_row_mode(row_mode)

    result = dict(payload)
    schema = _build_schema(headers, rows) if infer_schema or row_mode == "objects" else None

    if row_mode == "objects":
        result.pop("rows", None)
        result["records"] = _rows_to_records(rows, schema or [])
        result["row_mode"] = "objects"

    if infer_schema and schema is not None:
        result["schema"] = schema

    if not include_headers:
        result.pop("headers", None)

    if next_start_row is not None:
        result["next_start_row"] = next_start_row

    return result


def _get_header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        header_name = str(value)
        if header_name in header_map:
            raise DataError(f"Duplicate header '{header_name}' found in row {header_row}")
        header_map[header_name] = col

    if not header_map:
        raise DataError(f"No headers found in row {header_row}")

    return header_map


def _row_has_selected_data(ws: Worksheet, row: int, columns: List[int]) -> bool:
    return any(ws.cell(row=row, column=col).value is not None for col in columns)


def _detect_tabular_data_extent(
    ws: Worksheet,
    header_row: int,
    columns: List[int],
    *,
    blank_gap_tolerance: int = TABULAR_BLANK_GAP_TOLERANCE,
) -> Dict[str, Any]:
    # Worksheet-shaped readers favor the first contiguous data block after the header.
    # This keeps sparse footer notes or distant outlier cells from stretching a table read.
    non_empty_rows = [
        row
        for row in range(header_row + 1, ws.max_row + 1)
        if _row_has_selected_data(ws, row, columns)
    ]
    if not non_empty_rows:
        return {
            "first_data_row": None,
            "last_data_row": header_row,
            "ignored_trailing_rows": [],
            "ignored_trailing_row_count": 0,
        }

    last_data_row = non_empty_rows[0]
    cutoff_index = len(non_empty_rows)

    for index in range(1, len(non_empty_rows)):
        current_row = non_empty_rows[index]
        previous_row = non_empty_rows[index - 1]
        blank_gap = current_row - previous_row - 1
        if blank_gap > blank_gap_tolerance:
            cutoff_index = index
            break
        last_data_row = current_row

    ignored_trailing_rows = non_empty_rows[cutoff_index:]
    return {
        "first_data_row": non_empty_rows[0],
        "last_data_row": last_data_row,
        "ignored_trailing_rows": ignored_trailing_rows,
        "ignored_trailing_row_count": len(ignored_trailing_rows),
    }


def _find_last_data_row(ws: Worksheet, header_row: int, columns: List[int]) -> int:
    return _detect_tabular_data_extent(ws, header_row, columns)["last_data_row"]


def _native_table_append_conflict(
    ws: Worksheet,
    *,
    header_row: int,
    next_row: int,
    requested_columns: set[str],
) -> Optional[Dict[str, Any]]:
    if not hasattr(ws, "tables"):
        return None

    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row != header_row or next_row != max_row + 1:
            continue

        table_headers: set[str] = set()
        for column_index in range(min_col, max_col + 1):
            value = ws.cell(row=min_row, column=column_index).value
            if value is None or str(value).strip() == "":
                continue
            table_headers.add(str(value))

        if requested_columns and not requested_columns.issubset(table_headers):
            continue

        return {
            "table_name": table.displayName,
            "totals_row_shown": bool(getattr(table, "totalsRowShown", False)),
        }

    return None


def _build_cell_change(
    sheet_name: str,
    row: int,
    col: int,
    old_value: Any,
    new_value: Any,
    column_name: Optional[str] = None,
) -> Dict[str, Any]:
    change = {
        "sheet_name": sheet_name,
        "cell": _cell_address(row, col),
        "row": row,
        "column": col,
        "old_value": old_value,
        "new_value": new_value,
    }
    if column_name is not None:
        change["column_name"] = column_name
    return change


def _should_include_changes(dry_run: bool, include_changes: Optional[bool]) -> bool:
    if include_changes is None:
        return dry_run
    return include_changes


def _selected_columns(start_col_idx: int, end_col_idx: int) -> List[int]:
    return list(range(start_col_idx, end_col_idx + 1))


def _column_index(label: str, *, argument_name: str) -> int:
    try:
        return column_index_from_string(label.upper())
    except ValueError as exc:
        raise DataError(f"{argument_name} must be a valid Excel column label") from exc


def _sheet_type_name(ws: Any) -> str:
    return "chartsheet" if ws.__class__.__name__ == "Chartsheet" else "worksheet"


def _worksheet_is_empty(ws: Worksheet) -> bool:
    return ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None


def _worksheet_used_range(ws: Worksheet) -> Optional[str]:
    if _worksheet_is_empty(ws):
        return None
    return f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _stringify_value_for_uniqueness(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, sort_keys=True, default=str)
    if isinstance(value, (date, datetime, time, Decimal)):
        return str(value)
    return value


def _header_profile(headers: List[Any]) -> Dict[str, Any]:
    total = len(headers)
    if total == 0:
        return {
            "total_headers": 0,
            "non_empty_headers": 0,
            "blank_headers": 0,
            "string_headers": 0,
            "duplicate_headers": 0,
            "score": 0.0,
            "confidence": "low",
        }

    normalized_headers: List[str] = []
    non_empty_headers = 0
    string_headers = 0
    for header in headers:
        if _is_blank(header):
            continue
        non_empty_headers += 1
        if isinstance(header, str) and header.strip():
            string_headers += 1
            normalized_headers.append(header.strip().lower())
        else:
            normalized_headers.append(str(header).strip().lower())

    blank_headers = total - non_empty_headers
    duplicate_headers = len(normalized_headers) - len(set(normalized_headers))

    non_empty_ratio = non_empty_headers / total
    string_ratio = string_headers / non_empty_headers if non_empty_headers else 0.0
    duplicate_penalty = min(0.25, duplicate_headers * 0.1)
    score = max(0.0, min(1.0, non_empty_ratio * 0.65 + string_ratio * 0.35 - duplicate_penalty))

    if score >= 0.85:
        confidence = "high"
    elif score >= 0.6:
        confidence = "medium"
    else:
        confidence = "low"

    return {
        "total_headers": total,
        "non_empty_headers": non_empty_headers,
        "blank_headers": blank_headers,
        "string_headers": string_headers,
        "duplicate_headers": duplicate_headers,
        "score": round(score, 2),
        "confidence": confidence,
    }


def _infer_key_candidates(
    headers: List[Any],
    rows: List[List[Any]],
    schema: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    scan_rows = rows[:KEY_CANDIDATE_SCAN_LIMIT]
    if len(scan_rows) < 2:
        return []

    candidates: List[Dict[str, Any]] = []
    for index, column in enumerate(schema):
        column_values = [row[index] if index < len(row) else None for row in scan_rows]
        non_blank_values = [
            _stringify_value_for_uniqueness(value)
            for value in column_values
            if not _is_blank(value)
        ]
        if len(non_blank_values) < 2:
            continue

        unique_ratio = len(set(non_blank_values)) / len(non_blank_values)
        if unique_ratio < 0.95:
            continue

        confidence = (
            "high"
            if unique_ratio == 1.0 and len(non_blank_values) == len(scan_rows)
            else "medium"
        )
        candidates.append(
            {
                "field": column["field"],
                "header": headers[index] if index < len(headers) else None,
                "type": column["type"],
                "sample_unique_ratio": round(unique_ratio, 2),
                "confidence": confidence,
            }
        )

    candidates.sort(
        key=lambda item: (
            0 if item["confidence"] == "high" else 1,
            item["field"],
        )
    )
    return candidates[:5]


def _normalize_read_goal(goal: Optional[str]) -> str:
    if goal is None or not goal.strip():
        return "inspect"

    normalized = goal.strip().lower()
    layout_keywords = {"layout", "dashboard", "chart", "visual", "format", "canvas"}
    tabular_keywords = {
        "table",
        "rows",
        "records",
        "schema",
        "data",
        "extract",
        "analytics",
        "aggregate",
        "summary",
    }

    if any(keyword in normalized for keyword in layout_keywords):
        return "layout"
    if any(keyword in normalized for keyword in tabular_keywords):
        return "tabular"
    return "inspect"


def _table_dominates_sheet(summary: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    native_tables = summary.get("native_tables", [])
    if len(native_tables) != 1:
        return None

    table = native_tables[0]
    total_rows = summary.get("total_rows", 0)
    column_count = summary.get("column_count", 0)
    if total_rows <= 0 or column_count <= 0:
        return None

    row_coverage = table["data_row_count"] / total_rows if total_rows else 0.0
    col_coverage = table["column_count"] / column_count if column_count else 0.0
    if row_coverage >= 0.8 and col_coverage >= 0.8:
        return table
    return None


def _strategy_page_size(total_rows: int) -> Optional[int]:
    if total_rows > DEFAULT_DATASET_SAMPLE_ROWS:
        return DEFAULT_DATASET_SAMPLE_ROWS
    return None


def _worksheet_dataset_kind(
    *,
    used_range: Optional[str],
    total_rows: int,
    header_confidence: str,
    chart_count: int,
    merged_range_count: int,
) -> str:
    if used_range is None:
        return "empty_sheet"
    if (chart_count > 0 or merged_range_count > 0) and header_confidence == "low":
        return "layout_like_sheet"
    if total_rows > 0 and header_confidence in {"high", "medium"}:
        return "worksheet_table"
    return "mixed_sheet"


def _describe_chartsheet(
    *,
    filepath: str,
    sheet_name: str,
    auto_selected_sheet: bool,
    chart_count: int,
) -> Dict[str, Any]:
    return {
        "target_kind": "chartsheet",
        "dataset_kind": "chartsheet",
        "sheet_name": sheet_name,
        "auto_selected_sheet": auto_selected_sheet,
        "chart_count": chart_count,
        "observations": [
            "Chart sheets do not expose worksheet cells or tabular headers.",
            "Use chart-oriented workbook tools instead of worksheet table readers.",
        ],
        "recommended_read_tool": "list_charts",
        "recommended_args": {
            "filepath": filepath,
            "sheet_name": sheet_name,
        },
    }


def _describe_table_dataset(
    filepath: str,
    *,
    table_name: str,
    sheet_name: Optional[str] = None,
    sample_rows: int = DEFAULT_DATASET_SAMPLE_ROWS,
) -> Dict[str, Any]:
    from .tables import _build_table_metadata, _find_table

    with safe_workbook(str(filepath)) as wb:
        current_sheet_name, ws, table = _find_table(wb, table_name, sheet_name=sheet_name)
        metadata = _build_table_metadata(current_sheet_name, ws, table)
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        data_start_row = min_row + metadata["header_row_count"]
        data_end_row = max_row - metadata["totals_row_count"]
        row_limit = min(metadata["data_row_count"], sample_rows)
        sample_data_rows: List[List[Any]] = []
        for row_index in range(data_start_row, data_start_row + row_limit):
            if row_index > data_end_row:
                break
            sample_data_rows.append(
                [
                    ws.cell(row=row_index, column=column_index).value
                    for column_index in range(min_col, max_col + 1)
                ]
            )

        sample = {
            "headers": metadata["headers"],
            "rows": sample_data_rows,
            "schema": _build_schema(metadata["headers"], sample_data_rows),
            "truncated": metadata["data_row_count"] > sample_rows,
        }
        dataset_tokens = _table_dataset_tokens(
            ws,
            sheet_name=current_sheet_name,
            table_name=metadata["table_name"],
            table_range=metadata["range"],
            headers=metadata["headers"],
            header_row_count=metadata["header_row_count"],
            totals_row_count=metadata["totals_row_count"],
            totals_row_shown=metadata["totals_row_shown"],
        )

    key_candidates = _infer_key_candidates(
        sample["headers"],
        sample["rows"],
        sample.get("schema", []),
    )
    observations = [
        f"Native Excel table '{metadata['table_name']}' provides explicit row and column boundaries.",
    ]
    if metadata["totals_row_shown"]:
        observations.append("Table has a totals row, so append-like mutations stay more constrained.")
    if sample.get("truncated"):
        observations.append(
            f"Sample rows are truncated to {sample_rows}; continue with read_excel_table for deeper pagination."
        )

    recommended_args: Dict[str, Any] = {
        "filepath": filepath,
        "table_name": metadata["table_name"],
        "sheet_name": current_sheet_name,
        "row_mode": "objects",
        "infer_schema": True,
    }
    page_size = _strategy_page_size(metadata["data_row_count"])
    if page_size is not None:
        recommended_args["max_rows"] = page_size

    return _attach_dataset_identity({
        "target_kind": "excel_table",
        "dataset_kind": "structured_table",
        "sheet_name": current_sheet_name,
        "table_name": metadata["table_name"],
        "range": metadata["range"],
        "column_count": metadata["column_count"],
        "total_rows": metadata["data_row_count"],
        "sample_row_count": len(sample["rows"]),
        "headers": sample["headers"],
        "sample_rows": sample["rows"],
        "schema": sample.get("schema", []),
        "key_candidates": key_candidates,
        "native_table_count": 1,
        "native_tables": [
            {
                "table_name": metadata["table_name"],
                "range": metadata["range"],
                "data_row_count": metadata["data_row_count"],
                "column_count": metadata["column_count"],
            }
        ],
        "header_profile": _header_profile(sample["headers"]),
        "observations": observations,
        "recommended_read_tool": "read_excel_table",
        "recommended_args": recommended_args,
    }, dataset_tokens=dataset_tokens, filepath=filepath)


def _describe_worksheet_dataset(
    filepath: str,
    *,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    sample_rows: int = DEFAULT_DATASET_SAMPLE_ROWS,
) -> Dict[str, Any]:
    from .tables import _build_table_metadata

    with safe_workbook(str(filepath)) as wb:
        auto_selected_sheet = sheet_name is None
        if auto_selected_sheet:
            resolved_sheet_name, ws = first_worksheet(wb, error_cls=DataError)
        else:
            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found")
            resolved_sheet_name = sheet_name
            ws = wb[resolved_sheet_name]

        if _sheet_type_name(ws) == "chartsheet":
            return _describe_chartsheet(
                filepath=filepath,
                sheet_name=resolved_sheet_name,
                auto_selected_sheet=auto_selected_sheet,
                chart_count=len(getattr(ws, "_charts", [])),
            )

        worksheet = require_worksheet(
            wb,
            resolved_sheet_name,
            error_cls=DataError,
            operation="dataset description",
        )
        sample = _read_table_from_worksheet(
            worksheet,
            resolved_sheet_name,
            header_row=header_row,
            max_rows=sample_rows,
            include_headers=True,
            row_mode="arrays",
            infer_schema=True,
        )
        native_tables = [
            _build_table_metadata(resolved_sheet_name, worksheet, table)
            for table in worksheet.tables.values()
        ]
        selected_columns = _selected_columns(1, worksheet.max_column)
        data_extent = _detect_tabular_data_extent(
            worksheet,
            header_row,
            selected_columns,
        )

        used_range = _worksheet_used_range(worksheet)
        header_profile = _header_profile(sample["headers"])
        key_candidates = _infer_key_candidates(
            sample["headers"],
            sample["rows"],
            sample.get("schema", []),
        )
        chart_count = len(getattr(worksheet, "_charts", []))
        merged_range_count = len(worksheet.merged_cells.ranges)
        dataset_kind = _worksheet_dataset_kind(
            used_range=used_range,
            total_rows=sample["total_rows"],
            header_confidence=header_profile["confidence"],
            chart_count=chart_count,
            merged_range_count=merged_range_count,
        )

        observations: List[str] = []
        if native_tables:
            observations.append(
                f"Worksheet contains {len(native_tables)} native Excel table(s) that may be better read with read_excel_table."
            )
        if header_profile["blank_headers"] > 0:
            observations.append("Header row contains blank cells, which lowers tabular confidence.")
        if header_profile["duplicate_headers"] > 0:
            observations.append("Header row contains duplicate labels, so normalized field names may be deduped.")
        if chart_count > 0:
            observations.append("Worksheet contains embedded charts, which is common in dashboard-like layouts.")
        if merged_range_count > 0:
            observations.append("Worksheet contains merged cells, which often signals a layout-oriented sheet.")
        if sample.get("truncated"):
            observations.append(
                f"Sample rows are truncated to {sample_rows}; use pagination helpers for deeper reads."
            )
        if data_extent["ignored_trailing_row_count"] > 0:
            observations.append(
                "Later non-empty rows were detected after a large blank gap; worksheet table readers treat those rows as a separate block instead of stretching the main dataset."
            )

        dominated_table = _table_dominates_sheet(
            {
                "native_tables": native_tables,
                "total_rows": sample["total_rows"],
                "column_count": len(sample["headers"]),
            }
        )

        if dominated_table is not None:
            recommended_read_tool = "read_excel_table"
            recommended_args: Dict[str, Any] = {
                "filepath": filepath,
                "table_name": dominated_table["table_name"],
                "sheet_name": resolved_sheet_name,
                "row_mode": "objects",
                "infer_schema": True,
            }
            page_size = _strategy_page_size(dominated_table["data_row_count"])
            if page_size is not None:
                recommended_args["max_rows"] = page_size
        elif dataset_kind == "layout_like_sheet":
            recommended_read_tool = "profile_workbook"
            recommended_args = {"filepath": filepath}
        else:
            recommended_read_tool = "quick_read"
            recommended_args = {
                "filepath": filepath,
                "sheet_name": resolved_sheet_name,
                "row_mode": "objects",
                "infer_schema": True,
            }
            page_size = _strategy_page_size(sample["total_rows"])
            if page_size is not None:
                recommended_args["max_rows"] = page_size

        dataset_tokens = _worksheet_dataset_tokens(
            worksheet,
            sheet_name=resolved_sheet_name,
            header_row=header_row,
        )

        return _attach_dataset_identity({
            "target_kind": "worksheet",
            "dataset_kind": dataset_kind,
            "sheet_name": resolved_sheet_name,
            "auto_selected_sheet": auto_selected_sheet,
            "header_row": header_row,
            "used_range": used_range,
            "data_end_row": data_extent["last_data_row"],
            "column_count": len(sample["headers"]),
            "total_rows": sample["total_rows"],
            "sample_row_count": len(sample["rows"]),
            "headers": sample["headers"],
            "sample_rows": sample["rows"],
            "schema": sample.get("schema", []),
            "key_candidates": key_candidates,
            "header_profile": header_profile,
            "chart_count": chart_count,
            "merged_range_count": merged_range_count,
            "has_autofilter": bool(worksheet.auto_filter.ref),
            "freeze_panes": getattr(worksheet.freeze_panes, "coordinate", worksheet.freeze_panes),
            "ignored_trailing_row_count": data_extent["ignored_trailing_row_count"],
            "native_table_count": len(native_tables),
            "native_tables": [
                {
                    "table_name": table["table_name"],
                    "range": table["range"],
                    "data_row_count": table["data_row_count"],
                    "column_count": table["column_count"],
                }
                for table in native_tables
            ],
            "observations": observations,
            "recommended_read_tool": recommended_read_tool,
            "recommended_args": recommended_args,
        }, dataset_tokens=dataset_tokens, filepath=filepath)


def describe_dataset(
    filepath: str,
    sheet_name: Optional[str] = None,
    table_name: Optional[str] = None,
    header_row: int = 1,
    sample_rows: int = DEFAULT_DATASET_SAMPLE_ROWS,
) -> Dict[str, Any]:
    """Describe the most likely dataset shape for a worksheet or native Excel table."""
    try:
        _validate_positive_integer(header_row, argument_name="header_row")
        _validate_positive_integer(sample_rows, argument_name="sample_rows")

        if table_name is not None:
            return _describe_table_dataset(
                filepath,
                table_name=table_name,
                sheet_name=sheet_name,
                sample_rows=sample_rows,
            )

        return _describe_worksheet_dataset(
            filepath,
            sheet_name=sheet_name,
            header_row=header_row,
            sample_rows=sample_rows,
        )
    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to describe dataset: {e}")
        raise DataError(str(e))


def suggest_read_strategy(
    filepath: str,
    goal: Optional[str] = None,
    sheet_name: Optional[str] = None,
    table_name: Optional[str] = None,
    header_row: int = 1,
    sample_rows: int = DEFAULT_DATASET_SAMPLE_ROWS,
) -> Dict[str, Any]:
    """Recommend the best SheetForge read path for a workbook target."""
    try:
        summary = describe_dataset(
            filepath,
            sheet_name=sheet_name,
            table_name=table_name,
            header_row=header_row,
            sample_rows=sample_rows,
        )
        normalized_goal = _normalize_read_goal(goal)
        observations = list(summary.get("observations", []))

        if summary["target_kind"] == "chartsheet":
            recommended_tool = "list_charts"
            reason = "Chart sheets do not expose a worksheet grid, so tabular readers are the wrong fit."
            confidence = "high"
            suggested_args = {
                "filepath": filepath,
                "sheet_name": summary["sheet_name"],
            }
            alternatives = [
                {
                    "tool": "profile_workbook",
                    "reason": "Use workbook orientation first when you need broader chart and sheet context.",
                }
            ]
        elif summary["target_kind"] == "excel_table":
            recommended_tool = "read_excel_table"
            reason = (
                f"Native Excel table '{summary['table_name']}' already defines stable headers and row bounds."
            )
            confidence = "high"
            suggested_args = dict(summary["recommended_args"])
            alternatives = [
                {
                    "tool": "quick_read",
                    "reason": "Use only if you intentionally want worksheet-shaped reads instead of table semantics.",
                }
            ]
        elif summary["dataset_kind"] == "layout_like_sheet":
            if normalized_goal == "layout":
                recommended_tool = "read_data_from_excel"
                reason = (
                    "This sheet looks layout-oriented, so a compact cell-range preview is safer than assuming a clean table."
                )
                confidence = "high"
                suggested_args = {
                    "filepath": filepath,
                    "sheet_name": summary["sheet_name"],
                    "start_cell": "A1",
                    "end_cell": summary.get("used_range", "A1").split(":")[-1],
                    "values_only": True,
                    "preview_only": True,
                }
                alternatives = [
                    {
                        "tool": "profile_workbook",
                        "reason": "Use workbook orientation first if you need chart, table, or layout context before reading cells.",
                    }
                ]
            else:
                recommended_tool = "profile_workbook"
                reason = (
                    "This sheet looks layout-heavy, so workbook orientation is a safer first step than forcing a tabular read."
                )
                confidence = "high"
                suggested_args = {"filepath": filepath}
                alternatives = [
                    {
                        "tool": "read_data_from_excel",
                        "reason": "Use a compact cell-range preview next if you need exact grid values from the dashboard area.",
                    }
                ]
        elif _table_dominates_sheet(summary) is not None:
            dominated_table = _table_dominates_sheet(summary)
            recommended_tool = "read_excel_table"
            reason = (
                f"Worksheet '{summary['sheet_name']}' appears to be primarily driven by native table '{dominated_table['table_name']}'."
            )
            confidence = "high"
            suggested_args = {
                "filepath": filepath,
                "table_name": dominated_table["table_name"],
                "sheet_name": summary["sheet_name"],
                "row_mode": "objects",
                "infer_schema": True,
            }
            page_size = _strategy_page_size(dominated_table["data_row_count"])
            if page_size is not None:
                suggested_args["max_rows"] = page_size
            alternatives = [
                {
                    "tool": "quick_read",
                    "reason": "Use only if you explicitly want worksheet rows instead of table-bound reads.",
                }
            ]
        else:
            recommended_tool = "quick_read"
            reason = (
                f"Worksheet '{summary['sheet_name']}' looks like a regular header-based dataset without requiring native table semantics."
            )
            confidence = "medium" if summary["header_profile"]["confidence"] == "medium" else "high"
            suggested_args = {
                "filepath": filepath,
                "sheet_name": summary["sheet_name"],
                "row_mode": "objects",
                "infer_schema": True,
            }
            page_size = _strategy_page_size(summary["total_rows"])
            if page_size is not None:
                suggested_args["max_rows"] = page_size
            alternatives = [
                {
                    "tool": "read_excel_as_table",
                    "reason": "Use when you need the same worksheet read path but want explicit sheet-only table semantics.",
                },
                {
                    "tool": "read_data_from_excel",
                    "reason": "Use when cell addresses or sparse ranges matter more than header-based rows.",
                },
            ]

        return {
            "goal": goal,
            "normalized_goal": normalized_goal,
            "target_kind": summary["target_kind"],
            "dataset_kind": summary["dataset_kind"],
            "sheet_name": summary.get("sheet_name"),
            "table_name": summary.get("table_name"),
            "auto_selected_sheet": summary.get("auto_selected_sheet", False),
            "recommended_tool": recommended_tool,
            "confidence": confidence,
            "reason": reason,
            "suggested_args": suggested_args,
            "alternatives": alternatives,
            "observations": observations,
        }
    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to suggest read strategy: {e}")
        raise DataError(str(e))


def _encode_range_read_cursor(
    *,
    start_cell: str,
    end_cell: str,
    max_rows: Optional[int],
    max_cols: Optional[int],
    include_validation: bool,
    compact: bool,
    values_only: bool,
) -> str:
    payload: Dict[str, Any] = {
        "v": RANGE_READ_CURSOR_VERSION,
        "start_cell": start_cell,
        "end_cell": end_cell,
        "include_validation": include_validation,
        "compact": compact,
        "values_only": values_only,
    }
    if max_rows is not None:
        payload["max_rows"] = max_rows
    if max_cols is not None:
        payload["max_cols"] = max_cols
    encoded = base64.urlsafe_b64encode(
        json.dumps(payload, separators=(",", ":"), sort_keys=True).encode("utf-8")
    ).decode("ascii")
    return encoded.rstrip("=")


def _decode_range_read_cursor(cursor: str) -> Dict[str, Any]:
    if not cursor:
        raise DataError("cursor must not be empty")

    try:
        padded = cursor + "=" * (-len(cursor) % 4)
        decoded = base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
        payload = json.loads(decoded)
    except Exception as exc:
        raise DataError("Invalid cursor") from exc

    if not isinstance(payload, dict):
        raise DataError("Invalid cursor")
    if payload.get("v") not in {1, RANGE_READ_CURSOR_VERSION}:
        raise DataError("Unsupported cursor version")

    start_cell = payload.get("start_cell")
    end_cell = payload.get("end_cell")
    if not isinstance(start_cell, str) or not start_cell.strip():
        raise DataError("Invalid cursor")
    if not isinstance(end_cell, str) or not end_cell.strip():
        raise DataError("Invalid cursor")

    max_rows = payload.get("max_rows")
    max_cols = payload.get("max_cols")
    if max_rows is not None and (
        isinstance(max_rows, bool) or not isinstance(max_rows, int) or max_rows <= 0
    ):
        raise DataError("Invalid cursor")
    if max_cols is not None and (
        isinstance(max_cols, bool) or not isinstance(max_cols, int) or max_cols <= 0
    ):
        raise DataError("Invalid cursor")

    include_validation = payload.get("include_validation")
    compact = payload.get("compact")
    values_only = payload.get("values_only")
    if include_validation is not None and not isinstance(include_validation, bool):
        raise DataError("Invalid cursor")
    if compact is not None and not isinstance(compact, bool):
        raise DataError("Invalid cursor")
    if values_only is not None and not isinstance(values_only, bool):
        raise DataError("Invalid cursor")

    return {
        "start_cell": start_cell,
        "end_cell": end_cell,
        "max_rows": max_rows,
        "max_cols": max_cols,
        "include_validation": include_validation,
        "compact": compact,
        "values_only": values_only,
    }


def _build_range_continuation(
    *,
    start_row: int,
    start_col: int,
    requested_end_row: int,
    requested_end_col: int,
    max_rows: Optional[int],
    max_cols: Optional[int],
    include_validation: bool,
    compact: bool,
    values_only: bool,
) -> Dict[str, Any]:
    start_cell = _cell_address(start_row, start_col)
    end_cell = _cell_address(requested_end_row, requested_end_col)
    return {
        "cursor": _encode_range_read_cursor(
            start_cell=start_cell,
            end_cell=end_cell,
            max_rows=max_rows,
            max_cols=max_cols,
            include_validation=include_validation,
            compact=compact,
            values_only=values_only,
        ),
        "start_cell": start_cell,
        "end_cell": end_cell,
    }

def read_excel_range(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> List[Dict[str, Any]]:
    """Read data from Excel range with optional preview mode"""
    try:
        with safe_workbook(str(filepath)) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=DataError,
                operation="cell-based operations",
            )

            # Parse start cell
            if ':' in start_cell:
                start_cell, end_cell = start_cell.split(':')

            # Get start coordinates
            try:
                start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
                if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                    raise DataError(f"Invalid start cell reference: {start_cell}")
                start_row, start_col = start_coords[0], start_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid start cell format: {str(e)}")

            # Determine end coordinates
            if end_cell:
                try:
                    end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                    if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                        raise DataError(f"Invalid end cell reference: {end_cell}")
                    end_row, end_col = end_coords[0], end_coords[1]
                except ValueError as e:
                    raise DataError(f"Invalid end cell format: {str(e)}")
            else:
                # If no end_cell, use the full data range of the sheet
                if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                    # Handle empty sheet
                    end_row, end_col = start_row, start_col
                else:
                    end_row = ws.max_row
                    end_col = ws.max_column

            # Validate range bounds
            if start_row > ws.max_row or start_col > ws.max_column:
                logger.warning(
                    f"Start cell {start_cell} is outside the sheet's data boundary "
                    f"({get_column_letter(ws.min_column)}{ws.min_row}:{get_column_letter(ws.max_column)}{ws.max_row}). "
                    f"No data will be read."
                )
                return []

            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    row_data.append(cell.value)
                if any(v is not None for v in row_data):
                    data.append(row_data)

            return data
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range: {e}")
        raise DataError(str(e))

def write_data(
    filepath: str,
    sheet_name: Optional[str],
    data: Optional[List[List]],
    start_cell: str = "A1",
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Write data to Excel sheet with workbook handling

    Headers are handled intelligently based on context.
    """
    try:
        if not data:
            raise DataError("No data provided to write")

        with safe_workbook(filepath, save=not dry_run) as wb:
            # If no sheet specified, use active sheet
            if not sheet_name:
                active_sheet = wb.active
                if active_sheet is None:
                    raise DataError("No active sheet found in workbook")
                if not isinstance(active_sheet, Worksheet):
                    raise DataError(
                        f"Active sheet '{active_sheet.title}' is a chartsheet and cannot be used for cell-based operations"
                    )
                sheet_name = active_sheet.title
                sheet_created = False
                existing_ws = active_sheet
            else:
                sheet_created = sheet_name not in wb.sheetnames
                existing_ws = (
                    None
                    if sheet_created
                    else require_worksheet(
                        wb,
                        sheet_name,
                        error_cls=DataError,
                        operation="cell-based operations",
                    )
                )

            # Validate start cell
            try:
                start_coords = parse_cell_range(start_cell)
                if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                    raise DataError(f"Invalid start cell reference: {start_cell}")
            except ValueError as e:
                raise DataError(f"Invalid start cell format: {str(e)}")

            start_row, start_col = start_coords[0], start_coords[1]
            total_cells = sum(len(row) for row in data)
            if data:
                max_cols = max(len(row) for row in data)
                end_row = start_row + len(data) - 1
                end_col = start_col + max_cols - 1
                target_range = _range_string(start_row, start_col, end_row, end_col)
            else:
                target_range = start_cell

            changes: List[Dict[str, Any]] = []
            for i, row in enumerate(data):
                for j, val in enumerate(row):
                    row_idx = start_row + i
                    col_idx = start_col + j
                    old_value = None if existing_ws is None else existing_ws.cell(row=row_idx, column=col_idx).value
                    if old_value != val:
                        changes.append(
                            _build_cell_change(
                                sheet_name=sheet_name,
                                row=row_idx,
                                col=col_idx,
                                old_value=old_value,
                                new_value=val,
                            )
                        )

            if len(data) > 0:
                if sheet_created:
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = existing_ws
                if not dry_run:
                    _write_data_to_worksheet(ws, data, start_cell)

        result = {
            "message": f"{'Previewed' if dry_run else 'Wrote'} data to {sheet_name}",
            "active_sheet": sheet_name,
            "target_range": target_range,
            "sheet_created": sheet_created,
            "cells_written": total_cells,
            "changed_cells": len(changes),
            "dry_run": dry_run,
        }
        if _should_include_changes(dry_run, include_changes):
            result["changes"] = changes
        return result
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write data: {e}")
        raise DataError(str(e))

def _write_data_to_worksheet(
    worksheet: Worksheet, 
    data: List[List], 
    start_cell: str = "A1",
) -> None:
    """Write data to worksheet with intelligent header handling"""
    try:
        if not data:
            raise DataError("No data provided to write")

        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(x is not None for x in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Write data
        for i, row in enumerate(data):
            for j, val in enumerate(row):
                worksheet.cell(row=start_row + i, column=start_col + j, value=val)
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write worksheet data: {e}")
        raise DataError(str(e))

def read_excel_range_with_metadata(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    max_rows: Optional[int] = None,
    max_cols: Optional[int] = None,
    cursor: Optional[str] = None,
    include_validation: bool = True,
    compact: bool = False,
    values_only: bool = False,
) -> Dict[str, Any]:
    """Read data from Excel range with cell metadata including validation rules.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell address
        end_cell: Ending cell address (optional)
        include_validation: Whether to include validation metadata

    Returns:
        Dictionary containing structured cell data with metadata
    """
    try:
        if cursor is not None:
            cursor_state = _decode_range_read_cursor(cursor)
            start_cell = cursor_state["start_cell"]
            end_cell = cursor_state["end_cell"]
            if max_rows is None:
                max_rows = cursor_state["max_rows"]
            if max_cols is None:
                max_cols = cursor_state["max_cols"]
            if cursor_state["include_validation"] is not None:
                include_validation = cursor_state["include_validation"]
            if cursor_state["compact"] is not None:
                compact = cursor_state["compact"]
            if cursor_state["values_only"] is not None:
                values_only = cursor_state["values_only"]

        _validate_positive_integer(max_rows, argument_name="max_rows", allow_none=True)
        _validate_positive_integer(max_cols, argument_name="max_cols", allow_none=True)

        with safe_workbook(str(filepath)) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=DataError,
                operation="cell-based operations",
            )

            # Parse start cell
            if ':' in start_cell:
                start_cell, end_cell = start_cell.split(':')

            # Get start coordinates
            try:
                start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
                if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                    raise DataError(f"Invalid start cell reference: {start_cell}")
                start_row, start_col = start_coords[0], start_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid start cell format: {str(e)}")

            # Determine end coordinates
            if end_cell:
                try:
                    end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                    if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                        raise DataError(f"Invalid end cell reference: {end_cell}")
                    end_row, end_col = end_coords[0], end_coords[1]
                except ValueError as e:
                    raise DataError(f"Invalid end cell format: {str(e)}")
            else:
                # If no end_cell, use the full data range of the sheet
                if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                    # Handle empty sheet
                    end_row, end_col = start_row, start_col
                else:
                    # Expand to sheet's max boundaries; start_row/start_col from caller are preserved
                    end_row = ws.max_row
                    end_col = ws.max_column

            requested_end_row = end_row
            requested_end_col = end_col
            if max_rows is not None:
                end_row = min(end_row, start_row + max_rows - 1)
            if max_cols is not None:
                end_col = min(end_col, start_col + max_cols - 1)

            # Validate range bounds
            if start_row > ws.max_row or start_col > ws.max_column:
                logger.warning(
                    f"Start cell {start_cell} is outside the sheet's data boundary "
                    f"({get_column_letter(ws.min_column)}{ws.min_row}:{get_column_letter(ws.max_column)}{ws.max_row}). "
                    f"No data will be read."
                )
                requested_range = f"{start_cell}:{end_cell}" if end_cell else start_cell
                empty_key = "values" if values_only else "cells"
                return {"range": requested_range, "sheet_name": sheet_name, empty_key: []}

            range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            range_data = {"range": range_str, "sheet_name": sheet_name}
            total_rows = requested_end_row - start_row + 1
            total_cols = requested_end_col - start_col + 1
            truncated_rows = max_rows is not None and end_row < requested_end_row
            truncated_cols = max_cols is not None and end_col < requested_end_col
            truncated = truncated_rows or truncated_cols
            if max_rows is not None:
                range_data["total_rows"] = total_rows
            if max_cols is not None:
                range_data["total_cols"] = total_cols
            if max_rows is not None or max_cols is not None:
                range_data["truncated"] = truncated
            if max_rows is not None:
                if truncated_rows:
                    next_start_row = end_row + 1
                    range_data["next_start_row"] = next_start_row
                    range_data["next_start_cell"] = f"{get_column_letter(start_col)}{next_start_row}"
            if max_cols is not None and truncated_cols:
                next_start_col = end_col + 1
                range_data["next_start_col"] = get_column_letter(next_start_col)
                range_data["next_column_start_cell"] = f"{get_column_letter(next_start_col)}{start_row}"

            continuations: Dict[str, Dict[str, Any]] = {}
            if truncated_rows:
                continuations["down"] = _build_range_continuation(
                    start_row=end_row + 1,
                    start_col=start_col,
                    requested_end_row=requested_end_row,
                    requested_end_col=requested_end_col,
                    max_rows=max_rows,
                    max_cols=max_cols,
                    include_validation=include_validation,
                    compact=compact,
                    values_only=values_only,
                )
            if truncated_cols:
                continuations["right"] = _build_range_continuation(
                    start_row=start_row,
                    start_col=end_col + 1,
                    requested_end_row=requested_end_row,
                    requested_end_col=requested_end_col,
                    max_rows=max_rows,
                    max_cols=max_cols,
                    include_validation=include_validation,
                    compact=compact,
                    values_only=values_only,
                )
            if continuations:
                range_data["continuations"] = continuations
                if len(continuations) == 1:
                    range_data["next_cursor"] = next(iter(continuations.values()))["cursor"]

            if values_only:
                values: List[List[Any]] = []
                for row in range(start_row, end_row + 1):
                    row_values = []
                    for col in range(start_col, end_col + 1):
                        row_values.append(ws.cell(row=row, column=col).value)
                    values.append(row_values)
                range_data["values"] = values
                return range_data

            range_data["cells"] = []

            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell_address = f"{get_column_letter(col)}{row}"

                    cell_data = {
                        "address": cell_address,
                        "value": cell.value,
                        "row": row,
                        "column": col
                    }

                    # Add validation metadata if requested
                    if include_validation:
                        validation_info = get_data_validation_for_cell(ws, cell_address)
                        if validation_info:
                            cell_data["validation"] = validation_info
                        elif not compact:
                            cell_data["validation"] = {"has_validation": False}

                    range_data["cells"].append(cell_data)

            return range_data

    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range with metadata: {e}")
        raise DataError(str(e))


def read_as_table(
    filepath: str,
    sheet_name: str,
    header_row: int = 1,
    start_row: Optional[int] = None,
    start_col: str = "A",
    end_col: Optional[str] = None,
    max_rows: Optional[int] = None,
    compact: bool = False,
    include_headers: bool = True,
    row_mode: str = "arrays",
    infer_schema: bool = False,
) -> Dict[str, Any]:
    """Read Excel data as a compact table with headers.

    Returns a dict with headers plus either rows or records, optional schema
    hints, total_rows, truncated, and sheet_name.
    """
    try:
        with safe_workbook(str(filepath)) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=DataError,
                operation="cell-based operations",
            )
            return _read_table_from_worksheet(
                ws,
                sheet_name,
                filepath=filepath,
                header_row=header_row,
                start_row=start_row,
                start_col=start_col,
                end_col=end_col,
                max_rows=max_rows,
                compact=compact,
                include_headers=include_headers,
                row_mode=row_mode,
                infer_schema=infer_schema,
            )
    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to read as table: {e}")
        raise DataError(str(e))


def _read_table_from_worksheet(
    ws: Worksheet,
    sheet_name: str,
    *,
    filepath: Optional[str] = None,
    header_row: int = 1,
    start_row: Optional[int] = None,
    start_col: str = "A",
    end_col: Optional[str] = None,
    max_rows: Optional[int] = None,
    compact: bool = False,
    include_headers: bool = True,
    row_mode: str = "arrays",
    infer_schema: bool = False,
) -> Dict[str, Any]:
    _validate_positive_integer(header_row, argument_name="header_row")
    start_col_idx = _column_index(start_col, argument_name="start_col")
    if end_col:
        end_col_idx = _column_index(end_col, argument_name="end_col")
    else:
        end_col_idx = ws.max_column
    if end_col_idx < start_col_idx:
        raise DataError("end_col must be greater than or equal to start_col")
    _validate_positive_integer(max_rows, argument_name="max_rows", allow_none=True)

    effective_start_row = header_row + 1 if start_row is None else start_row
    _validate_positive_integer(effective_start_row, argument_name="start_row")
    if effective_start_row <= header_row:
        raise DataError("start_row must be greater than header_row")

    selected_columns = _selected_columns(start_col_idx, end_col_idx)
    last_data_row = _find_last_data_row(ws, header_row, selected_columns)

    headers = []
    for col in selected_columns:
        headers.append(ws.cell(row=header_row, column=col).value)

    total_rows = last_data_row - header_row
    if total_rows < 0:
        total_rows = 0

    rows = []
    available_rows = 0
    if effective_start_row <= last_data_row:
        available_rows = last_data_row - effective_start_row + 1
        row_limit = available_rows if max_rows is None else min(max_rows, available_rows)
        for row_idx in range(effective_start_row, effective_start_row + row_limit):
            row_data = []
            for col in selected_columns:
                row_data.append(ws.cell(row=row_idx, column=col).value)
            rows.append(row_data)

    next_start_row = None
    if max_rows is not None and available_rows > max_rows:
        next_start_row = effective_start_row + len(rows)

    result = {
        "headers": headers,
        "rows": rows,
        "total_rows": total_rows,
        "truncated": max_rows is not None and available_rows > max_rows,
        "sheet_name": sheet_name,
    }
    payload = _compact_table_payload(result) if compact else result
    payload = augment_tabular_payload(
        payload,
        headers=headers,
        rows=rows,
        include_headers=include_headers,
        row_mode=row_mode,
        infer_schema=infer_schema,
        next_start_row=next_start_row,
    )
    if filepath is not None:
        dataset_tokens = _worksheet_dataset_tokens(
            ws,
            sheet_name=sheet_name,
            header_row=header_row,
        )
        payload = _attach_dataset_identity(
            payload,
            dataset_tokens=dataset_tokens,
            filepath=filepath,
        )
    return payload


def quick_read(
    filepath: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    start_row: Optional[int] = None,
    start_col: str = "A",
    end_col: Optional[str] = None,
    max_rows: Optional[int] = None,
    include_headers: bool = True,
    row_mode: str = "arrays",
    infer_schema: bool = False,
) -> Dict[str, Any]:
    """Read a compact table from an explicit sheet or the first sheet automatically.

    Supports array rows by default, or object-shaped records plus lightweight
    inferred schema hints when requested.
    """
    try:
        with safe_workbook(str(filepath)) as wb:
            auto_selected_sheet = sheet_name is None
            if auto_selected_sheet:
                resolved_sheet_name, ws = first_worksheet(wb, error_cls=DataError)
            else:
                resolved_sheet_name = sheet_name
                ws = require_worksheet(
                    wb,
                    resolved_sheet_name,
                    error_cls=DataError,
                    operation="cell-based operations",
                )

            result = _read_table_from_worksheet(
                ws,
                resolved_sheet_name,
                filepath=filepath,
                header_row=header_row,
                start_row=start_row,
                start_col=start_col,
                end_col=end_col,
                max_rows=max_rows,
                include_headers=include_headers,
                row_mode=row_mode,
                infer_schema=infer_schema,
            )
            result["auto_selected_sheet"] = auto_selected_sheet
            return result
    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to quick read: {e}")
        raise DataError(str(e))


def _matches_exact_query(cell_value: Any, query: Any) -> bool:
    """Match typed values while still supporting string queries from clients."""
    if cell_value == query:
        return True

    if isinstance(query, str) and not isinstance(cell_value, str):
        if isinstance(cell_value, bool):
            return str(cell_value).lower() == query.lower()
        return str(cell_value) == query

    return False


def search_cells(
    filepath: str,
    sheet_name: str,
    query: Any,
    exact: bool = True,
    max_results: int = 50,
) -> List[Dict[str, Any]]:
    """Search for cells matching a value."""
    try:
        with safe_workbook(str(filepath)) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=DataError,
                operation="cell-based operations",
            )
            results = []

            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    val = cell.value
                    if val is None:
                        continue

                    matched = False
                    if exact:
                        matched = _matches_exact_query(val, query)
                    else:
                        matched = str(query).lower() in str(val).lower()

                    if matched:
                        results.append({
                            "cell": f"{get_column_letter(col)}{row}",
                            "value": val,
                            "row": row,
                            "column": col,
                        })
                        if len(results) >= max_results:
                            return results

            return results
    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to search cells: {e}")
        raise DataError(str(e))


def append_table_rows(
    filepath: str,
    sheet_name: str,
    rows: List[Dict[str, Any]],
    header_row: int = 1,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
    expected_structure_token: Optional[str] = None,
    allow_structure_change: bool = False,
) -> Dict[str, Any]:
    """Append dictionary-shaped rows using the worksheet's header row."""
    try:
        if not rows:
            raise DataError("No rows provided to append")
        if not all(isinstance(row, dict) for row in rows):
            raise DataError("Rows must be a list of objects keyed by column name")

        with safe_workbook(str(filepath), save=not dry_run) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=DataError,
                operation="appending tabular rows",
            )
            header_map = _get_header_map(ws, header_row)
            previous_dataset_tokens = _worksheet_dataset_tokens(
                ws,
                sheet_name=sheet_name,
                header_row=header_row,
            )
            _assert_expected_structure_token(
                expected_structure_token=expected_structure_token,
                dataset_tokens=previous_dataset_tokens,
            )
            _require_structure_change_intent(
                expected_structure_token=expected_structure_token,
                allow_structure_change=allow_structure_change,
                dataset_tokens=previous_dataset_tokens,
                reason=(
                    "Appending rows changes worksheet dataset boundaries; pass "
                    "allow_structure_change=True to confirm this intentional structure change."
                ),
            )
            unknown_columns = sorted(
                {
                    key
                    for row in rows
                    for key in row.keys()
                    if key not in header_map
                }
            )
            if unknown_columns:
                raise DataError(f"Unknown columns for append: {', '.join(unknown_columns)}")

            ordered_columns = sorted(header_map.items(), key=lambda item: item[1])
            last_data_row = _find_last_data_row(ws, header_row, list(header_map.values()))
            next_row = last_data_row + 1
            requested_columns = {key for row in rows for key in row.keys()}
            native_table_conflict = _native_table_append_conflict(
                ws,
                header_row=header_row,
                next_row=next_row,
                requested_columns=requested_columns,
            )
            if native_table_conflict is not None:
                table_name = native_table_conflict["table_name"]
                if native_table_conflict["totals_row_shown"]:
                    raise DataError(
                        "append_table_rows would write directly below native Excel table "
                        f"'{table_name}' without expanding its range. Use append_excel_table_rows "
                        "for native tables instead. This table currently has a totals row enabled, "
                        "so remove the totals row before appending."
                    )
                raise DataError(
                    "append_table_rows would write directly below native Excel table "
                    f"'{table_name}' without expanding its range. Use append_excel_table_rows "
                    "to keep the native table in sync."
                )

            changes: List[Dict[str, Any]] = []
            for row_offset, row_data in enumerate(rows):
                target_row = next_row + row_offset
                for column_name, col_idx in ordered_columns:
                    if column_name not in row_data:
                        continue
                    new_value = row_data[column_name]
                    old_value = ws.cell(row=target_row, column=col_idx).value
                    if old_value != new_value:
                        changes.append(
                            _build_cell_change(
                                sheet_name=sheet_name,
                                row=target_row,
                                col=col_idx,
                                old_value=old_value,
                                new_value=new_value,
                                column_name=column_name,
                            )
                        )
                    ws.cell(row=target_row, column=col_idx, value=new_value)

            end_row = next_row + len(rows) - 1
            target_range = _range_string(
                next_row,
                min(header_map.values()),
                end_row,
                max(header_map.values()),
            )
            new_dataset_tokens = _worksheet_dataset_tokens(
                ws,
                sheet_name=sheet_name,
                header_row=header_row,
            )

        result = {
            "message": f"{'Previewed' if dry_run else 'Appended'} {len(rows)} row(s) to {sheet_name}",
            "sheet_name": sheet_name,
            "header_row": header_row,
            "rows_appended": len(rows),
            "start_row": next_row,
            "target_range": target_range,
            "changed_cells": len(changes),
            "dry_run": dry_run,
            "previous_structure_token": previous_dataset_tokens["structure_token"],
            "new_structure_token": new_dataset_tokens["structure_token"],
            "previous_content_token": previous_dataset_tokens["content_token"],
            "new_content_token": new_dataset_tokens["content_token"],
            "snapshot_metadata": _snapshot_metadata(filepath),
        }
        if _should_include_changes(dry_run, include_changes):
            result["changes"] = changes
        return result
    except (DataError, PreconditionFailedError):
        raise
    except Exception as e:
        logger.error(f"Failed to append table rows: {e}")
        raise DataError(str(e))


def update_rows_by_key(
    filepath: str,
    sheet_name: str,
    key_column: str,
    updates: List[Dict[str, Any]],
    header_row: int = 1,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
    expected_structure_token: Optional[str] = None,
) -> Dict[str, Any]:
    """Update existing rows by matching a named key column."""
    try:
        if not updates:
            raise DataError("No updates provided")
        if not all(isinstance(update, dict) for update in updates):
            raise DataError("Updates must be a list of objects keyed by column name")

        with safe_workbook(str(filepath), save=not dry_run) as wb:
            ws = require_worksheet(
                wb,
                sheet_name,
                error_cls=DataError,
                operation="updating tabular rows",
            )
            header_map = _get_header_map(ws, header_row)
            previous_dataset_tokens = _worksheet_dataset_tokens(
                ws,
                sheet_name=sheet_name,
                header_row=header_row,
            )
            _assert_expected_structure_token(
                expected_structure_token=expected_structure_token,
                dataset_tokens=previous_dataset_tokens,
            )
            if key_column not in header_map:
                raise DataError(f"Key column '{key_column}' not found")

            unknown_columns = sorted(
                {
                    key
                    for update in updates
                    for key in update.keys()
                    if key not in header_map
                }
            )
            if unknown_columns:
                raise DataError(f"Unknown columns for update: {', '.join(unknown_columns)}")

            key_col_idx = header_map[key_column]
            last_data_row = _find_last_data_row(ws, header_row, list(header_map.values()))
            row_lookup: Dict[Any, int] = {}
            for row_idx in range(header_row + 1, last_data_row + 1):
                key_value = ws.cell(row=row_idx, column=key_col_idx).value
                if key_value is None:
                    continue
                if key_value in row_lookup:
                    raise DataError(f"Duplicate key '{key_value}' found in column '{key_column}'")
                row_lookup[key_value] = row_idx

            seen_update_keys = set()
            changes: List[Dict[str, Any]] = []
            missing_keys: List[Any] = []
            matched_keys: List[Any] = []

            for update in updates:
                if key_column not in update:
                    raise DataError(f"Update row is missing key column '{key_column}'")

                key_value = update[key_column]
                if key_value in seen_update_keys:
                    raise DataError(f"Duplicate update key '{key_value}' provided")
                seen_update_keys.add(key_value)

                target_row = row_lookup.get(key_value)
                if target_row is None:
                    missing_keys.append(key_value)
                    continue

                matched_keys.append(key_value)
                for column_name, new_value in update.items():
                    if column_name == key_column:
                        continue

                    col_idx = header_map[column_name]
                    cell = ws.cell(row=target_row, column=col_idx)
                    old_value = cell.value
                    if old_value == new_value:
                        continue

                    changes.append(
                        _build_cell_change(
                            sheet_name=sheet_name,
                            row=target_row,
                            col=col_idx,
                            old_value=old_value,
                            new_value=new_value,
                            column_name=column_name,
                        )
                    )
                    cell.value = new_value

            new_dataset_tokens = _worksheet_dataset_tokens(
                ws,
                sheet_name=sheet_name,
                header_row=header_row,
            )

        updated_rows = len(matched_keys)
        message = (
            f"{'Previewed' if dry_run else 'Updated'} {updated_rows} row(s) in {sheet_name}"
        )
        if missing_keys:
            message += f"; {len(missing_keys)} key(s) not found"

        result = {
            "message": message,
            "sheet_name": sheet_name,
            "key_column": key_column,
            "header_row": header_row,
            "updated_rows": updated_rows,
            "missing_keys": missing_keys,
            "changed_cells": len(changes),
            "dry_run": dry_run,
            "previous_structure_token": previous_dataset_tokens["structure_token"],
            "new_structure_token": new_dataset_tokens["structure_token"],
            "previous_content_token": previous_dataset_tokens["content_token"],
            "new_content_token": new_dataset_tokens["content_token"],
            "snapshot_metadata": _snapshot_metadata(filepath),
        }
        if _should_include_changes(dry_run, include_changes):
            result["changes"] = changes
        return result
    except (DataError, PreconditionFailedError):
        raise
    except Exception as e:
        logger.error(f"Failed to update rows by key: {e}")
        raise DataError(str(e))

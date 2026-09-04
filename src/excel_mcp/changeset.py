"""Preview and commit guarded multi-operation workbook changesets."""

from __future__ import annotations

from contextlib import suppress
from datetime import datetime, timezone
import hashlib
import hmac
import json
import logging
import math
import os
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from . import workbook as workbook_module
from .chart import (
    _chart_type_name,
    _extract_chart_anchor,
    _extract_chart_dimensions,
    _extract_title_text,
    create_chart_in_sheet,
)
from .data import write_data
from .exceptions import (
    PreconditionFailedError,
    ResponseTooLargeError,
    ValidationError,
    WorkbookError,
)
from .formatting import format_range, format_ranges
from .sheet import (
    autofit_columns,
    set_auto_filter,
    set_column_widths,
    set_freeze_panes,
    set_row_heights,
)
from .tables import create_excel_table
from .workbook import create_sheet, create_workbook_snapshot, diff_workbooks

logger = logging.getLogger(__name__)

CHANGESET_VERSION = 1
MAX_CHANGESET_OPERATIONS = 50
MAX_CHANGESET_ASSERTIONS = 50
MAX_CHANGESET_PLAN_BYTES = 2_000_000
MAX_WRITE_CELLS = 50_000
MAX_FORMAT_CELLS = 250_000
MAX_ASSERTED_RANGE_CELLS = 10_000
MAX_ASSERTION_SCAN_CELLS = 250_000
MAX_CHANGESET_RESPONSE_BYTES = 40_000
MAX_RESPONSE_STRING_LENGTH = 500
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384

_SHA256_RE = re.compile(r"^[0-9a-fA-F]{64}$")
_TOKEN_RE = re.compile(r"^changeset_v1_[0-9a-f]{64}$")
_FORMAT_FIELDS = {
    "start_cell",
    "end_cell",
    "bold",
    "italic",
    "underline",
    "font_size",
    "font_color",
    "bg_color",
    "border_style",
    "border_color",
    "number_format",
    "alignment",
    "wrap_text",
    "merge_cells",
    "protection",
    "conditional_format",
}

_OPERATION_SPECS: dict[str, tuple[set[str], set[str]]] = {
    "create_worksheet": ({"sheet_name"}, {"sheet_name"}),
    "write_data_to_excel": (
        {"sheet_name", "data", "start_cell"},
        {"data"},
    ),
    "format_range": (
        {"sheet_name", *_FORMAT_FIELDS},
        {"sheet_name", "start_cell"},
    ),
    "format_ranges": (
        {"sheet_name", "ranges"},
        {"sheet_name", "ranges"},
    ),
    "freeze_panes": ({"sheet_name", "cell"}, {"sheet_name", "cell"}),
    "set_autofilter": ({"sheet_name", "range_ref"}, {"sheet_name"}),
    "set_column_widths": ({"sheet_name", "widths"}, {"sheet_name", "widths"}),
    "set_row_heights": ({"sheet_name", "heights"}, {"sheet_name", "heights"}),
    "autofit_columns": (
        {"sheet_name", "columns", "min_width", "max_width", "padding"},
        {"sheet_name"},
    ),
    "create_table": (
        {"sheet_name", "data_range", "table_name", "table_style"},
        {"sheet_name", "data_range", "table_name"},
    ),
    "create_chart": (
        {
            "sheet_name",
            "data_range",
            "chart_type",
            "target_cell",
            "title",
            "x_axis",
            "y_axis",
            "style",
            "series",
            "categories_range",
            "width",
            "height",
            "placement",
        },
        {"sheet_name", "chart_type"},
    ),
}

_ASSERTION_SPECS: dict[str, tuple[set[str], set[str]]] = {
    "sheet_exists": (
        {"type", "sheet_name", "sheet_type"},
        {"type", "sheet_name"},
    ),
    "cell_equals": (
        {"type", "sheet_name", "cell", "expected"},
        {"type", "sheet_name", "cell", "expected"},
    ),
    "range_equals": (
        {"type", "sheet_name", "range_ref", "expected"},
        {"type", "sheet_name", "range_ref", "expected"},
    ),
    "range_values_unchanged": (
        {"type", "sheet_name", "range_ref"},
        {"type", "sheet_name", "range_ref"},
    ),
    "table_exists": (
        {"type", "table_name", "sheet_name", "range_ref"},
        {"type", "table_name"},
    ),
    "freeze_panes_equals": (
        {"type", "sheet_name", "cell"},
        {"type", "sheet_name", "cell"},
    ),
    "autofilter_equals": (
        {"type", "sheet_name", "range_ref"},
        {"type", "sheet_name", "range_ref"},
    ),
    "chart_exists": (
        {
            "type",
            "sheet_name",
            "chart_type",
            "target_cell",
            "title",
            "width",
            "height",
        },
        {"type", "sheet_name"},
    ),
    "no_cell_ref_errors": (
        {"type", "sheet_name", "range_ref"},
        {"type"},
    ),
}


def _canonical_json(value: Any) -> str:
    try:
        return json.dumps(
            value,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=False,
            allow_nan=False,
        )
    except (TypeError, ValueError) as exc:
        raise ValidationError(
            "ChangeSet operations and assertions must contain only finite JSON values"
        ) from exc


def _require_json_keys(value: Any, *, location: str) -> None:
    if isinstance(value, dict):
        for key, item in value.items():
            if not isinstance(key, str):
                raise ValidationError(f"{location} contains a non-string object key")
            _require_json_keys(item, location=location)
    elif isinstance(value, list):
        for item in value:
            _require_json_keys(item, location=location)
    elif isinstance(value, tuple):
        for item in value:
            _require_json_keys(item, location=location)


def _json_copy(value: Any, *, location: str) -> Any:
    _require_json_keys(value, location=location)
    return json.loads(_canonical_json(value))


def _validate_non_empty_string(value: Any, *, location: str) -> None:
    if not isinstance(value, str) or not value.strip():
        raise ValidationError(f"{location} must be a non-empty string")


def _range_bounds(range_ref: Any, *, location: str) -> tuple[int, int, int, int]:
    _validate_non_empty_string(range_ref, location=location)
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    except (TypeError, ValueError) as exc:
        raise ValidationError(f"{location} must be a valid A1 range") from exc
    if None in (min_col, min_row, max_col, max_row):
        raise ValidationError(f"{location} must include both row and column coordinates")
    if min_col > max_col or min_row > max_row:
        raise ValidationError(f"{location} must run from top-left to bottom-right")
    if (
        min_col < 1
        or min_row < 1
        or max_col > EXCEL_MAX_COLUMNS
        or max_row > EXCEL_MAX_ROWS
    ):
        raise ValidationError(
            f"{location} must stay inside Excel's A1:XFD{EXCEL_MAX_ROWS} grid"
        )
    return min_col, min_row, max_col, max_row


def _range_cell_count(range_ref: Any, *, location: str) -> int:
    min_col, min_row, max_col, max_row = _range_bounds(range_ref, location=location)
    return (max_col - min_col + 1) * (max_row - min_row + 1)


def _single_cell_bounds(cell_ref: Any, *, location: str) -> tuple[int, int]:
    min_col, min_row, max_col, max_row = _range_bounds(cell_ref, location=location)
    if min_col != max_col or min_row != max_row:
        raise ValidationError(f"{location} must identify exactly one cell")
    return min_col, min_row


def _canonical_range_ref(range_ref: Any, *, location: str) -> str:
    min_col, min_row, max_col, max_row = _range_bounds(
        range_ref, location=location
    )
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _canonical_freeze_cell(cell_ref: Any, *, location: str) -> str | None:
    if cell_ref is None or cell_ref == "":
        return None
    canonical = _canonical_range_ref(cell_ref, location=location)
    if ":" in canonical:
        raise ValidationError(f"{location} must identify exactly one cell")
    return None if canonical == "A1" else canonical


def _validate_positive_number(value: Any, *, location: str) -> float:
    if (
        isinstance(value, bool)
        or not isinstance(value, (int, float))
        or not math.isfinite(value)
        or value <= 0
    ):
        raise ValidationError(f"{location} must be a positive finite number")
    return float(value)


def _format_range_ref(operation: dict[str, Any], *, location: str) -> str:
    start_cell = operation.get("start_cell")
    _single_cell_bounds(start_cell, location=f"{location}.start_cell")
    end_cell = operation.get("end_cell")
    if end_cell is None:
        return str(start_cell)
    _single_cell_bounds(end_cell, location=f"{location}.end_cell")
    return f"{start_cell}:{end_cell}"


def _normalize_operations(operations: Any) -> list[dict[str, Any]]:
    if not isinstance(operations, list) or not operations:
        raise ValidationError("operations must be a non-empty list")
    if len(operations) > MAX_CHANGESET_OPERATIONS:
        raise ValidationError(
            f"A ChangeSet supports at most {MAX_CHANGESET_OPERATIONS} operations"
        )

    normalized: list[dict[str, Any]] = []
    write_cells = 0
    format_cells = 0
    for index, raw_operation in enumerate(operations, start=1):
        location = f"operations[{index - 1}]"
        if not isinstance(raw_operation, dict):
            raise ValidationError(f"{location} must be an object")
        unknown_envelope_fields = set(raw_operation) - {"tool", "args"}
        if unknown_envelope_fields:
            raise ValidationError(
                f"{location} has unsupported fields: "
                + ", ".join(sorted(unknown_envelope_fields))
            )

        tool = raw_operation.get("tool")
        _validate_non_empty_string(tool, location=f"{location}.tool")
        if tool not in _OPERATION_SPECS:
            raise ValidationError(
                f"Unsupported ChangeSet tool '{tool}'. Supported tools: "
                + ", ".join(sorted(_OPERATION_SPECS))
            )

        args = raw_operation.get("args", {})
        if not isinstance(args, dict):
            raise ValidationError(f"{location}.args must be an object")
        allowed_fields, required_fields = _OPERATION_SPECS[tool]
        unknown_fields = set(args) - allowed_fields
        if unknown_fields:
            raise ValidationError(
                f"{location}.args has unsupported fields for {tool}: "
                + ", ".join(sorted(unknown_fields))
            )
        missing_fields = required_fields - set(args)
        if missing_fields:
            raise ValidationError(
                f"{location}.args is missing required fields for {tool}: "
                + ", ".join(sorted(missing_fields))
            )

        normalized_args = _json_copy(args, location=f"{location}.args")
        if tool == "write_data_to_excel":
            data = normalized_args.get("data")
            if not isinstance(data, list) or not data:
                raise ValidationError(f"{location}.args.data must be a non-empty 2D list")
            for row_index, row in enumerate(data):
                if not isinstance(row, list):
                    raise ValidationError(
                        f"{location}.args.data[{row_index}] must be a list"
                    )
                write_cells += len(row)
            if "start_cell" in normalized_args:
                _single_cell_bounds(
                    normalized_args["start_cell"],
                    location=f"{location}.args.start_cell",
                )
        elif tool == "format_range":
            if normalized_args.get("merge_cells"):
                raise ValidationError(
                    "ChangeSet format_range does not support merge_cells; use an explicit "
                    "merge workflow outside the transaction"
                )
            format_cells += _range_cell_count(
                _format_range_ref(normalized_args, location=f"{location}.args"),
                location=f"{location}.args range",
            )
        elif tool == "format_ranges":
            ranges = normalized_args.get("ranges")
            if not isinstance(ranges, list) or not ranges:
                raise ValidationError(f"{location}.args.ranges must be a non-empty list")
            for range_index, range_operation in enumerate(ranges):
                range_location = f"{location}.args.ranges[{range_index}]"
                if not isinstance(range_operation, dict):
                    raise ValidationError(f"{range_location} must be an object")
                unknown_range_fields = set(range_operation) - _FORMAT_FIELDS
                if unknown_range_fields:
                    raise ValidationError(
                        f"{range_location} has unsupported fields: "
                        + ", ".join(sorted(unknown_range_fields))
                    )
                if "start_cell" not in range_operation:
                    raise ValidationError(f"{range_location}.start_cell is required")
                if range_operation.get("merge_cells"):
                    raise ValidationError(
                        "ChangeSet format_ranges does not support merge_cells; use an explicit "
                        "merge workflow outside the transaction"
                    )
                format_cells += _range_cell_count(
                    _format_range_ref(range_operation, location=range_location),
                    location=f"{range_location} range",
                )
        elif tool == "create_table":
            _validate_non_empty_string(
                normalized_args.get("table_name"),
                location=f"{location}.args.table_name",
            )
            _range_bounds(
                normalized_args.get("data_range"),
                location=f"{location}.args.data_range",
            )
        elif tool == "freeze_panes" and normalized_args.get("cell") not in (
            None,
            "",
        ):
            _single_cell_bounds(
                normalized_args["cell"], location=f"{location}.args.cell"
            )
        elif tool == "set_autofilter" and normalized_args.get("range_ref") is not None:
            _range_bounds(
                normalized_args["range_ref"],
                location=f"{location}.args.range_ref",
            )

        normalized.append({"tool": tool, "args": normalized_args})

    if write_cells > MAX_WRITE_CELLS:
        raise ValidationError(
            f"ChangeSet writes are limited to {MAX_WRITE_CELLS} cells; received {write_cells}"
        )
    if format_cells > MAX_FORMAT_CELLS:
        raise ValidationError(
            f"ChangeSet formatting is limited to {MAX_FORMAT_CELLS} cells; received {format_cells}"
        )
    return normalized


def _normalize_assertions(assertions: Any) -> list[dict[str, Any]]:
    if assertions is None:
        return []
    if not isinstance(assertions, list):
        raise ValidationError("assertions must be a list")
    if len(assertions) > MAX_CHANGESET_ASSERTIONS:
        raise ValidationError(
            f"A ChangeSet supports at most {MAX_CHANGESET_ASSERTIONS} assertions"
        )

    normalized: list[dict[str, Any]] = []
    for index, raw_assertion in enumerate(assertions):
        location = f"assertions[{index}]"
        if not isinstance(raw_assertion, dict):
            raise ValidationError(f"{location} must be an object")
        assertion_type = raw_assertion.get("type")
        _validate_non_empty_string(assertion_type, location=f"{location}.type")
        if assertion_type not in _ASSERTION_SPECS:
            raise ValidationError(
                f"Unsupported assertion type '{assertion_type}'. Supported assertions: "
                + ", ".join(sorted(_ASSERTION_SPECS))
            )

        allowed_fields, required_fields = _ASSERTION_SPECS[assertion_type]
        unknown_fields = set(raw_assertion) - allowed_fields
        if unknown_fields:
            raise ValidationError(
                f"{location} has unsupported fields: "
                + ", ".join(sorted(unknown_fields))
            )
        missing_fields = required_fields - set(raw_assertion)
        if missing_fields:
            raise ValidationError(
                f"{location} is missing required fields: "
                + ", ".join(sorted(missing_fields))
            )

        assertion = _json_copy(raw_assertion, location=location)
        if "sheet_name" in assertion:
            _validate_non_empty_string(
                assertion["sheet_name"], location=f"{location}.sheet_name"
            )
        if assertion_type == "sheet_exists":
            sheet_type = assertion.get("sheet_type", "any")
            if sheet_type not in {"any", "worksheet", "chartsheet"}:
                raise ValidationError(
                    f"{location}.sheet_type must be any, worksheet, or chartsheet"
                )
        elif assertion_type == "cell_equals":
            assertion["cell"] = _canonical_range_ref(
                assertion["cell"], location=f"{location}.cell"
            )
            if ":" in assertion["cell"]:
                raise ValidationError(f"{location}.cell must identify exactly one cell")
        elif assertion_type in {"range_equals", "range_values_unchanged"}:
            cell_count = _range_cell_count(
                assertion["range_ref"], location=f"{location}.range_ref"
            )
            if cell_count > MAX_ASSERTED_RANGE_CELLS:
                raise ValidationError(
                    f"{location}.range_ref exceeds the {MAX_ASSERTED_RANGE_CELLS}-cell assertion limit"
                )
            if assertion_type == "range_equals":
                expected = assertion["expected"]
                if not isinstance(expected, list) or not expected:
                    raise ValidationError(f"{location}.expected must be a non-empty 2D list")
                min_col, min_row, max_col, max_row = _range_bounds(
                    assertion["range_ref"], location=f"{location}.range_ref"
                )
                expected_rows = max_row - min_row + 1
                expected_cols = max_col - min_col + 1
                if len(expected) != expected_rows or any(
                    not isinstance(row, list) or len(row) != expected_cols
                    for row in expected
                ):
                    raise ValidationError(
                        f"{location}.expected shape must match {assertion['range_ref']} "
                        f"({expected_rows} row(s) x {expected_cols} column(s))"
                    )
            assertion["range_ref"] = _canonical_range_ref(
                assertion["range_ref"], location=f"{location}.range_ref"
            )
        elif assertion_type == "table_exists":
            _validate_non_empty_string(
                assertion["table_name"], location=f"{location}.table_name"
            )
            if "range_ref" in assertion:
                assertion["range_ref"] = _canonical_range_ref(
                    assertion["range_ref"], location=f"{location}.range_ref"
                )
        elif assertion_type == "freeze_panes_equals":
            assertion["cell"] = _canonical_freeze_cell(
                assertion["cell"], location=f"{location}.cell"
            )
        elif assertion_type == "autofilter_equals":
            assertion["range_ref"] = _canonical_range_ref(
                assertion["range_ref"], location=f"{location}.range_ref"
            )
        elif assertion_type == "chart_exists":
            if "chart_type" in assertion:
                _validate_non_empty_string(
                    assertion["chart_type"], location=f"{location}.chart_type"
                )
                assertion["chart_type"] = assertion["chart_type"].strip().lower()
            if "target_cell" in assertion:
                assertion["target_cell"] = _canonical_range_ref(
                    assertion["target_cell"], location=f"{location}.target_cell"
                )
                if ":" in assertion["target_cell"]:
                    raise ValidationError(
                        f"{location}.target_cell must identify exactly one cell"
                    )
            if "title" in assertion and not isinstance(assertion["title"], str):
                raise ValidationError(f"{location}.title must be a string")
            for dimension in ("width", "height"):
                if dimension in assertion:
                    assertion[dimension] = _validate_positive_number(
                        assertion[dimension], location=f"{location}.{dimension}"
                    )
        elif assertion_type == "no_cell_ref_errors" and "range_ref" in assertion:
            cell_count = _range_cell_count(
                assertion["range_ref"], location=f"{location}.range_ref"
            )
            if cell_count > MAX_ASSERTED_RANGE_CELLS:
                raise ValidationError(
                    f"{location}.range_ref exceeds the {MAX_ASSERTED_RANGE_CELLS}-cell assertion limit"
                )
            assertion["range_ref"] = _canonical_range_ref(
                assertion["range_ref"], location=f"{location}.range_ref"
            )

        normalized.append(assertion)
    return normalized


def _normalize_sha256(value: Any, *, required: bool) -> str | None:
    if value is None:
        if required:
            raise ValidationError("expected_workbook_sha256 is required in commit mode")
        return None
    if not isinstance(value, str) or not _SHA256_RE.fullmatch(value.strip()):
        raise ValidationError("expected_workbook_sha256 must be a 64-character SHA-256 hex digest")
    return value.strip().lower()


def _normalize_changeset_token(value: Any, *, required: bool) -> str | None:
    if value is None:
        if required:
            raise ValidationError("changeset_token is required in commit mode")
        return None
    if not isinstance(value, str) or not _TOKEN_RE.fullmatch(value.strip()):
        raise ValidationError("changeset_token must be a valid changeset_v1 token")
    return value.strip().lower()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _new_temp_workbook(parent: Path, *, source: Path, label: str) -> Path:
    fd, name = tempfile.mkstemp(
        prefix=f".{source.stem}.sheetforge-changeset-{label}-",
        suffix=source.suffix or ".xlsx",
        dir=str(parent),
    )
    os.close(fd)
    return Path(name)


def _stage_workbook(source: Path, baseline: Path, candidate: Path) -> str:
    source_stat_before = source.stat()
    digest = hashlib.sha256()
    with source.open("rb") as source_handle, baseline.open("wb") as baseline_handle:
        while chunk := source_handle.read(1024 * 1024):
            baseline_handle.write(chunk)
            digest.update(chunk)
        baseline_handle.flush()
        os.fsync(baseline_handle.fileno())

    source_stat_after = source.stat()
    if workbook_module._file_identity(source_stat_before) != workbook_module._file_identity(
        source_stat_after
    ):
        raise PreconditionFailedError(
            "Workbook changed while the ChangeSet baseline was being prepared; retry from a fresh preview",
            code="stale_workbook",
            suggested_next_tool="apply_workbook_changeset",
        )

    with suppress(OSError):
        os.chmod(baseline, source_stat_before.st_mode)
    workbook_module._verify_saved_workbook(str(baseline))
    shutil.copy2(baseline, candidate)
    workbook_module._fsync_file(candidate)
    workbook_module._verify_saved_workbook(str(candidate))
    return digest.hexdigest()


def _plan_payload(
    *,
    source: Path,
    source_sha256: str,
    operations: list[dict[str, Any]],
    assertions: list[dict[str, Any]],
    create_snapshot: bool,
    snapshot_filepath: Path | None,
) -> dict[str, Any]:
    return {
        "version": CHANGESET_VERSION,
        "target_path": str(source),
        "source_sha256": source_sha256,
        "operations": operations,
        "assertions": assertions,
        "snapshot": {
            "enabled": create_snapshot,
            "path": str(snapshot_filepath) if snapshot_filepath is not None else None,
        },
    }


def _changeset_token(plan: dict[str, Any]) -> str:
    encoded = _canonical_json(plan).encode("utf-8")
    if len(encoded) > MAX_CHANGESET_PLAN_BYTES:
        raise ValidationError(
            f"ChangeSet plan exceeds the {MAX_CHANGESET_PLAN_BYTES:,}-byte limit"
        )
    digest = hashlib.sha256(b"sheetforge-changeset-v1\0" + encoded).hexdigest()
    return f"changeset_v1_{digest}"


def _snapshot_destination(
    source: Path,
    *,
    create_snapshot: bool,
    requested_path: Path | None,
    changeset_token: str,
) -> Path | None:
    if not create_snapshot:
        return None
    if requested_path is not None:
        return requested_path
    token_suffix = changeset_token.removeprefix("changeset_v1_")[:12]
    return source.with_name(
        f"{source.stem}.sheetforge-before-{token_suffix}{source.suffix or '.xlsx'}"
    )


def _snapshot_status(path: Path | None, *, source_sha256: str) -> dict[str, Any]:
    if path is None:
        return {"enabled": False, "path": None, "status": "disabled"}
    if not path.exists() and not path.is_symlink():
        return {"enabled": True, "path": str(path), "status": "available"}
    if path.is_symlink():
        return {
            "enabled": True,
            "path": str(path),
            "status": "conflict",
            "reason": "Snapshot destination must not be a symbolic link",
        }
    if not path.is_file():
        return {
            "enabled": True,
            "path": str(path),
            "status": "conflict",
            "reason": "Snapshot destination exists but is not a regular file",
        }
    try:
        workbook_module._verify_saved_workbook(str(path))
        existing_sha256 = _sha256_file(path)
    except Exception as exc:
        return {
            "enabled": True,
            "path": str(path),
            "status": "conflict",
            "reason": f"Unable to inspect existing snapshot: {exc!s}",
        }
    if existing_sha256 == source_sha256:
        return {
            "enabled": True,
            "path": str(path),
            "status": "existing_matching",
            "sha256": existing_sha256,
        }
    return {
        "enabled": True,
        "path": str(path),
        "status": "conflict",
        "sha256": existing_sha256,
        "reason": "Snapshot destination already exists with different content",
    }


def _compact_operation_result(result: Any, *, sample_limit: int) -> Any:
    if not isinstance(result, dict):
        return result
    compact = {
        key: value
        for key, value in result.items()
        if key not in {"message", "changes", "preview", "warnings", "errors", "dry_run"}
    }
    for key in ("ranges",):
        items = compact.get(key)
        if isinstance(items, list) and len(items) > sample_limit:
            compact[key] = items[:sample_limit]
            compact[f"{key}_count"] = len(items)
            compact[f"{key}_truncated"] = True
    for key in ("widths", "heights"):
        items = compact.get(key)
        if isinstance(items, dict) and len(items) > sample_limit:
            sampled_keys = sorted(items, key=str)[:sample_limit]
            compact[key] = {item_key: items[item_key] for item_key in sampled_keys}
            compact[f"{key}_count"] = len(items)
            compact[f"{key}_truncated"] = True
    return _bound_response_value(compact)


def _bound_response_value(value: Any) -> Any:
    if isinstance(value, str) and len(value) > MAX_RESPONSE_STRING_LENGTH:
        return {
            "value_preview": value[:MAX_RESPONSE_STRING_LENGTH],
            "value_length": len(value),
            "value_sha256": hashlib.sha256(value.encode("utf-8")).hexdigest(),
            "truncated": True,
        }
    if isinstance(value, list):
        return [_bound_response_value(item) for item in value]
    if isinstance(value, dict):
        return {key: _bound_response_value(item) for key, item in value.items()}
    return value


def _assert_response_budget(payload: dict[str, Any]) -> None:
    encoded = json.dumps(
        payload,
        separators=(",", ":"),
        ensure_ascii=False,
        default=str,
    ).encode("utf-8")
    if len(encoded) > MAX_CHANGESET_RESPONSE_BYTES:
        raise ResponseTooLargeError(
            (
                f"ChangeSet result would be {len(encoded):,} bytes, exceeding the "
                f"pre-commit safety limit of {MAX_CHANGESET_RESPONSE_BYTES:,}."
            ),
            estimated_size=len(encoded),
            limit=MAX_CHANGESET_RESPONSE_BYTES,
            hints=[
                "set sample_limit to a smaller value",
                "split the ChangeSet into smaller verified transactions",
            ],
        )


def _record_cell_changes(
    aggregate: dict[tuple[str, str], dict[str, Any]],
    result: Any,
) -> None:
    if not isinstance(result, dict) or not isinstance(result.get("changes"), list):
        return
    for change in result["changes"]:
        if not isinstance(change, dict):
            continue
        sheet_name = change.get("sheet_name")
        cell = change.get("cell")
        if not isinstance(sheet_name, str) or not isinstance(cell, str):
            continue
        if "old_value" not in change or "new_value" not in change:
            continue
        key = (sheet_name, cell)
        if key not in aggregate:
            aggregate[key] = {
                "sheet_name": sheet_name,
                "cell": cell,
                "before": change["old_value"],
                "after": change["new_value"],
            }
        else:
            aggregate[key]["after"] = change["new_value"]


def _dispatch_operation(candidate: Path, operation: dict[str, Any]) -> dict[str, Any]:
    tool = operation["tool"]
    args = dict(operation["args"])
    filepath = str(candidate)
    if tool == "create_chart":
        args.setdefault("data_range", None)

    dispatch: dict[str, Callable[[], dict[str, Any]]] = {
        "create_worksheet": lambda: create_sheet(filepath, **args),
        "write_data_to_excel": lambda: write_data(
            filepath,
            dry_run=False,
            include_changes=True,
            **args,
        ),
        "format_range": lambda: format_range(
            filepath,
            dry_run=False,
            include_changes=True,
            **args,
        ),
        "format_ranges": lambda: format_ranges(
            filepath,
            dry_run=False,
            include_changes=True,
            **args,
        ),
        "freeze_panes": lambda: set_freeze_panes(
            filepath,
            dry_run=False,
            include_changes=True,
            **args,
        ),
        "set_autofilter": lambda: set_auto_filter(
            filepath,
            dry_run=False,
            include_changes=True,
            **args,
        ),
        "set_column_widths": lambda: set_column_widths(
            filepath,
            dry_run=False,
            include_changes=True,
            **args,
        ),
        "set_row_heights": lambda: set_row_heights(
            filepath,
            dry_run=False,
            include_changes=True,
            **args,
        ),
        "autofit_columns": lambda: autofit_columns(filepath, dry_run=False, **args),
        "create_table": lambda: create_excel_table(filepath, **args),
        "create_chart": lambda: create_chart_in_sheet(filepath, **args),
    }
    result = dispatch[tool]()
    if tool == "format_ranges" and int(result.get("ranges_failed", 0)) > 0:
        errors = result.get("errors", [])
        sampled_errors = errors[:10]
        error_summary = "; ".join(
            f"range {item.get('range')}: {item.get('error')}"
            for item in sampled_errors
        )
        if len(errors) > len(sampled_errors):
            error_summary += f"; and {len(errors) - len(sampled_errors)} more failure(s)"
        raise ValidationError(
            "ChangeSet format_ranges failed atomically: " + error_summary
        )
    return result


def _apply_operations(
    candidate: Path,
    operations: list[dict[str, Any]],
    *,
    sample_limit: int,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    operation_results: list[dict[str, Any]] = []
    cell_change_map: dict[tuple[str, str], dict[str, Any]] = {}
    for index, operation in enumerate(operations, start=1):
        result = _dispatch_operation(candidate, operation)
        _record_cell_changes(cell_change_map, result)
        operation_results.append(
            {
                "index": index,
                "tool": operation["tool"],
                "result": _compact_operation_result(
                    result, sample_limit=sample_limit
                ),
            }
        )

    effective_changes = [
        _bound_response_value(change)
        for change in cell_change_map.values()
        if change["before"] != change["after"]
    ]
    return operation_results, {
        "count": len(effective_changes),
        "changes": effective_changes,
    }


def _worksheet_or_none(wb: Any, sheet_name: str) -> Worksheet | None:
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    return ws if isinstance(ws, Worksheet) else None


def _range_values(ws: Worksheet, range_ref: str) -> list[list[Any]]:
    min_col, min_row, max_col, max_row = _range_bounds(
        range_ref, location="assertion range"
    )
    return [
        [ws.cell(row=row, column=column).value for column in range(min_col, max_col + 1)]
        for row in range(min_row, max_row + 1)
    ]


def _cell_values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, bool) or isinstance(right, bool):
        return type(left) is type(right) and left == right
    return left == right


def _assert_sheet_exists(wb: Any, assertion: dict[str, Any]) -> tuple[bool, dict[str, Any]]:
    sheet_name = assertion["sheet_name"]
    if sheet_name not in wb.sheetnames:
        return False, {"reason": f"Sheet '{sheet_name}' does not exist"}
    ws = wb[sheet_name]
    actual_type = "worksheet" if isinstance(ws, Worksheet) else "chartsheet"
    expected_type = assertion.get("sheet_type", "any")
    passed = expected_type == "any" or expected_type == actual_type
    return passed, {
        "sheet_name": sheet_name,
        "expected_type": expected_type,
        "actual_type": actual_type,
    }


def _assert_cell_equals(wb: Any, assertion: dict[str, Any]) -> tuple[bool, dict[str, Any]]:
    ws = _worksheet_or_none(wb, assertion["sheet_name"])
    if ws is None:
        return False, {"reason": f"Worksheet '{assertion['sheet_name']}' does not exist"}
    min_col, min_row, _, _ = _range_bounds(
        assertion["cell"], location="assertion cell"
    )
    actual = ws.cell(row=min_row, column=min_col).value
    return _cell_values_equal(actual, assertion["expected"]), {
        "sheet_name": assertion["sheet_name"],
        "cell": assertion["cell"],
        "expected": _bound_response_value(assertion["expected"]),
        "actual": _bound_response_value(actual),
    }


def _range_comparison_details(
    before: list[list[Any]],
    after: list[list[Any]],
    *,
    sheet_name: str,
    range_ref: str,
    before_label: str,
    after_label: str,
    sample_limit: int,
) -> tuple[bool, dict[str, Any]]:
    min_col, min_row, _, _ = _range_bounds(range_ref, location="assertion range")
    mismatch_count = 0
    mismatches: list[dict[str, Any]] = []
    for row_offset, (before_row, after_row) in enumerate(zip(before, after)):
        for column_offset, (before_value, after_value) in enumerate(
            zip(before_row, after_row)
        ):
            if _cell_values_equal(before_value, after_value):
                continue
            mismatch_count += 1
            if len(mismatches) < sample_limit:
                cell = get_column_letter(min_col + column_offset)
                mismatches.append(
                    {
                        "cell": f"{cell}{min_row + row_offset}",
                        before_label: _bound_response_value(before_value),
                        after_label: _bound_response_value(after_value),
                    }
                )
    return mismatch_count == 0, {
        "sheet_name": sheet_name,
        "range": range_ref,
        "mismatch_count": mismatch_count,
        "mismatches": mismatches,
        "truncated": mismatch_count > sample_limit,
    }


def _assert_range_equals(
    wb: Any,
    assertion: dict[str, Any],
    *,
    sample_limit: int,
) -> tuple[bool, dict[str, Any]]:
    ws = _worksheet_or_none(wb, assertion["sheet_name"])
    if ws is None:
        return False, {"reason": f"Worksheet '{assertion['sheet_name']}' does not exist"}
    actual = _range_values(ws, assertion["range_ref"])
    return _range_comparison_details(
        assertion["expected"],
        actual,
        sheet_name=assertion["sheet_name"],
        range_ref=assertion["range_ref"],
        before_label="expected",
        after_label="actual",
        sample_limit=sample_limit,
    )


def _assert_range_values_unchanged(
    baseline_wb: Any,
    candidate_wb: Any,
    assertion: dict[str, Any],
    *,
    sample_limit: int,
) -> tuple[bool, dict[str, Any]]:
    before_ws = _worksheet_or_none(baseline_wb, assertion["sheet_name"])
    after_ws = _worksheet_or_none(candidate_wb, assertion["sheet_name"])
    if before_ws is None or after_ws is None:
        return False, {
            "reason": f"Worksheet '{assertion['sheet_name']}' must exist in both workbook states"
        }
    before = _range_values(before_ws, assertion["range_ref"])
    after = _range_values(after_ws, assertion["range_ref"])
    return _range_comparison_details(
        before,
        after,
        sheet_name=assertion["sheet_name"],
        range_ref=assertion["range_ref"],
        before_label="before",
        after_label="after",
        sample_limit=sample_limit,
    )


def _assert_table_exists(wb: Any, assertion: dict[str, Any]) -> tuple[bool, dict[str, Any]]:
    expected_name = assertion["table_name"]
    expected_sheet = assertion.get("sheet_name")
    expected_range = assertion.get("range_ref")
    matches: list[dict[str, str]] = []
    for ws in wb.worksheets:
        if expected_sheet is not None and ws.title != expected_sheet:
            continue
        for table in ws.tables.values():
            if table.displayName == expected_name:
                matches.append(
                    {
                        "sheet_name": ws.title,
                        "table_name": table.displayName,
                        "range": table.ref,
                    }
                )
    passed = bool(matches) and (
        expected_range is None or any(item["range"] == expected_range for item in matches)
    )
    return passed, {
        "table_name": expected_name,
        "expected_sheet": expected_sheet,
        "expected_range": expected_range,
        "matches": matches,
    }


def _assert_freeze_panes_equals(
    wb: Any,
    assertion: dict[str, Any],
) -> tuple[bool, dict[str, Any]]:
    ws = _worksheet_or_none(wb, assertion["sheet_name"])
    if ws is None:
        return False, {"reason": f"Worksheet '{assertion['sheet_name']}' does not exist"}
    actual_value = ws.freeze_panes
    if hasattr(actual_value, "coordinate"):
        actual_value = actual_value.coordinate
    actual = _canonical_freeze_cell(actual_value, location="worksheet freeze panes")
    expected = assertion["cell"]
    return actual == expected, {
        "sheet_name": assertion["sheet_name"],
        "expected_cell": expected,
        "actual_cell": actual,
    }


def _assert_autofilter_equals(
    wb: Any,
    assertion: dict[str, Any],
) -> tuple[bool, dict[str, Any]]:
    ws = _worksheet_or_none(wb, assertion["sheet_name"])
    if ws is None:
        return False, {"reason": f"Worksheet '{assertion['sheet_name']}' does not exist"}
    actual_ref = ws.auto_filter.ref
    actual = (
        _canonical_range_ref(actual_ref, location="worksheet autofilter")
        if actual_ref
        else None
    )
    expected = assertion["range_ref"]
    return actual == expected, {
        "sheet_name": assertion["sheet_name"],
        "expected_range": expected,
        "actual_range": actual,
    }


def _chart_matches_assertion(
    chart: Any,
    assertion: dict[str, Any],
) -> tuple[bool, dict[str, Any]]:
    width, height = _extract_chart_dimensions(chart)
    metadata = {
        "chart_type": _chart_type_name(chart),
        "target_cell": _extract_chart_anchor(chart),
        "title": _extract_title_text(getattr(chart, "title", None)),
        "width": width,
        "height": height,
    }
    for field in ("chart_type", "target_cell"):
        if field in assertion and metadata[field] != assertion[field]:
            return False, metadata
    if "title" in assertion:
        expected_title = assertion["title"] or None
        if metadata["title"] != expected_title:
            return False, metadata
    for dimension in ("width", "height"):
        if dimension not in assertion:
            continue
        actual_dimension = metadata[dimension]
        if actual_dimension is None or not math.isclose(
            actual_dimension,
            assertion[dimension],
            rel_tol=1e-6,
            abs_tol=0.01,
        ):
            return False, metadata
    return True, metadata


def _assert_chart_exists(
    wb: Any,
    assertion: dict[str, Any],
    *,
    sample_limit: int,
) -> tuple[bool, dict[str, Any]]:
    ws = _worksheet_or_none(wb, assertion["sheet_name"])
    if ws is None:
        return False, {"reason": f"Worksheet '{assertion['sheet_name']}' does not exist"}

    inspected: list[dict[str, Any]] = []
    matches: list[dict[str, Any]] = []
    for chart_index, chart in enumerate(getattr(ws, "_charts", []), start=1):
        matched, metadata = _chart_matches_assertion(chart, assertion)
        item = {"chart_index": chart_index, **metadata}
        inspected.append(item)
        if matched:
            matches.append(item)

    expected = {
        key: assertion[key]
        for key in ("chart_type", "target_cell", "title", "width", "height")
        if key in assertion
    }
    return bool(matches), {
        "sheet_name": assertion["sheet_name"],
        "expected": expected,
        "match_count": len(matches),
        "matches": matches[:sample_limit],
        "inspected_chart_count": len(inspected),
        "inspected": inspected[:sample_limit] if not matches else [],
        "truncated": len(inspected) > sample_limit,
    }


def _assert_no_cell_ref_errors(
    wb: Any,
    assertion: dict[str, Any],
    *,
    sample_limit: int,
) -> tuple[bool, dict[str, Any]]:
    sheet_name = assertion.get("sheet_name")
    if sheet_name is not None:
        ws = _worksheet_or_none(wb, sheet_name)
        if ws is None:
            return False, {"reason": f"Worksheet '{sheet_name}' does not exist"}
        worksheets = [ws]
    else:
        worksheets = list(wb.worksheets)

    bounds = None
    if assertion.get("range_ref") is not None:
        bounds = _range_bounds(assertion["range_ref"], location="assertion range_ref")
    findings: list[dict[str, Any]] = []
    finding_count = 0
    scanned_cells = 0
    for ws in worksheets:
        for (row, column), cell in ws._cells.items():
            if bounds is not None:
                min_col, min_row, max_col, max_row = bounds
                if not (min_row <= row <= max_row and min_col <= column <= max_col):
                    continue
            scanned_cells += 1
            if scanned_cells > MAX_ASSERTION_SCAN_CELLS:
                return False, {
                    "reason": (
                        "no_cell_ref_errors exceeded the "
                        f"{MAX_ASSERTION_SCAN_CELLS:,}-cell scan limit"
                    ),
                    "scanned_cells": scanned_cells,
                }
            value = cell.value
            if isinstance(value, str) and "#REF!" in value.upper():
                finding_count += 1
                if len(findings) < sample_limit:
                    findings.append(
                        {
                            "sheet_name": ws.title,
                            "cell": cell.coordinate,
                            "value": _bound_response_value(value),
                        }
                    )
    return finding_count == 0, {
        "scanned_cells": scanned_cells,
        "finding_count": finding_count,
        "findings": findings,
        "truncated": finding_count > sample_limit,
    }


def _evaluate_assertions(
    baseline: Path,
    candidate: Path,
    assertions: list[dict[str, Any]],
    *,
    sample_limit: int,
) -> list[dict[str, Any]]:
    if not assertions:
        return []
    baseline_wb = load_workbook(baseline, data_only=False)
    candidate_wb = load_workbook(candidate, data_only=False)
    try:
        results: list[dict[str, Any]] = []
        for index, assertion in enumerate(assertions, start=1):
            assertion_type = assertion["type"]
            if assertion_type == "sheet_exists":
                passed, details = _assert_sheet_exists(candidate_wb, assertion)
            elif assertion_type == "cell_equals":
                passed, details = _assert_cell_equals(candidate_wb, assertion)
            elif assertion_type == "range_equals":
                passed, details = _assert_range_equals(
                    candidate_wb, assertion, sample_limit=sample_limit
                )
            elif assertion_type == "range_values_unchanged":
                passed, details = _assert_range_values_unchanged(
                    baseline_wb,
                    candidate_wb,
                    assertion,
                    sample_limit=sample_limit,
                )
            elif assertion_type == "table_exists":
                passed, details = _assert_table_exists(candidate_wb, assertion)
            elif assertion_type == "freeze_panes_equals":
                passed, details = _assert_freeze_panes_equals(
                    candidate_wb, assertion
                )
            elif assertion_type == "autofilter_equals":
                passed, details = _assert_autofilter_equals(
                    candidate_wb, assertion
                )
            elif assertion_type == "chart_exists":
                passed, details = _assert_chart_exists(
                    candidate_wb,
                    assertion,
                    sample_limit=sample_limit,
                )
            elif assertion_type == "no_cell_ref_errors":
                passed, details = _assert_no_cell_ref_errors(
                    candidate_wb, assertion, sample_limit=sample_limit
                )
            else:  # pragma: no cover - normalization guarantees a known assertion type
                raise AssertionError(f"Unhandled assertion type: {assertion_type}")
            results.append(
                {
                    "index": index,
                    "type": assertion_type,
                    "passed": passed,
                    "details": details,
                }
            )
        return results
    finally:
        baseline_wb.close()
        candidate_wb.close()


def _preview_diff(
    baseline: Path,
    candidate: Path,
    *,
    cell_changes: dict[str, Any],
    sample_limit: int,
) -> dict[str, Any]:
    workbook_diff = diff_workbooks(
        str(baseline),
        str(candidate),
        sample_limit=sample_limit,
        include_cell_changes=False,
    )
    workbook_diff.pop("before_file", None)
    workbook_diff.pop("after_file", None)
    changes = cell_changes["changes"]
    workbook_diff["cell_changes"] = {
        "count": cell_changes["count"],
        "sample": changes[:sample_limit],
        "truncated": len(changes) > sample_limit,
    }
    return _bound_response_value(workbook_diff)


def _create_or_reuse_snapshot(
    baseline: Path,
    snapshot_status: dict[str, Any],
    *,
    source_sha256: str,
) -> dict[str, Any]:
    if not snapshot_status["enabled"]:
        return snapshot_status
    current_status = _snapshot_status(
        Path(snapshot_status["path"]), source_sha256=source_sha256
    )
    if current_status["status"] == "existing_matching":
        return {**current_status, "reused": True}
    if current_status["status"] != "available":
        raise PreconditionFailedError(
            current_status.get("reason", "Snapshot destination is unavailable"),
            code="snapshot_conflict",
            details=current_status,
            suggested_next_tool="apply_workbook_changeset",
        )
    result = create_workbook_snapshot(str(baseline), current_status["path"])
    return {
        "enabled": True,
        "path": current_status["path"],
        "status": "created",
        "sha256": result["sha256"],
        "reused": False,
    }


def _discard_uncommitted_snapshot(snapshot_result: dict[str, Any] | None) -> dict[str, Any]:
    if snapshot_result is None or snapshot_result.get("status") != "created":
        return {"attempted": False, "status": "not_created"}

    path = Path(snapshot_result["path"])
    cleanup = {"attempted": True, "path": str(path)}
    try:
        if not path.exists() and not path.is_symlink():
            return {**cleanup, "status": "already_absent"}
        if path.is_symlink() or not path.is_file():
            return {
                **cleanup,
                "status": "retained",
                "reason": "snapshot path changed type before cleanup",
            }

        stat_before = path.stat()
        actual_sha256 = _sha256_file(path)
        stat_after = path.stat()
        if workbook_module._file_identity(stat_before) != workbook_module._file_identity(
            stat_after
        ):
            return {
                **cleanup,
                "status": "retained",
                "reason": "snapshot changed while cleanup was being verified",
            }
        if actual_sha256 != snapshot_result.get("sha256"):
            return {
                **cleanup,
                "status": "retained",
                "reason": "snapshot content changed before cleanup",
                "actual_sha256": actual_sha256,
            }

        path.unlink()
        workbook_module._fsync_directory(path.parent)
        return {**cleanup, "status": "removed"}
    except FileNotFoundError:
        return {**cleanup, "status": "already_absent"}
    except OSError as exc:
        logger.warning("Unable to remove uncommitted ChangeSet snapshot '%s': %s", path, exc)
        return {
            **cleanup,
            "status": "retained",
            "reason": f"snapshot cleanup failed: {exc!s}",
        }


def _restore_baseline(source: Path, baseline: Path, *, expected_sha256: str) -> None:
    rollback = _new_temp_workbook(source.parent, source=source, label="rollback")
    try:
        shutil.copy2(baseline, rollback)
        workbook_module._fsync_file(rollback)
        workbook_module._verify_saved_workbook(str(rollback))
        os.replace(rollback, source)
        workbook_module._fsync_directory(source.parent)
        workbook_module._verify_saved_workbook(str(source))
        restored_sha256 = _sha256_file(source)
        if restored_sha256 != expected_sha256:
            raise WorkbookError(
                "Rollback verification produced a workbook with an unexpected SHA-256 digest"
            )
    finally:
        workbook_module._remove_save_artifact_best_effort(
            rollback, label="ChangeSet rollback workbook"
        )


def _validate_snapshot_request(
    source: Path,
    *,
    create_snapshot: Any,
    snapshot_filepath: Any,
) -> tuple[bool, Path | None]:
    if not isinstance(create_snapshot, bool):
        raise ValidationError("create_snapshot must be a boolean")
    if not create_snapshot and snapshot_filepath is not None:
        raise ValidationError(
            "snapshot_filepath cannot be provided when create_snapshot is false"
        )
    if snapshot_filepath is None:
        return create_snapshot, None
    _validate_non_empty_string(snapshot_filepath, location="snapshot_filepath")
    requested = workbook_module._canonical_workbook_path(
        snapshot_filepath, must_exist=False
    )
    if requested.suffix.lower() != ".xlsx":
        raise ValidationError("snapshot_filepath must end with .xlsx")
    if requested == source:
        raise ValidationError("snapshot_filepath must be different from filepath")
    if requested.exists() and requested.resolve(strict=True) == source:
        raise ValidationError("snapshot_filepath must be different from filepath")
    return create_snapshot, requested


def _validate_common_arguments(
    *,
    filepath: str,
    operations: Any,
    assertions: Any,
    mode: Any,
    expected_workbook_sha256: Any,
    changeset_token: Any,
    create_snapshot: Any,
    snapshot_filepath: Any,
    sample_limit: Any,
) -> tuple[
    Path,
    list[dict[str, Any]],
    list[dict[str, Any]],
    str,
    str | None,
    str | None,
    bool,
    Path | None,
]:
    if not isinstance(mode, str) or mode.strip().lower() not in {"preview", "commit"}:
        raise ValidationError("mode must be 'preview' or 'commit'")
    normalized_mode = mode.strip().lower()
    if not isinstance(sample_limit, int) or isinstance(sample_limit, bool) or sample_limit <= 0:
        raise ValidationError("sample_limit must be a positive integer")
    if sample_limit > 100:
        raise ValidationError("sample_limit cannot exceed 100")

    source = workbook_module._canonical_workbook_path(filepath, must_exist=True)
    if source.suffix.lower() != ".xlsx":
        raise ValidationError("filepath must end with .xlsx")
    normalized_operations = _normalize_operations(operations)
    normalized_assertions = _normalize_assertions(assertions)
    expected_sha256 = _normalize_sha256(
        expected_workbook_sha256, required=normalized_mode == "commit"
    )
    normalized_token = _normalize_changeset_token(
        changeset_token, required=normalized_mode == "commit"
    )
    snapshot_enabled, requested_snapshot = _validate_snapshot_request(
        source,
        create_snapshot=create_snapshot,
        snapshot_filepath=snapshot_filepath,
    )
    return (
        source,
        normalized_operations,
        normalized_assertions,
        normalized_mode,
        expected_sha256,
        normalized_token,
        snapshot_enabled,
        requested_snapshot,
    )


def apply_workbook_changeset(
    filepath: str,
    operations: list[dict[str, Any]],
    assertions: list[dict[str, Any]] | None = None,
    mode: str = "preview",
    expected_workbook_sha256: str | None = None,
    changeset_token: str | None = None,
    create_snapshot: bool = True,
    snapshot_filepath: str | None = None,
    sample_limit: int = 25,
) -> dict[str, Any]:
    """Preview or atomically commit an assertion-backed workbook mutation plan."""
    (
        source,
        normalized_operations,
        normalized_assertions,
        normalized_mode,
        expected_sha256,
        supplied_token,
        snapshot_enabled,
        requested_snapshot,
    ) = _validate_common_arguments(
        filepath=filepath,
        operations=operations,
        assertions=assertions,
        mode=mode,
        expected_workbook_sha256=expected_workbook_sha256,
        changeset_token=changeset_token,
        create_snapshot=create_snapshot,
        snapshot_filepath=snapshot_filepath,
        sample_limit=sample_limit,
    )

    if normalized_mode == "preview":
        work_dir = Path(tempfile.mkdtemp(prefix="sheetforge-changeset-preview-"))
        baseline = work_dir / "baseline.xlsx"
        candidate = work_dir / "candidate.xlsx"
        try:
            with workbook_module._exclusive_workbook_lock(source):
                source_sha256 = _stage_workbook(source, baseline, candidate)
            if expected_sha256 is not None and expected_sha256 != source_sha256:
                raise PreconditionFailedError(
                    "Workbook SHA-256 does not match the caller's expected baseline",
                    code="stale_workbook",
                    details={
                        "expected_workbook_sha256": expected_sha256,
                        "actual_workbook_sha256": source_sha256,
                    },
                    suggested_next_tool="apply_workbook_changeset",
                )

            plan = _plan_payload(
                source=source,
                source_sha256=source_sha256,
                operations=normalized_operations,
                assertions=normalized_assertions,
                create_snapshot=snapshot_enabled,
                snapshot_filepath=requested_snapshot,
            )
            computed_token = _changeset_token(plan)
            if supplied_token is not None and not hmac.compare_digest(
                supplied_token, computed_token
            ):
                raise PreconditionFailedError(
                    "ChangeSet token does not match the current workbook and plan",
                    code="changeset_token_mismatch",
                    suggested_next_tool="apply_workbook_changeset",
                )

            operation_results, cell_changes = _apply_operations(
                candidate,
                normalized_operations,
                sample_limit=sample_limit,
            )
            workbook_module._verify_saved_workbook(str(candidate))
            assertion_results = _evaluate_assertions(
                baseline,
                candidate,
                normalized_assertions,
                sample_limit=sample_limit,
            )
            snapshot_path = _snapshot_destination(
                source,
                create_snapshot=snapshot_enabled,
                requested_path=requested_snapshot,
                changeset_token=computed_token,
            )
            snapshot = _snapshot_status(snapshot_path, source_sha256=source_sha256)
            assertions_passed = all(item["passed"] for item in assertion_results)
            ready_to_commit = assertions_passed and snapshot["status"] != "conflict"
            preview_result = {
                "message": (
                    "ChangeSet preview is ready to commit"
                    if ready_to_commit
                    else "ChangeSet preview requires attention before commit"
                ),
                "mode": "preview",
                "persisted": False,
                "ready_to_commit": ready_to_commit,
                "expected_workbook_sha256": source_sha256,
                "changeset_token": computed_token,
                "operation_count": len(normalized_operations),
                "operations": operation_results,
                "assertion_count": len(normalized_assertions),
                "assertions_passed": sum(
                    1 for item in assertion_results if item["passed"]
                ),
                "assertions": assertion_results,
                "snapshot": snapshot,
                "diff": _preview_diff(
                    baseline,
                    candidate,
                    cell_changes=cell_changes,
                    sample_limit=sample_limit,
                ),
            }
            _assert_response_budget(preview_result)
            return preview_result
        finally:
            try:
                shutil.rmtree(work_dir)
            except OSError as exc:
                logger.warning("Unable to remove ChangeSet preview directory '%s': %s", work_dir, exc)

    baseline: Path | None = None
    candidate: Path | None = None
    preserve_baseline = False
    try:
        baseline = _new_temp_workbook(source.parent, source=source, label="baseline")
        candidate = _new_temp_workbook(source.parent, source=source, label="candidate")
        with workbook_module._exclusive_workbook_lock(source):
            source_sha256 = _stage_workbook(source, baseline, candidate)
            if expected_sha256 != source_sha256:
                raise PreconditionFailedError(
                    "Workbook changed after preview; no ChangeSet operations were committed",
                    code="stale_workbook",
                    details={
                        "expected_workbook_sha256": expected_sha256,
                        "actual_workbook_sha256": source_sha256,
                    },
                    suggested_next_tool="apply_workbook_changeset",
                )

            plan = _plan_payload(
                source=source,
                source_sha256=source_sha256,
                operations=normalized_operations,
                assertions=normalized_assertions,
                create_snapshot=snapshot_enabled,
                snapshot_filepath=requested_snapshot,
            )
            computed_token = _changeset_token(plan)
            if not hmac.compare_digest(supplied_token or "", computed_token):
                raise PreconditionFailedError(
                    "ChangeSet token does not match the previewed workbook and plan",
                    code="changeset_token_mismatch",
                    suggested_next_tool="apply_workbook_changeset",
                )

            snapshot_path = _snapshot_destination(
                source,
                create_snapshot=snapshot_enabled,
                requested_path=requested_snapshot,
                changeset_token=computed_token,
            )
            snapshot = _snapshot_status(snapshot_path, source_sha256=source_sha256)
            if snapshot["status"] == "conflict":
                raise PreconditionFailedError(
                    snapshot["reason"],
                    code="snapshot_conflict",
                    details=snapshot,
                    suggested_next_tool="apply_workbook_changeset",
                )

            operation_results, cell_changes = _apply_operations(
                candidate,
                normalized_operations,
                sample_limit=sample_limit,
            )
            workbook_module._verify_saved_workbook(str(candidate))
            assertion_results = _evaluate_assertions(
                baseline,
                candidate,
                normalized_assertions,
                sample_limit=sample_limit,
            )
            failed_assertions = [
                item for item in assertion_results if not item["passed"]
            ]
            if failed_assertions:
                _assert_response_budget(
                    {
                        "failed_assertion_count": len(failed_assertions),
                        "failed_assertions": failed_assertions,
                    }
                )
                raise PreconditionFailedError(
                    "ChangeSet assertions failed; the original workbook was not modified",
                    code="changeset_assertion_failed",
                    details={"failed_assertions": failed_assertions},
                    suggested_next_tool="apply_workbook_changeset",
                )

            preview_diff = _preview_diff(
                baseline,
                candidate,
                cell_changes=cell_changes,
                sample_limit=sample_limit,
            )
            _assert_response_budget(
                {
                    "mode": "commit",
                    "operations": operation_results,
                    "assertions": assertion_results,
                    "snapshot": snapshot,
                    "diff": preview_diff,
                }
            )

            live_sha256 = _sha256_file(source)
            if live_sha256 != source_sha256:
                raise PreconditionFailedError(
                    "Workbook changed while the ChangeSet candidate was being prepared; no changes were committed",
                    code="stale_workbook",
                    details={
                        "expected_workbook_sha256": source_sha256,
                        "actual_workbook_sha256": live_sha256,
                    },
                    suggested_next_tool="apply_workbook_changeset",
                )
            snapshot_result: dict[str, Any] | None = None
            destination_replaced = False
            try:
                snapshot_result = _create_or_reuse_snapshot(
                    baseline,
                    snapshot,
                    source_sha256=source_sha256,
                )
                live_sha256 = _sha256_file(source)
                if live_sha256 != source_sha256:
                    raise PreconditionFailedError(
                        "Workbook changed while the ChangeSet snapshot was being prepared; no changes were committed",
                        code="stale_workbook",
                        details={
                            "expected_workbook_sha256": source_sha256,
                            "actual_workbook_sha256": live_sha256,
                        },
                        suggested_next_tool="apply_workbook_changeset",
                    )

                os.replace(candidate, source)
                destination_replaced = True
                workbook_module._fsync_directory(source.parent)
                workbook_module._verify_saved_workbook(str(source))
                post_assertions = _evaluate_assertions(
                    baseline,
                    source,
                    normalized_assertions,
                    sample_limit=sample_limit,
                )
                if any(not item["passed"] for item in post_assertions):
                    raise WorkbookError(
                        "Post-commit assertion verification failed"
                    )
                after_sha256 = _sha256_file(source)
            except Exception as commit_error:
                if not destination_replaced:
                    snapshot_cleanup = _discard_uncommitted_snapshot(snapshot_result)
                    if isinstance(commit_error, PreconditionFailedError):
                        if snapshot_cleanup["attempted"]:
                            commit_error.details = {
                                **commit_error.details,
                                "snapshot_cleanup": snapshot_cleanup,
                            }
                        if snapshot_cleanup["status"] == "retained":
                            raise PreconditionFailedError(
                                f"{commit_error!s}. Uncommitted snapshot retained at "
                                f"{snapshot_cleanup['path']}: {snapshot_cleanup['reason']}",
                                code=commit_error.code,
                                details=commit_error.details,
                                suggested_next_tool=commit_error.suggested_next_tool,
                            ) from commit_error
                        raise
                    cleanup_note = ""
                    if snapshot_cleanup["status"] == "retained":
                        cleanup_note = (
                            f" Uncommitted snapshot retained at {snapshot_cleanup['path']}: "
                            f"{snapshot_cleanup['reason']}"
                        )
                    raise WorkbookError(
                        "ChangeSet commit failed before replacing the original workbook: "
                        f"{commit_error!s}.{cleanup_note}"
                    ) from commit_error
                try:
                    _restore_baseline(
                        source, baseline, expected_sha256=source_sha256
                    )
                except Exception as rollback_error:
                    preserve_baseline = True
                    raise WorkbookError(
                        "ChangeSet verification failed and rollback could not be verified. "
                        f"Recovery baseline retained at {baseline}: {rollback_error!s}"
                    ) from rollback_error
                snapshot_note = ""
                if snapshot_result is not None and snapshot_result.get("enabled"):
                    snapshot_note = (
                        f" Verified snapshot retained at {snapshot_result['path']}."
                    )
                raise WorkbookError(
                    "ChangeSet commit verification failed; the original workbook was restored: "
                    f"{commit_error!s}.{snapshot_note}"
                ) from commit_error

            committed_at = datetime.now(timezone.utc).isoformat()
            receipt_digest = hashlib.sha256(
                f"{computed_token}:{after_sha256}:{committed_at}".encode("utf-8")
            ).hexdigest()[:24]
            return {
                "message": "ChangeSet committed and verified",
                "mode": "commit",
                "persisted": True,
                "ready_to_commit": False,
                "changeset_token": computed_token,
                "receipt": {
                    "receipt_id": f"changeset_receipt_{receipt_digest}",
                    "committed_at": committed_at,
                    "before_sha256": source_sha256,
                    "after_sha256": after_sha256,
                    "operation_count": len(normalized_operations),
                    "assertion_count": len(normalized_assertions),
                    "snapshot": snapshot_result,
                },
                "operations": operation_results,
                "assertions_passed": len(assertion_results),
                "assertions": assertion_results,
                "diff": preview_diff,
            }
    finally:
        if candidate is not None:
            workbook_module._remove_save_artifact_best_effort(
                candidate, label="ChangeSet candidate workbook"
            )
        if baseline is not None and not preserve_baseline:
            workbook_module._remove_save_artifact_best_effort(
                baseline, label="ChangeSet baseline workbook"
            )

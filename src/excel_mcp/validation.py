import logging
import re
from typing import Any

from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .cell_utils import (
    MAX_EXCEL_COLUMN,
    MAX_EXCEL_ROW,
    parse_cell_range,
    validate_cell_reference,
)
from .exceptions import ValidationError
from .workbook import require_worksheet, safe_workbook

logger = logging.getLogger(__name__)
UNSAFE_FORMULA_FUNCTIONS = {"INDIRECT", "HYPERLINK", "WEBSERVICE", "DGET", "RTD"}
_CELL_TOKEN_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([0-9]+)$")
_COLUMN_TOKEN_RE = re.compile(r"^\$?([A-Za-z]{1,3})$")
_ROW_TOKEN_RE = re.compile(r"^\$?([0-9]+)$")

def validate_formula_in_cell_operation(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str
) -> dict[str, Any]:
    """Validate Excel formula before writing"""
    try:
        with safe_workbook(filepath) as wb:
            if not validate_cell_reference(cell):
                raise ValidationError(f"Invalid cell reference: {cell}")

            # First validate the provided formula's syntax
            is_valid, message = validate_formula(formula)
            if not is_valid:
                raise ValidationError(f"Invalid formula syntax: {message}")

            # Now check if there's a formula in the cell and compare
            sheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="formula validation",
            )
            cell_obj = sheet[cell]
            current_formula = cell_obj.value

            # If cell has a formula (starts with =)
            if isinstance(current_formula, str) and current_formula.startswith('='):
                normalized_formula = formula if formula.startswith('=') else f"={formula}"
                if current_formula != normalized_formula:
                    return {
                        "message": "Formula is valid but doesn't match cell content",
                        "valid": True,
                        "matches": False,
                        "cell": cell,
                        "provided_formula": formula,
                        "current_formula": current_formula
                    }
                return {
                    "message": "Formula is valid and matches cell content",
                    "valid": True,
                    "matches": True,
                    "cell": cell,
                    "formula": formula
                }
            else:
                return {
                    "message": "Formula is valid but cell contains no formula",
                    "valid": True,
                    "matches": False,
                    "cell": cell,
                    "provided_formula": formula,
                    "current_content": str(current_formula) if current_formula else ""
                }

    except ValidationError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to validate formula: {e}")
        raise ValidationError(str(e))

def validate_range_in_sheet_operation(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
) -> dict[str, Any]:
    """Validate if a range exists in a worksheet and return data range info."""
    try:
        with safe_workbook(filepath) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="range validation",
            )

            # Get actual data dimensions
            data_max_row = worksheet.max_row
            data_max_col = worksheet.max_column

            # Validate range
            try:
                start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            except ValueError as e:
                raise ValidationError(f"Invalid range: {str(e)}")

            # If end not specified, use start
            if end_row is None:
                end_row = start_row
            if end_col is None:
                end_col = start_col

            # Validate bounds against maximum possible Excel limits
            is_valid, message = validate_range_bounds(
                worksheet, start_row, start_col, end_row, end_col
            )
            if not is_valid:
                raise ValidationError(message)

            range_str = f"{start_cell}" if end_cell is None else f"{start_cell}:{end_cell}"
            data_range_str = f"A1:{get_column_letter(data_max_col)}{data_max_row}"

            # Check if range is within data or extends beyond
            extends_beyond_data = (
                end_row > data_max_row or
                end_col > data_max_col
            )

            return {
                "message": (
                    f"Range '{range_str}' is valid. "
                    f"Sheet contains data in range '{data_range_str}'"
                ),
                "valid": True,
                "range": range_str,
                "data_range": data_range_str,
                "extends_beyond_data": extends_beyond_data,
                "data_dimensions": {
                    "max_row": data_max_row,
                    "max_col": data_max_col,
                    "max_col_letter": get_column_letter(data_max_col)
                }
            }
    except ValidationError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to validate range: {e}")
        raise ValidationError(str(e))

def validate_formula(formula: str) -> tuple[bool, str]:
    """Validate formula structure and reject functions unsafe for agent writes."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return False, "Formula must start with '='"
    if not formula[1:].strip():
        return False, "Formula expression is empty"

    try:
        tokenizer_formula = _normalize_spill_operators_for_tokenizer(formula)
        tokens = list(Tokenizer(tokenizer_formula).items)
    except Exception as exc:
        return False, f"Formula tokenization failed: {exc!s}"
    if not tokens:
        return False, "Formula expression is empty"

    valid_sequence, sequence_message = _validate_formula_token_sequence(tokens)
    if not valid_sequence:
        return False, sequence_message

    for token in tokens:
        if token.type == "FUNC" and token.subtype == "OPEN":
            raw_name = str(token.value).rstrip("(").strip().lstrip("@").upper()
            function_name = raw_name.rsplit(".", 1)[-1]
            if function_name in UNSAFE_FORMULA_FUNCTIONS:
                return False, f"Unsafe function: {function_name}"

        if token.type == "OPERAND" and token.subtype == "RANGE":
            valid_reference, reference_message = _validate_range_token(str(token.value))
            if not valid_reference:
                return False, reference_message

    return True, "Formula passed structural validation"


def _normalize_spill_operators_for_tokenizer(formula: str) -> str:
    """Represent dynamic-array spill markers with a tokenizer-supported postfix."""
    return re.sub(
        r"(?<=[A-Za-z0-9_\]\)])#(?=$|[,+\-*/^&=<>%)])",
        "%",
        formula,
    )


def _validate_formula_token_sequence(tokens: list[Any]) -> tuple[bool, str]:
    meaningful_tokens = [token for token in tokens if token.type != "WHITE-SPACE"]
    if not meaningful_tokens:
        return False, "Formula expression is empty"

    stack: list[str] = []
    expecting_operand = True
    previous: Any = None

    for index, token in enumerate(tokens):
        if token.type == "WHITE-SPACE":
            previous_nonspace = next(
                (candidate for candidate in reversed(tokens[:index]) if candidate.type != "WHITE-SPACE"),
                None,
            )
            next_nonspace = next(
                (candidate for candidate in tokens[index + 1:] if candidate.type != "WHITE-SPACE"),
                None,
            )
            if (
                previous_nonspace is not None
                and next_nonspace is not None
                and previous_nonspace.type == "OPERAND"
                and previous_nonspace.subtype == "RANGE"
                and next_nonspace.type == "OPERAND"
                and next_nonspace.subtype == "RANGE"
            ):
                expecting_operand = True
                previous = token
            continue

        if token.subtype == "OPEN":
            lambda_invocation = (
                token.type == "PAREN"
                and previous is not None
                and previous.type == "FUNC"
                and previous.subtype == "CLOSE"
            )
            if not expecting_operand and not lambda_invocation:
                return False, f"Missing operator before '{token.value}'"
            stack.append(token.type)
            expecting_operand = True
        elif token.subtype == "CLOSE":
            if not stack or stack[-1] != token.type:
                return False, "Unmatched closing parenthesis"
            if expecting_operand:
                empty_function = token.type == "FUNC" and previous is not None and previous.subtype == "OPEN"
                omitted_argument = token.type == "FUNC" and previous is not None and previous.type == "SEP"
                if not (empty_function or omitted_argument):
                    return False, f"Missing expression before '{token.value}'"
            stack.pop()
            expecting_operand = False
        elif token.type == "OPERAND":
            if not expecting_operand:
                return False, f"Missing operator before '{token.value}'"
            expecting_operand = False
        elif token.type == "OPERATOR-PREFIX":
            if not expecting_operand:
                return False, f"Unexpected prefix operator '{token.value}'"
        elif token.type == "OPERATOR-INFIX":
            if expecting_operand:
                return False, f"Unexpected operator '{token.value}'"
            expecting_operand = True
        elif token.type == "OPERATOR-POSTFIX":
            if expecting_operand:
                return False, f"Unexpected postfix operator '{token.value}'"
            expecting_operand = False
        elif token.type == "SEP":
            if not stack:
                return False, f"Unexpected separator '{token.value}'"
            expecting_operand = True

        previous = token

    if stack:
        return False, "Unclosed parenthesis"
    if expecting_operand:
        return False, "Formula cannot end with an operator or separator"
    return True, "Formula token sequence is valid"


def _validate_range_token(token_value: str) -> tuple[bool, str]:
    local_reference = token_value.rsplit("!", 1)[-1]
    if "[" in local_reference or "]" in local_reference:
        return True, "Structured or external reference"

    endpoints = local_reference.split(":")
    if len(endpoints) > 2:
        return False, f"Invalid range reference: {token_value}"

    for endpoint in endpoints:
        endpoint = endpoint.strip()
        cell_match = _CELL_TOKEN_RE.fullmatch(endpoint)
        if cell_match:
            if not validate_cell_reference(endpoint.replace("$", "")):
                return False, f"Cell reference outside Excel limits: {token_value}"
            continue

        column_match = _COLUMN_TOKEN_RE.fullmatch(endpoint)
        if len(endpoints) == 2 and column_match:
            try:
                _, column = parse_cell_range(f"{column_match.group(1)}1")[:2]
            except ValueError:
                return False, f"Column reference outside Excel limits: {token_value}"
            if column > MAX_EXCEL_COLUMN:
                return False, f"Column reference outside Excel limits: {token_value}"
            continue

        row_match = _ROW_TOKEN_RE.fullmatch(endpoint)
        if len(endpoints) == 2 and row_match:
            row_number = int(row_match.group(1))
            if not 1 <= row_number <= MAX_EXCEL_ROW:
                return False, f"Row reference outside Excel limits: {token_value}"

    return True, "Range reference is valid"


def validate_range_bounds(
    worksheet: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int | None = None,
    end_col: int | None = None,
) -> tuple[bool, str]:
    """Validate that a cell range is within Excel's physical worksheet limits."""

    try:
        # Check start cell bounds
        if start_row < 1 or start_row > MAX_EXCEL_ROW:
            return False, f"Start row {start_row} out of bounds (1-{MAX_EXCEL_ROW})"
        if start_col < 1 or start_col > MAX_EXCEL_COLUMN:
            return False, (
                f"Start column {get_column_letter(start_col)} "
                f"out of bounds (A-XFD)"
            )

        # If end cell specified, check its bounds
        if end_row is not None and end_col is not None:
            if end_row < start_row:
                return False, "End row cannot be before start row"
            if end_col < start_col:
                return False, "End column cannot be before start column"
            if end_row > MAX_EXCEL_ROW:
                return False, f"End row {end_row} out of bounds (1-{MAX_EXCEL_ROW})"
            if end_col > MAX_EXCEL_COLUMN:
                return False, (
                    f"End column {get_column_letter(end_col)} "
                    f"out of bounds (A-XFD)"
                )

        return True, "Range is valid"
    except Exception as e:
        return False, f"Invalid range: {e!s}"

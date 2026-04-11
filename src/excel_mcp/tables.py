import uuid
import logging
from typing import Any, Optional

from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from .data import _build_cell_change, _should_include_changes, augment_tabular_payload
from .exceptions import DataError
from .workbook import safe_workbook

logger = logging.getLogger(__name__)


def _get_table_style_details(table: Table) -> dict[str, Any]:
    style_name = None
    show_first_column = None
    show_last_column = None
    show_row_stripes = None
    show_column_stripes = None
    if table.tableStyleInfo is not None:
        style_name = table.tableStyleInfo.name
        show_first_column = table.tableStyleInfo.showFirstColumn
        show_last_column = table.tableStyleInfo.showLastColumn
        show_row_stripes = table.tableStyleInfo.showRowStripes
        show_column_stripes = table.tableStyleInfo.showColumnStripes

    return {
        "style": style_name,
        "show_first_column": show_first_column,
        "show_last_column": show_last_column,
        "show_row_stripes": show_row_stripes,
        "show_column_stripes": show_column_stripes,
    }


def _build_table_metadata(
    current_sheet_name: str,
    ws: Any,
    table: Table,
) -> dict[str, Any]:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    header_row_count = int(table.headerRowCount or 0)
    totals_row_count = int(table.totalsRowCount or 0)
    total_row_span = max_row - min_row + 1
    data_row_count = max(total_row_span - header_row_count - totals_row_count, 0)

    headers: list[Any] = []
    if header_row_count > 0:
        headers = [
            ws.cell(row=min_row, column=column_index).value
            for column_index in range(min_col, max_col + 1)
        ]

    metadata = {
        "sheet_name": current_sheet_name,
        "table_name": table.displayName,
        "range": table.ref,
        "headers": headers,
        "column_count": max_col - min_col + 1,
        "data_row_count": data_row_count,
        "header_row_count": header_row_count,
        "totals_row_count": totals_row_count,
        "totals_row_shown": bool(table.totalsRowShown),
    }
    metadata.update(_get_table_style_details(table))
    return metadata


def _range_from_bounds(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _table_header_map(ws: Any, table: Table) -> dict[str, int]:
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    header_row_count = int(table.headerRowCount or 0)
    if header_row_count < 1:
        raise DataError(f"Table '{table.displayName}' does not have a header row")

    header_map: dict[str, int] = {}
    for column_index in range(min_col, max_col + 1):
        value = ws.cell(row=min_row, column=column_index).value
        if value is None or str(value).strip() == "":
            raise DataError(
                f"Table '{table.displayName}' has an empty header in column "
                f"{get_column_letter(column_index)}"
            )

        header_name = str(value)
        if header_name in header_map:
            raise DataError(
                f"Duplicate header '{header_name}' found in table '{table.displayName}'"
            )
        header_map[header_name] = column_index

    return header_map


def _table_data_bounds(table: Table) -> tuple[int, int, int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    header_row_count = int(table.headerRowCount or 0)
    totals_row_count = int(table.totalsRowCount or 0)
    data_start_row = min_row + header_row_count
    data_end_row = max_row - totals_row_count
    return min_col, min_row, max_col, max_row, data_start_row, data_end_row


def _ensure_table_append_space_clear(
    ws: Any,
    *,
    start_row: int,
    row_count: int,
    min_col: int,
    max_col: int,
) -> None:
    for row_index in range(start_row, start_row + row_count):
        for column_index in range(min_col, max_col + 1):
            if ws.cell(row=row_index, column=column_index).value is not None:
                raise DataError(
                    "Cannot expand table into occupied cells; "
                    f"found existing value at {get_column_letter(column_index)}{row_index}"
                )


def _find_table(
    wb: Any,
    table_name: str,
    sheet_name: Optional[str] = None,
) -> tuple[str, Any, Table]:
    if sheet_name is not None and sheet_name not in wb.sheetnames:
        raise DataError(f"Sheet '{sheet_name}' not found.")

    sheet_names = [sheet_name] if sheet_name is not None else list(wb.sheetnames)
    for current_sheet_name in sheet_names:
        ws = wb[current_sheet_name]
        for table in ws.tables.values():
            if table.displayName == table_name:
                return current_sheet_name, ws, table

    if sheet_name is not None:
        raise DataError(f"Table '{table_name}' not found in sheet '{sheet_name}'.")
    raise DataError(f"Table '{table_name}' not found.")

def create_excel_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: str | None = None,
    table_style: str = "TableStyleMedium9"
) -> dict:
    """Creates a native Excel table for the given data range.
    
    Args:
        filepath: Path to the Excel file.
        sheet_name: Name of the worksheet.
        data_range: The cell range for the table (e.g., "A1:D5").
        table_name: A unique name for the table. If not provided, a unique name is generated.
        table_style: The visual style to apply to the table.
        
    Returns:
        A dictionary with a success message and table details.
    """
    try:
        with safe_workbook(filepath, save=True) as wb:
            if sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found.")

            ws = wb[sheet_name]

            # If no table name is provided, generate a unique one
            if not table_name:
                table_name = f"Table_{uuid.uuid4().hex[:8]}"

            # Check if table name already exists
            if table_name in ws.parent.defined_names:
                raise DataError(f"Table name '{table_name}' already exists.")

            # Create the table
            table = Table(displayName=table_name, ref=data_range)

            # Apply style
            style = TableStyleInfo(
                name=table_style,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style

            ws.add_table(table)

        return {
            "message": f"Successfully created table '{table_name}' in sheet '{sheet_name}'.",
            "table_name": table_name,
            "range": data_range
        }

    except Exception as e:
        logger.error(f"Failed to create table: {e}")
        raise DataError(str(e))


def list_excel_tables(
    filepath: str,
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    """List native Excel tables for one sheet or the whole workbook."""
    try:
        with safe_workbook(filepath) as wb:
            if sheet_name is not None and sheet_name not in wb.sheetnames:
                raise DataError(f"Sheet '{sheet_name}' not found.")

            sheet_names = [sheet_name] if sheet_name is not None else list(wb.sheetnames)
            tables: list[dict[str, Any]] = []

            for current_sheet_name in sheet_names:
                ws = wb[current_sheet_name]
                for table in ws.tables.values():
                    tables.append(_build_table_metadata(current_sheet_name, ws, table))

            return tables

    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to list tables: {e}")
        raise DataError(str(e))


def read_excel_table(
    filepath: str,
    table_name: str,
    sheet_name: Optional[str] = None,
    max_rows: Optional[int] = None,
    compact: bool = False,
    row_mode: str = "arrays",
    infer_schema: bool = False,
) -> dict[str, Any]:
    """Read rows from a native Excel table by its table name.

    Supports the same array-vs-record row modes and lightweight inferred schema
    hints as the compact worksheet table readers.
    """
    try:
        with safe_workbook(filepath) as wb:
            current_sheet_name, ws, table = _find_table(wb, table_name, sheet_name=sheet_name)
            metadata = _build_table_metadata(current_sheet_name, ws, table)
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)

            data_start_row = min_row + metadata["header_row_count"]
            data_end_row = max_row - metadata["totals_row_count"]
            total_rows = metadata["data_row_count"]
            row_limit = min(total_rows, max_rows) if max_rows is not None else total_rows

            rows: list[list[Any]] = []
            for row_index in range(data_start_row, data_start_row + row_limit):
                if row_index > data_end_row:
                    break
                rows.append(
                    [
                        ws.cell(row=row_index, column=column_index).value
                        for column_index in range(min_col, max_col + 1)
                    ]
                )

            result = {
                "sheet_name": current_sheet_name,
                "table_name": table.displayName,
                "range": table.ref,
                "style": metadata["style"],
                "headers": metadata["headers"],
                "rows": rows,
                "total_rows": total_rows,
                "truncated": max_rows is not None and total_rows > max_rows,
                "header_row_count": metadata["header_row_count"],
                "totals_row_count": metadata["totals_row_count"],
                "totals_row_shown": metadata["totals_row_shown"],
            }
            if compact:
                payload = {
                    "sheet_name": result["sheet_name"],
                    "table_name": result["table_name"],
                    "range": result["range"],
                    "headers": result["headers"],
                    "rows": result["rows"],
                }
                if result["truncated"]:
                    payload["total_rows"] = result["total_rows"]
                    payload["truncated"] = True
            else:
                payload = result

            return augment_tabular_payload(
                payload,
                headers=result["headers"],
                rows=result["rows"],
                row_mode=row_mode,
                infer_schema=infer_schema,
            )

    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to read table: {e}")
        raise DataError(str(e))


def upsert_excel_table_rows(
    filepath: str,
    table_name: str,
    key_column: str,
    rows: list[dict[str, Any]],
    sheet_name: Optional[str] = None,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> dict[str, Any]:
    """Update matching rows in a native Excel table and append missing keys."""
    try:
        if not rows:
            raise DataError("No rows provided to upsert")
        if not all(isinstance(row, dict) for row in rows):
            raise DataError("Rows must be a list of objects keyed by column name")

        with safe_workbook(filepath, save=not dry_run) as wb:
            current_sheet_name, ws, table = _find_table(wb, table_name, sheet_name=sheet_name)
            header_map = _table_header_map(ws, table)
            if key_column not in header_map:
                raise DataError(
                    f"Key column '{key_column}' not found in table '{table.displayName}'"
                )

            unknown_columns = sorted(
                {
                    key
                    for row in rows
                    for key in row.keys()
                    if key not in header_map
                }
            )
            if unknown_columns:
                raise DataError(f"Unknown columns for upsert: {', '.join(unknown_columns)}")

            min_col, min_row, max_col, max_row, data_start_row, data_end_row = _table_data_bounds(
                table
            )
            totals_row_count = int(table.totalsRowCount or 0)
            key_col_idx = header_map[key_column]

            row_lookup: dict[Any, int] = {}
            if data_start_row <= data_end_row:
                for row_index in range(data_start_row, data_end_row + 1):
                    key_value = ws.cell(row=row_index, column=key_col_idx).value
                    if key_value is None:
                        continue
                    if key_value in row_lookup:
                        raise DataError(
                            f"Duplicate key '{key_value}' found in table '{table.displayName}'"
                        )
                    row_lookup[key_value] = row_index

            seen_update_keys: set[Any] = set()
            update_rows: list[tuple[int, dict[str, Any]]] = []
            append_rows: list[dict[str, Any]] = []
            changes: list[dict[str, Any]] = []

            for row_data in rows:
                if key_column not in row_data:
                    raise DataError(f"Row is missing key column '{key_column}'")

                key_value = row_data[key_column]
                if key_value is None:
                    raise DataError(f"Key column '{key_column}' cannot be null")
                if key_value in seen_update_keys:
                    raise DataError(f"Duplicate input key '{key_value}' provided")
                seen_update_keys.add(key_value)

                target_row = row_lookup.get(key_value)
                if target_row is None:
                    append_rows.append(row_data)
                    continue

                update_rows.append((target_row, row_data))
                for column_name, new_value in row_data.items():
                    if column_name == key_column:
                        continue
                    col_idx = header_map[column_name]
                    old_value = ws.cell(row=target_row, column=col_idx).value
                    if old_value == new_value:
                        continue
                    changes.append(
                        _build_cell_change(
                            sheet_name=current_sheet_name,
                            row=target_row,
                            col=col_idx,
                            old_value=old_value,
                            new_value=new_value,
                            column_name=column_name,
                        )
                    )

            append_start_row = data_end_row + 1
            if totals_row_count == 0 and append_rows:
                _ensure_table_append_space_clear(
                    ws,
                    start_row=append_start_row,
                    row_count=len(append_rows),
                    min_col=min_col,
                    max_col=max_col,
                )

            ordered_columns = sorted(header_map.items(), key=lambda item: item[1])
            for row_offset, row_data in enumerate(append_rows):
                target_row = append_start_row + row_offset
                for column_name, col_idx in ordered_columns:
                    if column_name not in row_data:
                        continue
                    changes.append(
                        _build_cell_change(
                            sheet_name=current_sheet_name,
                            row=target_row,
                            col=col_idx,
                            old_value=None,
                            new_value=row_data[column_name],
                            column_name=column_name,
                        )
                    )

            previous_table_range = table.ref
            new_max_row = max_row + len(append_rows)
            table_range = _range_from_bounds(min_col, min_row, max_col, new_max_row)

            if not dry_run:
                if append_rows and totals_row_count > 0:
                    ws.insert_rows(append_start_row, amount=len(append_rows))

                for target_row, row_data in update_rows:
                    for column_name, new_value in row_data.items():
                        if column_name == key_column:
                            continue
                        ws.cell(row=target_row, column=header_map[column_name], value=new_value)

                for row_offset, row_data in enumerate(append_rows):
                    target_row = append_start_row + row_offset
                    for column_name, col_idx in ordered_columns:
                        if column_name not in row_data:
                            continue
                        ws.cell(row=target_row, column=col_idx, value=row_data[column_name])

                if append_rows:
                    table.ref = table_range

        appended_keys = [row[key_column] for row in append_rows]
        updated_keys = [row_data[key_column] for _, row_data in update_rows]
        result = {
            "message": (
                f"{'Previewed' if dry_run else 'Upserted'} {len(rows)} row(s) in table "
                f"'{table_name}' on '{current_sheet_name}'"
            ),
            "sheet_name": current_sheet_name,
            "table_name": table_name,
            "key_column": key_column,
            "updated_rows": len(update_rows),
            "appended_rows": len(append_rows),
            "updated_keys": updated_keys,
            "appended_keys": appended_keys,
            "changed_cells": len(changes),
            "previous_table_range": previous_table_range,
            "table_range": table_range,
            "dry_run": dry_run,
        }
        if _should_include_changes(dry_run, include_changes):
            result["changes"] = changes
        return result

    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to upsert table rows: {e}")
        raise DataError(str(e))
